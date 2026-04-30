import io, os, re, copy, base64, json, time, requests
import pandas as pd
from openpyxl import load_workbook
import streamlit as st
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
import streamlit.components.v1 as components
from PIL import Image, ImageEnhance, ImageFilter, ExifTags
from datetime import datetime
import openpyxl
from openpyxl import Workbook
import random
import pyodbc
import socket

# ══════════════════════════════════════════════════════════════════════════════
# SQL SERVER CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════════
SQL_SERVER   = r'localhost\SQLEXPRESS'
SQL_DATABASE = 'sintex_quote_db'
SQL_DRIVER   = '{ODBC Driver 17 for SQL Server}'

def get_db_connection():
    """Return a live pyodbc connection using Windows Authentication."""
    conn_str = (
        f'DRIVER={SQL_DRIVER};'
        f'SERVER={SQL_SERVER};'
        f'DATABASE={SQL_DATABASE};'
        f'Trusted_Connection=yes;'
    )
    return pyodbc.connect(conn_str, autocommit=False)

def ensure_tables_exist():
    ddl_ocr = """
    IF NOT EXISTS (
        SELECT 1 FROM sys.tables
        WHERE object_id = OBJECT_ID(N'dbo.ocr_logs') AND type = 'U'
    )
    BEGIN
        CREATE TABLE dbo.ocr_logs (
            ocr_id              INT IDENTITY(1,1) PRIMARY KEY,
            image_path          NVARCHAR(500)   NULL,
            quantities_detected INT             NOT NULL DEFAULT 0,
            ocr_start_time      DATETIME2       NOT NULL DEFAULT GETDATE(),
            ocr_end_time        DATETIME2       NULL,
            status              TINYINT         NOT NULL DEFAULT 0,
            matched_count       INT             NOT NULL DEFAULT 0,
            unmatched_count     INT             NOT NULL DEFAULT 0,
            selected_state      NVARCHAR(100)   NULL
        );
    END
    """

    ddl_quotation = """
    IF NOT EXISTS (
        SELECT 1 FROM sys.tables
        WHERE object_id = OBJECT_ID(N'dbo.quotation_logs') AND type = 'U'
    )
    BEGIN
        CREATE TABLE dbo.quotation_logs (
            qid                 INT IDENTITY(1,1) PRIMARY KEY,
            quotation_id        NVARCHAR(50)    NOT NULL,
            ocr_id              INT             NULL
                REFERENCES dbo.ocr_logs(ocr_id) ON DELETE SET NULL,
            pdf_link            NVARCHAR(500)   NULL,
            detection_pdf_link  NVARCHAR(500)   NULL,
            customer_name       NVARCHAR(200)   NULL,
            customer_no         NVARCHAR(100)   NULL,
            customer_location   NVARCHAR(500)   NULL,
            customer_phone      NVARCHAR(20)    NULL,
            customer_mobile     NVARCHAR(20)    NULL,
            customer_gst        NVARCHAR(50)    NULL,
            customer_pan        NVARCHAR(20)    NULL,
            customer_state      NVARCHAR(100)   NULL,
            db_name             NVARCHAR(200)   NULL,
            db_no               NVARCHAR(100)   NULL,
            db_state            NVARCHAR(100)   NULL,
            db_phone            NVARCHAR(20)    NULL,
            db_mobile           NVARCHAR(20)    NULL,
            db_gst              NVARCHAR(50)    NULL,
            db_pan              NVARCHAR(20)    NULL,
            distributor_name    NVARCHAR(200)   NULL,
            margin_percent      FLOAT           NULL,
            gross_mrp           FLOAT           NULL,
            distributor_discount FLOAT          NULL,
            net_taxable         FLOAT           NULL,
            line_item_count     INT             NULL,
            created_date        DATETIME2       NOT NULL DEFAULT GETDATE(),
            ip_address          NVARCHAR(50)    NULL
        );
    END
    """

    ddl_session = """
    IF NOT EXISTS (
        SELECT 1 FROM sys.tables
        WHERE object_id = OBJECT_ID(N'dbo.session_logs') AND type = 'U'
    )
    BEGIN
        CREATE TABLE dbo.session_logs (
            sr_no               INT IDENTITY(1,1) PRIMARY KEY,
            session_id          NVARCHAR(100)   NULL,
            ocr_id              INT             NULL,
            quotation_id        NVARCHAR(50)    NULL,
            ip_address          NVARCHAR(50)    NULL,
            latitude            FLOAT           NULL,
            longitude           FLOAT           NULL,
            session_start_dt    DATETIME2       NOT NULL DEFAULT GETDATE(),
            session_end_dt      DATETIME2       NULL
        );
    END
    """

    try:
        with get_db_connection() as conn:
            cur = conn.cursor()
            cur.execute(ddl_ocr)
            cur.execute(ddl_quotation)
            cur.execute(ddl_session)
            conn.commit()
    except Exception as exc:
        st.warning(f"⚠️ DB table-init warning: {exc}")


def db_insert_ocr_log(
    image_path: str,
    selected_state: str,
    ocr_start_time: datetime,
) -> "int | None":
    sql = """
    INSERT INTO dbo.ocr_logs
        (image_path, quantities_detected, ocr_start_time, status,
         matched_count, unmatched_count, selected_state)
    OUTPUT INSERTED.ocr_id
    VALUES (?, 0, ?, 0, 0, 0, ?)
    """
    try:
        with get_db_connection() as conn:
            cur = conn.cursor()
            cur.execute(sql, image_path, ocr_start_time, selected_state)
            row = cur.fetchone()
            conn.commit()
            return int(row[0]) if row else None
    except Exception as exc:
        st.warning(f"⚠️ DB ocr_logs insert failed: {exc}")
        return None


def db_update_ocr_log(
    ocr_id: int,
    quantities_detected: int,
    matched_count: int,
    unmatched_count: int,
    status: int,
):
    sql = """
    UPDATE dbo.ocr_logs SET
        quantities_detected = ?,
        ocr_end_time        = GETDATE(),
        status              = ?,
        matched_count       = ?,
        unmatched_count     = ?
    WHERE ocr_id = ?
    """
    try:
        with get_db_connection() as conn:
            cur = conn.cursor()
            cur.execute(sql,
                        quantities_detected, status,
                        matched_count, unmatched_count,
                        ocr_id)
            conn.commit()
    except Exception as exc:
        st.warning(f"⚠️ DB ocr_logs update failed: {exc}")


def db_insert_quotation_log(
    quotation_id: str,
    ocr_id: "int | None",
    pdf_link: str,
    detection_pdf_link: str,
    bill_to: dict,
    ship_to: dict,
    distributor_name: str,
    margin_percent: float,
    gross_mrp: float,
    distributor_discount: float,
    net_taxable: float,
    line_item_count: int,
    ip_address: str,
):
    sql = """
    INSERT INTO dbo.quotation_logs (
        quotation_id, ocr_id,
        pdf_link, detection_pdf_link,
        customer_name, customer_no, customer_location,
        customer_phone, customer_mobile, customer_gst, customer_pan, customer_state,
        db_name, db_no, db_state, db_phone, db_mobile, db_gst, db_pan,
        distributor_name, margin_percent,
        gross_mrp, distributor_discount, net_taxable,
        line_item_count, ip_address, created_date
    ) VALUES (
        ?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,GETDATE()
    )
    """
    try:
        with get_db_connection() as conn:
            cur = conn.cursor()
            cur.execute(sql,
                quotation_id,
                ocr_id if ocr_id else None,
                pdf_link,
                detection_pdf_link,
                ship_to.get("name","").strip(),
                ship_to.get("party_no","").strip(),
                (ship_to.get("name","").strip() + ", " + ship_to.get("address","").strip()).strip(", "),
                ship_to.get("phone","").strip(),
                ship_to.get("mobile","").strip(),
                ship_to.get("gst","").strip(),
                ship_to.get("pan","").strip(),
                ship_to.get("state","").strip(),
                bill_to.get("name","").strip(),
                bill_to.get("party_no","").strip(),
                bill_to.get("state","").strip(),
                bill_to.get("phone","").strip(),
                bill_to.get("mobile","").strip(),
                bill_to.get("gst","").strip(),
                bill_to.get("pan","").strip(),
                distributor_name.strip(),
                float(margin_percent),
                float(gross_mrp),
                float(distributor_discount),
                float(net_taxable),
                int(line_item_count),
                ip_address,
            )
            conn.commit()
    except Exception as exc:
        st.warning(f"⚠️ DB quotation_logs insert failed: {exc}")


def _build_session_id(
    ocr_id: "int | None",
    quotation_id: str,
    latitude: "float | None",
    longitude: "float | None",
) -> str:
    def _first4(val: str) -> str:
        s = re.sub(r'[^0-9]', '', str(val))
        return (s + "0000")[:4]

    p1 = _first4(ocr_id if ocr_id else "0")
    p2 = _first4(quotation_id)
    p3 = _first4(str(latitude).replace(".", "").replace("-", "") if latitude is not None else "0")
    p4 = _first4(str(longitude).replace(".", "").replace("-", "") if longitude is not None else "0")
    return f"{p1}{p2}{p3}{p4}"


def db_insert_session_log(
    ocr_id: "int | None",
    quotation_id: str,
    ip_address: str,
    latitude: "float | None",
    longitude: "float | None",
    session_start_dt: datetime,
) -> "int | None":
    session_id = _build_session_id(ocr_id, quotation_id, latitude, longitude)
    sql = """
    INSERT INTO dbo.session_logs
        (session_id, ocr_id, quotation_id, ip_address,
         latitude, longitude, session_start_dt, session_end_dt)
    OUTPUT INSERTED.sr_no
    VALUES (?, ?, ?, ?, ?, ?, ?, NULL)
    """
    try:
        with get_db_connection() as conn:
            cur = conn.cursor()
            cur.execute(sql,
                session_id,
                ocr_id if ocr_id else None,
                quotation_id if quotation_id else None,
                ip_address,
                float(latitude)  if latitude  is not None else None,
                float(longitude) if longitude is not None else None,
                session_start_dt,
            )
            row = cur.fetchone()
            conn.commit()
            return int(row[0]) if row else None
    except Exception as exc:
        st.warning(f"⚠️ DB session_logs insert failed: {exc}")
        return None


def db_update_session_log_end(sr_no: int):
    """Stamp session_end_dt when user downloads PDF or shares via WhatsApp."""
    sql = """
    UPDATE dbo.session_logs
    SET session_end_dt = GETDATE()
    WHERE sr_no = ?
    """
    try:
        with get_db_connection() as conn:
            cur = conn.cursor()
            cur.execute(sql, sr_no)
            conn.commit()
    except Exception as exc:
        st.warning(f"⚠️ DB session_logs end-stamp failed: {exc}")


def db_update_session_log_quotation(sr_no: int, quotation_id: str,
                                     ocr_id: "int | None",
                                     latitude: "float | None",
                                     longitude: "float | None"):
    session_id = _build_session_id(ocr_id, quotation_id, latitude, longitude)
    sql = """
    UPDATE dbo.session_logs
    SET quotation_id = ?,
        ocr_id       = ?,
        latitude     = ?,
        longitude    = ?,
        session_id   = ?
    WHERE sr_no = ?
    """
    try:
        with get_db_connection() as conn:
            cur = conn.cursor()
            cur.execute(sql,
                quotation_id,
                ocr_id if ocr_id else None,
                float(latitude)  if latitude  is not None else None,
                float(longitude) if longitude is not None else None,
                session_id,
                sr_no,
            )
            conn.commit()
    except Exception as exc:
        st.warning(f"⚠️ DB session_logs quotation update failed: {exc}")


# ══════════════════════════════════════════════════════════════════════════════
# FOLDER SETUP
# ══════════════════════════════════════════════════════════════════════════════
_BASE_DIR       = os.path.dirname(os.path.abspath(__file__))
IMAGES_DIR      = os.path.join(_BASE_DIR, "images")
QUOTATIONS_DIR  = os.path.join(_BASE_DIR, "quotations")
DETECTIONS_DIR  = os.path.join(_BASE_DIR, "detections")

def _ensure_folders():
    for d in (IMAGES_DIR, QUOTATIONS_DIR, DETECTIONS_DIR):
        os.makedirs(d, exist_ok=True)

_ensure_folders()

LOG_XLSX = os.path.join(_BASE_DIR, "quotation_log.xlsx")

_LOGO_PATH = os.path.join(_BASE_DIR, "sintex-logo.jpg")
def _load_logo_b64() -> str:
    try:
        with open(_LOGO_PATH, "rb") as f:
            return base64.b64encode(f.read()).decode()
    except Exception:
        return ""
_LOGO_B64 = _load_logo_b64()

if _LOGO_B64:
    _LOGO_HTML = (
        f'<img src="data:image/jpeg;base64,{_LOGO_B64}" '
        f'style="height:52px;width:auto;object-fit:contain;display:block;border-radius:6px;"/>'
    )
else:
    _LOGO_HTML = '<span style="font-size:26px;">🔴</span>'

_PAGE_ICON = _LOGO_PATH if os.path.exists(_LOGO_PATH) else "🔴"

st.set_page_config(
    page_title="Sintex BAPL – Quotation Generator",
    page_icon=_PAGE_ICON, layout="centered", initial_sidebar_state="collapsed",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&family=JetBrains+Mono:wght@400;500&display=swap');
:root{--red:#C0211F;--dred:#8B1514;--lred:#FDEAEA;--lgray:#F7F7F7;--mgray:#E8E8E8;
  --text:#1A1A1A;--muted:#6B6B6B;--border:#DEDEDE;--green:#1E7E4A;
  --radius:12px;--shadow:0 2px 12px rgba(0,0,0,0.08);}
html,body,[class*="css"]{font-family:'Inter',-apple-system,sans-serif!important;
  color:var(--text);background:#F2F2F2;}
#MainMenu,footer,header{visibility:hidden;}
.block-container{padding:1rem 0.75rem 5rem!important;max-width:920px!important;margin:0 auto!important;}

.app-header{background:linear-gradient(135deg,var(--red),var(--dred));color:white;
  padding:18px 20px;border-radius:0;margin-bottom:24px;
  display:flex;align-items:center;gap:16px;box-shadow:0 4px 20px rgba(192,33,31,.35);}
.app-header-badge{background:rgba(255,255,255,.18);border-radius:10px;width:52px;height:52px;
  display:flex;align-items:center;justify-content:contain;font-size:26px;flex-shrink:0;overflow:hidden;}
.app-header-text h1{font-size:17px;font-weight:800;margin:0;}
.app-header-text p{font-size:11.5px;margin:3px 0 0;opacity:.75;}

.step-navbar{display:flex;background:white;border-radius:var(--radius);
  box-shadow:var(--shadow);margin-bottom:20px;overflow:hidden;
  border:1px solid var(--border);}
.step-nav-item{flex:1;display:flex;flex-direction:column;align-items:center;
  padding:12px 6px 10px;cursor:default;position:relative;
  border-right:1px solid var(--border);transition:background .2s;}
.step-nav-item:last-child{border-right:none;}
.step-nav-item.locked{opacity:.45;cursor:not-allowed;}
.step-nav-item.active{background:var(--lred);}
.step-nav-item.done{background:#ECFDF5;cursor:pointer;}
.step-nav-item.active::after{content:'';position:absolute;bottom:0;left:0;right:0;
  height:3px;background:var(--red);}
.step-nav-item.done::after{content:'';position:absolute;bottom:0;left:0;right:0;
  height:3px;background:var(--green);}
.step-nav-dot{width:26px;height:26px;border-radius:50%;display:flex;align-items:center;
  justify-content:center;font-size:11px;font-weight:700;margin-bottom:4px;
  background:var(--mgray);color:var(--muted);}
.step-nav-item.active .step-nav-dot{background:var(--red);color:white;}
.step-nav-item.done .step-nav-dot{background:var(--green);color:white;}
.step-nav-label{font-size:10px;font-weight:600;color:var(--muted);text-align:center;}
.step-nav-item.active .step-nav-label{color:var(--red);}
.step-nav-item.done .step-nav-label{color:var(--green);}

.step-card{background:white;border-radius:var(--radius);box-shadow:var(--shadow);
  margin-bottom:20px;overflow:hidden;border:1px solid var(--border);}
.step-card-header{background:linear-gradient(90deg,var(--red),#D94644);padding:14px 18px;
  display:flex;align-items:center;gap:12px;}
.step-number{background:rgba(255,255,255,.25);border-radius:8px;width:30px;height:30px;
  min-width:30px;display:flex;align-items:center;justify-content:center;
  font-size:13px;font-weight:700;color:white;}
.step-number.done{background:#1E7E4A;}
.step-title{font-size:14px;font-weight:700;color:white;margin:0;}
.step-subtitle{font-size:11px;color:rgba(255,255,255,.75);margin:2px 0 0;}
.step-body{padding:18px;}

.stButton>button{width:100%;background:linear-gradient(135deg,var(--red),var(--dred))!important;
  color:white!important;border:none!important;border-radius:10px!important;padding:13px 20px!important;
  font-family:'Inter',sans-serif!important;font-size:14.5px!important;font-weight:600!important;
  box-shadow:0 4px 14px rgba(192,33,31,.28)!important;transition:all .2s!important;}
.stButton>button:hover{transform:translateY(-1px)!important;}
.btn-secondary>.stButton>button{background:white!important;color:var(--red)!important;
  border:2px solid var(--red)!important;box-shadow:none!important;margin-top:0!important;}
[data-testid="column"]{display:flex;flex-direction:column;justify-content:center;}
.btn-green>.stButton>button{background:linear-gradient(135deg,var(--green),#155d38)!important;
  color:white!important;border:none!important;
  box-shadow:0 4px 14px rgba(30,126,74,.28)!important;}
.btn-blue>.stButton>button{background:linear-gradient(135deg,#1d4ed8,#1e40af)!important;
  color:white!important;border:none!important;
  box-shadow:0 4px 14px rgba(29,78,216,.28)!important;}

.rotate-group-wrapper {
  display: flex;
  flex-direction: column;
  align-items: center;
  width: 100%;
  margin-top: 18px;
  margin-bottom: 4px;
}
.rotate-inner {
  width: 60%;
  min-width: 260px;
  max-width: 400px;
}
.rotate-bar-label {
  background: linear-gradient(135deg, #1a1a1a, #2a2a2a);
  border-radius: 12px 12px 0 0;
  padding: 13px 0 12px;
  text-align: center;
  width: 100%;
  color: #e0e0e0;
  font-size: 12.5px;
  font-weight: 700;
  letter-spacing: 0.6px;
  margin-bottom: 0 !important;
  border-top: 1.5px solid #3a3a3a;
  border-left: 1.5px solid #3a3a3a;
  border-right: 1.5px solid #3a3a3a;
  text-transform: uppercase;
}
div[data-testid="stHorizontalBlock"]:has(.rotate-btn-col) {
  gap: 0 !important;
  border-left: 1.5px solid #3a3a3a;
  border-right: 1.5px solid #3a3a3a;
  border-bottom: 1.5px solid #3a3a3a;
  border-radius: 0 0 12px 12px;
  overflow: hidden;
  background: #111111;
  box-shadow: 0 6px 24px rgba(0,0,0,0.45), 0 1.5px 0 #C0211F inset;
}
.rotate-btn-col .stButton>button {
  background: #111111 !important;
  color: #ffffff !important;
  border: none !important;
  border-top: 1.5px solid #2e2e2e !important;
  border-radius: 0 !important;
  padding: 16px 8px !important;
  font-size: 14px !important;
  font-weight: 700 !important;
  box-shadow: none !important;
  width: 100% !important;
  letter-spacing: 0.3px;
  transition: background 0.15s, color 0.15s !important;
}
.rotate-btn-col .stButton>button:hover {
  background: #C0211F !important;
  color: #ffffff !important;
  transform: none !important;
}
.rotate-btn-col .stButton>button:active {
  background: #8B1514 !important;
}
.rotate-btn-col-left .stButton>button {
  border-radius: 0 !important;
  border-right: 1px solid #2e2e2e !important;
}
.rotate-btn-col-right .stButton>button {
  border-radius: 0 !important;
}
.submit-btn-col .stButton>button {
  background: linear-gradient(135deg, #1E7E4A, #155d38) !important;
  color: white !important;
  border: none !important;
  border-radius: 10px !important;
  padding: 16px 20px !important;
  font-size: 15px !important;
  font-weight: 700 !important;
  box-shadow: 0 4px 16px rgba(30,126,74,0.38) !important;
  width: 100% !important;
}
.submit-btn-col .stButton>button:hover {
  filter: brightness(1.08) !important;
  transform: translateY(-1px) !important;
}

/* ── Step-4 action buttons ── */
.s4-btn-red>.stButton>button {
  background: linear-gradient(135deg,#C0211F,#8B1514) !important;
  color: white !important;
  border: none !important;
  border-radius: 10px !important;
  padding: 15px 20px !important;
  font-size: 14.5px !important;
  font-weight: 700 !important;
  box-shadow: 0 4px 14px rgba(192,33,31,.32) !important;
  width: 100% !important;
}
.s4-btn-red>.stButton>button:hover {
  filter: brightness(1.08) !important;
  transform: translateY(-1px) !important;
}
.s4-btn-wa>.stButton>button {
  background: linear-gradient(135deg,#25D366,#128C7E) !important;
  color: white !important;
  border: none !important;
  border-radius: 10px !important;
  padding: 15px 20px !important;
  font-size: 14.5px !important;
  font-weight: 700 !important;
  box-shadow: 0 4px 14px rgba(37,211,102,.32) !important;
  width: 100% !important;
}
.s4-btn-wa>.stButton>button:hover {
  filter: brightness(1.08) !important;
  transform: translateY(-1px) !important;
}
.s4-btn-new>.stButton>button {
  background: linear-gradient(135deg,#1d4ed8,#1e40af) !important;
  color: white !important;
  border: none !important;
  border-radius: 10px !important;
  padding: 15px 20px !important;
  font-size: 14.5px !important;
  font-weight: 700 !important;
  box-shadow: 0 4px 14px rgba(29,78,216,.28) !important;
  width: 100% !important;
}
.s4-btn-new>.stButton>button:hover {
  filter: brightness(1.08) !important;
  transform: translateY(-1px) !important;
}

/* Hide st.download_button default styling — we restyle via s4-btn-red */
[data-testid="stDownloadButton"]>button {
  width: 100% !important;
  background: linear-gradient(135deg,#C0211F,#8B1514) !important;
  color: white !important;
  border: none !important;
  border-radius: 10px !important;
  padding: 15px 20px !important;
  font-size: 14.5px !important;
  font-weight: 700 !important;
  box-shadow: 0 4px 14px rgba(192,33,31,.32) !important;
  font-family: 'Inter', sans-serif !important;
  transition: all 0.2s !important;
}
[data-testid="stDownloadButton"]>button:hover {
  filter: brightness(1.08) !important;
  transform: translateY(-1px) !important;
}

.stTabs [data-baseweb="tab-list"]{gap:4px;background:var(--mgray);
  padding:4px;border-radius:10px;border:none!important;}
.stTabs [data-baseweb="tab"]{border-radius:8px!important;font-family:'Inter',sans-serif!important;
  font-weight:500!important;font-size:13px!important;color:var(--muted)!important;}
.stTabs [aria-selected="true"]{background:var(--red)!important;color:white!important;}

.info-box{background:#EEF6FF;border:1.5px solid #BFDBFE;border-radius:9px;
  padding:11px 15px;font-size:12.5px;color:#1E40AF;margin:12px 0;}
.warn-box{background:#FFFBEB;border:1.5px solid #FDE68A;border-radius:9px;
  padding:11px 15px;font-size:12.5px;color:#92400E;margin:12px 0;}
.success-box{background:#ECFDF5;border:1.5px solid #A7F3D0;border-radius:9px;
  padding:11px 15px;font-size:12.5px;color:#065F46;margin:12px 0;}
.error-box{background:#FEF2F2;border:1.5px solid #FECACA;border-radius:9px;
  padding:11px 15px;font-size:12.5px;color:#991B1B;margin:6px 0;}

.totals-box{background:white;border:1.5px solid var(--border);border-radius:var(--radius);
  padding:16px 18px;margin:16px 0;}
.total-row{display:flex;justify-content:space-between;padding:7px 0;font-size:13.5px;
  border-bottom:1px solid var(--mgray);}
.total-row:last-child{border:none;padding-top:10px;}
.total-row.grand .total-lbl{font-weight:700;color:var(--red);font-size:15px;}
.total-row.grand .total-val{font-weight:800;color:var(--red);font-size:15px;
  font-family:'JetBrains Mono',monospace;}
.total-lbl{color:var(--muted);font-weight:500;}
.total-val{font-family:'JetBrains Mono',monospace;font-weight:600;}
.total-val.neg{color:#c0392b;}

.ocr-wrap{overflow-x:auto;-webkit-overflow-scrolling:touch;margin:10px 0;}
.ocr-tbl{width:100%;border-collapse:collapse;font-family:'Inter',sans-serif;font-size:12px;}
.ocr-tbl th{background:var(--red);color:white;padding:8px 7px;text-align:center;
  font-size:10.5px;font-weight:600;white-space:nowrap;}
.ocr-tbl th.L{text-align:left;min-width:120px;}
.ocr-tbl td{padding:6px 7px;border-bottom:1px solid var(--mgray);vertical-align:middle;text-align:center;color:var(--text)!important;}
.ocr-tbl td.L{text-align:left;font-weight:500;font-size:11.5px;color:var(--text)!important;}
.ocr-tbl td.M{font-family:'JetBrains Mono',monospace;font-size:10px;color:#555!important;}
.ocr-tbl tr:nth-child(even) td{background:var(--lgray);}
.ocr-tbl .ok td{background:#ECFDF5!important;color:#065F46!important;}
.ocr-tbl .no td{background:#FFFBEB!important;color:#92400E!important;}
.raw-tbl td{padding:5px 8px;border:1px solid #eee;white-space:nowrap;color:#1A1A1A!important;}

.raw-wrap{overflow-x:auto;-webkit-overflow-scrolling:touch;margin:8px 0;
  border:1.5px solid var(--border);border-radius:8px;}
.raw-tbl{border-collapse:collapse;font-size:11px;font-family:'JetBrains Mono',monospace;min-width:500px;}
.raw-tbl th{background:#2D2D2D;color:white;padding:6px 8px;white-space:nowrap;text-align:center;}
.raw-tbl td{padding:5px 8px;border:1px solid #eee;white-space:nowrap;}
.raw-tbl tr:nth-child(even) td{background:#fafafa;}
.raw-tbl .sku{background:#FFF3E0!important;font-weight:700;color:#8B1514;}
.raw-tbl .qty{color:#1E7E4A;font-weight:600;}
.raw-tbl .hdr td{background:#F0F0F0;font-weight:700;color:#333;font-family:'Inter',sans-serif;}

.party-section{background:var(--lgray);border-radius:10px;padding:14px;
  border:1px solid var(--border);margin-bottom:14px;}
.party-title{font-size:12px;font-weight:700;color:var(--red);text-transform:uppercase;
  letter-spacing:.6px;margin-bottom:10px;}
.fsl{font-size:11.5px;font-weight:700;text-transform:uppercase;letter-spacing:.8px;
  color:var(--muted);margin:14px 0 8px;padding-bottom:4px;border-bottom:1px solid var(--mgray);}
.dl-btn{display:flex;align-items:center;justify-content:center;gap:8px;width:100%;
  background:linear-gradient(135deg,var(--green),#155d38);color:white!important;
  text-align:center;padding:15px 20px;border-radius:10px;font-weight:700;font-size:14.5px;
  text-decoration:none!important;box-shadow:0 4px 18px rgba(30,126,74,.3);
  margin:8px 0;box-sizing:border-box;}

.pdf-preview-wrap{border:2px solid var(--border);border-radius:var(--radius);
  overflow:hidden;margin:16px 0;background:#525659;}

.state-inner-card{background:var(--lgray);border-radius:10px;padding:14px 16px;
  border:1px solid var(--border);margin-bottom:16px;}
.state-inner-title{font-size:11.5px;font-weight:700;color:var(--red);text-transform:uppercase;
  letter-spacing:.6px;margin-bottom:10px;}

.stTextArea textarea {
    font-family: 'Inter', sans-serif !important;
    font-size: 13.5px !important;
    border-radius: 8px !important;
    resize: none !important;
}

/* Geolocation status badge */
.geo-badge {
    display: inline-flex;
    align-items: center;
    gap: 6px;
    padding: 5px 12px;
    border-radius: 20px;
    font-size: 11.5px;
    font-weight: 600;
    margin: 8px 0 4px;
}
.geo-badge.acquired {
    background: #ECFDF5;
    color: #065F46;
    border: 1.5px solid #A7F3D0;
}
.geo-badge.pending {
    background: #FFFBEB;
    color: #92400E;
    border: 1.5px solid #FDE68A;
}
.geo-badge.denied {
    background: #FEF2F2;
    color: #991B1B;
    border: 1.5px solid #FECACA;
}

/* Hide geo bridge input in all Streamlit versions */
[data-testid="stTextInput"]:has(input[aria-label="__geo_bridge_hidden__"]) {
    display: none !important;
    visibility: hidden !important;
    height: 0 !important;
    overflow: hidden !important;
    margin: 0 !important;
    padding: 0 !important;
}
div[data-geo-bridge-wrapper="true"] {
    display: none !important;
    visibility: hidden !important;
    height: 0 !important;
    overflow: hidden !important;
    position: absolute !important;
    pointer-events: none !important;
}

/* ── OCR edit bridge hidden input ─────────────────────────────────────────── */
[data-testid="stTextInput"]:has(input[aria-label="__ocr_edit_bridge__"]) {
    display: none !important;
    visibility: hidden !important;
    height: 0 !important;
    overflow: hidden !important;
    margin: 0 !important;
    padding: 0 !important;
}

/* Log table colour rows */
.log-row-green td { background: #ECFDF5 !important; color: #065F46 !important; }
.log-row-orange td { background: #FFF7ED !important; color: #92400E !important; }
.log-row-red td { background: #FEF2F2 !important; color: #991B1B !important; }
.log-row-green td.M { color: #065F46 !important; }
.log-row-orange td.M { color: #B45309 !important; }
.log-row-red td.M { color: #991B1B !important; }

/* ── OCR Edit Table ───────────────────────────────────────────────────────── */
.ocr-edit-wrap {
    overflow-x: auto;
    -webkit-overflow-scrolling: touch;
    margin: 10px 0 16px;
    border: 1.5px solid var(--border);
    border-radius: 10px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.05);
}
.ocr-edit-tbl {
    width: 100%;
    border-collapse: collapse;
    font-family: 'Inter', sans-serif;
    font-size: 12.5px;
    min-width: 700px;
}
.ocr-edit-tbl thead tr {
    background: #1A1A1A;
}
.ocr-edit-tbl thead th {
    color: white;
    padding: 10px 10px;
    text-align: left;
    font-size: 11px;
    font-weight: 700;
    letter-spacing: 0.5px;
    text-transform: uppercase;
    white-space: nowrap;
    border-right: 1px solid #2e2e2e;
}
.ocr-edit-tbl thead th:last-child { border-right: none; }
.ocr-edit-tbl thead th.C { text-align: center; }
.ocr-edit-tbl thead th.R { text-align: right; }
.ocr-edit-tbl tbody tr {
    border-bottom: 1px solid #EFEFEF;
    transition: background 0.1s;
}
.ocr-edit-tbl tbody tr:nth-child(even) { background: #FAFAFA; }
.ocr-edit-tbl tbody tr:hover { background: #FFF5F5; }
.ocr-edit-tbl tbody td {
    padding: 8px 10px;
    color: #1A1A1A;
    vertical-align: middle;
    border-right: 1px solid #EFEFEF;
}
.ocr-edit-tbl tbody td:last-child { border-right: none; }
.ocr-edit-tbl tbody td.C { text-align: center; }
.ocr-edit-tbl tbody td.R { text-align: right; font-family: 'JetBrains Mono', monospace; font-size: 12px; }
.ocr-edit-tbl tbody td.sku-cell {
    font-family: 'JetBrains Mono', monospace;
    font-size: 10.5px;
    color: #555;
}
.ocr-edit-tbl tbody td.price-cell {
    font-family: 'JetBrains Mono', monospace;
    font-size: 12px;
    font-weight: 600;
    color: #1A1A1A;
    text-align: right;
}
.ocr-edit-tbl tbody td.total-price-cell {
    font-family: 'JetBrains Mono', monospace;
    font-size: 12px;
    font-weight: 700;
    color: #C0211F;
    text-align: right;
}
/* Quantity input inside table */
.ocr-qty-input {
    width: 72px;
    padding: 5px 8px;
    border: 1.5px solid #D0D0D0;
    border-radius: 6px;
    font-family: 'Inter', sans-serif;
    font-size: 13px;
    font-weight: 700;
    color: #1A1A1A;
    text-align: center;
    background: white;
    transition: border-color 0.15s, box-shadow 0.15s;
    outline: none;
}
.ocr-qty-input:focus {
    border-color: #C0211F;
    box-shadow: 0 0 0 3px rgba(192,33,31,0.12);
}
.ocr-qty-input:disabled {
    background: #F2F2F2;
    color: #999;
    cursor: not-allowed;
    border-color: #E0E0E0;
}
/* Apply changes button */
.apply-changes-wrap {
    display: flex;
    align-items: center;
    justify-content: flex-end;
    gap: 12px;
    margin-top: 12px;
    padding: 12px 0 4px;
    border-top: 1.5px solid #EFEFEF;
}
.apply-hint {
    font-size: 11.5px;
    color: #92400E;
    background: #FFFBEB;
    border: 1.5px solid #FDE68A;
    border-radius: 8px;
    padding: 7px 14px;
    font-weight: 500;
}
/* Summary footer row in edit table */
.ocr-edit-tbl tfoot tr {
    background: #F0F0F0;
    border-top: 2px solid #D0D0D0;
}
.ocr-edit-tbl tfoot td {
    padding: 9px 10px;
    font-weight: 700;
    font-size: 12.5px;
    color: #1A1A1A;
    border-right: 1px solid #DEDEDE;
}
.ocr-edit-tbl tfoot td:last-child { border-right: none; }
.ocr-edit-tbl tfoot td.R {
    text-align: right;
    font-family: 'JetBrains Mono', monospace;
    color: #C0211F;
}

/* Session-end status badge */
.session-end-badge {
    display: inline-flex;
    align-items: center;
    gap: 6px;
    padding: 4px 11px;
    border-radius: 20px;
    font-size: 11px;
    font-weight: 600;
    margin-top: 4px;
}
.session-end-badge.stamped {
    background: #ECFDF5;
    color: #065F46;
    border: 1.5px solid #A7F3D0;
}
.session-end-badge.pending {
    background: #F9FAFB;
    color: #6B7280;
    border: 1.5px solid #E5E7EB;
}

@media(max-width:600px){
  .block-container{padding:.75rem 0 5rem!important;}
  .step-body{padding:14px 12px;}
  .step-nav-label{font-size:9px;}
  .rotate-inner{width:90%;min-width:0;}
}
</style>
""", unsafe_allow_html=True)

# ── Path constants ─────────────────────────────────────────────────────────────
XLSX_PATH      = os.path.join(_BASE_DIR, "Sample form for Product list.xlsx")
MRP_PATH       = os.path.join(_BASE_DIR, "MRP_State_chhattisghar.csv")
CUST_PATH      = os.path.join(_BASE_DIR, "ZSD_CUST.csv")
AZURE_ENDPOINT = os.environ.get("AZURE_OCR_ENDPOINT", "")
AZURE_KEY      = os.environ.get("AZURE_OCR_KEY", "")

ensure_tables_exist()


def generate_quotation_id() -> str:
    now = datetime.now()
    date_part = now.strftime("%d%m%Y")
    time_part = now.strftime("%H%M%S")
    ms_part   = f"{now.microsecond // 1000:03d}"
    rand_part = f"{random.randint(100, 999)}"
    return f"{date_part}{time_part}{ms_part}{rand_part}"


def _get_local_ip() -> str:
    try:
        return socket.gethostbyname(socket.gethostname())
    except Exception:
        return "127.0.0.1"


# ══════════════════════════════════════════════════════════════════════════════
# GEOLOCATION — Robust multi-method JS → Python bridge
# ══════════════════════════════════════════════════════════════════════════════

_GEO_BRIDGE_KEY   = "_geo_bridge"
_GEO_BRIDGE_LABEL = "__geo_bridge_hidden__"


def _parse_geo_bridge():
    if st.session_state.get("geo_status") in ("ok", "denied", "unavailable", "timeout"):
        try:
            if "_geo_status" in st.query_params:
                st.query_params.clear()
        except Exception:
            pass
        return

    already_processed = st.session_state.get("_geo_qp_processed", False)
    if not already_processed:
        try:
            qp     = st.query_params
            status = str(qp.get("_geo_status", "")).strip()
            if status:
                lat_raw = qp.get("_geo_lat")
                lng_raw = qp.get("_geo_lng")
                st.session_state["_geo_qp_processed"] = True
                try:
                    st.query_params.clear()
                except Exception:
                    pass

                if status == "ok" and lat_raw is not None and lng_raw is not None:
                    st.session_state["user_latitude"]  = float(lat_raw)
                    st.session_state["user_longitude"] = float(lng_raw)
                    st.session_state["geo_status"]     = "ok"
                    st.rerun()
                elif status in ("denied", "unavailable", "timeout"):
                    st.session_state["user_latitude"]  = None
                    st.session_state["user_longitude"] = None
                    st.session_state["geo_status"]     = status
                    st.rerun()
        except Exception:
            pass

    raw = st.session_state.get(_GEO_BRIDGE_KEY, "").strip()
    if not raw:
        return
    try:
        payload = json.loads(raw)
        status  = payload.get("status", "")
        lat     = payload.get("lat")
        lng     = payload.get("lng")
        if status == "ok" and lat is not None and lng is not None:
            st.session_state["user_latitude"]  = float(lat)
            st.session_state["user_longitude"] = float(lng)
            st.session_state["geo_status"]     = "ok"
            st.rerun()
        elif status in ("denied", "unavailable", "timeout"):
            st.session_state["user_latitude"]  = None
            st.session_state["user_longitude"] = None
            st.session_state["geo_status"]     = status
            st.rerun()
    except (json.JSONDecodeError, TypeError, ValueError):
        pass


# ══════════════════════════════════════════════════════════════════════════════
# SESSION-END STAMPING
# ══════════════════════════════════════════════════════════════════════════════
# The old approach used a JS-bridge with components.html(height=0) which is
# unreliable because zero-height iframes are suppressed by browsers.
#
# NEW APPROACH:
#   • st.download_button has an on_click callback → we call
#     _stamp_session_end_once() directly from Python on the same rerun.
#   • The WhatsApp "Open" button is a native st.button → same callback.
#   • No JS bridge, no hidden text_input, no iframe required.
# ══════════════════════════════════════════════════════════════════════════════

def _stamp_session_end_once():
    """
    Stamp session_end_dt in DB exactly once per session.
    Called directly from Python on the rerun triggered by the download button
    or the WhatsApp button click.
    """
    if st.session_state.get("session_end_stamped"):
        return  # already stamped — do nothing

    sr_no = st.session_state.get("session_log_sr_no")
    if sr_no:
        db_update_session_log_end(sr_no)

    st.session_state["session_end_stamped"] = True


def render_geolocation_component():
    """
    Dual-bridge geolocation:
      Primary  → sets st.query_params (?_geo_status=ok&_geo_lat=X&_geo_lng=Y)
      Fallback → writes to hidden text_input
    """
    st.markdown('<div data-geo-bridge-wrapper="true">', unsafe_allow_html=True)
    st.text_input(
        label=_GEO_BRIDGE_LABEL,
        key=_GEO_BRIDGE_KEY,
        label_visibility="hidden",
        value=st.session_state.get(_GEO_BRIDGE_KEY, ""),
    )
    st.markdown('</div>', unsafe_allow_html=True)

    if st.session_state.get("geo_status", "pending") != "pending":
        return

    geo_js = """<!DOCTYPE html>
<html>
<head>
<style>*{margin:0;padding:0;box-sizing:border-box;}body{background:transparent;overflow:hidden;}</style>
</head>
<body>
<script>
(function () {
    if (window._sintexGeoFired) return;
    window._sintexGeoFired  = true;
    window._sintexGeoResolved = false;

    function sendViaQueryParams(status, lat, lng) {
        if (window._sintexGeoResolved) return;
        window._sintexGeoResolved = true;
        try {
            var url = new URL(window.parent.location.href);
            url.searchParams.set('_geo_status', status);
            if (lat !== null && lat !== undefined) url.searchParams.set('_geo_lat', String(lat));
            if (lng !== null && lng !== undefined) url.searchParams.set('_geo_lng', String(lng));
            window.parent.history.replaceState({}, '', url.toString());
        } catch(e) { }
    }

    function sendViaTextInput(status, lat, lng) {
        var payload = JSON.stringify({ status: status, lat: lat, lng: lng });
        var writeAttempts = 0;
        function tryWrite() {
            writeAttempts++;
            var written = false;
            try {
                var doc = window.parent.document;
                var inputs = doc.querySelectorAll('input[type="text"]');
                var el = null;
                for (var i = 0; i < inputs.length; i++) {
                    var lbl = inputs[i].getAttribute('aria-label') || '';
                    if (lbl.indexOf('__geo_bridge_hidden__') !== -1) { el = inputs[i]; break; }
                }
                if (el) {
                    var setter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype,'value').set;
                    setter.call(el, payload);
                    el.dispatchEvent(new Event('input',  {bubbles:true}));
                    el.dispatchEvent(new Event('change', {bubbles:true}));
                    el.dispatchEvent(new Event('blur',   {bubbles:true}));
                    written = true;
                }
            } catch(e) { written = false; }
            if (!written && writeAttempts < 8) {
                setTimeout(tryWrite, Math.min(500 * Math.pow(2, writeAttempts - 1), 8000));
            }
        }
        tryWrite();
    }

    function sendOk(lat, lng) {
        sendViaQueryParams('ok', lat, lng);
        sendViaTextInput('ok', lat, lng);
    }
    function sendFail(reason) {
        sendViaQueryParams(reason || 'unavailable', null, null);
        sendViaTextInput(reason || 'unavailable', null, null);
    }

    function ipFallbackPrimary(onSuccess, onFail) {
        try {
            var xhr = new XMLHttpRequest();
            xhr.open('GET', 'https://ipapi.co/json/', true);
            xhr.timeout = 8000;
            xhr.onreadystatechange = function() {
                if (xhr.readyState !== 4) return;
                try {
                    var d = JSON.parse(xhr.responseText);
                    if (d && d.latitude && d.longitude) onSuccess(parseFloat(d.latitude), parseFloat(d.longitude));
                    else onFail();
                } catch(e) { onFail(); }
            };
            xhr.ontimeout = onFail; xhr.onerror = onFail; xhr.send();
        } catch(e) { onFail(); }
    }
    function ipFallbackSecondary(onSuccess, onFail) {
        try {
            var xhr = new XMLHttpRequest();
            xhr.open('GET', 'http://ip-api.com/json/', true);
            xhr.timeout = 8000;
            xhr.onreadystatechange = function() {
                if (xhr.readyState !== 4) return;
                try {
                    var d = JSON.parse(xhr.responseText);
                    if (d && d.status === 'success' && d.lat && d.lon) onSuccess(parseFloat(d.lat), parseFloat(d.lon));
                    else onFail();
                } catch(e) { onFail(); }
            };
            xhr.ontimeout = onFail; xhr.onerror = onFail; xhr.send();
        } catch(e) { onFail(); }
    }
    function tryIpFallback() {
        ipFallbackPrimary(
            function(lat,lng){ sendOk(lat,lng); },
            function(){
                ipFallbackSecondary(
                    function(lat,lng){ sendOk(lat,lng); },
                    function(){ sendFail('unavailable'); }
                );
            }
        );
    }

    if (!navigator.geolocation) { tryIpFallback(); return; }

    navigator.geolocation.getCurrentPosition(
        function(pos) { sendOk(pos.coords.latitude, pos.coords.longitude); },
        function(err) { tryIpFallback(); },
        { enableHighAccuracy: true, timeout: 12000, maximumAge: 0 }
    );

    setTimeout(function() {
        if (!window._sintexGeoResolved) { tryIpFallback(); }
    }, 20000);
})();
</script>
</body>
</html>"""
    components.html(geo_js, height=0, scrolling=False)

def _geo_status_badge() -> str:
    status = st.session_state.get("geo_status", "pending")
    lat    = st.session_state.get("user_latitude")
    lng    = st.session_state.get("user_longitude")

    if status == "ok" and lat is not None and lng is not None:
        return
    elif status == "denied":
        return (
            '<div class="geo-badge denied">'
            '⛔ Browser location denied — attempting IP-based location…'
            '</div>'
        )
    elif status in ("unavailable", "timeout"):
        return (
            '<div class="geo-badge denied">'
            '⚠️ Location unavailable — quotation will be logged without coordinates'
            '</div>'
        )
    else:
        return (
            '<div class="geo-badge pending">'
            '⏳ Acquiring location (GPS / IP fallback)…'
            '</div>'
        )


# ── File helpers ───────────────────────────────────────────────────────────────
def save_image_to_disk(raw_bytes: bytes, rotation: int = 0) -> str:
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    path = os.path.join(IMAGES_DIR, f"order_{ts}.jpg")
    try:
        img = Image.open(io.BytesIO(raw_bytes))
        img = fix_exif_orientation(img)
        if img.mode not in ("RGB", "L"):
            img = img.convert("RGB")
        if rotation != 0:
            img = img.rotate(-rotation, expand=True)
        img.save(path, format="JPEG", quality=92)
    except Exception:
        with open(path, "wb") as f:
            f.write(raw_bytes)
    return path


def save_quotation_pdf_to_disk(pdf_bytes: bytes, quotation_id: str) -> str:
    path = os.path.join(QUOTATIONS_DIR, f"{quotation_id}.pdf")
    with open(path, "wb") as f:
        f.write(pdf_bytes)
    return path


def save_detection_pdf_to_disk(pdf_bytes: bytes, quotation_id: str) -> str:
    path = os.path.join(DETECTIONS_DIR, f"detection_{quotation_id}.pdf")
    with open(path, "wb") as f:
        f.write(pdf_bytes)
    return path


def _ensure_log_xlsx():
    if not os.path.exists(LOG_XLSX):
        wb = Workbook()
        ws = wb.active
        ws.title = "Quotation Log"
        headers = [
            "quotation_id", "pdf_path", "customer_name", "customer_no",
            "customer_location", "customer_date",
            "db_name", "db_no", "db_state", "distributor_name", "margin_percent",
        ]
        ws.append(headers)
        wb.save(LOG_XLSX)


def append_log_entry(quotation_id: str, pdf_path: str, bill_to: dict, ship_to: dict,
                     distributor_name: str, margin_percent: float):
    _ensure_log_xlsx()
    wb = openpyxl.load_workbook(LOG_XLSX)
    ws = wb.active
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row = [
        quotation_id,
        pdf_path,
        ship_to.get("name", "").strip(),
        ship_to.get("party_no", "").strip(),
        ship_to.get("state", "").strip(),
        now_str,
        bill_to.get("name", "").strip(),
        bill_to.get("party_no", "").strip(),
        bill_to.get("state", "").strip(),
        distributor_name.strip(),
        margin_percent,
    ]
    ws.append(row)
    wb.save(LOG_XLSX)


# ── Aliases / constants ────────────────────────────────────────────────────────
SIZE_ALIASES: dict[str, str] = {}
for _mm in ["15","20","25","32","40","50","63","75","90","110"]:
    SIZE_ALIASES[_mm + "MM"]      = _mm + "MM"
    SIZE_ALIASES[_mm + " MM"]     = _mm + "MM"
    SIZE_ALIASES[_mm]             = _mm + "MM"
SIZE_ALIASES.update({
    '1/2"':"15MM",'1/2':"15MM",'½"':"15MM",
    '3/4"':"20MM",'3/4':"20MM",'¾"':"20MM",
    '1"'  :"25MM",
    '1.25"':"32MM",'1-1/4"':"32MM",'1¼"':"32MM",
    '1.5"':"40MM",'1-1/2"':"40MM",'1½"':"40MM",
    '2"'  :"50MM",
})

ALL_INDIA_STATES = [
    "", "Andhra Pradesh", "Arunachal Pradesh", "Assam", "Bihar", "Chhattisgarh",
    "Goa", "Gujarat", "Haryana", "Himachal Pradesh", "Jharkhand", "Karnataka",
    "Kerala", "Madhya Pradesh", "Maharashtra", "Manipur", "Meghalaya", "Mizoram",
    "Nagaland", "Odisha", "Punjab", "Rajasthan", "Sikkim", "Tamil Nadu",
    "Telangana", "Tripura", "Uttar Pradesh", "Uttarakhand", "West Bengal",
    "Andaman and Nicobar Islands", "Chandigarh", "Dadra and Nagar Haveli and Daman and Diu",
    "Delhi", "Jammu and Kashmir", "Ladakh", "Lakshadweep", "Puducherry",
]

# ── Image utilities ────────────────────────────────────────────────────────────
def fix_exif_orientation(img: Image.Image) -> Image.Image:
    try:
        exif_data = img._getexif()
        if exif_data is None:
            return img
        orientation_key = next(
            (k for k, v in ExifTags.TAGS.items() if v == "Orientation"), None
        )
        if orientation_key is None or orientation_key not in exif_data:
            return img
        orientation = exif_data[orientation_key]
        rotation_map = {3: 180, 6: 270, 8: 90}
        if orientation in rotation_map:
            img = img.rotate(rotation_map[orientation], expand=True)
    except Exception:
        pass
    return img

def enhance_image_for_ocr(raw_bytes: bytes, rotation_degrees: int = 0) -> bytes:
    img = Image.open(io.BytesIO(raw_bytes))
    img = fix_exif_orientation(img)
    if img.mode not in ("RGB", "L"):
        img = img.convert("RGB")
    max_dim = 3000
    w, h = img.size
    if max(w, h) > max_dim:
        ratio = max_dim / max(w, h)
        img = img.resize((int(w * ratio), int(h * ratio)), Image.LANCZOS)
    elif max(w, h) < 1200:
        ratio = 1200 / max(w, h)
        img = img.resize((int(w * ratio), int(h * ratio)), Image.LANCZOS)
    if rotation_degrees != 0:
        img = img.rotate(-rotation_degrees, expand=True)
    enhancer_contrast = ImageEnhance.Contrast(img)
    img = enhancer_contrast.enhance(1.35)
    enhancer_sharpness = ImageEnhance.Sharpness(img)
    img = enhancer_sharpness.enhance(2.2)
    enhancer_brightness = ImageEnhance.Brightness(img)
    img = enhancer_brightness.enhance(1.08)
    out = io.BytesIO()
    img.save(out, format="JPEG", quality=97, optimize=True, subsampling=0)
    return out.getvalue()

def get_image_dimensions(raw_bytes: bytes) -> tuple:
    try:
        img = Image.open(io.BytesIO(raw_bytes))
        return img.size
    except Exception:
        return (0, 0)

# ── Data loaders ───────────────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def load_sku_sheets() -> dict:
    wb = load_workbook(XLSX_PATH, read_only=True)
    result = {}
    for sn in wb.sheetnames:
        ws  = wb[sn]
        all_rows = list(ws.iter_rows(values_only=True))
        r1  = list(all_rows[1]) if len(all_rows) > 1 else []
        r2  = list(all_rows[2]) if len(all_rows) > 2 else []
        cis = [ci for ci,v in enumerate(r1) if ci>=2 and v is not None]
        if not cis:
            cis = [ci for ci,v in enumerate(r2) if ci>=2 and v is not None]
        od  = {ci: str(r1[ci]).strip() if ci<len(r1) and r1[ci] else "" for ci in cis}
        inch= {ci: str(r2[ci]).strip() if ci<len(r2) and r2[ci] else "" for ci in cis}
        rows, sec = [], "PIPES"
        for raw in all_rows[3:]:
            row   = list(raw)
            if not any(v for v in row if v is not None): continue
            label = str(row[0]).strip() if row[0] else ""
            if not label or label == "\xa0": continue
            up = label.upper()
            if up in {"FITTINGS","PIPES","FITTING SCH 80","FITTING SCH80"}:
                sec = up; rows.append({"sh":True,"label":label}); continue
            if (len(row)>1 and (row[1] is None or str(row[1]).strip()=="\xa0")
                    and not any(str(v).strip().startswith("CP") for v in row[2:] if v)):
                sec = label; rows.append({"sh":True,"label":label}); continue
            cells = {}
            for ci in cis:
                if ci<len(row) and row[ci] is not None:
                    v = str(row[ci]).strip()
                    if v and v not in("-","\xa0") and v.startswith("CP"):
                        cells[ci] = v
            rows.append({"sh":False,"label":label,"sec":sec,"cells":cells})
        result[sn] = {"od":od,"inch":inch,"cis":cis,"rows":rows}
    return result

@st.cache_data(show_spinner=False)
def build_sku_master(sku_sheets: dict) -> dict:
    master = {}
    for sn, sd in sku_sheets.items():
        sec = "PIPES"
        for rd in sd["rows"]:
            if rd.get("sh"):
                sec = rd["label"]; continue
            for ci, sku in rd["cells"].items():
                master[sku] = {
                    "sheet": sn, "product": rd["label"], "section": sec,
                    "od_size": sd["od"].get(ci,""), "inch_size": sd["inch"].get(ci,""),
                }
    return master

@st.cache_data(show_spinner=False)
def build_prefix_index(sku_master: dict) -> dict:
    bps, bp = {}, {}
    for sku, info in sku_master.items():
        sz = info.get("od_size","").upper().replace(" ","")
        for l in range(4, len(sku)+1):
            p = sku[:l]
            bp.setdefault(p, []).append(sku)
            if sz:
                bps[(p, sz)] = sku
    return {"bps": bps, "bp": bp}

@st.cache_data(show_spinner=False)
def load_mrp_data_for_state(state_name: str) -> dict:
    df = pd.read_csv(MRP_PATH, encoding="latin-1")
    df["_m"] = pd.to_numeric(
        df["MRP(ZPR1-933)"].astype(str).str.replace(",","").str.strip(), errors="coerce"
    ).fillna(0)
    df["_d"] = pd.to_numeric(
        df["Distributor Landing"].astype(str).str.replace(",","").str.strip(), errors="coerce"
    ).fillna(0)
    if "State Name" in df.columns and state_name:
        state_df = df[df["State Name"].str.strip().str.lower() == state_name.strip().lower()]
        if not state_df.empty:
            df = state_df
    r = {}
    for _, row in df.iterrows():
        mat = str(row["Material Number"]).strip()
        r[mat] = {
            "mrp": float(row["_m"]),
            "distributor_landing": float(row["_d"]),
            "description": str(row["Material Description"]).strip(),
        }
    return r

@st.cache_data(show_spinner=False)
def load_state_names_from_csv() -> list:
    try:
        df = pd.read_csv(MRP_PATH, encoding="latin-1")
        if "State Name" in df.columns:
            states = (
                df["State Name"]
                .dropna()
                .str.strip()
                .unique()
                .tolist()
            )
            states = [s for s in states if s]
            if states:
                return sorted(states)
    except Exception:
        pass
    return ALL_INDIA_STATES

@st.cache_data(show_spinner=False)
def load_customers() -> list:
    df  = pd.read_csv(CUST_PATH, encoding="latin-1")
    out = []
    for _, r in df.iterrows():
        name = str(r.get("Customer Name","")).strip()
        code = str(r.get("Customer Code","")).strip()
        if name and name != "nan":
            out.append({
                "code": code, "name": name,
                "address": " ".join(filter(None,[
                    str(r.get("Address 1","") or "").strip(),
                    str(r.get("Address 2","") or "").strip(),
                    str(r.get("City","") or "").strip(),
                    str(r.get("State Code Desc.","") or "").strip()])),
                "phone":  str(r.get("Telephone","")  or "").strip(),
                "mobile": str(r.get("Mobile No.","") or "").strip(),
                "state_code": str(r.get("State Code","") or "").strip(),
                "state": str(r.get("State Code Desc.","") or "").strip(),
                "gst": str(r.get("GST Number","") or "").strip(),
                "pan": str(r.get("PAN No.","") or "").strip(),
                "display": f"{code} – {name}",
            })
    return out

# ── OCR helpers ────────────────────────────────────────────────────────────────
def _poly_bbox(poly):
    if not poly:
        return 0, 0, 0, 0
    if isinstance(poly[0], dict):
        xs, ys = [p["x"] for p in poly], [p["y"] for p in poly]
    else:
        xs, ys = poly[0::2], poly[1::2]
    return min(xs), min(ys), max(xs)-min(xs), max(ys)-min(ys)

def _words_v3v4(data: dict) -> list:
    out = []
    ar  = data.get("analyzeResult", data)
    for page in ar.get("pages", []):
        pw = page.get("width", 1) or 1
        ph = page.get("height", 1) or 1
        for w in page.get("words", []):
            x, y, ww, hh = _poly_bbox(w.get("polygon", w.get("boundingBox",[])))
            out.append({"text": w.get("content", w.get("text","")),
                        "x":x,"y":y,"w":ww,"h":hh,"cx":x+ww/2,"cy":y+hh/2,
                        "pw":pw,"ph":ph})
    return out

def _words_v2(data: dict) -> list:
    out = []
    for page in data.get("analyzeResult",{}).get("readResults",[]):
        pw = page.get("width",1) or 1
        ph = page.get("height",1) or 1
        for line in page.get("lines",[]):
            for w in line.get("words",[]):
                x,y,ww,hh = _poly_bbox(w.get("boundingBox",[]))
                out.append({"text":w.get("text",""),
                            "x":x,"y":y,"w":ww,"h":hh,"cx":x+ww/2,"cy":y+hh/2,
                            "pw":pw,"ph":ph})
    return out

def run_azure_ocr(img: bytes, endpoint: str, key: str) -> list:
    ep  = endpoint.rstrip("/")
    hdr = {"Ocp-Apim-Subscription-Key": key, "Content-Type": "application/octet-stream"}

    def post(url):
        return requests.post(url, headers=hdr, data=img, timeout=60)

    def poll_async(resp) -> dict:
        if resp.status_code != 202:
            resp.raise_for_status()
            return resp.json()
        op = resp.headers.get("Operation-Location") or resp.headers.get("operation-location","")
        if not op:
            raise ValueError("No Operation-Location header")
        for _ in range(30):
            time.sleep(2)
            p = requests.get(op, headers={"Ocp-Apim-Subscription-Key": key}, timeout=30)
            p.raise_for_status()
            d = p.json()
            s = d.get("status","")
            if s == "succeeded": return d
            if s == "failed":    raise ValueError("Azure failed: " + json.dumps(d.get("error",{})))
        raise TimeoutError("Azure OCR timed out")

    r = post(f"{ep}/documentintelligence/documentModels/prebuilt-read:analyze?api-version=2024-02-29-preview")
    if r.status_code not in (200, 202):
        r = post(f"{ep}/formrecognizer/documentModels/prebuilt-read:analyze?api-version=2023-07-31")
    if r.status_code not in (200, 202):
        r = post(f"{ep}/formrecognizer/v2.1/layout/analyze")
        r.raise_for_status()
        op = r.headers.get("Operation-Location","")
        if not op: raise ValueError("No Operation-Location in v2 response")
        for _ in range(20):
            time.sleep(2)
            p = requests.get(op, headers={"Ocp-Apim-Subscription-Key": key}, timeout=30)
            p.raise_for_status()
            d = p.json()
            if d.get("status") == "succeeded": return _words_v2(d)
            if d.get("status") == "failed":    raise ValueError("Azure v2 failed")
        raise TimeoutError("Azure v2 timed out")

    return _words_v3v4(poll_async(r))

def reconstruct_table(words: list) -> list:
    if not words:
        return []

    heights = sorted(w["h"] for w in words if w["h"] > 0)
    med_h   = heights[len(heights)//2] if heights else 8
    row_tol = max(med_h * 0.55, 2)
    pw      = words[0]["pw"] or 1
    merge_x = pw * 0.016
    col_tol = pw * 0.025

    by_y = sorted(words, key=lambda w: w["cy"])
    raw_rows = []
    cur = []
    cy_avg = None
    for w in by_y:
        if cy_avg is None or abs(w["cy"] - cy_avg) <= row_tol:
            cur.append(w)
            cy_avg = (cy_avg + w["cy"]) / 2 if cy_avg else w["cy"]
        else:
            raw_rows.append(sorted(cur, key=lambda ww: ww["cx"]))
            cur    = [w]
            cy_avg = w["cy"]
    if cur:
        raw_rows.append(sorted(cur, key=lambda ww: ww["cx"]))

    cell_rows = []
    for row in raw_rows:
        cells = []
        buf, buf_cx, buf_rx = row[0]["text"], row[0]["cx"], row[0]["x"]+row[0]["w"]
        for w in row[1:]:
            if w["x"] - buf_rx <= merge_x:
                buf    += " " + w["text"]
                buf_rx  = w["x"] + w["w"]
                buf_cx  = (buf_cx + w["cx"]) / 2
            else:
                cells.append({"text": buf.strip(), "cx": buf_cx})
                buf, buf_cx, buf_rx = w["text"], w["cx"], w["x"]+w["w"]
        cells.append({"text": buf.strip(), "cx": buf_cx})
        cell_rows.append(cells)

    all_cx  = sorted(c["cx"] for row in cell_rows for c in row)
    col_cxs = []
    for cx in all_cx:
        placed = False
        for i, cc in enumerate(col_cxs):
            if abs(cx - cc) <= col_tol:
                col_cxs[i] = (cc + cx) / 2
                placed = True; break
        if not placed:
            col_cxs.append(cx)
    col_cxs.sort()
    nc = len(col_cxs)

    def nearest(cx):
        return min(range(nc), key=lambda i: abs(col_cxs[i] - cx))

    grid = []
    for row in cell_rows:
        arr = [""] * nc
        for cell in row:
            ci = nearest(cell["cx"])
            arr[ci] = (arr[ci] + " " + cell["text"]).strip() if arr[ci] else cell["text"]
        grid.append(arr)

    return grid

def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", s.strip().upper())

def _size_of(cell: str):
    n = _norm(cell).replace(" ","").rstrip(".")
    v = SIZE_ALIASES.get(n)
    if not v:
        v = SIZE_ALIASES.get(re.sub(r'[^0-9]','',n) + "MM") if re.sub(r'[^0-9]','',n) else None
    return v

def _is_sku_prefix(cell: str) -> bool:
    return bool(re.match(r'^CP[A-Z0-9]{3,}', cell.strip().upper()))

def _clean_ocr_num(cell: str):
    c = cell.strip().replace(",","")
    c = c.replace("O","0").replace("o","0")
    c = c.replace("l","1").replace("I","1").replace("|","1")
    c = c.replace("S","5").replace("s","5").replace("G","6")
    m = re.search(r'\d+', c)
    if m:
        v = int(m.group())
        return v if 1 <= v <= 9999 else None
    return None

def _is_qty(cell: str) -> bool:
    return _clean_ocr_num(cell) is not None


def analyze_table(grid: list) -> dict:
    if not grid:
        return {"rows":[], "meta":{}, "grid":grid}

    nc = max(len(r) for r in grid)

    best_hdr, best_szc = 0, {}
    for ri, row in enumerate(grid[:15]):
        szc = {ci: _size_of(cell) for ci,cell in enumerate(row) if _size_of(cell)}
        if len(szc) > len(best_szc):
            best_szc, best_hdr = szc, ri

    if not best_szc:
        for ri, row in enumerate(grid[:15]):
            szc = {}
            for ci, cell in enumerate(row):
                n = _norm(cell).replace(" ","")
                if re.fullmatch(r'\d{2,3}', n) and (n+"MM") in SIZE_ALIASES:
                    szc[ci] = n+"MM"
            if len(szc) > len(best_szc):
                best_szc, best_hdr = szc, ri

    size_cols  = best_szc
    header_row = best_hdr

    sku_votes = {}
    for row in grid[header_row+1:]:
        for ci, cell in enumerate(row):
            if _is_sku_prefix(cell):
                sku_votes[ci] = sku_votes.get(ci,0) + 1
    sku_col = max(sku_votes, key=sku_votes.get) if sku_votes else None

    skip = set(size_cols) | ({sku_col} if sku_col is not None else set())
    name_col = next((c for c in range(nc) if c not in skip), None)

    rows = []
    for ri in range(header_row+1, len(grid)):
        row       = grid[ri]
        sku_pfx   = row[sku_col].strip().upper() if sku_col is not None and sku_col < len(row) else ""
        product   = row[name_col].strip()        if name_col is not None and name_col < len(row) else ""

        if not sku_pfx and not product:
            continue
        if _norm(product) in {"SKU NAME","PRODUCT","DESCRIPTION","ITEM","S NO","SR NO",
                               "SKU CODE","SIZE","CATEGORY","SECTION"}:
            continue
        if not sku_pfx and not any(_is_qty(row[ci]) for ci in size_cols if ci < len(row)):
            continue

        sizes = {}
        for ci, sz_label in size_cols.items():
            if ci < len(row):
                q = _clean_ocr_num(row[ci])
                if q:
                    sizes[sz_label] = q

        if not sizes:
            continue

        rows.append({"product": product, "sku_prefix": sku_pfx, "sizes": sizes, "ri": ri})

    return {
        "rows": rows,
        "meta": {"header_row": header_row, "sku_col": sku_col,
                 "name_col": name_col, "size_cols": size_cols},
        "grid": grid,
    }

_OCR_FIXES = str.maketrans({
    "O":"0","o":"0","Q":"0",
    "I":"1","l":"1","|":"1",
    "Z":"2","z":"2",
    "S":"5","s":"5",
    "G":"6","b":"6",
    "B":"8",
})

def _fix_sku(raw: str) -> str:
    s = raw.strip().upper()
    return s[:2] + s[2:].translate(_OCR_FIXES) if len(s) > 2 else s

def match_sku(prefix: str, size: str, pidx: dict, master: dict):
    if not prefix:
        return None

    sz   = size.upper().replace(" ","")
    pfx  = _fix_sku(prefix)
    bps  = pidx["bps"]
    bp   = pidx["bp"]

    if pfx in master and master[pfx].get("od_size","").upper().replace(" ","") == sz:
        return pfx

    for l in range(min(len(pfx),20), 3, -1):
        key = (pfx[:l], sz)
        if key in bps:
            return bps[key]

    for l in range(min(len(pfx),20), 3, -1):
        for sku in bp.get(pfx[:l], []):
            if master[sku].get("od_size","").upper().replace(" ","") == sz:
                return sku

    best_sku, best_score = None, 0
    for sku, info in master.items():
        if info.get("od_size","").upper().replace(" ","") != sz:
            continue
        min_l = min(len(pfx), len(sku))
        if min_l < 4:
            continue
        score = sum(a==b for a,b in zip(pfx[:min_l], sku[:min_l]))
        if score >= max(4, min_l - 2) and score > best_score:
            best_score, best_sku = score, sku

    return best_sku

def build_quantities(extracted_rows: list, pidx: dict, master: dict):
    line_items = []
    quantities = {}
    log        = []
    for row in extracted_rows:
        for sz, qty in row["sizes"].items():
            full = match_sku(row["sku_prefix"], sz, pidx, master)
            if full:
                quantities[full] = quantities.get(full, 0) + qty
                line_items.append({"sku": full, "qty": qty})
            log.append({
                "product":  row["product"],
                "prefix":   row["sku_prefix"],
                "size":     sz,
                "qty":      qty,
                "full_sku": full or "—",
                "status":   "matched" if full else "unmatched",
            })
    return quantities, log, line_items


# ══════════════════════════════════════════════════════════════════════════════
# HELPER — Build a live OCR detection log from current quantities + match_log
# ══════════════════════════════════════════════════════════════════════════════
def build_live_log(
    match_log: list,
    quantities: dict,
    line_items: list,
    sku_master: dict,
    mrp_data: dict,
) -> list:
    live: list = []
    ocr_sku_set: set = set()

    for entry in match_log:
        sku    = entry["full_sku"]
        status = entry["status"]

        if status == "unmatched":
            live.append(dict(entry))
            continue

        current_qty = quantities.get(sku, 0)
        if current_qty <= 0:
            continue

        updated = dict(entry)
        updated["qty"] = current_qty
        live.append(updated)
        ocr_sku_set.add(sku)

    seen_manual: set = set()
    for item in line_items:
        sku = item["sku"]
        if sku in ocr_sku_set:
            continue
        if sku in seen_manual:
            continue
        seen_manual.add(sku)

        current_qty = quantities.get(sku, item["qty"])
        if current_qty <= 0:
            continue

        info = sku_master.get(sku, {})
        live.append({
            "product":  info.get("product", sku),
            "prefix":   sku[:6] if len(sku) >= 6 else sku,
            "size":     info.get("od_size", ""),
            "qty":      current_qty,
            "full_sku": sku,
            "status":   "matched",
        })

    return live


# ── PDF builder ────────────────────────────────────────────────────────────────
def build_pdf(quantities: dict, mrp_data: dict, bill_to: dict,
              ship_to: dict, sku_master: dict, quotation_id: str,
              line_items: list = None) -> bytes:
    buf = io.BytesIO()
    PAGE = landscape(A4)
    LEFT_M = RIGHT_M = 12 * mm
    TOP_M = BOT_M = 12 * mm
    USABLE_W = PAGE[0] - LEFT_M - RIGHT_M

    doc = SimpleDocTemplate(buf, pagesize=PAGE,
                            leftMargin=LEFT_M, rightMargin=RIGHT_M,
                            topMargin=TOP_M, bottomMargin=BOT_M)

    BLACK   = colors.HexColor("#000000")
    DKGRAY  = colors.HexColor("#1A1A1A")
    MDGRAY  = colors.HexColor("#4A4A4A")
    GRAY    = colors.HexColor("#6B6B6B")
    LGRAY   = colors.HexColor("#F0F0F0")
    MLGRAY  = colors.HexColor("#D0D0D0")
    XLGRAY  = colors.HexColor("#E8E8E8")
    WHITE   = colors.white
    HDRFILL = colors.HexColor("#1A1A1A")
    ALTFILL = colors.HexColor("#F5F5F5")
    ACCLINE = colors.HexColor("#555555")

    sty = getSampleStyleSheet()
    def ps(n, **kw): return ParagraphStyle(n, parent=sty["Normal"], **kw)

    story = []

    hdr_left = ps("hl", fontName="Times-Bold", fontSize=9, textColor=WHITE, leading=14)
    hdr_ctr  = ps("hc", fontName="Times-Bold", fontSize=20, textColor=WHITE, alignment=TA_CENTER)
    hdr_ctr_sub = ps("hcs", fontName="Times-Roman", fontSize=8, textColor=WHITE,
                     alignment=TA_CENTER, leading=13)
    hdr_rt   = ps("hr", fontName="Times-Roman", fontSize=8, textColor=WHITE,
                  alignment=TA_RIGHT, leading=13)

    selected_state = st.session_state.get("selected_state", "")

    center_content = Paragraph(
        "SALES QUOTATION<br/>"
        f'<font size="8" color="#CCCCCC">Quotation ID: {quotation_id}</font>',
        hdr_ctr,
    )

    hdr = Table([[
        Paragraph("Sintex BAPL Limited<br/>"
                  "<font size='7.5'>Kutesar Road, Raipur, CG 492101</font><br/>"
                  "<font size='7.5'>GSTIN: 22AADCB1921F1ZE</font>", hdr_left),
        center_content,
        Paragraph(f"State: <b>{selected_state}</b><br/>"
                  "<font size='7'>CPVC / UPVC Pipes &amp; Fittings</font>", hdr_rt),
    ]], colWidths=[USABLE_W*0.30, USABLE_W*0.40, USABLE_W*0.30])
    hdr.setStyle(TableStyle([
        ("BACKGROUND",   (0,0), (-1,-1), HDRFILL),
        ("TEXTCOLOR",    (0,0), (-1,-1), WHITE),
        ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",   (0,0), (-1,-1), 12),
        ("BOTTOMPADDING",(0,0), (-1,-1), 12),
        ("LEFTPADDING",  (0,0), (-1,-1), 14),
        ("RIGHTPADDING", (0,0), (-1,-1), 14),
        ("LINEBELOW",    (0,0), (-1,-1), 2, ACCLINE),
    ]))
    story += [hdr, Spacer(1, 4*mm)]

    plbl = ps("plbl", fontName="Times-Bold", fontSize=7.5, textColor=WHITE)
    pval = ps("pval", fontName="Times-Roman", fontSize=7.5, textColor=BLACK, leading=11)

    def party_cell(d):
        ph_raw  = d.get('phone','').strip()
        mb_raw  = d.get('mobile','').strip()
        ph_disp = f"+91 {ph_raw}" if ph_raw else ""
        mb_disp = f"+91 {mb_raw}" if mb_raw else ""
        name_val    = d.get('name', '').strip()
        address_val = d.get('address', '').strip()
        if name_val and address_val:
            name_address = f"{name_val}, {address_val}"
        elif name_val:
            name_address = name_val
        else:
            name_address = address_val
        lines = [
            f"<b>Party No.:</b> {d.get('party_no','')}",
            f"<b>Name &amp; Address:</b> {name_address}",
            f"<b>Phone:</b> {ph_disp}   <b>Mobile:</b> {mb_disp}",
            f"<b>State:</b> {d.get('state_code','')} {d.get('state','')}",
            f"<b>GST:</b> {d.get('gst','')}   <b>PAN:</b> {d.get('pan','')}",
        ]
        return Paragraph("<br/>".join(lines), pval)

    HALF = USABLE_W / 2
    pt = Table([
        [Paragraph("BILL TO PARTY", plbl), Paragraph("SHIP TO PARTY", plbl)],
        [party_cell(bill_to),              party_cell(ship_to)],
    ], colWidths=[HALF, HALF])
    pt.setStyle(TableStyle([
        ("BACKGROUND",   (0,0), (-1,0), MDGRAY),
        ("TEXTCOLOR",    (0,0), (-1,0), WHITE),
        ("BACKGROUND",   (0,1), (-1,1), XLGRAY),
        ("BOX",          (0,0), (-1,-1), 0.8, MLGRAY),
        ("INNERGRID",    (0,0), (-1,-1), 0.5, MLGRAY),
        ("VALIGN",       (0,0), (-1,-1), "TOP"),
        ("TOPPADDING",   (0,0), (-1,-1), 7),
        ("BOTTOMPADDING",(0,0), (-1,-1), 7),
        ("LEFTPADDING",  (0,0), (-1,-1), 10),
        ("RIGHTPADDING", (0,0), (-1,-1), 10),
    ]))
    story += [pt, Spacer(1, 5*mm)]

    sch  = ps("sch",  fontName="Times-Bold",   fontSize=7,   textColor=WHITE, alignment=TA_CENTER)
    cell = ps("c",    fontName="Times-Roman",  fontSize=7.5, textColor=DKGRAY)
    celc = ps("cc",   fontName="Times-Roman",  fontSize=7.5, textColor=DKGRAY, alignment=TA_CENTER)
    celr = ps("cr",   fontName="Times-Roman",  fontSize=7.5, textColor=DKGRAY, alignment=TA_RIGHT)
    celb = ps("cb",   fontName="Times-Bold",   fontSize=7.5, textColor=BLACK,  alignment=TA_RIGHT)
    skuc = ps("sk",   fontName="Courier",      fontSize=6.5, textColor=MDGRAY)
    qtys_ps = ps("qp",fontName="Times-Bold",   fontSize=8,   textColor=BLACK,  alignment=TA_CENTER)

    heads = ["S.No", "Product", "SKU Code", "OD", "Inch", "MRP (₹)", "Qty", "Total MRP (₹)", "Dist. Landing (₹)", "Taxable (₹)"]
    cw_mm = [8, 68, 48, 14, 12, 22, 10, 26, 26, 26]
    cw = [x * mm for x in cw_mm]
    scale = USABLE_W / sum(cw)
    cw = [x * scale for x in cw]

    trows = [[Paragraph(h, sch) for h in heads]]
    ln = gm = gd = gt = 0

    if line_items:
        items_iter = line_items
    else:
        items_iter = [{"sku": s, "qty": q} for s, q in quantities.items() if q > 0]

    for item in items_iter:
        sku = item["sku"]; qty = item["qty"]
        if qty <= 0:
            continue
        info  = sku_master.get(sku, {})
        mi    = mrp_data.get(sku, {})
        mrp   = mi.get("mrp", 0.)
        dist  = mi.get("distributor_landing", 0.)
        tot   = round(mrp * qty, 2)
        tax   = round(dist * qty, 2)
        ln += 1; gm += tot; gd += (mrp - dist) * qty; gt += tax
        trows.append([
            Paragraph(str(ln),                    celc),
            Paragraph(info.get("product",""),     cell),
            Paragraph(sku,                        skuc),
            Paragraph(info.get("od_size",""),     celc),
            Paragraph(info.get("inch_size",""),   celc),
            Paragraph(f"{mrp:,.2f}",              celr),
            Paragraph(str(qty),                   qtys_ps),
            Paragraph(f"{tot:,.2f}",              celb),
            Paragraph(f"{dist:,.2f}",             celr),
            Paragraph(f"{tax:,.2f}",              celb),
        ])
    if ln == 0:
        trows.append([Paragraph("No items", cell)] + [""] * 9)

    t2 = Table(trows, colWidths=cw, repeatRows=1)
    ts = [
        ("BACKGROUND",    (0,0),  (-1,0),  HDRFILL),
        ("TEXTCOLOR",     (0,0),  (-1,0),  WHITE),
        ("BOX",           (0,0),  (-1,-1), 0.8,  DKGRAY),
        ("INNERGRID",     (0,1),  (-1,-1), 0.3,  MLGRAY),
        ("LINEBELOW",     (0,0),  (-1,0),  1.5,  ACCLINE),
        ("VALIGN",        (0,0),  (-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0),  (-1,-1), 4),
        ("BOTTOMPADDING", (0,0),  (-1,-1), 4),
        ("LEFTPADDING",   (0,0),  (-1,-1), 4),
        ("RIGHTPADDING",  (0,0),  (-1,-1), 4),
    ]
    for ri in range(1, len(trows)):
        if ri % 2 == 0:
            ts.append(("BACKGROUND", (0, ri), (-1, ri), ALTFILL))
    t2.setStyle(TableStyle(ts))
    story += [t2, Spacer(1, 0)]

    lbl_style = ps("tlbl", fontName="Times-Roman", fontSize=8,  textColor=GRAY,   alignment=TA_RIGHT)
    lbl_bold  = ps("tbb",  fontName="Times-Bold",  fontSize=8,  textColor=BLACK,  alignment=TA_RIGHT)
    val_style = ps("tval", fontName="Times-Bold",  fontSize=8,  textColor=BLACK,  alignment=TA_RIGHT)
    val_emph  = ps("tvr",  fontName="Times-Bold",  fontSize=9,  textColor=BLACK,  alignment=TA_RIGHT)
    val_neg   = ps("tvn",  fontName="Times-Roman", fontSize=8,  textColor=MDGRAY, alignment=TA_RIGHT)

    LABEL_W = 55 * mm
    VALUE_W = 38 * mm
    FILL_W  = USABLE_W - LABEL_W - VALUE_W

    tot_rows = [
        ["", Paragraph("Gross Total (MRP):", lbl_style),
              Paragraph(f"INR {gm:,.2f}/-", val_style)],
        ["", Paragraph("Less Distributor Discount:", lbl_style),
              Paragraph(f"– INR {gd:,.2f}/-", val_neg)],
        ["", Paragraph("Net Taxable Value:", lbl_bold),
              Paragraph(f"INR {gt:,.2f}/-", val_emph)],
    ]
    tot_tbl = Table(tot_rows, colWidths=[FILL_W, LABEL_W, VALUE_W])
    tot_tbl.setStyle(TableStyle([
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0), (-1,-1), 5),
        ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ("LEFTPADDING",   (0,0), (-1,-1), 6),
        ("RIGHTPADDING",  (0,0), (-1,-1), 6),
        ("BOX",           (1,0), (2,-1),  0.8, MLGRAY),
        ("INNERGRID",     (1,0), (2,-1),  0.4, XLGRAY),
        ("BACKGROUND",    (1,2), (2,2),   XLGRAY),
        ("LINEABOVE",     (1,2), (2,2),   1.5, DKGRAY),
        ("LINEBELOW",     (1,2), (2,2),   1.5, DKGRAY),
        ("ALIGN",         (1,0), (2,-1),  "RIGHT"),
    ]))
    story += [tot_tbl, Spacer(1, 5*mm)]

    ft = Table([[
        Paragraph("<i>Computer-generated quotation. Subject to change without notice.</i>",
                  ps("ft", fontName="Times-Italic", fontSize=7, textColor=GRAY)),
        Paragraph("<b>Authorised Signatory</b>",
                  ps("sg", fontName="Times-Bold", fontSize=8, textColor=BLACK,
                     alignment=TA_RIGHT)),
    ]], colWidths=[USABLE_W * 0.65, USABLE_W * 0.35])
    ft.setStyle(TableStyle([
        ("LINEABOVE",     (0,0), (-1,0), 0.8, MLGRAY),
        ("VALIGN",        (0,0), (-1,0), "MIDDLE"),
        ("TOPPADDING",    (0,0), (-1,0), 6),
        ("BOTTOMPADDING", (0,0), (-1,0), 6),
    ]))
    story.append(ft)

    doc.build(story)
    buf.seek(0)
    return buf.read()


# ── Session-state defaults ─────────────────────────────────────────────────────
def _ss(k, v):
    if k not in st.session_state: st.session_state[k] = v

for k, v in [
    ("step", 1), ("image_bytes", None), ("ocr_done", False),
    ("quantities", {}), ("match_log", []), ("ocr_grid", []),
    ("ocr_meta", {}), ("ocr_extracted", []), ("line_items", []),
    ("bill_to", {}), ("ship_to", {}), ("party_confirmed", False), ("pdf_bytes", None),
    ("azure_endpoint", AZURE_ENDPOINT), ("azure_key", AZURE_KEY),
    ("image_submitted", False),
    ("editor_open", False),
    ("raw_image_bytes", None),
    ("ocr_reviewed", False),
    ("upload_key", 0),
    ("selected_state", ""),
    ("state_confirmed", False),
    ("cam_preview_bytes", None),
    ("image_rotation", 0),
    ("capture_mode", "camera"),
    ("image_saved_to_disk", False),
    ("log_saved_to_disk", False),
    ("current_quotation_id", ""),
    ("distributor_name", ""),
    ("margin_percent", 0.0),
    ("db_ocr_id", None),
    ("ocr_db_written", False),
    ("quotation_db_written", False),
    ("image_disk_path", ""),
    ("quotation_disk_path", ""),
    ("detection_disk_path", ""),
    ("user_latitude", None),
    ("user_longitude", None),
    ("geo_status", "pending"),
    ("_geo_qp_processed", False),
    # Session log tracking
    ("session_log_sr_no", None),
    ("session_log_written", False),
    ("session_start_dt", datetime.now()),
    # Session-end tracking (stamped via Python on_click, NOT via JS bridge)
    ("session_end_stamped", False),
]:
    _ss(k, v)

if _GEO_BRIDGE_KEY not in st.session_state:
    st.session_state[_GEO_BRIDGE_KEY] = ""

sku_sheets   = load_sku_sheets()
sku_master   = build_sku_master(sku_sheets)
prefix_index = build_prefix_index(sku_master)
customers    = load_customers()

mrp_data = load_mrp_data_for_state(st.session_state.selected_state)
available_states = load_state_names_from_csv()

@st.cache_data(show_spinner=False)
def build_product_size_map(sku_master_frozen: str) -> dict:
    master = json.loads(sku_master_frozen)
    pmap = {}
    for sku, info in master.items():
        prod = info.get("product", "").strip()
        sz   = info.get("od_size", "").strip()
        if prod and sz:
            pmap.setdefault(prod, {})[sz] = sku
    return pmap

def validate_party(d: dict, prefix: str) -> list:
    errors = []
    phone  = d.get("phone", "").strip()
    mobile = d.get("mobile", "").strip()
    pno    = d.get("party_no", "").strip()
    if phone and not re.fullmatch(r'\d{10}', phone):
        errors.append(f"{prefix} Phone must be exactly 10 digits (no spaces/symbols).")
    if mobile and not re.fullmatch(r'\d{10}', mobile):
        errors.append(f"{prefix} Mobile must be exactly 10 digits (no spaces/symbols).")
    if pno and not re.fullmatch(r'\d+', pno):
        errors.append(f"{prefix} Party No. must contain digits only.")
    return errors

def is_party_complete(d: dict) -> bool:
    return bool(
        d.get("party_no","").strip() and
        d.get("name","").strip() and
        d.get("phone","").strip() and
        d.get("mobile","").strip() and
        d.get("state_code","").strip() and
        d.get("state","").strip() and
        d.get("gst","").strip() and
        d.get("pan","").strip()
    )


# ══════════════════════════════════════════════════════════════════════════════
# SESSION LOG — Insert row on first load (once per session)
# ══════════════════════════════════════════════════════════════════════════════
def _ensure_session_log_created():
    if st.session_state.get("session_log_written"):
        return

    sr_no = db_insert_session_log(
        ocr_id           = None,
        quotation_id     = None,
        ip_address       = _get_local_ip(),
        latitude         = st.session_state.get("user_latitude"),
        longitude        = st.session_state.get("user_longitude"),
        session_start_dt = st.session_state["session_start_dt"],
    )
    if sr_no is not None:
        st.session_state["session_log_sr_no"]   = sr_no
        st.session_state["session_log_written"] = True


# ── Parse geo bridge BEFORE rendering anything ─────────────────────────────────
_parse_geo_bridge()

render_geolocation_component()

# Insert session log row on first meaningful load
_ensure_session_log_created()

# If geo just resolved this run AND session log already exists, update it with coordinates
if (
    st.session_state.get("geo_status") == "ok"
    and st.session_state.get("session_log_sr_no")
    and not st.session_state.get("session_geo_updated")
    and st.session_state.get("user_latitude") is not None
):
    db_update_session_log_quotation(
        sr_no        = st.session_state["session_log_sr_no"],
        quotation_id = st.session_state.get("current_quotation_id", "") or "",
        ocr_id       = st.session_state.get("db_ocr_id"),
        latitude     = st.session_state.get("user_latitude"),
        longitude    = st.session_state.get("user_longitude"),
    )
    st.session_state["session_geo_updated"] = True

# ── App header ─────────────────────────────────────────────────────────────────
st.markdown(f"""
<div class="app-header">
  <div class="app-header-badge">{_LOGO_HTML}</div>
  <div class="app-header-text">
    <h1>Sintex BAPL — Quotation Generator</h1>
    <p>CPVC / UPVC Pipes &amp; Fittings</p>
  </div>
</div>""", unsafe_allow_html=True)

step_unlocked = [
    True,
    st.session_state.image_submitted,
    st.session_state.ocr_reviewed,
    st.session_state.party_confirmed,
]

cur_active = 1
if st.session_state.party_confirmed:
    cur_active = 4
elif st.session_state.ocr_reviewed:
    cur_active = 3
elif st.session_state.image_submitted:
    cur_active = 2

nav_html = '<div class="step-navbar">'
for i, lbl in enumerate(["Capture", "Edit/Review Prices", "User Information", "Download and Send Quotation"], 1):
    is_active = (i == cur_active)
    is_done   = step_unlocked[i-1] and (i < cur_active)
    is_locked = not step_unlocked[i-1]
    cls = "locked" if is_locked else ("active" if is_active else ("done" if is_done else ""))
    dot_txt = "✓" if is_done else str(i)
    nav_html += f"""
    <div class="step-nav-item {cls}">
      <div class="step-nav-dot">{dot_txt}</div>
      <div class="step-nav-label">{lbl}</div>
    </div>"""
nav_html += "</div>"
st.markdown(nav_html, unsafe_allow_html=True)

def render_rotated_preview(raw_bytes: bytes, rotation: int) -> str:
    try:
        img = Image.open(io.BytesIO(raw_bytes))
        img = fix_exif_orientation(img)
        if img.mode not in ("RGB", "L"):
            img = img.convert("RGB")
        max_dim = 1200
        w, h = img.size
        if max(w, h) > max_dim:
            ratio = max_dim / max(w, h)
            img = img.resize((int(w * ratio), int(h * ratio)), Image.LANCZOS)
        if rotation != 0:
            img = img.rotate(-rotation, expand=True)
        preview_buf = io.BytesIO()
        img.save(preview_buf, format="JPEG", quality=85)
        return base64.b64encode(preview_buf.getvalue()).decode()
    except Exception:
        return base64.b64encode(raw_bytes).decode()


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — Capture
# ══════════════════════════════════════════════════════════════════════════════
def render_step1():
    done = st.session_state.image_submitted and st.session_state.ocr_done

    st.markdown(f"""
    <div class="step-card">
      <div class="step-card-header">
        <div class="step-number {'done' if done else ''}">{'✓' if done else '1'}</div>
        <div>
          <div class="step-title">Step 1 — Capture Order Form</div>
          <div class="step-subtitle">Select your state, upload image</div>
        </div>
      </div>
      <div class="step-body">
    """, unsafe_allow_html=True)

    geo_badge_html = _geo_status_badge()
    if geo_badge_html:
        st.markdown(geo_badge_html, unsafe_allow_html=True)

    st.markdown(
        '<div class="state-inner-card"><div class="state-inner-title">🗺️ Select State / Region</div>',
        unsafe_allow_html=True,
    )
    try:
        current_idx = available_states.index(st.session_state.selected_state)
    except ValueError:
        current_idx = 0

    chosen = st.selectbox(
        "State", options=available_states, index=current_idx,
        key="state_selectbox", label_visibility="collapsed",
    )
    if chosen != st.session_state.selected_state:
        st.session_state.selected_state = chosen
        load_mrp_data_for_state.clear()
        if not st.session_state.image_submitted:
            st.rerun()
        else:
            st.markdown(
                '<div class="warn-box">⚠️ State changed. New pricing applies on next PDF.</div>',
                unsafe_allow_html=True,
            )

    st.markdown('</div>', unsafe_allow_html=True)

    with st.expander("🔧 Azure OCR Settings", expanded=not st.session_state.azure_key):
        st.session_state.azure_endpoint = st.text_input(
            "Azure Endpoint", value=st.session_state.azure_endpoint,
            placeholder="https://YOUR-RESOURCE.cognitiveservices.azure.com",
        )
        st.session_state.azure_key = st.text_input(
            "Azure API Key", value=st.session_state.azure_key,
            type="password", placeholder="••••••••••••••••••••••••••••••••",
        )

    if st.session_state.image_submitted:
        if st.session_state.ocr_done:
            pass
        else:
            raw = st.session_state.raw_image_bytes
            if raw:
                b64s = render_rotated_preview(raw, st.session_state.image_rotation)
                w, h = get_image_dimensions(raw)
                st.markdown(f"""
                <div style="border:1.5px solid #333;border-radius:10px;overflow:hidden;margin-bottom:14px;">
                  <div style="background:#1a1a1a;padding:7px 14px;display:flex;align-items:center;
                      justify-content:space-between;">
                    <span style="color:#aaa;font-size:12px;font-weight:600;">📷 Submitted Image</span>
                    <span style="color:#555;font-size:10px;">{w}×{h}px</span>
                  </div>
                  <img src="data:image/jpeg;base64,{b64s}"
                    style="width:100%;max-height:320px;object-fit:contain;display:block;background:#000;"/>
                </div>""", unsafe_allow_html=True)

                st.markdown("<div style='height:18px'></div>", unsafe_allow_html=True)
                _, sub_center, _ = st.columns([1, 2, 1])
                with sub_center:
                    st.markdown(
                        '<div class="rotate-bar-label">🔄&nbsp;&nbsp;Rotate Image</div>',
                        unsafe_allow_html=True,
                    )
                    sr1, sr2 = st.columns(2)
                    with sr1:
                        st.markdown('<div class="rotate-btn-col rotate-btn-col-left">', unsafe_allow_html=True)
                        if st.button("↺  90° Left", key="srot_ccw"):
                            st.session_state.image_rotation = (st.session_state.image_rotation - 90) % 360
                            st.rerun()
                        st.markdown('</div>', unsafe_allow_html=True)
                    with sr2:
                        st.markdown('<div class="rotate-btn-col rotate-btn-col-right">', unsafe_allow_html=True)
                        if st.button("↻  90° Right", key="srot_cw"):
                            st.session_state.image_rotation = (st.session_state.image_rotation + 90) % 360
                            st.rerun()
                        st.markdown('</div>', unsafe_allow_html=True)
                st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)

            st.markdown(
                '<div class="success-box" style="margin-top:16px;">✅ Image submitted — proceed for OCR.</div>',
                unsafe_allow_html=True,
            )

        st.markdown("</div></div>", unsafe_allow_html=True)
        return

    st.markdown("""
    <style>
    div[data-testid="stRadio"] > label { display: none !important; }
    div[data-testid="stRadio"] > div {
        display: flex !important;
        flex-direction: row !important;
        gap: 0 !important;
        background: #1a1a1a !important;
        border-radius: 10px 10px 0 0 !important;
        border: 2px solid #2a2a2a !important;
        border-bottom: none !important;
        overflow: hidden !important;
        padding: 0 !important;
        width: 100% !important;
        box-sizing: border-box !important;
    }
    div[data-testid="stRadio"] > div > label {
        display: flex !important;
        flex: 1 !important;
        justify-content: center !important;
        align-items: center !important;
        padding: 14px 8px !important;
        font-size: 14px !important;
        font-weight: 700 !important;
        font-family: 'Inter', sans-serif !important;
        cursor: pointer !important;
        color: #ffffff !important;
        border-bottom: 3px solid transparent !important;
        transition: all 0.15s !important;
        margin: 0 !important;
        user-select: none !important;
    }
    div[data-testid="stRadio"] > div > label:hover {
        background: #222 !important;
        color: #ffffff !important;
    }
    div[data-testid="stRadio"] > div > label[data-checked="true"] {
        background: #1f0a0a !important;
        color: #C0211F !important;
        border-bottom: 3px solid #C0211F !important;
    }
    div[data-testid="stRadio"] > div > label > div:first-child {
        display: none !important;
    }
    .sx-capture-panel {
        background: #111;
        border: 2px solid #2a2a2a;
        border-top: none;
        border-radius: 0 0 10px 10px;
        padding: 20px 16px 24px;
        margin-bottom: 0;
    }
    </style>
    """, unsafe_allow_html=True)

    st.markdown("""
    <style>
    div[data-testid="stRadio"],
    div[data-testid="stRadio"] > div[role="radiogroup"],
    div[data-testid="stHorizontalBlock"]:has(div[data-testid="stRadio"]),
    div[data-testid="stElementContainer"]:has(div[data-testid="stRadio"]) {
        width: 100% !important;
        max-width: 100% !important;
        min-width: 0 !important;
        flex: 1 1 100% !important;
        box-sizing: border-box !important;
    }
    div[data-testid="stRadio"] > div {
        width: 100% !important;
        max-width: 100% !important;
        min-width: 0 !important;
        box-sizing: border-box !important;
    }
    div[data-testid="stRadio"] > div > label {
        flex: 1 1 50% !important;
        max-width: 50% !important;
        min-width: 0 !important;
        color: #ffffff !important;
        opacity: 1 !important;
    }
    div[data-testid="stRadio"] > div > label:not([data-checked="true"]) {
        color: #ffffff !important;
        background: #1a1a1a !important;
        opacity: 1 !important;
    }
    div[data-testid="stRadio"] > div > label[data-checked="true"] {
        color: #ffffff !important;
        background: #2a0808 !important;
        border-bottom: 3px solid #C0211F !important;
        opacity: 1 !important;
    }
    div[data-testid="stRadio"] > div > label p,
    div[data-testid="stRadio"] > div > label span {
        color: #ffffff !important;
    }
    </style>
    """, unsafe_allow_html=True)

    tab_choice = st.radio(
        "Capture mode",
        options=["📷  Camera", "📁  Upload File"],
        index=0 if st.session_state.capture_mode == "camera" else 1,
        horizontal=True,
        key="capture_mode_radio",
        label_visibility="collapsed",
    )

    new_mode = "camera" if tab_choice == "📷  Camera" else "upload"
    if new_mode != st.session_state.capture_mode:
        st.session_state.capture_mode = new_mode
        st.session_state.raw_image_bytes = None
        st.session_state.image_rotation = 0
        st.rerun()

    is_camera = (st.session_state.capture_mode == "camera")
    upload_key_suffix = st.session_state.upload_key

    st.markdown('<div class="sx-capture-panel">', unsafe_allow_html=True)

    if is_camera:
        cam_html = """<!DOCTYPE html>
<html>
<head>
<meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0">
<style>
*{box-sizing:border-box;margin:0;padding:0;}
body{background:#111;font-family:'Inter',-apple-system,sans-serif;}
.wrap{background:#111;padding:20px 14px 16px;display:flex;flex-direction:column;
    align-items:center;gap:14px;}
#file-inp{display:none;}
.open-btn{
    display:flex;align-items:center;justify-content:center;gap:10px;
    background:linear-gradient(135deg,#C0211F,#8B1514);color:white;border:none;
    border-radius:12px;padding:18px 20px;font-size:16px;font-weight:700;
    font-family:'Inter',sans-serif;cursor:pointer;
    box-shadow:0 4px 20px rgba(192,33,31,.45);width:100%;
    -webkit-tap-highlight-color:transparent;
}
.tip{color:#666;font-size:12px;text-align:center;line-height:1.6;}
#preview-sec{display:none;width:100%;flex-direction:column;align-items:center;gap:10px;}
.prev-bar{width:100%;background:#1a1a1a;padding:8px 12px;display:flex;
    align-items:center;justify-content:space-between;}
.prev-lbl{color:#aaa;font-size:12px;font-weight:600;}
.badge{background:#1E7E4A;color:white;font-size:10px;font-weight:700;
    padding:3px 10px;border-radius:4px;}
#prev-img{width:100%;max-height:340px;object-fit:contain;display:block;background:#000;}
.act-btns{display:flex;flex-direction:column;align-items:center;gap:8px;width:100%;}
.use-btn{
    display:flex;align-items:center;justify-content:center;gap:8px;
    background:linear-gradient(135deg,#1E7E4A,#155d38);color:white;border:none;
    border-radius:10px;padding:16px 20px;font-size:15px;font-weight:700;
    font-family:'Inter',sans-serif;cursor:pointer;width:100%;
    box-shadow:0 4px 16px rgba(30,126,74,.4);
    -webkit-tap-highlight-color:transparent;
}
.retake-btn{
    background:transparent;color:#888;border:1.5px solid #333;border-radius:10px;
    padding:12px 20px;font-size:13px;font-weight:600;font-family:'Inter',sans-serif;
    cursor:pointer;width:100%;
    -webkit-tap-highlight-color:transparent;
}
#msg{color:#4ade80;font-size:12px;font-weight:600;text-align:center;
    padding:8px 12px;background:rgba(30,126,74,0.15);border-radius:6px;
    width:100%;display:none;}
#errmsg{color:#f87171;font-size:12px;font-weight:600;text-align:center;
    padding:8px 12px;background:rgba(192,33,31,0.15);border-radius:6px;
    width:100%;display:none;}
</style>
</head>
<body>
<div class="wrap">
    <input type="file" id="file-inp" accept="image/*" capture="environment"/>
    <div id="shoot-sec" style="display:flex;flex-direction:column;align-items:center;gap:12px;width:100%;">
        <button class="open-btn" id="open-btn">📷&nbsp; Open Camera</button>
        <p class="tip">Tap above to open your phone camera.<br/>Take a clear, well-lit photo of the order form.</p>
    </div>
    <div id="preview-sec">
        <div class="prev-bar">
            <span class="prev-lbl">📸 Preview</span>
            <span class="badge">✓ CAPTURED</span>
        </div>
        <img id="prev-img" src="" alt="preview"/>
        <div class="act-btns">
            <button class="use-btn" id="use-btn">✅&nbsp; Use This Photo</button>
            <button class="retake-btn" id="retake-btn">🔄&nbsp; Retake</button>
        </div>
    </div>
    <div id="msg">Photo sent! Scroll down for rotate &amp; submit controls.</div>
    <div id="errmsg"></div>
</div>
<script>
(function(){
    var inp=document.getElementById('file-inp');
    var openBtn=document.getElementById('open-btn');
    var shootSec=document.getElementById('shoot-sec');
    var prevSec=document.getElementById('preview-sec');
    var prevImg=document.getElementById('prev-img');
    var useBtn=document.getElementById('use-btn');
    var retakeBtn=document.getElementById('retake-btn');
    var msgDiv=document.getElementById('msg');
    var errDiv=document.getElementById('errmsg');
    var b64=null, mime='image/jpeg';

    openBtn.addEventListener('click',function(e){e.preventDefault();inp.click();});

    inp.addEventListener('change',function(){
        var file=inp.files[0];
        if(!file) return;
        mime=file.type||'image/jpeg';
        var reader=new FileReader();
        reader.onload=function(ev){
            b64=ev.target.result;
            prevImg.src=b64;
            shootSec.style.display='none';
            prevSec.style.display='flex';
            errDiv.style.display='none';
        };
        reader.readAsDataURL(file);
    });

    retakeBtn.addEventListener('click',function(){
        inp.value=''; b64=null; prevImg.src='';
        prevSec.style.display='none';
        shootSec.style.display='flex';
        msgDiv.style.display='none';
        errDiv.style.display='none';
    });

    useBtn.addEventListener('click',function(){
        if(!b64){
            errDiv.textContent='No photo captured yet.';
            errDiv.style.display='block';
            return;
        }
        useBtn.disabled=true;
        useBtn.textContent='Sending...';
        try{
            var raw=b64.split(',')[1];
            var ext=(mime.split('/')[1]||'jpg').replace('jpeg','jpg');
            var bin=atob(raw);
            var arr=new Uint8Array(bin.length);
            for(var i=0;i<bin.length;i++) arr[i]=bin.charCodeAt(i);
            var blob=new Blob([arr],{type:mime});
            var f=new File([blob],'photo.'+ext,{type:mime});
            var dt=new DataTransfer();
            dt.items.add(f);
            var doc=window.parent.document;
            var target=null;
            var uploaders=doc.querySelectorAll('[data-testid="stFileUploader"]');
            for(var i=0;i<uploaders.length;i++){
                var lbl=uploaders[i].querySelector('label[data-testid="stWidgetLabel"]');
                if(!lbl) continue;
                var cs=window.parent.getComputedStyle(lbl);
                if(cs.visibility==='hidden'||lbl.innerText.trim()===''||lbl.innerText.trim()==='cam_bridge'){
                    var fi=uploaders[i].querySelector('input[type="file"]');
                    if(fi){target=fi;break;}
                }
            }
            if(!target){
                var all=doc.querySelectorAll('input[type="file"]');
                for(var j=0;j<all.length;j++){
                    var pu=all[j].closest('[data-testid="stFileUploader"]');
                    if(!pu) continue;
                    var pl=pu.querySelector('label[data-testid="stWidgetLabel"]');
                    if(pl&&window.parent.getComputedStyle(pl).visibility==='hidden'){
                        target=all[j]; break;
                    }
                }
            }
            if(!target){
                errDiv.textContent='Bridge not found — please use the Upload File tab instead.';
                errDiv.style.display='block';
                useBtn.disabled=false;
                useBtn.textContent='✅ Use This Photo';
                return;
            }
            target.files=dt.files;
            target.dispatchEvent(new Event('change',{bubbles:true}));
            msgDiv.style.display='block';
            useBtn.textContent='✅ Sent!';
        }catch(e){
            errDiv.textContent='Error: '+e.message;
            errDiv.style.display='block';
            useBtn.disabled=false;
            useBtn.textContent='✅ Use This Photo';
        }
    });
})();
</script>
</body>
</html>"""
        components.html(cam_html, height=520, scrolling=False)

        _bridge_container = st.container()
        with _bridge_container:
            cam_file = st.file_uploader(
                "cam_bridge",
                type=["jpg", "jpeg", "png", "webp"],
                key=f"cam_native_{upload_key_suffix}",
                label_visibility="hidden",
            )
        st.markdown("""
        <style>
        [data-testid="stFileUploader"] {
            display: none !important;
        }
        </style>
        """, unsafe_allow_html=True)

        if cam_file is not None:
            nb = cam_file.getvalue()
            if nb != st.session_state.get("raw_image_bytes"):
                st.session_state.raw_image_bytes = nb
                st.session_state.image_rotation = 0
                st.rerun()

    else:
        st.markdown("""
        <style>
        [data-testid="stFileUploaderDropzone"] {
            border: 2px dashed #C0211F !important;
            background: rgba(192,33,31,0.06) !important;
            border-radius: 10px !important;
        }
        [data-testid="stFileUploaderDropzoneInstructions"] span {
            color: #C0211F !important; font-weight: 700 !important; font-size: 14px !important;
        }
        [data-testid="stFileUploaderDropzoneInstructions"] small {
            color: #888 !important; font-size: 12px !important;
        }
        </style>
        """, unsafe_allow_html=True)

        up_file = st.file_uploader(
            "📁 Choose image from device",
            type=["jpg", "jpeg", "png", "webp"],
            key=f"up_{upload_key_suffix}",
            label_visibility="visible",
        )

        if up_file is not None:
            nb = up_file.read()
            if nb != st.session_state.get("raw_image_bytes"):
                st.session_state.raw_image_bytes = nb
                st.session_state.image_rotation = 0

    st.markdown('</div>', unsafe_allow_html=True)

    raw_bytes_pending = st.session_state.raw_image_bytes

    if raw_bytes_pending is not None:
        rotation    = st.session_state.image_rotation
        b64_prev    = render_rotated_preview(raw_bytes_pending, rotation)
        w, h        = get_image_dimensions(raw_bytes_pending)
        badge_label = "CAPTURED" if is_camera else "UPLOADED"
        badge_icon  = "📷" if is_camera else "📁"
        source_lbl  = "Camera Photo" if is_camera else "Uploaded Image"

        st.markdown(f"""
        <div style="margin-top:14px;border:1.5px solid #333;border-radius:10px;overflow:hidden;">
          <div style="background:#1a1a1a;padding:8px 14px;display:flex;
              align-items:center;justify-content:space-between;">
            <span style="color:#aaa;font-size:12px;font-weight:600;">
              {badge_icon} {source_lbl}&nbsp;
              <span style="color:#555;font-size:10px;">{w}×{h}px</span>
            </span>
            <span style="background:#1E7E4A;color:white;font-size:10px;
                font-weight:700;padding:3px 10px;border-radius:4px;">✓ {badge_label}</span>
          </div>
          <img src="data:image/jpeg;base64,{b64_prev}"
            style="width:100%;max-height:360px;object-fit:contain;display:block;background:#000;"/>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<div style='height:18px'></div>", unsafe_allow_html=True)

        _, center_col, _ = st.columns([1, 2, 1])
        with center_col:
            st.markdown(
                '<div class="rotate-bar-label">🔄&nbsp;&nbsp;Rotate Image</div>',
                unsafe_allow_html=True,
            )
            rot_left_col, rot_right_col = st.columns(2)
            with rot_left_col:
                st.markdown('<div class="rotate-btn-col rotate-btn-col-left">', unsafe_allow_html=True)
                if st.button("↺  90° Left", key="btn_rotate_ccw"):
                    st.session_state.image_rotation = (rotation - 90) % 360
                    st.rerun()
                st.markdown('</div>', unsafe_allow_html=True)
            with rot_right_col:
                st.markdown('<div class="rotate-btn-col rotate-btn-col-right">', unsafe_allow_html=True)
                if st.button("↻  90° Right", key="btn_rotate_cw"):
                    st.session_state.image_rotation = (rotation + 90) % 360
                    st.rerun()
                st.markdown('</div>', unsafe_allow_html=True)
        st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)

        st.markdown("<div style='margin-top:14px;'>", unsafe_allow_html=True)
        st.markdown('<div class="submit-btn-col">', unsafe_allow_html=True)
        if st.button("✅  Submit & Process Image", key="btn_submit_image"):
            for k in ["quantities", "match_log", "ocr_grid", "ocr_meta", "ocr_extracted", "line_items"]:
                st.session_state[k] = [] if isinstance(st.session_state.get(k), list) else {}
            st.session_state.image_bytes              = raw_bytes_pending
            st.session_state.raw_image_bytes          = raw_bytes_pending
            st.session_state.ocr_done                 = False
            st.session_state.ocr_reviewed             = False
            st.session_state.party_confirmed          = False
            st.session_state.pdf_bytes                = None
            st.session_state.image_submitted          = True
            st.session_state.image_saved_to_disk      = False
            st.session_state.log_saved_to_disk        = False
            st.session_state.current_quotation_id     = ""
            st.session_state.db_ocr_id                = None
            st.session_state.ocr_db_written           = False
            st.session_state.quotation_db_written     = False
            st.session_state.image_disk_path          = ""
            st.session_state.quotation_disk_path      = ""
            st.session_state.detection_disk_path      = ""
            st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

    st.markdown("</div></div>", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — OCR · Spatial Reconstruction · SKU Matching
# ══════════════════════════════════════════════════════════════════════════════
def render_step2():
    if not step_unlocked[1]:
        return

    if (st.session_state.image_submitted
            and not st.session_state.image_saved_to_disk
            and st.session_state.raw_image_bytes):
        try:
            img_path = save_image_to_disk(
                st.session_state.raw_image_bytes,
                st.session_state.get("image_rotation", 0),
            )
            st.session_state.image_disk_path     = img_path
            st.session_state.image_saved_to_disk = True

            if not st.session_state.ocr_db_written:
                ocr_id = db_insert_ocr_log(
                    image_path    = img_path,
                    selected_state= st.session_state.get("selected_state", ""),
                    ocr_start_time= datetime.now(),
                )
                st.session_state.db_ocr_id      = ocr_id
                st.session_state.ocr_db_written = True

                if st.session_state.get("session_log_sr_no") and ocr_id:
                    db_update_session_log_quotation(
                        sr_no        = st.session_state["session_log_sr_no"],
                        quotation_id = st.session_state.get("current_quotation_id", "") or "",
                        ocr_id       = ocr_id,
                        latitude     = st.session_state.get("user_latitude"),
                        longitude    = st.session_state.get("user_longitude"),
                    )
        except Exception as exc:
            st.warning(f"⚠️ File/DB write warning (Step 2 init): {exc}")

    done = st.session_state.ocr_done
    st.markdown(f"""
    <div class="step-card"><div class="step-card-header">
      <div class="step-number {'done' if done else ''}">{'✓' if done else '2'}</div>
      <div><div class="step-title">Step 2 — OCR · Spatial Reconstruction · SKU Matching</div>
      <div class="step-subtitle">Detects table structure from bounding boxes — works with any form layout</div></div>
    </div><div class="step-body">""", unsafe_allow_html=True)

    if not done:
        c1, c2 = st.columns(2)
        with c1:
            run = st.button("Run OCR & Match", key="run_ocr",
                            disabled=not (st.session_state.azure_key and st.session_state.azure_endpoint and st.session_state.image_bytes))
        with c2:
            skip = st.button("Enter SKUs Manually", key="skip_ocr")

        if not st.session_state.image_bytes:
            st.markdown('<div class="error-box">❌ No image captured — go back to Step 1 and capture or upload an image first.</div>',
                        unsafe_allow_html=True)
        if not st.session_state.azure_endpoint:
            st.markdown('<div class="warn-box">⚠️ Azure Endpoint not set — configure in Step 1 → Azure OCR Settings.</div>',
                        unsafe_allow_html=True)
        if not st.session_state.azure_key:
            st.markdown('<div class="warn-box">⚠️ Azure API Key not set — configure in Step 1 → Azure OCR Settings.</div>',
                        unsafe_allow_html=True)

        if run:
            ph = st.empty()
            ocr_start = datetime.now()
            try:
                ph.info("⏳ 1/4 - Loading…")
                enhanced_bytes = enhance_image_for_ocr(
                    st.session_state.image_bytes,
                    st.session_state.get("image_rotation", 0)
                )

                ph.info("⏳ 2/4 - Loading…")
                words = run_azure_ocr(enhanced_bytes,
                                      st.session_state.azure_endpoint,
                                      st.session_state.azure_key)

                ph.info(f"🔲 3/4 - Loading…")
                grid = reconstruct_table(words)

                ph.info(f"🔍 4/4 - Loading…")
                res  = analyze_table(grid)
                qtys, log, line_items = build_quantities(res["rows"], prefix_index, sku_master)

                st.session_state.ocr_grid      = grid
                st.session_state.ocr_meta      = res["meta"]
                st.session_state.ocr_extracted = res["rows"]
                st.session_state.quantities    = qtys
                st.session_state.match_log     = log
                st.session_state.line_items    = line_items
                st.session_state.ocr_done      = True

                nm = sum(1 for m in log if m["status"] == "matched")
                nu = sum(1 for m in log if m["status"] == "unmatched")

                if st.session_state.db_ocr_id:
                    db_update_ocr_log(
                        ocr_id              = st.session_state.db_ocr_id,
                        quantities_detected = len(line_items),
                        matched_count       = nm,
                        unmatched_count     = nu,
                        status              = 1,
                    )

                if nm == 0:
                    ph.warning(f"⚠️ Table reconstructed ({len(grid)} rows, {len(res['rows'])} data rows) "
                               f"but 0 SKUs matched. Check the Edit / Add tab below.")
                else:
                    ph.success(f"✅ Done — {nm} detections matched, {nu} unmatched. Review below.")
                time.sleep(0.8); st.rerun()

            except requests.exceptions.HTTPError as e:
                ph.empty()
                status_code = e.response.status_code if e.response else "?"
                try:    body = e.response.json().get("error", {}).get("message", "")
                except: body = (e.response.text or "")[:300]
                if st.session_state.db_ocr_id:
                    db_update_ocr_log(ocr_id=st.session_state.db_ocr_id,
                                      quantities_detected=0, matched_count=0,
                                      unmatched_count=0, status=2)
                st.error(f"❌ Azure HTTP {status_code}: {body or str(e)}")
            except requests.exceptions.ConnectionError:
                ph.empty()
                if st.session_state.db_ocr_id:
                    db_update_ocr_log(ocr_id=st.session_state.db_ocr_id,
                                      quantities_detected=0, matched_count=0,
                                      unmatched_count=0, status=2)
                st.error("❌ Cannot connect to Azure — check endpoint URL.")
            except TimeoutError as e:
                ph.empty()
                if st.session_state.db_ocr_id:
                    db_update_ocr_log(ocr_id=st.session_state.db_ocr_id,
                                      quantities_detected=0, matched_count=0,
                                      unmatched_count=0, status=2)
                st.error(f"❌ Timeout: {e}")
            except Exception as e:
                ph.empty()
                if st.session_state.db_ocr_id:
                    db_update_ocr_log(ocr_id=st.session_state.db_ocr_id,
                                      quantities_detected=0, matched_count=0,
                                      unmatched_count=0, status=2)
                st.error(f"❌ {type(e).__name__}: {e}")

        if skip:
            st.session_state.quantities = {}
            st.session_state.match_log  = []
            st.session_state.ocr_done   = True
            if st.session_state.db_ocr_id:
                db_update_ocr_log(ocr_id=st.session_state.db_ocr_id,
                                  quantities_detected=0, matched_count=0,
                                  unmatched_count=0, status=1)
            st.rerun()

    if done:
        log  = st.session_state.match_log
        qtys = st.session_state.quantities

        nm = sum(1 for m in log if m["status"] == "matched")
        nu = sum(1 for m in log if m["status"] == "unmatched")

        if nm: st.markdown(f'<div class="success-box">✅ {nm} SKU-size detections matched → {nm} line items in document</div>',
                           unsafe_allow_html=True)
        if nu: st.markdown(f'<div class="warn-box">⚠️ {nu} pair(s) unmatched — review in Edit / Add tab</div>',
                           unsafe_allow_html=True)
        if not nm and not nu:
            st.markdown('<div class="warn-box">No OCR data — use Edit tab to add SKUs manually.</div>',
                        unsafe_allow_html=True)

        tab_edit, tab1 = st.tabs(["✏️ Edit / Add", "📋 OCR Detection Log"])

        # ── EDIT / ADD TAB ────────────────────────────────────────────────────
        with tab_edit:

            # ── Section A: OCR-Detected Items — inline editable table ─────────
            items_existing = {s: q for s, q in st.session_state.quantities.items() if q > 0}

            if items_existing:
                st.markdown('<div class="fsl">OCR-Detected Items — Edit Quantities</div>', unsafe_allow_html=True)
                st.markdown(
                    '<div class="info-box" style="margin-bottom:10px;">'
                    'ℹ️ Quantities detected by OCR are shown below. '
                    'Edit the <b>Qty</b> column as needed. '
                    'Prices update automatically. '
                    'Click <b>Apply Changes</b> to save.'
                    '</div>',
                    unsafe_allow_html=True,
                )

                # Build row data for the embedded JS table
                ocr_rows_data = []
                for sku, qty in items_existing.items():
                    info = sku_master.get(sku, {})
                    mi   = mrp_data.get(sku, {})
                    ocr_rows_data.append({
                        "sku":      sku,
                        "product":  info.get("product", sku),
                        "od":       info.get("od_size", ""),
                        "inch":     info.get("inch_size", ""),
                        "mrp":      mi.get("mrp", 0.0),
                        "dist":     mi.get("distributor_landing", 0.0),
                        "qty":      qty,
                    })

                rows_json = json.dumps(ocr_rows_data)

                # Hidden bridge text_input — JS writes updated quantities here
                _EDIT_BRIDGE_KEY   = "_ocr_edit_bridge"
                _EDIT_BRIDGE_LABEL = "__ocr_edit_bridge__"

                if _EDIT_BRIDGE_KEY not in st.session_state:
                    st.session_state[_EDIT_BRIDGE_KEY] = ""

                st.text_input(
                    label=_EDIT_BRIDGE_LABEL,
                    key=_EDIT_BRIDGE_KEY,
                    label_visibility="hidden",
                    value=st.session_state.get(_EDIT_BRIDGE_KEY, ""),
                )

                # Check if bridge has data from a previous "Apply" click
                _bridge_raw = st.session_state.get(_EDIT_BRIDGE_KEY, "").strip()
                if _bridge_raw:
                    try:
                        updated_from_js = json.loads(_bridge_raw)
                        new_qtys = copy.deepcopy(st.session_state.quantities)
                        new_items = []
                        for entry in updated_from_js:
                            sku_e = entry.get("sku", "")
                            qty_e = int(entry.get("qty", 0))
                            if sku_e:
                                new_qtys[sku_e] = qty_e
                                if qty_e > 0:
                                    new_items.append({"sku": sku_e, "qty": qty_e})
                        # Preserve manually-added items (those not in OCR list)
                        ocr_skus = set(r["sku"] for r in ocr_rows_data)
                        for item in st.session_state.line_items:
                            if item["sku"] not in ocr_skus:
                                new_items.append(item)
                        st.session_state.quantities = new_qtys
                        st.session_state.line_items = new_items
                        st.session_state[_EDIT_BRIDGE_KEY] = ""
                        st.success("✅ Quantities updated successfully.")
                        st.rerun()
                    except (json.JSONDecodeError, TypeError, ValueError):
                        pass

                # Render the interactive HTML table
                edit_table_html = f"""<!DOCTYPE html>
<html>
<head>
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<style>
*{{box-sizing:border-box;margin:0;padding:0;}}
body{{font-family:'Inter',-apple-system,sans-serif;background:transparent;padding:0 0 8px;}}
.tbl-wrap{{overflow-x:auto;border:1.5px solid #DEDEDE;border-radius:10px;box-shadow:0 2px 8px rgba(0,0,0,0.05);}}
table{{width:100%;border-collapse:collapse;font-size:12.5px;min-width:680px;}}
thead tr{{background:#1A1A1A;}}
thead th{{color:white;padding:10px 10px;text-align:left;font-size:11px;font-weight:700;
    letter-spacing:0.5px;text-transform:uppercase;white-space:nowrap;
    border-right:1px solid #2e2e2e;}}
thead th:last-child{{border-right:none;}}
thead th.C{{text-align:center;}}
thead th.R{{text-align:right;}}
tbody tr{{border-bottom:1px solid #EFEFEF;transition:background 0.1s;}}
tbody tr:nth-child(even){{background:#FAFAFA;}}
tbody tr:hover{{background:#FFF5F5;}}
tbody td{{padding:8px 10px;color:#1A1A1A;vertical-align:middle;
    border-right:1px solid #EFEFEF;}}
tbody td:last-child{{border-right:none;}}
tbody td.C{{text-align:center;}}
tbody td.R{{text-align:right;font-family:'JetBrains Mono',monospace;font-size:12px;}}
tbody td.sku-cell{{font-family:'JetBrains Mono',monospace;font-size:10.5px;color:#555;}}
tbody td.price-cell{{font-family:'JetBrains Mono',monospace;font-size:12px;
    font-weight:600;color:#1A1A1A;text-align:right;}}
tbody td.total-cell{{font-family:'JetBrains Mono',monospace;font-size:12px;
    font-weight:700;color:#C0211F;text-align:right;}}
.qty-inp{{width:72px;padding:5px 8px;border:1.5px solid #D0D0D0;border-radius:6px;
    font-family:'Inter',sans-serif;font-size:13px;font-weight:700;color:#1A1A1A;
    text-align:center;background:white;outline:none;transition:border-color 0.15s,box-shadow 0.15s;}}
.qty-inp:focus{{border-color:#C0211F;box-shadow:0 0 0 3px rgba(192,33,31,0.12);}}
.qty-inp.changed{{border-color:#1E7E4A;background:#F0FFF6;}}
tfoot tr{{background:#F0F0F0;border-top:2px solid #D0D0D0;}}
tfoot td{{padding:9px 10px;font-weight:700;font-size:12.5px;color:#1A1A1A;
    border-right:1px solid #DEDEDE;}}
tfoot td:last-child{{border-right:none;}}
tfoot td.R{{text-align:right;font-family:'JetBrains Mono',monospace;color:#C0211F;}}
.actions-bar{{display:flex;align-items:center;justify-content:space-between;
    gap:12px;margin-top:12px;padding:10px 0 0;border-top:1.5px solid #EFEFEF;
    flex-wrap:wrap;}}
.hint-text{{font-size:11.5px;color:#92400E;background:#FFFBEB;
    border:1.5px solid #FDE68A;border-radius:8px;padding:7px 14px;font-weight:500;}}
.apply-btn{{background:linear-gradient(135deg,#1E7E4A,#155d38);color:white;
    border:none;border-radius:9px;padding:11px 24px;font-family:'Inter',sans-serif;
    font-size:13.5px;font-weight:700;cursor:pointer;
    box-shadow:0 4px 14px rgba(30,126,74,0.3);transition:all 0.15s;
    white-space:nowrap;}}
.apply-btn:hover{{filter:brightness(1.08);transform:translateY(-1px);}}
.apply-btn:active{{transform:none;}}
.applied-msg{{display:none;color:#065F46;font-size:12px;font-weight:600;
    background:#ECFDF5;border:1.5px solid #A7F3D0;border-radius:6px;
    padding:7px 14px;}}
</style>
</head>
<body>
<div class="tbl-wrap">
<table id="edit-tbl">
  <thead>
    <tr>
      <th style="width:36px;" class="C">#</th>
      <th>Product</th>
      <th class="sku-cell">SKU Code</th>
      <th class="C">OD</th>
      <th class="C">Inch</th>
      <th class="R">MRP / Unit (₹)</th>
      <th class="C" style="width:90px;">Qty</th>
      <th class="R">Total MRP (₹)</th>
      <th class="R">Dist. Landing / Unit (₹)</th>
      <th class="R">Taxable Value (₹)</th>
    </tr>
  </thead>
  <tbody id="edit-tbody"></tbody>
  <tfoot>
    <tr>
      <td colspan="7" style="text-align:right;font-size:12px;color:#555;font-weight:600;font-style:italic;">
        Grand Totals →
      </td>
      <td class="R" id="foot-mrp">₹ 0.00</td>
      <td class="R">—</td>
      <td class="R" id="foot-tax">₹ 0.00</td>
    </tr>
  </tfoot>
</table>
</div>
<div class="actions-bar">
  <span class="hint-text">✏️ Edit quantities above, then click Apply to save.</span>
  <div style="display:flex;gap:10px;align-items:center;">
    <span class="applied-msg" id="applied-msg">✅ Saved — page refreshing…</span>
    <button class="apply-btn" id="apply-btn" onclick="applyChanges()">
      ✅&nbsp; Apply Changes
    </button>
  </div>
</div>

<script>
var ROWS = {rows_json};
var origQtys = {{}};

function fmt(n){{
    return '₹ ' + n.toLocaleString('en-IN',{{minimumFractionDigits:2,maximumFractionDigits:2}});
}}

function buildTable(){{
    var tbody = document.getElementById('edit-tbody');
    tbody.innerHTML = '';
    ROWS.forEach(function(r, idx){{
        origQtys[r.sku] = r.qty;
        var tr = document.createElement('tr');
        tr.innerHTML =
            '<td class="C" style="color:#999;font-size:10px;">' + (idx+1) + '</td>' +
            '<td style="font-weight:500;">' + escHtml(r.product) + '</td>' +
            '<td class="sku-cell"><code style="font-size:10px;">' + escHtml(r.sku) + '</code></td>' +
            '<td class="C"><b>' + escHtml(r.od) + '</b></td>' +
            '<td class="C">' + escHtml(r.inch) + '</td>' +
            '<td class="price-cell">' + fmt(r.mrp) + '</td>' +
            '<td class="C"><input type="number" class="qty-inp" id="qty-' + idx + '" ' +
                'data-sku="' + escHtml(r.sku) + '" ' +
                'data-mrp="' + r.mrp + '" ' +
                'data-dist="' + r.dist + '" ' +
                'data-orig="' + r.qty + '" ' +
                'value="' + r.qty + '" min="0" max="9999" step="1" ' +
                'oninput="onQtyChange(this,' + idx + ')"/></td>' +
            '<td class="total-cell" id="tmrp-' + idx + '">' + fmt(r.mrp * r.qty) + '</td>' +
            '<td class="price-cell">' + fmt(r.dist) + '</td>' +
            '<td class="total-cell" id="ttax-' + idx + '">' + fmt(r.dist * r.qty) + '</td>';
        tbody.appendChild(tr);
    }});
    updateFooter();
}}

function escHtml(s){{
    return String(s)
        .replace(/&/g,'&amp;')
        .replace(/</g,'&lt;')
        .replace(/>/g,'&gt;')
        .replace(/"/g,'&quot;');
}}

function onQtyChange(inp, idx){{
    var qty = parseInt(inp.value) || 0;
    if(qty < 0) {{ inp.value = 0; qty = 0; }}
    var mrp  = parseFloat(inp.getAttribute('data-mrp'))  || 0;
    var dist = parseFloat(inp.getAttribute('data-dist')) || 0;
    var orig = parseInt(inp.getAttribute('data-orig'))   || 0;
    document.getElementById('tmrp-' + idx).textContent = fmt(mrp  * qty);
    document.getElementById('ttax-' + idx).textContent = fmt(dist * qty);
    if(qty !== orig){{
        inp.classList.add('changed');
    }} else {{
        inp.classList.remove('changed');
    }}
    updateFooter();
}}

function updateFooter(){{
    var inputs = document.querySelectorAll('.qty-inp');
    var totalMrp = 0, totalTax = 0;
    inputs.forEach(function(inp){{
        var qty  = parseInt(inp.value) || 0;
        var mrp  = parseFloat(inp.getAttribute('data-mrp'))  || 0;
        var dist = parseFloat(inp.getAttribute('data-dist')) || 0;
        totalMrp += mrp  * qty;
        totalTax += dist * qty;
    }});
    document.getElementById('foot-mrp').textContent = fmt(totalMrp);
    document.getElementById('foot-tax').textContent = fmt(totalTax);
}}

function applyChanges(){{
    var inputs = document.querySelectorAll('.qty-inp');
    var result = [];
    inputs.forEach(function(inp){{
        result.push({{
            sku: inp.getAttribute('data-sku'),
            qty: parseInt(inp.value) || 0
        }});
    }});
    var payload = JSON.stringify(result);

    // Write into hidden Streamlit text_input bridge
    var writeAttempts = 0;
    function tryWrite(){{
        writeAttempts++;
        var written = false;
        try{{
            var doc = window.parent.document;
            var inputs2 = doc.querySelectorAll('input[type="text"]');
            var el = null;
            for(var i=0; i<inputs2.length; i++){{
                var lbl = inputs2[i].getAttribute('aria-label') || '';
                if(lbl.indexOf('__ocr_edit_bridge__') !== -1){{ el = inputs2[i]; break; }}
            }}
            if(el){{
                var setter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype,'value').set;
                setter.call(el, payload);
                el.dispatchEvent(new Event('input',  {{bubbles:true}}));
                el.dispatchEvent(new Event('change', {{bubbles:true}}));
                el.dispatchEvent(new Event('blur',   {{bubbles:true}}));
                written = true;
            }}
        }} catch(e){{ written = false; }}
        if(!written && writeAttempts < 8){{
            setTimeout(tryWrite, Math.min(400 * Math.pow(2, writeAttempts-1), 6000));
        }}
    }}
    tryWrite();

    document.getElementById('applied-msg').style.display = 'inline-block';
    document.getElementById('apply-btn').disabled = true;
    document.getElementById('apply-btn').textContent = 'Applying…';
}}

buildTable();
</script>
</body>
</html>"""

                table_height = 42 + len(ocr_rows_data) * 48 + 42 + 60 + 40
                table_height = max(table_height, 300)
                components.html(edit_table_html, height=table_height, scrolling=False)

            else:
                st.markdown(
                    '<div class="info-box" style="margin-bottom:12px;">ℹ️ No OCR-detected quantities yet. '
                    'Add items manually using the section below.</div>',
                    unsafe_allow_html=True,
                )

            # ── Section B: Add Item by Product & Size ─────────────────────────
            st.markdown('<div class="fsl">Add Item by Product &amp; Size</div>', unsafe_allow_html=True)

            _master_json = json.dumps(sku_master)
            pmap = build_product_size_map(_master_json)
            product_names = sorted(pmap.keys())

            col_prod, col_sz, col_qty, col_add = st.columns([4, 2, 1, 1])

            with col_prod:
                sel_product = st.selectbox(
                    "Product Name",
                    options=["— select product —"] + product_names,
                    key="add_product_select",
                    label_visibility="collapsed",
                )

            if sel_product and sel_product != "— select product —":
                avail_sizes = sorted(pmap[sel_product].keys(),
                                     key=lambda s: int(re.sub(r'\D','',s) or 0))
            else:
                avail_sizes = []

            with col_sz:
                sel_size = st.selectbox(
                    "Size",
                    options=["— size —"] + avail_sizes,
                    key="add_size_select",
                    label_visibility="collapsed",
                    disabled=(not avail_sizes),
                )

            with col_qty:
                add_qty = st.number_input("Qty", min_value=1, value=1, step=1,
                                          key="add_qty_input",
                                          label_visibility="collapsed")

            resolved_sku = None
            if (sel_product and sel_product != "— select product —"
                    and sel_size and sel_size != "— size —"):
                resolved_sku = pmap[sel_product].get(sel_size)

            with col_add:
                st.markdown("<br/>", unsafe_allow_html=True)
                if st.button("ADD", key="add_by_product", disabled=(resolved_sku is None)):
                    updated2 = copy.deepcopy(st.session_state.quantities)
                    updated2[resolved_sku] = updated2.get(resolved_sku, 0) + add_qty
                    new_items = copy.deepcopy(st.session_state.line_items)
                    new_items.append({"sku": resolved_sku, "qty": add_qty})
                    st.session_state.quantities = updated2
                    st.session_state.line_items = new_items
                    info_added = sku_master.get(resolved_sku, {})
                    st.success(f"Added: {info_added.get('product',resolved_sku)} "
                               f"({sel_size}) × {add_qty} → SKU: {resolved_sku}")
                    st.rerun()

            if resolved_sku:
                st.markdown(
                    f'<div class="info-box" style="margin-top:6px;padding:8px 12px;">'
                    f'🔗 Resolved SKU: <code><b>{resolved_sku}</b></code></div>',
                    unsafe_allow_html=True,
                )
            elif sel_product != "— select product —" and sel_size != "— size —" and avail_sizes:
                st.markdown(
                    '<div class="warn-box" style="margin-top:6px;">⚠️ No SKU found for this combination.</div>',
                    unsafe_allow_html=True,
                )

        # ── OCR DETECTION LOG TAB ─────────────────────────────────────────────
        with tab1:
            live_log = build_live_log(
                match_log  = st.session_state.match_log,
                quantities = st.session_state.quantities,
                line_items = st.session_state.line_items,
                sku_master = sku_master,
                mrp_data   = mrp_data,
            )

            if live_log:
                st.markdown("""
                <div style="display:flex;gap:10px;flex-wrap:wrap;margin-bottom:12px;align-items:center;">
                  <span style="font-size:11.5px;font-weight:700;color:#6B6B6B;text-transform:uppercase;
                      letter-spacing:.6px;">Legend:</span>
                  <span style="background:#ECFDF5;color:#065F46;border:1.5px solid #A7F3D0;
                      border-radius:6px;padding:3px 10px;font-size:11.5px;font-weight:600;">
                      🟢 Matched &amp; Priced
                  </span>
                  <span style="background:#FFF7ED;color:#92400E;border:1.5px solid #FED7AA;
                      border-radius:6px;padding:3px 10px;font-size:11.5px;font-weight:600;">
                      🟠 Matched — No Price
                  </span>
                  <span style="background:#FEF2F2;color:#991B1B;border:1.5px solid #FECACA;
                      border-radius:6px;padding:3px 10px;font-size:11.5px;font-weight:600;">
                      🔴 Unmatched
                  </span>
                </div>
                """, unsafe_allow_html=True)

                h = """<div class="ocr-wrap"><table class="ocr-tbl">
                <thead><tr>
                  <th style="text-align:center;width:36px">#</th>
                  <th class="L">OCR Product Name</th>
                  <th class="L">OCR Prefix</th>
                  <th>Size</th>
                  <th>Qty</th>
                  <th class="L">Resolved SKU</th>
                  <th>MRP (₹)</th>
                  <th>Dist. Landing (₹)</th>
                  <th>Status</th>
                </tr></thead><tbody>"""

                for idx, entry in enumerate(live_log, 1):
                    sku       = entry["full_sku"]
                    status    = entry["status"]
                    mi        = mrp_data.get(sku, {}) if sku != "—" else {}
                    mrp_val   = mi.get("mrp", 0)
                    dist_val  = mi.get("distributor_landing", 0)

                    if status == "unmatched":
                        row_cls    = "log-row-red"
                        status_lbl = "🔴 Unmatched"
                        mrp_disp   = "—"
                        dist_disp  = "—"
                    elif mrp_val == 0:
                        row_cls    = "log-row-orange"
                        status_lbl = "🟠 No Price"
                        mrp_disp   = "₹ 0.00"
                        dist_disp  = "₹ 0.00"
                    else:
                        row_cls    = "log-row-green"
                        status_lbl = "🟢 Matched"
                        mrp_disp   = f"₹ {mrp_val:,.2f}"
                        dist_disp  = f"₹ {dist_val:,.2f}"

                    sku_display = (
                        f'<code style="font-size:10px">{sku}</code>'
                        if sku != "—"
                        else '<span style="color:#ccc">—</span>'
                    )

                    h += (
                        f'<tr class="{row_cls}">'
                        f'<td style="color:#999;font-size:10px;text-align:center">{idx}</td>'
                        f'<td class="L">{entry["product"] or "—"}</td>'
                        f'<td class="M">{entry["prefix"] or "—"}</td>'
                        f'<td><b>{entry["size"]}</b></td>'
                        f'<td><b>{entry["qty"]}</b></td>'
                        f'<td class="L">{sku_display}</td>'
                        f'<td>{mrp_disp}</td>'
                        f'<td>{dist_disp}</td>'
                        f'<td><b>{status_lbl}</b></td>'
                        f'</tr>'
                    )

                h += "</tbody></table></div>"

                n_green  = sum(1 for e in live_log if e["status"] == "matched" and mrp_data.get(e["full_sku"], {}).get("mrp", 0) > 0)
                n_orange = sum(1 for e in live_log if e["status"] == "matched" and mrp_data.get(e["full_sku"], {}).get("mrp", 0) == 0)
                n_red    = sum(1 for e in live_log if e["status"] == "unmatched")

                h += f"""
                <div style="display:flex;gap:12px;flex-wrap:wrap;margin-top:10px;padding:10px 0;">
                  <div style="background:#ECFDF5;border:1.5px solid #A7F3D0;border-radius:8px;
                      padding:8px 16px;font-size:12.5px;font-weight:700;color:#065F46;">
                      🟢 {n_green} Matched &amp; Priced
                  </div>
                  <div style="background:#FFF7ED;border:1.5px solid #FED7AA;border-radius:8px;
                      padding:8px 16px;font-size:12.5px;font-weight:700;color:#92400E;">
                      🟠 {n_orange} Matched — No Price
                  </div>
                  <div style="background:#FEF2F2;border:1.5px solid #FECACA;border-radius:8px;
                      padding:8px 16px;font-size:12.5px;font-weight:700;color:#991B1B;">
                      🔴 {n_red} Unmatched
                  </div>
                </div>
                """
                st.markdown(h, unsafe_allow_html=True)
            else:
                st.markdown(
                    '<div class="warn-box">No OCR detection log available. '
                    'Run OCR first, or add items manually via the Edit / Add tab.</div>',
                    unsafe_allow_html=True,
                )

        # ── Totals ────────────────────────────────────────────────────────────
        iord  = {s: q for s, q in st.session_state.quantities.items() if q > 0}
        gmrp  = sum(mrp_data.get(s, {}).get("mrp", 0) * q        for s, q in iord.items())
        gdist = sum(mrp_data.get(s, {}).get("distributor_landing", 0) * q for s, q in iord.items())
        disc  = gmrp - gdist
        st.markdown(f"""
        <div class="totals-box">
          <div class="total-row"><span class="total-lbl">Line items</span><span class="total-val">{len(st.session_state.line_items) if st.session_state.line_items else len(iord)}</span></div>
          <div class="total-row"><span class="total-lbl">Gross MRP</span><span class="total-val">₹ {gmrp:,.2f}</span></div>
          <div class="total-row"><span class="total-lbl">Distributor Discount</span><span class="total-val neg">− ₹ {disc:,.2f}</span></div>
          <div class="total-row grand"><span class="total-lbl">Net Distributor Landing</span><span class="total-val">₹ {gdist:,.2f}</span></div>
        </div>""", unsafe_allow_html=True)

        if st.button("Proceed to fill Customer Details", key="goto3"):
            st.session_state.ocr_reviewed = True
            st.rerun()

    st.markdown("</div></div>", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# STEP 3 — Dealer / Distributor Details
# ══════════════════════════════════════════════════════════════════════════════
def render_step3():
    if not step_unlocked[2]:
        return

    done = st.session_state.party_confirmed
    st.markdown(f"""
    <div class="step-card"><div class="step-card-header">
      <div class="step-number {'done' if done else ''}">{'✓' if done else '3'}</div>
      <div><div class="step-title">Step 3 — Dealer / Distributor Details</div>
      <div class="step-subtitle">Bill To &amp; Ship To party information</div></div>
    </div><div class="step-body">""", unsafe_allow_html=True)

    st.markdown(
        '<div class="party-section"><div class="party-title">🏢 Dealer / Distributor Details (Bill To)</div>',
        unsafe_allow_html=True,
    )

    bill_name    = st.text_input("Distributor Name", key="bill_na", placeholder="Full name / business name")
    bill_address = st.text_area("Address", key="bill_addr", placeholder="Street, area, city, PIN code", height=80)

    bc1, bc2 = st.columns(2)
    with bc1:
        bill_pno = st.text_input("Party No.", key="bill_pno", placeholder="Digits only")
    with bc2:
        bill_gst = st.text_input("GST No.", key="bill_gst")

    bc3, bc4 = st.columns(2)
    with bc3:
        st.markdown("**Phone** (+91)")
        bill_ph = st.text_input("+91 Phone", key="bill_ph", label_visibility="collapsed", placeholder="10-digit number")
    with bc4:
        st.markdown("**Mobile** (+91)")
        bill_mb = st.text_input("+91 Mobile", key="bill_mb", label_visibility="collapsed", placeholder="10-digit number")

    bc5, bc6 = st.columns(2)
    with bc5:
        bill_st = st.text_input("State", key="bill_st")
    with bc6:
        bill_pan = st.text_input("PAN No.", key="bill_pan")

    st.markdown("</div>", unsafe_allow_html=True)

    bill_to = {
        "party_no":   bill_pno,
        "name":       bill_name,
        "address":    bill_address,
        "phone":      bill_ph,
        "mobile":     bill_mb,
        "state":      bill_st,
        "state_code": "",
        "gst":        bill_gst,
        "pan":        bill_pan,
    }

    st.markdown(
        '<div class="party-section"><div class="party-title">🚚 Customer Details (Ship To)</div>',
        unsafe_allow_html=True,
    )

    ship_name    = st.text_input("Customer Name", key="ship_na", placeholder="Full name / business name")
    ship_address = st.text_area("Address", key="ship_addr", placeholder="Street, area, city, PIN code", height=80)

    sc1, sc2 = st.columns(2)
    with sc1:
        ship_pno = st.text_input("Party No.", key="ship_pno", placeholder="Digits only")
    with sc2:
        ship_gst = st.text_input("GST No.", key="ship_gst")

    sc3, sc4 = st.columns(2)
    with sc3:
        st.markdown("**Phone** (+91)")
        ship_ph = st.text_input("+91 Phone", key="ship_ph", label_visibility="collapsed", placeholder="10-digit number")
    with sc4:
        st.markdown("**Mobile** (+91)")
        ship_mb = st.text_input("+91 Mobile", key="ship_mb", label_visibility="collapsed", placeholder="10-digit number")

    sc5, sc6 = st.columns(2)
    with sc5:
        ship_st = st.text_input("State", key="ship_st")
    with sc6:
        ship_pan = st.text_input("PAN No.", key="ship_pan")

    sm1, _ = st.columns(2)
    with sm1:
        margin_pct_field = st.number_input(
            "Add Margin (%)",
            min_value=0.0, max_value=100.0,
            value=float(st.session_state.get("margin_percent", 0.0)),
            step=0.5,
            key="extra_margin_percent",
        )

    st.markdown("</div>", unsafe_allow_html=True)

    ship_to = {
        "party_no":   ship_pno,
        "name":       ship_name,
        "address":    ship_address,
        "phone":      ship_ph,
        "mobile":     ship_mb,
        "state":      ship_st,
        "state_code": "",
        "gst":        ship_gst,
        "pan":        ship_pan,
    }

    all_errors = validate_party(bill_to, "Bill To") + validate_party(ship_to, "Ship To")

    if all_errors:
        for e in all_errors:
            st.markdown(f'<div class="error-box">⚠️ {e}</div>', unsafe_allow_html=True)

    def do_confirm(b: dict, s: dict):
        st.session_state.margin_percent  = margin_pct_field
        st.session_state.bill_to         = b
        st.session_state.ship_to         = s
        st.session_state.party_confirmed = True

        if not st.session_state.current_quotation_id:
            st.session_state.current_quotation_id = generate_quotation_id()

        qid = st.session_state.current_quotation_id

        with st.spinner("Generating PDF…"):
            try:
                pdf_bytes = build_pdf(
                    st.session_state.quantities,
                    mrp_data,
                    b, s,
                    sku_master,
                    qid,
                    st.session_state.get("line_items", []),
                )
                st.session_state.pdf_bytes = pdf_bytes

                q_path = save_quotation_pdf_to_disk(pdf_bytes, qid)
                st.session_state.quotation_disk_path = q_path

                d_path = save_detection_pdf_to_disk(pdf_bytes, qid)
                st.session_state.detection_disk_path = d_path

                line_items = st.session_state.get("line_items", [])
                if line_items:
                    gmrp  = sum(mrp_data.get(it["sku"],{}).get("mrp",0) * it["qty"] for it in line_items)
                    gdist = sum(mrp_data.get(it["sku"],{}).get("distributor_landing",0) * it["qty"] for it in line_items)
                    n_ln  = len(line_items)
                else:
                    iord  = {sk: q for sk, q in st.session_state.quantities.items() if q > 0}
                    gmrp  = sum(mrp_data.get(sk,{}).get("mrp",0) * q for sk, q in iord.items())
                    gdist = sum(mrp_data.get(sk,{}).get("distributor_landing",0) * q for sk, q in iord.items())
                    n_ln  = len(iord)

                disc = gmrp - gdist

                if not st.session_state.quotation_db_written:
                    db_insert_quotation_log(
                        quotation_id        = qid,
                        ocr_id              = st.session_state.db_ocr_id,
                        pdf_link            = q_path,
                        detection_pdf_link  = d_path,
                        bill_to             = b,
                        ship_to             = s,
                        distributor_name    = b.get("name",""),
                        margin_percent      = margin_pct_field,
                        gross_mrp           = gmrp,
                        distributor_discount= disc,
                        net_taxable         = gdist,
                        line_item_count     = n_ln,
                        ip_address          = _get_local_ip(),
                    )
                    st.session_state.quotation_db_written = True

                    if st.session_state.get("session_log_sr_no"):
                        db_update_session_log_quotation(
                            sr_no        = st.session_state["session_log_sr_no"],
                            quotation_id = qid,
                            ocr_id       = st.session_state.db_ocr_id,
                            latitude     = st.session_state.get("user_latitude"),
                            longitude    = st.session_state.get("user_longitude"),
                        )

                try:
                    append_log_entry(qid, q_path, b, s, b.get("name",""), margin_pct_field)
                except Exception:
                    pass

                st.rerun()

            except Exception as exc:
                st.error(f"PDF / DB error: {exc}")

    if is_party_complete(bill_to) and is_party_complete(ship_to) and not all_errors:
        st.markdown('<div class="success-box">✅ All fields filled — generating PDF automatically…</div>',
                    unsafe_allow_html=True)
        if not st.session_state.party_confirmed:
            do_confirm(bill_to, ship_to)

    btn_label = "Generate Quotation"
    if not all_errors:
        if st.button(btn_label, key="confirm"):
            do_confirm(bill_to, ship_to)
    else:
        st.markdown(
            "<div style='font-size:12px;color:#92400E;margin-top:4px;'>"
            "Fix the validation errors above, then confirm.</div>", unsafe_allow_html=True)
        if st.button("⚡ Skip Validation & Generate PDF", key="confirm_force"):
            do_confirm(bill_to, ship_to)

    st.markdown("</div></div>", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# STEP 4 — Preview & Download
#
# FIX: session_end_dt is now stamped reliably via Python on_click callbacks
#      on st.download_button and st.button — no JS bridge, no hidden input,
#      no zero-height iframe needed.
#
#      st.download_button fires its on_click callback on the SAME server-side
#      rerun as the download, so db_update_session_log_end() is called directly
#      in Python, guaranteed every time.
# ══════════════════════════════════════════════════════════════════════════════
def render_step4():
    if not step_unlocked[3]:
        return

    quotation_id = st.session_state.get("current_quotation_id", "")

    st.markdown("""
    <div class="step-card"><div class="step-card-header">
      <div class="step-number done">✓</div>
      <div><div class="step-title">Step 4 — Preview &amp; Download Quotation</div>
      <div class="step-subtitle">Review your Sales Quotation PDF before downloading</div></div>
    </div><div class="step-body">""", unsafe_allow_html=True)

    line_items = st.session_state.get("line_items", [])
    if line_items:
        items_for_total = line_items
        gmrp  = sum(mrp_data.get(it["sku"], {}).get("mrp", 0) * it["qty"] for it in items_for_total)
        gdist = sum(mrp_data.get(it["sku"], {}).get("distributor_landing", 0) * it["qty"] for it in items_for_total)
        n_lines = len(items_for_total)
    else:
        iord   = {s: q for s, q in st.session_state.quantities.items() if q > 0}
        gmrp   = sum(mrp_data.get(s, {}).get("mrp", 0) * q        for s, q in iord.items())
        gdist  = sum(mrp_data.get(s, {}).get("distributor_landing", 0) * q for s, q in iord.items())
        n_lines = len(iord)
    disc = gmrp - gdist

    if quotation_id:
        lat    = st.session_state.get("user_latitude")
        lng    = st.session_state.get("user_longitude")
        geo_st = st.session_state.get("geo_status", "pending")

        if geo_st == "ok" and lat is not None and lng is not None:
            maps_url = f"https://maps.google.com/?q={lat:.6f},{lng:.6f}"
            geo_info_html = (
                f'<div style="font-size:11px;color:#065F46;margin-top:4px;">'
                f'📍 Location recorded: '
                f'<a href="{maps_url}" target="_blank" rel="noopener noreferrer" '
                f'style="color:#065F46;font-family:\'JetBrains Mono\',monospace;">'
                f'{lat:.5f}, {lng:.5f}</a></div>'
            )
        else:
            geo_info_html = (
                '<div style="font-size:11px;color:#92400E;margin-top:4px;">'
                '⚠️ Location not recorded for this quotation</div>'
            )

        # Session-end status — now driven purely by Python state
        session_ended = st.session_state.get("session_end_stamped", False)
        if session_ended:
            session_end_badge = (
                '<span class="session-end-badge stamped">'
                '✅ Session closed — download / share confirmed'
                '</span>'
            )
        else:
            session_end_badge = (
                '<span class="session-end-badge pending">'
                '⏳ Session end stamped on first download or WhatsApp share'
                '</span>'
            )

        st.markdown(f"""
        <div style="background:#1A1A1A;border-radius:9px;padding:12px 16px;margin-bottom:12px;
            display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:8px;">
          <div>
            <div style="font-size:10px;font-weight:700;color:#888;text-transform:uppercase;
                letter-spacing:.8px;margin-bottom:2px;">Quotation ID</div>
            <div style="font-family:'JetBrains Mono',monospace;font-size:13px;
                font-weight:700;color:#FFFFFF;letter-spacing:1px;">{quotation_id}</div>
            {geo_info_html}
            {session_end_badge}
          </div>
          <div style="background:#C0211F;color:white;font-size:10px;font-weight:700;
              padding:4px 12px;border-radius:6px;">GENERATED</div>
        </div>
        """, unsafe_allow_html=True)

    if st.session_state.pdf_bytes:
        # ── PDF Preview ───────────────────────────────────────────────────────
        st.markdown("""
        <div style="margin:20px 0 8px;">
          <div style="font-size:13px;font-weight:700;color:#1A1A1A;margin-bottom:4px;">
            📄 Document Preview
          </div>
        </div>""", unsafe_allow_html=True)

        b64 = base64.b64encode(st.session_state.pdf_bytes).decode()
        pdf_display = f"""
        <div class="pdf-preview-wrap">
          <iframe
            src="data:application/pdf;base64,{b64}#toolbar=1&navpanes=1&scrollbar=1&view=FitH"
            width="100%"
            height="620px"
            style="border:none;display:block;"
            type="application/pdf"
          >
            <p style="color:white;padding:20px;text-align:center;">
              Your browser does not support embedded PDF preview.<br/>
              Please use the download button below.
            </p>
          </iframe>
        </div>"""
        st.markdown(pdf_display, unsafe_allow_html=True)
        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

        pdf_filename = f"{quotation_id}.pdf" if quotation_id else "sintex_quotation.pdf"

        # ── Totals summary strip ──────────────────────────────────────────────
        st.markdown(f"""
        <div class="totals-box" style="margin-bottom:20px;">
          <div class="total-row">
            <span class="total-lbl">Line Items</span>
            <span class="total-val">{n_lines}</span>
          </div>
          <div class="total-row">
            <span class="total-lbl">Gross MRP</span>
            <span class="total-val">₹ {gmrp:,.2f}</span>
          </div>
          <div class="total-row">
            <span class="total-lbl">Distributor Discount</span>
            <span class="total-val neg">− ₹ {disc:,.2f}</span>
          </div>
          <div class="total-row grand">
            <span class="total-lbl">Net Taxable Value</span>
            <span class="total-val">₹ {gdist:,.2f}</span>
          </div>
        </div>
        """, unsafe_allow_html=True)

        # ── Action Buttons ────────────────────────────────────────────────────
        # THREE COLUMNS:
        #   Col 1 — st.download_button  (PDF download, triggers session_end stamp)
        #   Col 2 — st.button           (Open WhatsApp, triggers session_end stamp)
        #   Col 3 — st.button           (New Quotation, resets session)

        col_dl, col_wa, col_new = st.columns(3)

        with col_dl:
            # st.download_button fires on_click on the SAME rerun as the download.
            # This is the reliable Python-native replacement for the JS bridge.
            st.download_button(
                label="📥  Save PDF",
                data=st.session_state.pdf_bytes,
                file_name=pdf_filename,
                mime="application/pdf",
                key="dl_pdf_btn",
                on_click=_stamp_session_end_once,
                use_container_width=True,
            )

        with col_wa:
            wa_text = (
                f"Sintex BAPL Sales Quotation {quotation_id} — "
                "Please find the attached PDF for your reference."
            )
            wa_url  = f"https://wa.me/?text={requests.utils.quote(wa_text)}"

            # st.button fires on_click on the SAME rerun as the button press.
            # The user then opens WhatsApp in a new tab via the JS redirect below.
            def _wa_click():
                _stamp_session_end_once()
                # Store the URL so we can inject the redirect JS after rerun
                st.session_state["_wa_open_url"] = wa_url

            st.markdown('<div class="s4-btn-wa">', unsafe_allow_html=True)
            if st.button("📲  Open WhatsApp", key="wa_btn", use_container_width=True,
                         on_click=_wa_click):
                pass  # action handled in on_click
            st.markdown('</div>', unsafe_allow_html=True)

            # If the WA button was just clicked, open the URL in a new tab
            if st.session_state.get("_wa_open_url"):
                _open_url = st.session_state.pop("_wa_open_url")
                wa_redirect_js = f"""<!DOCTYPE html>
<html><head></head><body>
<script>
window.parent.open({json.dumps(_open_url)}, '_blank', 'noopener,noreferrer');
</script>
</body></html>"""
                components.html(wa_redirect_js, height=0, scrolling=False)

        with col_new:
            st.markdown('<div class="s4-btn-new">', unsafe_allow_html=True)
            if st.button("🆕  New Quotation", key="new_quote_btn", use_container_width=True):
                _reset_session()
            st.markdown('</div>', unsafe_allow_html=True)

    st.markdown("</div></div>", unsafe_allow_html=True)


def _reset_session():
    """Reset all session state for a fresh quotation."""
    for k in ["image_bytes", "pdf_bytes"]:
        st.session_state[k] = None
    for k in ["quantities", "bill_to", "ship_to", "ocr_meta"]:
        st.session_state[k] = {}
    for k in ["match_log", "ocr_grid", "ocr_extracted", "line_items"]:
        st.session_state[k] = []
    st.session_state["ocr_done"]                 = False
    st.session_state["ocr_reviewed"]             = False
    st.session_state["party_confirmed"]          = False
    st.session_state["image_submitted"]          = False
    st.session_state["upload_key"]              += 1
    st.session_state["image_rotation"]           = 0
    st.session_state["capture_mode"]             = "camera"
    st.session_state["image_saved_to_disk"]      = False
    st.session_state["log_saved_to_disk"]        = False
    st.session_state["current_quotation_id"]     = ""
    st.session_state["distributor_name"]         = ""
    st.session_state["margin_percent"]           = 0.0
    st.session_state["db_ocr_id"]               = None
    st.session_state["ocr_db_written"]           = False
    st.session_state["quotation_db_written"]     = False
    st.session_state["image_disk_path"]          = ""
    st.session_state["quotation_disk_path"]      = ""
    st.session_state["detection_disk_path"]      = ""
    st.session_state["session_log_sr_no"]        = None
    st.session_state["session_log_written"]      = False
    st.session_state["session_start_dt"]         = datetime.now()
    st.session_state["session_end_stamped"]      = False
    # NOTE: Geolocation intentionally preserved across quotations.
    st.rerun()


# ── Render all steps ───────────────────────────────────────────────────────────
render_step1()
render_step2()
render_step3()
render_step4()

st.markdown("""
<div style="text-align:center;padding:32px 0 8px;font-size:11px;color:#AAA;">
  Sintex BAPL Limited &nbsp;·&nbsp; CPVC / UPVC Quotation System
</div>""", unsafe_allow_html=True)