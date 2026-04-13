"""
Sintex BAPL – CPVC Pipes Quotation Generator
Self-contained Streamlit app. No FastAPI / backend required.

Dependencies:
    pip install streamlit pandas openpyxl pillow requests reportlab

Run:
    streamlit run app.py
"""

import io
import os
import copy
import base64
import textwrap
from datetime import date

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

# ─── optional: Azure OCR (requests only used here) ───────────────────────────
try:
    import requests as _requests
    _HAS_REQUESTS = True
except ImportError:
    _HAS_REQUESTS = False

# ─── optional: PDF generation ────────────────────────────────────────────────
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, HRFlowable)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    _HAS_REPORTLAB = True
except ImportError:
    _HAS_REPORTLAB = False

# ─────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Sintex BAPL – Quotation Generator",
    page_icon="🔧",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ─────────────────────────────────────────────────────────────────────────────
# GLOBAL CSS
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@300;400;500;600;700&family=IBM+Plex+Mono:wght@400;500&display=swap');
:root {
  --navy:   #0A2342;
  --blue:   #1565C0;
  --sky:    #1E88E5;
  --teal:   #00796B;
  --gold:   #F9A825;
  --danger: #C62828;
  --border: #DEE3EC;
  --surface:#F4F6FA;
  --text:   #1A1F36;
  --muted:  #5A6880;
  --radius: 10px;
}
html, body, [class*="css"] {
  font-family: 'IBM Plex Sans', sans-serif !important;
  background: #F0F2F8 !important;
  color: var(--text);
}
#MainMenu, footer, header { visibility: hidden; }
.block-container { padding: 1.2rem 1.5rem 4rem !important; max-width: 1400px !important; }

/* ── Cards ── */
.card {
  background: #fff;
  border: 1px solid var(--border);
  border-radius: var(--radius);
  padding: 20px 24px;
  margin-bottom: 18px;
  box-shadow: 0 1px 4px rgba(10,35,66,.07);
}
.card-title {
  font-size: 13px;
  font-weight: 700;
  text-transform: uppercase;
  letter-spacing: .7px;
  color: var(--navy);
  margin-bottom: 14px;
  display: flex;
  align-items: center;
  gap: 8px;
}
.step-badge {
  display: inline-flex;
  align-items: center;
  justify-content: center;
  background: var(--navy);
  color: #fff;
  border-radius: 50%;
  width: 24px; height: 24px;
  font-size: 11px;
  font-weight: 700;
  flex-shrink: 0;
}
.step-badge.done { background: var(--teal); }

/* ── App header ── */
.app-header {
  background: linear-gradient(135deg, #0A2342 0%, #1565C0 100%);
  border-radius: 12px;
  padding: 18px 26px;
  display: flex; align-items: center; gap: 18px;
  margin-bottom: 22px;
  box-shadow: 0 4px 16px rgba(10,35,66,.25);
}
.app-header h1 { font-size: 20px; font-weight: 700; color: #fff; margin: 0; }
.app-header p  { font-size: 12px; color: rgba(255,255,255,.65); margin: 3px 0 0; }

/* ── SKU table ── */
.sku-table { width: 100%; border-collapse: collapse; font-size: 11.5px; }
.sku-table th {
  background: var(--navy); color: #fff;
  padding: 6px 8px; text-align: center;
  font-weight: 600; font-size: 10.5px;
  border: 1px solid rgba(255,255,255,.15);
  white-space: nowrap;
}
.sku-table th.row-hdr { text-align: left; min-width: 180px; }
.sku-table tr:nth-child(even) td { background: #F7F9FC; }
.sku-table td {
  padding: 5px 8px; border: 1px solid var(--border);
  text-align: center; vertical-align: middle;
}
.sku-table td.row-label {
  text-align: left; font-weight: 500;
  color: var(--text); white-space: nowrap;
}
.sku-table td.section-hdr {
  background: #E8EDF6; font-weight: 700;
  color: var(--navy); font-size: 11px;
  text-align: left; padding: 4px 8px;
}
.sku-code { font-family: 'IBM Plex Mono', monospace; font-size: 10px; color: var(--blue); }

/* ── Input cells ── */
.stNumberInput > div > div > input {
  border-radius: 6px !important;
  border: 1.5px solid var(--border) !important;
  font-family: 'IBM Plex Mono', monospace !important;
  font-size: 13px !important;
  padding: 4px 6px !important;
  text-align: center !important;
}

/* ── Buttons ── */
.stButton > button {
  font-family: 'IBM Plex Sans', sans-serif !important;
  font-weight: 600 !important; font-size: 14px !important;
  border-radius: 8px !important; border: none !important;
  background: var(--navy) !important; color: #fff !important;
  padding: 10px 20px !important;
  box-shadow: 0 2px 8px rgba(10,35,66,.2) !important;
  transition: all .15s !important;
}
.stButton > button:hover { background: var(--sky) !important; transform: translateY(-1px) !important; }

/* ── Tabs ── */
.stTabs [data-baseweb="tab-list"] {
  background: var(--surface); padding: 4px; border-radius: 10px;
  border: 1px solid var(--border); gap: 4px;
}
.stTabs [data-baseweb="tab"] {
  border-radius: 8px !important; font-weight: 600 !important;
  font-size: 13px !important; color: var(--muted) !important;
}
.stTabs [aria-selected="true"] {
  background: var(--navy) !important; color: #fff !important;
}

/* ── Info / success pills ── */
.pill {
  display: inline-block; padding: 2px 10px;
  border-radius: 20px; font-size: 11px; font-weight: 600;
}
.pill-blue   { background: #EBF3FF; color: var(--blue); }
.pill-green  { background: #E8F5E9; color: var(--teal); }
.pill-gold   { background: #FFF8E1; color: #7B5B00; }

/* ── Totals box ── */
.totals-row {
  display: flex; justify-content: space-between; align-items: center;
  padding: 9px 0; border-bottom: 1px solid var(--border);
  font-size: 14px;
}
.totals-row:last-child { border: none; padding-top: 12px; }
.totals-row.net { font-weight: 700; font-size: 16px; color: var(--navy); }
.totals-row .val { font-family: 'IBM Plex Mono', monospace; font-weight: 600; }
.totals-row .val.neg { color: var(--danger); }

/* ── Party details ── */
.party-box {
  border: 1px solid var(--border); border-radius: 8px;
  padding: 14px 16px; background: var(--surface);
}
.party-box h4 { margin: 0 0 10px; font-size: 12px; font-weight: 700;
  text-transform: uppercase; letter-spacing: .5px; color: var(--navy); }
.party-row { display: flex; gap: 6px; margin-bottom: 5px; font-size: 12.5px; }
.party-lbl { color: var(--muted); min-width: 80px; flex-shrink: 0; }
.party-val { font-weight: 500; color: var(--text); word-break: break-word; }

/* ── stTextInput labels ── */
label { font-size: 12px !important; font-weight: 600 !important; color: var(--muted) !important; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# PATHS – update these to your local paths
# ─────────────────────────────────────────────────────────────────────────────
_HERE = os.path.dirname(os.path.abspath(__file__))
MRP_CSV_PATH   = os.path.join(_HERE, "MRP_State_chhattisghar.csv")
ZSD_CSV_PATH   = os.path.join(_HERE, "ZSD_CUST.csv")
SKU_XLSX_PATH  = os.path.join(_HERE, "Sample form for Product list.xlsx")

# ─────────────────────────────────────────────────────────────────────────────
# DATA LOADING (cached)
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_data
def load_mrp():
    df = pd.read_csv(MRP_CSV_PATH)
    df["MRP_clean"] = (
        df["MRP(ZPR1-933)"]
        .astype(str).str.replace(",", "", regex=False)
        .pipe(pd.to_numeric, errors="coerce")
    )
    df["Distributor Landing"] = pd.to_numeric(
        df["Distributor Landing"].astype(str).str.replace(",", "", regex=False),
        errors="coerce",
    )
    # Build lookup: Material Number → row dict
    lookup = {}
    for _, row in df.iterrows():
        mat = str(row["Material Number"]).strip()
        lookup[mat] = row.to_dict()
    return lookup

@st.cache_data
def load_zsd():
    df = pd.read_csv(ZSD_CSV_PATH, encoding="latin1")
    return df

@st.cache_data
def load_sku_sheets():
    """
    Returns a dict:
      sheet_name → {
        "label": str,
        "col_ids":   [str, ...],          # e.g. "15MM", "20MM" …
        "col_labels":[str, ...],          # e.g. '½"', '¾"' …
        "rows": [
          {"label": str, "section": str, "skus": [str|None, ...]},
          ...
        ]
      }
    """
    wb  = load_workbook(SKU_XLSX_PATH, read_only=True)
    out = {}
    for sname in wb.sheetnames:
        ws   = wb[sname]
        data = [row for row in ws.iter_rows(values_only=True)
                if any(c is not None for c in row)]
        if len(data) < 3:
            continue
        # Row 0: title row (skip)
        # Row 1: col ids
        # Row 2: col labels (inch notation)
        col_ids    = [str(c).strip() if c is not None else "" for c in data[1][2:]]
        col_labels = [str(c).strip() if c is not None else "" for c in data[2][2:]]
        # Trim trailing empty
        while col_ids and col_ids[-1] == "":
            col_ids.pop(); col_labels.pop()

        rows   = []
        section = "General"
        for raw in data[3:]:
            if not any(c is not None for c in raw):
                continue
            label = str(raw[0]).strip() if raw[0] is not None else ""
            # Section separator row (e.g. "FITTINGS", "Fittings")
            if label.upper() in ("FITTINGS", "PIPES", "FITTING SCH 80") and all(
                c is None for c in raw[2:]
            ):
                section = label.title()
                continue
            if not label:
                continue
            skus = []
            for i in range(len(col_ids)):
                idx = i + 2
                val = raw[idx] if idx < len(raw) else None
                skus.append(str(val).strip() if val is not None else None)
            rows.append({"label": label, "section": section, "skus": skus})

        out[sname] = {
            "label":      sname,
            "col_ids":    col_ids,
            "col_labels": col_labels,
            "rows":       rows,
        }
    return out


# ─────────────────────────────────────────────────────────────────────────────
# AZURE OCR  (numbers only, returns list of number strings)
# Tries Computer Vision Read API first; falls back to Document Intelligence
# (Form Recognizer) prebuilt-read if Computer Vision returns 401/404.
# ─────────────────────────────────────────────────────────────────────────────
def _extract_numbers_from_lines(lines: list) -> list:
    """Pull pure numeric strings out of a list of line-text strings."""
    numbers = []
    for txt in lines:
        clean = txt.strip().replace(",", "").replace(" ", "")
        if clean.replace(".", "").isdigit():
            numbers.append(clean)
    return numbers


def _ocr_computer_vision(image_bytes: bytes, endpoint: str, key: str) -> list:
    """Azure Computer Vision v3.2 Read API."""
    import time
    url     = endpoint.rstrip("/") + "/vision/v3.2/read/analyze"
    headers = {"Ocp-Apim-Subscription-Key": key, "Content-Type": "application/octet-stream"}
    resp    = _requests.post(url, headers=headers, data=image_bytes, timeout=30)
    resp.raise_for_status()
    op_url  = resp.headers["Operation-Location"]
    result  = {}
    for _ in range(20):
        time.sleep(1.5)
        r      = _requests.get(op_url, headers={"Ocp-Apim-Subscription-Key": key}, timeout=15)
        result = r.json()
        if result.get("status") == "succeeded":
            break
    lines = [
        line["text"]
        for page in result.get("analyzeResult", {}).get("readResults", [])
        for line in page.get("lines", [])
    ]
    return _extract_numbers_from_lines(lines)


def _ocr_document_intelligence(image_bytes: bytes, endpoint: str, key: str) -> list:
    """Azure Document Intelligence (Form Recognizer) prebuilt-read model."""
    import time
    url     = endpoint.rstrip("/") + "/formrecognizer/documentModels/prebuilt-read:analyze?api-version=2023-07-31"
    headers = {"Ocp-Apim-Subscription-Key": key, "Content-Type": "application/octet-stream"}
    resp    = _requests.post(url, headers=headers, data=image_bytes, timeout=30)
    resp.raise_for_status()
    op_url  = resp.headers["Operation-Location"]
    result  = {}
    for _ in range(20):
        time.sleep(1.5)
        r      = _requests.get(op_url, headers={"Ocp-Apim-Subscription-Key": key}, timeout=15)
        result = r.json()
        if result.get("status") == "succeeded":
            break
    lines = [
        line["content"]
        for page in result.get("analyzeResult", {}).get("pages", [])
        for line in page.get("lines", [])
    ]
    return _extract_numbers_from_lines(lines)


def run_azure_ocr(image_bytes: bytes, azure_endpoint: str, azure_key: str) -> list:
    """
    Tries Document Intelligence first, then falls back to Computer Vision.
    Returns a flat list of detected number strings.
    """
    if not _HAS_REQUESTS:
        return []
    try:
        return _ocr_document_intelligence(image_bytes, azure_endpoint, azure_key)
    except Exception as di_err:
        di_status = getattr(getattr(di_err, "response", None), "status_code", None)
        if di_status in (401, 403, 404):
            # Not a Document Intelligence resource — try Computer Vision instead
            return _ocr_computer_vision(image_bytes, azure_endpoint, azure_key)
        raise  # Any other error (network, timeout, etc.) — re-raise as before


# ─────────────────────────────────────────────────────────────────────────────
# PDF GENERATION  (reportlab)
# ─────────────────────────────────────────────────────────────────────────────
def generate_pdf(
    sheet_name: str,
    sheet_data: dict,
    qty_map: dict,           # {(row_idx, col_idx): qty}
    mrp_lookup: dict,
    bill_to: dict,
    ship_to: dict,
) -> bytes:
    """
    Returns PDF bytes for the given filled quotation.
    """
    buf    = io.BytesIO()
    doc    = SimpleDocTemplate(buf, pagesize=A4,
                               leftMargin=10*mm, rightMargin=10*mm,
                               topMargin=12*mm, bottomMargin=12*mm)
    styles = getSampleStyleSheet()
    navy   = colors.HexColor("#0A2342")
    sky    = colors.HexColor("#1565C0")
    light  = colors.HexColor("#EBF3FF")
    gold   = colors.HexColor("#F9A825")

    title_style = ParagraphStyle("title", fontName="Helvetica-Bold",
                                 fontSize=14, textColor=colors.white,
                                 alignment=TA_CENTER)
    sub_style   = ParagraphStyle("sub",   fontName="Helvetica",
                                 fontSize=8, textColor=colors.white,
                                 alignment=TA_CENTER)
    normal_sm   = ParagraphStyle("nsm",   fontName="Helvetica", fontSize=7.5)
    bold_sm     = ParagraphStyle("bsm",   fontName="Helvetica-Bold", fontSize=7.5)
    cell_style  = ParagraphStyle("cell",  fontName="Helvetica", fontSize=7,
                                 alignment=TA_CENTER)
    header_cell = ParagraphStyle("hcell", fontName="Helvetica-Bold", fontSize=7,
                                 textColor=colors.white, alignment=TA_CENTER)

    story = []

    # ── 1. Header banner ─────────────────────────────────────────────────────
    hdr_data = [[
        Paragraph("SINTEX BAPL LIMITED", title_style),
        Paragraph("Kutesar Road, Raipur, Chhattisgarh – 492101<br/>GSTIN: 22AADCB1921F1ZE", sub_style),
    ]]
    hdr_table = Table(hdr_data, colWidths=["40%", "60%"])
    hdr_table.setStyle(TableStyle([
        ("BACKGROUND",   (0,0), (-1,-1), navy),
        ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",   (0,0), (-1,-1), 8),
        ("BOTTOMPADDING",(0,0), (-1,-1), 8),
        ("LEFTPADDING",  (0,0), (-1,-1), 10),
    ]))
    story.append(hdr_table)
    story.append(Spacer(1, 4*mm))

    # ── 2. Title + date ──────────────────────────────────────────────────────
    title_data = [[
        Paragraph(f"<b>QUOTATION</b>", ParagraphStyle("qt", fontName="Helvetica-Bold",
                  fontSize=13, textColor=navy)),
        Paragraph(f"<b>Date:</b> {date.today().strftime('%d-%m-%Y')}<br/>"
                  f"<b>Sheet:</b> {sheet_name}",
                  ParagraphStyle("qd", fontName="Helvetica", fontSize=8,
                                 alignment=TA_RIGHT)),
    ]]
    t = Table(title_data, colWidths=["60%", "40%"])
    t.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"MIDDLE"),
                           ("LINEBELOW",(0,0),(-1,-1),0.5,navy)]))
    story.append(t)
    story.append(Spacer(1, 4*mm))

    # ── 3. Bill-to / Ship-to ─────────────────────────────────────────────────
    def party_lines(d: dict, title: str):
        lines = [f"<b>{title}</b>"]
        for k, v in d.items():
            if v:
                lines.append(f"<b>{k}:</b> {v}")
        return "<br/>".join(lines)

    pt_data = [[
        Paragraph(party_lines(bill_to, "BILL TO PARTY"),
                  ParagraphStyle("pt", fontName="Helvetica", fontSize=7.5,
                                 leading=11)),
        Paragraph(party_lines(ship_to, "SHIP TO PARTY"),
                  ParagraphStyle("pt", fontName="Helvetica", fontSize=7.5,
                                 leading=11)),
    ]]
    pt = Table(pt_data, colWidths=["50%", "50%"])
    pt.setStyle(TableStyle([
        ("BOX",         (0,0),(-1,-1), 0.5, navy),
        ("INNERGRID",   (0,0),(-1,-1), 0.5, colors.HexColor("#C5D0E0")),
        ("VALIGN",      (0,0),(-1,-1), "TOP"),
        ("TOPPADDING",  (0,0),(-1,-1), 6),
        ("BOTTOMPADDING",(0,0),(-1,-1), 6),
        ("LEFTPADDING", (0,0),(-1,-1), 8),
        ("BACKGROUND",  (0,0),(-1,-1), colors.HexColor("#F4F6FA")),
    ]))
    story.append(pt)
    story.append(Spacer(1, 5*mm))

    # ── 4. Quotation line table ───────────────────────────────────────────────
    col_ids    = sheet_data["col_ids"]
    col_labels = sheet_data["col_labels"]
    rows       = sheet_data["rows"]

    # Collect active columns (those with at least 1 entry)
    active_cols = [i for i, cid in enumerate(col_ids) if cid]

    tbl_header1 = ["#", "Item / Description", "SKU Code", "Size", "MRP (₹)", "Qty", "Rate (₹)", "Amount (₹)"]
    tbl_rows    = [tbl_header1]
    
    grand_total = 0.0
    sno         = 0
    current_sec = None

    for ri, row in enumerate(rows):
        sec = row["section"]
        if sec != current_sec:
            current_sec = sec
            tbl_rows.append([Paragraph(f"<b>{sec.upper()}</b>",
                              ParagraphStyle("s", fontName="Helvetica-Bold", fontSize=7,
                                             textColor=colors.white)),
                             "", "", "", "", "", "", ""])
        for ci in active_cols:
            sku = row["skus"][ci] if ci < len(row["skus"]) else None
            if not sku or sku in ("-", "None", ""):
                continue
            qty = qty_map.get((ri, ci), 0)
            if qty <= 0:
                continue
            sno += 1
            mrp_info = mrp_lookup.get(sku, {})
            mrp      = mrp_info.get("MRP_clean") or 0.0
            rate     = mrp_info.get("Distributor Landing") or mrp
            amount   = round(rate * qty, 2)
            grand_total += amount
            tbl_rows.append([
                str(sno),
                Paragraph(row["label"], ParagraphStyle("rl", fontName="Helvetica", fontSize=7)),
                Paragraph(f'<font name="Courier" size="6.5">{sku}</font>',
                          ParagraphStyle("sk", fontName="Helvetica", fontSize=7)),
                col_labels[ci] if ci < len(col_labels) else col_ids[ci],
                f"{mrp:,.2f}",
                str(qty),
                f"{rate:,.2f}",
                f"{amount:,.2f}",
            ])

    # Grand total row
    tbl_rows.append(["", "", "", "", "", "", Paragraph("<b>GRAND TOTAL</b>",
                     ParagraphStyle("gt", fontName="Helvetica-Bold", fontSize=8)),
                     Paragraph(f"<b>₹ {grand_total:,.2f}</b>",
                     ParagraphStyle("gv", fontName="Helvetica-Bold", fontSize=8,
                                    alignment=TA_RIGHT))])

    col_widths = [8*mm, 55*mm, 38*mm, 16*mm, 18*mm, 12*mm, 18*mm, 22*mm]
    lt = Table(tbl_rows, colWidths=col_widths, repeatRows=1)

    n = len(tbl_rows)
    style_cmds = [
        ("BACKGROUND",   (0,0), (-1,0),  navy),
        ("TEXTCOLOR",    (0,0), (-1,0),  colors.white),
        ("FONTNAME",     (0,0), (-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",     (0,0), (-1,0),  7.5),
        ("ALIGN",        (0,0), (-1,-1), "CENTER"),
        ("ALIGN",        (1,1), (1,-1),  "LEFT"),
        ("ALIGN",        (2,1), (2,-1),  "LEFT"),
        ("FONTSIZE",     (0,1), (-1,-1), 7),
        ("ROWBACKGROUNDS",(0,1),(-1,-2), [colors.white, colors.HexColor("#F4F8FF")]),
        ("GRID",         (0,0), (-1,-1), 0.35, colors.HexColor("#C5D0E0")),
        ("TOPPADDING",   (0,0), (-1,-1), 3),
        ("BOTTOMPADDING",(0,0), (-1,-1), 3),
        ("BACKGROUND",   (0,n-1),(-1,n-1), colors.HexColor("#EBF3FF")),
        ("LINEABOVE",    (0,n-1),(-1,n-1), 1, navy),
    ]
    # Mark section rows
    for idx, r in enumerate(tbl_rows):
        if idx == 0:
            continue
        if isinstance(r[0], Paragraph) or (isinstance(r[0], str) and r[1] == "" and r[2] == ""):
            if idx > 0 and r[1] == "":
                style_cmds += [
                    ("BACKGROUND",  (0,idx),(-1,idx), sky),
                    ("SPAN",        (0,idx),(-1,idx)),
                    ("TEXTCOLOR",   (0,idx),(-1,idx), colors.white),
                ]

    lt.setStyle(TableStyle(style_cmds))
    story.append(lt)

    # ── 5. Footer ────────────────────────────────────────────────────────────
    story.append(Spacer(1, 6*mm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=navy))
    story.append(Spacer(1, 2*mm))
    story.append(Paragraph(
        "This is a computer-generated quotation. Prices are subject to change without notice. "
        "Taxes applicable as per prevailing rates.",
        ParagraphStyle("footer", fontName="Helvetica", fontSize=7, textColor=colors.grey,
                       alignment=TA_CENTER)
    ))

    doc.build(story)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# SESSION STATE
# ─────────────────────────────────────────────────────────────────────────────
DEFAULTS = {
    "step":              1,
    "active_sheet":      None,
    "image_bytes":       None,
    "ocr_numbers":       [],
    "ocr_done":          False,
    "qty_map":           {},
    "qty_keys_seeded":   False,
    "bill_to":           {},
    "ship_to":           {},
    "pdf_bytes":         None,
    "azure_endpoint":    "",
    "azure_key":         "",
    "zsd_customer_code": "",
}
for k, v in DEFAULTS.items():
    if k not in st.session_state:
        st.session_state[k] = v


# ─────────────────────────────────────────────────────────────────────────────
# LOAD DATA
# ─────────────────────────────────────────────────────────────────────────────
mrp_lookup  = load_mrp()
zsd_df      = load_zsd()
sku_sheets  = load_sku_sheets()
sheet_names = list(sku_sheets.keys())


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def badge(n, done=False):
    cls = "step-badge done" if done else "step-badge"
    return f'<span class="{cls}">{n}</span>'

def pill(text, kind="blue"):
    return f'<span class="pill pill-{kind}">{text}</span>'

def party_html(d, title):
    rows = "".join(
        f'<div class="party-row"><span class="party-lbl">{k}</span>'
        f'<span class="party-val">{v or "—"}</span></div>'
        for k, v in d.items()
    )
    return f'<div class="party-box"><h4>{title}</h4>{rows}</div>'


# ─────────────────────────────────────────────────────────────────────────────
# APP HEADER
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="app-header">
  <div>
    <h1>🔧 Sintex BAPL – Quotation Generator</h1>
    <p>CPVC / UPVC Pipes &amp; Fittings · Chhattisgarh Price List</p>
  </div>
</div>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# STEP INDICATOR
# ─────────────────────────────────────────────────────────────────────────────
step = st.session_state.step
cols_steps = st.columns(5)
STEP_LABELS = [
    "1  Select Form",
    "2  Capture & OCR",
    "3  Fill Quantities",
    "4  Party Details",
    "5  Download",
]
for i, (col, lbl) in enumerate(zip(cols_steps, STEP_LABELS), 1):
    done   = step > i
    active = step == i
    bg = "#0A2342" if active else ("#00796B" if done else "#DEE3EC")
    tc = "#fff" if (active or done) else "#5A6880"
    col.markdown(f"""
    <div style="background:{bg};color:{tc};border-radius:8px;
                padding:8px 12px;text-align:center;font-size:12px;font-weight:600;">
      {'✓ ' if done else ''}{lbl}
    </div>""", unsafe_allow_html=True)

st.markdown("<div style='height:18px'></div>", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 1 — Select Form Sheet + Preview
# ═══════════════════════════════════════════════════════════════════════════════
if step == 1:
    st.markdown(f"""
    <div class="card">
      <div class="card-title">{badge(1)} Select Product Form Sheet</div>
    </div>""", unsafe_allow_html=True)

    chosen = st.selectbox(
        "Choose which form (sheet) to fill:",
        sheet_names,
        index=sheet_names.index(st.session_state.active_sheet)
               if st.session_state.active_sheet in sheet_names else 0,
        key="sheet_select",
    )
    # Reset OCR + qty state when sheet changes
    if st.session_state.active_sheet != chosen:
        st.session_state.active_sheet    = chosen
        st.session_state.ocr_done        = False
        st.session_state.ocr_numbers     = []
        st.session_state.qty_map         = {}
        st.session_state.qty_keys_seeded = False
        for k in list(st.session_state.keys()):
            if k.startswith("qty_"):
                del st.session_state[k]
    st.session_state.active_sheet = chosen

    sheet          = sku_sheets[chosen]
    col_ids        = sheet["col_ids"]
    col_labels     = sheet["col_labels"]
    rows           = sheet["rows"]
    active_col_idx = [i for i, c in enumerate(col_ids) if c]

    st.markdown(f"""
    <div class="card">
      <div class="card-title">{badge(1)} Form Preview — {chosen}</div>
    </div>""", unsafe_allow_html=True)

    header_cells = "".join(
        f'<th title="{col_ids[i]}">{col_labels[i] if i < len(col_labels) else col_ids[i]}</th>'
        for i in active_col_idx
    )
    table_html = f"""
    <div style="overflow-x:auto;max-height:400px;overflow-y:auto;">
    <table class="sku-table">
      <thead>
        <tr>
          <th class="row-hdr">Item</th>
          <th style="min-width:70px;">Section</th>
          {header_cells}
        </tr>
      </thead>
      <tbody>
    """
    current_sec = None
    for row in rows:
        if row["section"] != current_sec:
            current_sec = row["section"]
            span = 2 + len(active_col_idx)
            table_html += (f'<tr><td class="section-hdr" colspan="{span}">'
                           f'{current_sec.upper()}</td></tr>')
        cells = "".join(
            (f'<td class="sku-code">{row["skus"][ci]}</td>'
             if ci < len(row["skus"]) and row["skus"][ci] and row["skus"][ci] not in ("-","None","")
             else '<td style="color:#ccc;">—</td>')
            for ci in active_col_idx
        )
        table_html += (f'<tr><td class="row-label">{row["label"]}</td>'
                       f'<td style="font-size:10px;color:var(--muted);">{row["section"]}</td>'
                       f'{cells}</tr>')
    table_html += "</tbody></table></div>"
    st.markdown(table_html, unsafe_allow_html=True)

    st.markdown("<div style='height:14px'></div>", unsafe_allow_html=True)
    if st.button("▶  Next: Capture & OCR", key="go_step2"):
        st.session_state.step = 2
        st.rerun()


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 2 — Capture Image & Run OCR
# ═══════════════════════════════════════════════════════════════════════════════
elif step == 2:
    st.markdown(f"""
    <div class="card">
      <div class="card-title">{badge(2)} Capture Handwritten Order Form</div>
    </div>""", unsafe_allow_html=True)

    st.markdown(
        "📸 **Photograph or upload** the handwritten order form. "
        "Azure OCR will read the numbers and pre-fill quantities in Step 3. "
        "You can skip this and fill manually.",
        unsafe_allow_html=False,
    )

    # ── Azure credentials — always rendered (never inside a collapsed expander) ─
    # We use session_state as the single source of truth so values survive reruns.
    with st.expander("🔑 Azure OCR Credentials (required to run OCR)", expanded=True):
        ep = st.text_input(
            "Azure Endpoint",
            value=st.session_state.azure_endpoint,
            placeholder="https://YOUR_RESOURCE.cognitiveservices.azure.com",
            key="az_ep_input",
        )
        key = st.text_input(
            "Azure Key",
            value=st.session_state.azure_key,
            type="password",
            placeholder="32-character subscription key",
            key="az_key_input",
        )
        # Always persist typed values back to session_state
        st.session_state.azure_endpoint = ep
        st.session_state.azure_key      = key

    # ── Image source ──────────────────────────────────────────────────────────
    # Camera tab uses a full-width custom JS component that:
    #   • requests max device resolution via getUserMedia constraints
    #   • applies an unsharp-mask (sharpening) + contrast boost on a Canvas
    #     before encoding to JPEG — this fixes the browser blur problem
    #   • sends the processed base64 image back via postMessage → Streamlit
    # Upload tab is a standard fallback.

    def _on_file_upload():
        uf = st.session_state.get("file_upload_input")
        if uf is not None:
            st.session_state.image_bytes     = uf.getvalue()
            st.session_state.ocr_done        = False
            st.session_state.ocr_numbers     = []
            st.session_state.qty_keys_seeded = False

    img_mode = st.radio(
        "Image source",
        ["📷  Camera (recommended)", "📁  Upload File"],
        horizontal=True,
        label_visibility="collapsed",
        key="img_mode_radio",
    )

    if img_mode == "📷  Camera (recommended)":
        # ── Full-width high-res camera with canvas sharpening ─────────────────
        CAMERA_HTML = """
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { background: transparent; }
  #cam-wrap {
    width: 100%; background: #0A2342; border-radius: 10px;
    overflow: hidden; display: flex; flex-direction: column; align-items: center;
  }
  #video {
    width: 100%; max-height: 70vh; object-fit: cover; display: block;
  }
  #canvas { display: none; }
  #preview {
    width: 100%; display: none; border-top: 3px solid #1E88E5;
  }
  .cam-btn-row {
    display: flex; gap: 12px; padding: 12px 16px; width: 100%;
    background: #0A2342; justify-content: center; flex-wrap: wrap;
  }
  .cam-btn {
    flex: 1; max-width: 220px;
    padding: 12px 0; border: none; border-radius: 8px;
    font-size: 15px; font-weight: 700; cursor: pointer; transition: all .15s;
  }
  #btn-capture { background: #1E88E5; color: #fff; }
  #btn-capture:hover { background: #1565C0; }
  #btn-retake  { background: #fff; color: #0A2342; display: none; }
  #btn-retake:hover  { background: #EBF3FF; }
  #btn-use     { background: #00796B; color: #fff; display: none; }
  #btn-use:hover     { background: #00695C; }
  #status {
    color: rgba(255,255,255,.7); font-size: 12px;
    padding: 6px 16px 10px; text-align: center; width: 100%;
  }
</style>

<div id="cam-wrap">
  <video id="video" autoplay playsinline muted></video>
  <canvas id="canvas"></canvas>
  <img id="preview" alt="captured"/>
  <div class="cam-btn-row">
    <button class="cam-btn" id="btn-capture">📷 Capture</button>
    <button class="cam-btn" id="btn-retake">🔄 Retake</button>
    <button class="cam-btn" id="btn-use">✅ Use This Photo</button>
  </div>
  <div id="status">Starting camera…</div>
</div>

<script>
(function() {
  const video    = document.getElementById('video');
  const canvas   = document.getElementById('canvas');
  const preview  = document.getElementById('preview');
  const btnCap   = document.getElementById('btn-capture');
  const btnRet   = document.getElementById('btn-retake');
  const btnUse   = document.getElementById('btn-use');
  const status   = document.getElementById('status');
  let stream     = null;
  let capturedB64 = null;

  // ── Start camera at highest possible resolution ──────────────────────────
  async function startCamera() {
    try {
      stream = await navigator.mediaDevices.getUserMedia({
        video: {
          facingMode: { ideal: 'environment' },   // rear camera on phones
          width:  { ideal: 4096 },
          height: { ideal: 3072 },
          focusMode: { ideal: 'continuous' }
        },
        audio: false
      });
      video.srcObject = stream;
      await video.play();
      status.textContent = 'Camera ready — hold steady and tap Capture.';
    } catch(e) {
      status.textContent = 'Camera error: ' + e.message;
    }
  }

  // ── Unsharp mask on ImageData ─────────────────────────────────────────────
  // A standard convolution-based sharpen: subtract blurred copy from original.
  function convolve3x3(data, w, h, kernel) {
    const out = new Uint8ClampedArray(data.length);
    for (let y = 1; y < h - 1; y++) {
      for (let x = 1; x < w - 1; x++) {
        for (let c = 0; c < 3; c++) {
          let sum = 0;
          for (let ky = -1; ky <= 1; ky++) {
            for (let kx = -1; kx <= 1; kx++) {
              const idx = ((y + ky) * w + (x + kx)) * 4 + c;
              sum += data[idx] * kernel[(ky+1)*3 + (kx+1)];
            }
          }
          out[(y * w + x) * 4 + c] = Math.max(0, Math.min(255, sum));
        }
        out[(y * w + x) * 4 + 3] = 255; // alpha
      }
    }
    return out;
  }

  function sharpenAndEnhance(ctx, w, h) {
    const imgData = ctx.getImageData(0, 0, w, h);
    const d = imgData.data;

    // 1. Contrast boost: stretch histogram slightly
    for (let i = 0; i < d.length; i += 4) {
      for (let c = 0; c < 3; c++) {
        let v = d[i+c];
        // S-curve contrast: pull shadows down, highlights up
        v = Math.round(((v / 255 - 0.5) * 1.25 + 0.5) * 255);
        d[i+c] = Math.max(0, Math.min(255, v));
      }
    }

    // 2. Unsharp mask kernel (sharpen)
    const kernel = [
       0, -1,  0,
      -1,  5, -1,
       0, -1,  0
    ];
    const sharpened = convolve3x3(d, w, h, kernel);
    const out = ctx.createImageData(w, h);
    out.data.set(sharpened);
    ctx.putImageData(out, 0, 0);
  }

  // ── Capture button ────────────────────────────────────────────────────────
  btnCap.addEventListener('click', () => {
    const w = video.videoWidth  || 1280;
    const h = video.videoHeight || 720;
    canvas.width  = w;
    canvas.height = h;
    const ctx = canvas.getContext('2d');
    ctx.drawImage(video, 0, 0, w, h);

    // Apply sharpening + contrast
    sharpenAndEnhance(ctx, w, h);

    // Encode at high quality
    capturedB64 = canvas.toDataURL('image/jpeg', 0.97);

    // Show preview
    preview.src = capturedB64;
    video.style.display   = 'none';
    preview.style.display = 'block';
    btnCap.style.display  = 'none';
    btnRet.style.display  = 'block';
    btnUse.style.display  = 'block';
    status.textContent    = 'Preview — tap "Use This Photo" to confirm.';
  });

  // ── Retake button ─────────────────────────────────────────────────────────
  btnRet.addEventListener('click', () => {
    preview.style.display = 'none';
    video.style.display   = 'block';
    btnCap.style.display  = 'block';
    btnRet.style.display  = 'none';
    btnUse.style.display  = 'none';
    capturedB64           = null;
    status.textContent    = 'Camera ready — hold steady and tap Capture.';
  });

  // ── Use This Photo — send to Streamlit via postMessage ───────────────────
  btnUse.addEventListener('click', () => {
    if (!capturedB64) return;
    status.textContent = 'Sending to app…';
    // Post to parent Streamlit frame
    window.parent.postMessage({ type: 'CAMERA_CAPTURE', data: capturedB64 }, '*');
    status.textContent = '✅ Photo sent! Scroll down to Run OCR.';
    btnUse.disabled = true;
  });

  startCamera();
})();
</script>
"""
        import streamlit.components.v1 as components

        # Render the full-width camera component
        components.html(CAMERA_HTML, height=620, scrolling=False)

        # ── Receive the base64 image from JS via a hidden text_input ─────────
        # JS posts a message; we intercept it with a small injected listener
        # that writes the value into a hidden Streamlit text_input via DOM.
        RECEIVER_HTML = """
<script>
window.addEventListener('message', function(e) {
  if (e.data && e.data.type === 'CAMERA_CAPTURE') {
    // Find the hidden textarea Streamlit rendered and set its value
    const inputs = window.parent.document.querySelectorAll('input[type="text"]');
    for (const inp of inputs) {
      if (inp.getAttribute('data-cam-receiver') === 'true') {
        const nativeInputValueSetter = Object.getOwnPropertyDescriptor(
          window.HTMLInputElement.prototype, 'value').set;
        nativeInputValueSetter.call(inp, e.data.data);
        inp.dispatchEvent(new Event('input', { bubbles: true }));
        break;
      }
    }
    // Simpler fallback: store in sessionStorage and trigger rerun
    window.parent.sessionStorage.setItem('cam_capture_b64', e.data.data);
    // Trigger a Streamlit rerun by clicking a hidden button if present
    const btn = window.parent.document.querySelector('[data-cam-trigger="true"]');
    if (btn) btn.click();
  }
});
</script>
"""
        components.html(RECEIVER_HTML, height=0)

        # Read from a text_input that JS can target
        cam_b64 = st.text_input(
            "cam_b64_hidden",
            value=st.session_state.get("cam_b64_val", ""),
            label_visibility="hidden",
            key="cam_b64_input",
        )

        # Alternative: let user paste or use st.session_state trick —
        # Actually the cleanest Streamlit-native approach is: after JS fires,
        # user clicks a visible "📥 Load Captured Photo" button which reads
        # sessionStorage via a components.html query.
        st.markdown(
            "<div style='margin-top:8px;font-size:12px;color:#5A6880;'>"
            "After capturing, click <b>✅ Use This Photo</b> in the camera above, "
            "then click the button below to load it into the app.</div>",
            unsafe_allow_html=True,
        )

        load_col, _ = st.columns([1, 3])
        with load_col:
            load_clicked = st.button("📥  Load Captured Photo", key="load_cam_photo")

        # JS bridge: on button click, read sessionStorage and push into Streamlit
        LOAD_BRIDGE = """
<script>
(function() {
  // Run after a short delay to let Streamlit render the button
  function tryLoad() {
    const b64 = window.parent.sessionStorage.getItem('cam_capture_b64');
    if (!b64) return;
    // Find the hidden text input by its aria-label and set value
    const doc = window.parent.document;
    const allInputs = doc.querySelectorAll('input');
    for (const inp of allInputs) {
      const label = doc.querySelector('label[for="' + inp.id + '"]');
      if (label && label.textContent.trim() === 'cam_b64_hidden') {
        const setter = Object.getOwnPropertyDescriptor(
          window.HTMLInputElement.prototype, 'value').set;
        setter.call(inp, b64);
        inp.dispatchEvent(new Event('input', { bubbles: true }));
        window.parent.sessionStorage.removeItem('cam_capture_b64');
        break;
      }
    }
  }
  setTimeout(tryLoad, 300);
})();
</script>
"""
        if load_clicked:
            components.html(LOAD_BRIDGE, height=0)

        # Process the base64 value if it arrived
        b64_val = st.session_state.get("cam_b64_input", "")
        if b64_val and b64_val.startswith("data:image"):
            import base64 as _b64
            header, encoded = b64_val.split(",", 1)
            raw = _b64.b64decode(encoded)
            if raw != st.session_state.image_bytes:
                st.session_state.image_bytes     = raw
                st.session_state.ocr_done        = False
                st.session_state.ocr_numbers     = []
                st.session_state.qty_keys_seeded = False
                st.session_state["cam_b64_val"]  = b64_val
                st.rerun()

    else:
        st.file_uploader(
            "Upload image of order form",
            type=["jpg", "jpeg", "png"],
            label_visibility="collapsed",
            key="file_upload_input",
            on_change=_on_file_upload,
        )

    # Show image preview
    if st.session_state.image_bytes:
        st.image(st.session_state.image_bytes,
                 caption="Captured image", width="stretch")
    else:
        st.info("No image loaded yet. Capture or upload one above.", icon="🖼️")

    # OCR status display
    if st.session_state.ocr_done and st.session_state.ocr_numbers:
        nums = st.session_state.ocr_numbers
        st.success(
            f"✅ OCR complete — **{len(nums)} numbers** detected: "
            + ", ".join(str(n) for n in nums[:20])
            + (" …" if len(nums) > 20 else ""),
            icon="🔍",
        )
    elif st.session_state.ocr_done and not st.session_state.ocr_numbers:
        st.warning("OCR ran but found no numbers. You can fill quantities manually in the next step.",
                   icon="⚠️")

    st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

    # ── Action buttons — NONE are ever disabled ───────────────────────────────
    btn_cols = st.columns([1, 1, 1, 3])

    with btn_cols[0]:
        if st.button("◀  Back", key="back_step1"):
            st.session_state.step = 1
            st.rerun()

    with btn_cols[1]:
        if st.button("🔍  Run OCR", key="run_ocr_btn"):
            if not st.session_state.image_bytes:
                st.error("Please upload or capture an image first.", icon="🖼️")
            else:
                _ep  = st.session_state.azure_endpoint.strip()
                _key = st.session_state.azure_key.strip()
                if not _ep or not _key:
                    st.error(
                        "Please enter your **Azure Endpoint** and **Azure Key** above.",
                        icon="🔑",
                    )
                else:
                    with st.spinner("Sending image to Azure OCR… this may take 10–20 seconds"):
                        try:
                            nums = run_azure_ocr(
                                st.session_state.image_bytes,
                                _ep,
                                _key,
                            )
                            st.session_state.ocr_numbers     = nums
                            st.session_state.ocr_done        = True
                            st.session_state.qty_keys_seeded = False
                            st.rerun()
                        except Exception as exc:
                            st.error(f"OCR failed: {exc}", icon="❌")

    # Next is always clickable — OCR is optional
    with btn_cols[2]:
        if st.button("▶  Next: Fill Quantities", key="go_step3"):
            st.session_state.step = 3
            st.rerun()


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 3 — Fill Quantities  (OCR pre-seeded, fully editable)
# ═══════════════════════════════════════════════════════════════════════════════
elif step == 3:
    sheet     = sku_sheets[st.session_state.active_sheet]
    col_ids   = sheet["col_ids"]
    col_labels= sheet["col_labels"]
    rows      = sheet["rows"]
    active_ci = [i for i, c in enumerate(col_ids) if c]

    # ── ONE-TIME: seed qty_map from OCR numbers before any widget renders ─────
    if not st.session_state.qty_keys_seeded:
        ocr_nums = list(st.session_state.ocr_numbers)
        ocr_ptr  = 0
        for ri, row in enumerate(rows):
            for ci in active_ci:
                sku = row["skus"][ci] if ci < len(row["skus"]) else None
                if not sku or sku in ("-", "None", ""):
                    continue
                widget_key = f"qty_{ri}_{ci}"
                if ocr_nums and ocr_ptr < len(ocr_nums):
                    try:
                        val = max(0, int(float(ocr_nums[ocr_ptr])))
                        ocr_ptr += 1
                    except Exception:
                        val = 0
                else:
                    val = 0
                st.session_state[widget_key] = val
        st.session_state.qty_keys_seeded = True

    # ── Header ────────────────────────────────────────────────────────────────
    st.markdown(f"""
    <div class="card">
      <div class="card-title">{badge(3, step > 3)} Fill Order Quantities — {st.session_state.active_sheet}</div>
    </div>""", unsafe_allow_html=True)

    if st.session_state.ocr_done and st.session_state.ocr_numbers:
        st.info(
            f"🔍 OCR detected **{len(st.session_state.ocr_numbers)} numbers** — "
            "quantities have been pre-filled below. Please verify and correct as needed.",
            icon="✅",
        )
    else:
        st.caption("Enter quantities manually (0 = not ordering). Only cells with a valid SKU are editable.")

    st.markdown(
        "<div style='background:#FFF8E1;border:1px solid #FFE082;border-radius:8px;"
        "padding:8px 14px;font-size:12px;margin-bottom:12px;'>"
        "📝 <b>Top 3 header rows and the first column (item names) are fixed.</b> "
        "Edit quantities in the white cells only."
        "</div>",
        unsafe_allow_html=True,
    )

    # ── Column header row ─────────────────────────────────────────────────────
    hdr_cols = st.columns([3] + [1] * len(active_ci))
    hdr_cols[0].markdown(
        '<div style="font-size:11px;font-weight:700;color:#0A2342;">Item</div>',
        unsafe_allow_html=True,
    )
    for j, ci in enumerate(active_ci):
        lbl = col_labels[ci] if ci < len(col_labels) else col_ids[ci]
        hdr_cols[j + 1].markdown(
            f'<div style="text-align:center;font-size:11px;font-weight:700;color:#0A2342;'
            f'background:#EBF3FF;border-radius:4px;padding:3px 1px;">'
            f'{lbl}<br/><span style="font-size:9px;color:#5A6880;">{col_ids[ci]}</span></div>',
            unsafe_allow_html=True,
        )

    # ── Data rows ─────────────────────────────────────────────────────────────
    current_sec = None
    new_qty_map = {}

    for ri, row in enumerate(rows):
        sec = row["section"]
        if sec != current_sec:
            current_sec = sec
            st.markdown(
                f'<div style="background:#0A2342;color:#fff;font-size:11px;font-weight:700;'
                f'padding:5px 10px;border-radius:6px;margin:10px 0 4px;">▸ {sec.upper()}</div>',
                unsafe_allow_html=True,
            )

        row_cols = st.columns([3] + [1] * len(active_ci))
        row_cols[0].markdown(
            f'<div style="font-size:12px;font-weight:500;padding:4px 0;">{row["label"]}</div>',
            unsafe_allow_html=True,
        )

        for j, ci in enumerate(active_ci):
            sku = row["skus"][ci] if ci < len(row["skus"]) else None
            if not sku or sku in ("-", "None", ""):
                row_cols[j + 1].markdown(
                    '<div style="text-align:center;color:#ccc;font-size:11px;padding:6px 0;">—</div>',
                    unsafe_allow_html=True,
                )
                continue

            widget_key = f"qty_{ri}_{ci}"
            qty = row_cols[j + 1].number_input(
                label=sku,
                label_visibility="collapsed",
                min_value=0,
                step=1,
                key=widget_key,
            )
            if qty > 0:
                new_qty_map[(ri, ci)] = qty

    st.session_state.qty_map = new_qty_map

    # ── Live summary ──────────────────────────────────────────────────────────
    total_items = len(new_qty_map)
    grand_mrp = grand_land = 0.0
    for (ri, ci), qty in new_qty_map.items():
        sku = rows[ri]["skus"][ci] if ci < len(rows[ri]["skus"]) else None
        if sku and sku not in ("-", "None", ""):
            info       = mrp_lookup.get(sku, {})
            mrp        = info.get("MRP_clean") or 0.0
            land       = info.get("Distributor Landing") or mrp
            grand_mrp  += mrp  * qty
            grand_land += land * qty

    st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)
    st.markdown(f"""
    <div class="card" style="max-width:420px;">
      <div class="card-title">📊 Live Summary</div>
      <div class="totals-row">
        <span>Active Line Items</span><span class="val">{total_items}</span>
      </div>
      <div class="totals-row">
        <span>Gross MRP Value</span><span class="val">₹ {grand_mrp:,.2f}</span>
      </div>
      <div class="totals-row">
        <span>Distributor Landing</span><span class="val">₹ {grand_land:,.2f}</span>
      </div>
      <div class="totals-row net">
        <span>Net Payable (Landing)</span><span class="val">₹ {grand_land:,.2f}</span>
      </div>
    </div>
    """, unsafe_allow_html=True)

    col_b1, col_b2, _ = st.columns([1, 1, 3])
    with col_b1:
        if st.button("◀  Back", key="back_step2"):
            st.session_state.step = 2
            st.rerun()
    with col_b2:
        if st.button("▶  Next: Party Details", key="go_step4"):
            st.session_state.step = 4
            st.rerun()


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 4 — Dealer / Distributor & Customer Details
# ═══════════════════════════════════════════════════════════════════════════════
elif step == 4:
    st.markdown(f"""
    <div class="card">
      <div class="card-title">{badge(4, step > 4)} Party Details</div>
    </div>""", unsafe_allow_html=True)

    # ── Auto-fill from ZSD_CUST lookup ───────────────────────────────────────
    st.markdown("##### 🔍 Auto-fill from Customer Master (ZSD_CUST)")
    cust_options = ["— Select or type below —"] + [
        f'{row["Customer Code"]} | {row["Customer Name"]}'
        for _, row in zsd_df.iterrows()
    ]
    sel = st.selectbox("Search Customer", cust_options, key="zsd_search")

    def zsd_fill(prefix: str, row: pd.Series):
        addr = " ".join(filter(None, [
            str(row.get("Address 1","") or ""),
            str(row.get("Address 2","") or ""),
            str(row.get("Address 3","") or ""),
            str(row.get("City","") or ""),
        ]))
        return {
            f"{prefix}_party_no":   str(row.get("Customer Code","") or ""),
            f"{prefix}_party_name": str(row.get("Customer Name","") or ""),
            f"{prefix}_address":    addr.strip(),
            f"{prefix}_phone":      str(row.get("Telephone","") or ""),
            f"{prefix}_mobile":     str(row.get("Mobile No.","") or ""),
            f"{prefix}_state_code": str(row.get("State Code","") or ""),
            f"{prefix}_state":      str(row.get("State Code Desc.","") or ""),
            f"{prefix}_gst":        str(row.get("GST Number","") or ""),
            f"{prefix}_pan":        str(row.get("PAN No.","") or ""),
        }

    if sel != cust_options[0]:
        code = sel.split("|")[0].strip()
        matched = zsd_df[zsd_df["Customer Code"].astype(str) == code]
        if not matched.empty:
            row = matched.iloc[0]
            for k, v in zsd_fill("bill", row).items():
                if k not in st.session_state or not st.session_state[k]:
                    st.session_state[k] = v
            st.toast("Bill-to details filled from customer master.", icon="✅")

    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

    # ── Bill-to ───────────────────────────────────────────────────────────────
    st.markdown("#### 📋 Bill To Party")
    bc1, bc2 = st.columns(2)
    bill_party_no   = bc1.text_input("Bill to Party No.",   key="bill_party_no")
    bill_party_name = bc2.text_input("Bill to Party Name",  key="bill_party_name")
    bill_address    = st.text_input("Bill to Address",      key="bill_address")
    bc3, bc4 = st.columns(2)
    bill_phone      = bc3.text_input("Phone",               key="bill_phone")
    bill_mobile     = bc4.text_input("Mobile",              key="bill_mobile")
    bc5, bc6 = st.columns(2)
    bill_sc         = bc5.text_input("State Code",          key="bill_state_code")
    bill_state      = bc6.text_input("State",               key="bill_state")
    bc7, bc8 = st.columns(2)
    bill_gst        = bc7.text_input("GST No.",             key="bill_gst")
    bill_pan        = bc8.text_input("PAN No.",             key="bill_pan")

    st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)

    # ── Copy bill → ship ──────────────────────────────────────────────────────
    same_as_bill = st.checkbox("Ship-to same as Bill-to", value=False)

    st.markdown("#### 🚚 Ship To Party")
    if same_as_bill:
        st.session_state["ship_party_no"]   = bill_party_no
        st.session_state["ship_party_name"] = bill_party_name
        st.session_state["ship_address"]    = bill_address
        st.session_state["ship_phone"]      = bill_phone
        st.session_state["ship_mobile"]     = bill_mobile
        st.session_state["ship_state_code"] = bill_sc
        st.session_state["ship_state"]      = bill_state
        st.session_state["ship_gst"]        = bill_gst
        st.session_state["ship_pan"]        = bill_pan

    sc1, sc2 = st.columns(2)
    ship_party_no   = sc1.text_input("Ship to Party No.",   key="ship_party_no")
    ship_party_name = sc2.text_input("Ship to Party Name",  key="ship_party_name")
    ship_address    = st.text_input("Ship to Address",      key="ship_address")
    sc3, sc4 = st.columns(2)
    ship_phone      = sc3.text_input("Phone ",              key="ship_phone")
    ship_mobile     = sc4.text_input("Mobile ",             key="ship_mobile")
    sc5, sc6 = st.columns(2)
    ship_sc         = sc5.text_input("State Code ",         key="ship_state_code")
    ship_state      = sc6.text_input("State ",              key="ship_state")
    sc7, sc8 = st.columns(2)
    ship_gst        = sc7.text_input("GST No. ",            key="ship_gst")
    ship_pan        = sc8.text_input("PAN No. ",            key="ship_pan")

    # Save to session
    st.session_state.bill_to = {
        "Party No.":   bill_party_no,
        "Name":        bill_party_name,
        "Address":     bill_address,
        "Phone":       bill_phone,
        "Mobile":      bill_mobile,
        "State Code":  bill_sc,
        "State":       bill_state,
        "GST No.":     bill_gst,
        "PAN No.":     bill_pan,
    }
    st.session_state.ship_to = {
        "Party No.":   ship_party_no,
        "Name":        ship_party_name,
        "Address":     ship_address,
        "Phone":       ship_phone,
        "Mobile":      ship_mobile,
        "State Code":  ship_sc,
        "State":       ship_state,
        "GST No.":     ship_gst,
        "PAN No.":     ship_pan,
    }

    # ── Preview ───────────────────────────────────────────────────────────────
    st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)
    with st.expander("👁 Preview Party Details", expanded=False):
        pc1, pc2 = st.columns(2)
        with pc1:
            st.markdown(party_html(st.session_state.bill_to, "BILL TO PARTY"),
                        unsafe_allow_html=True)
        with pc2:
            st.markdown(party_html(st.session_state.ship_to, "SHIP TO PARTY"),
                        unsafe_allow_html=True)

    col_b1, col_b2, _ = st.columns([1, 1, 3])
    with col_b1:
        if st.button("◀  Back", key="back_step3"):
            st.session_state.step = 3
            st.rerun()
    with col_b2:
        if st.button("▶  Generate & Download", key="go_step5"):
            st.session_state.step = 5
            st.rerun()


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 5 — Preview & Download
# ═══════════════════════════════════════════════════════════════════════════════
elif step == 5:
    st.markdown(f"""
    <div class="card">
      <div class="card-title">{badge(5, False)} Review & Download Quotation</div>
    </div>""", unsafe_allow_html=True)

    sheet      = sku_sheets[st.session_state.active_sheet]
    col_ids    = sheet["col_ids"]
    col_labels = sheet["col_labels"]
    rows       = sheet["rows"]
    active_ci  = [i for i, c in enumerate(col_ids) if c]
    qty_map    = st.session_state.qty_map

    # ── Party summary ─────────────────────────────────────────────────────────
    pc1, pc2 = st.columns(2)
    with pc1:
        st.markdown(party_html(st.session_state.bill_to, "BILL TO PARTY"),
                    unsafe_allow_html=True)
    with pc2:
        st.markdown(party_html(st.session_state.ship_to, "SHIP TO PARTY"),
                    unsafe_allow_html=True)

    st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

    # ── Line items table ──────────────────────────────────────────────────────
    st.markdown("#### 📦 Ordered Items")

    line_rows   = []
    grand_mrp   = 0.0
    grand_land  = 0.0

    for (ri, ci), qty in sorted(qty_map.items()):
        if qty <= 0:
            continue
        row = rows[ri]
        sku = row["skus"][ci] if ci < len(row["skus"]) else ""
        if not sku or sku in ("-", "None", ""):
            continue
        info  = mrp_lookup.get(sku, {})
        mrp   = info.get("MRP_clean") or 0.0
        land  = info.get("Distributor Landing") or mrp
        desc  = info.get("Material Description","") or ""
        size  = col_labels[ci] if ci < len(col_labels) else col_ids[ci]
        grand_mrp  += mrp  * qty
        grand_land += land * qty
        line_rows.append({
            "Item":         row["label"],
            "SKU Code":     sku,
            "Size":         size,
            "Description":  desc[:55] + ("…" if len(desc) > 55 else ""),
            "MRP (₹)":      round(mrp, 2),
            "Qty":          qty,
            "Rate (₹)":     round(land, 2),
            "Amount (₹)":   round(land * qty, 2),
        })

    if line_rows:
        df_lines = pd.DataFrame(line_rows)
        st.dataframe(df_lines, use_container_width=True, hide_index=True,
                     column_config={
                         "MRP (₹)":    st.column_config.NumberColumn(format="₹ %.2f"),
                         "Rate (₹)":   st.column_config.NumberColumn(format="₹ %.2f"),
                         "Amount (₹)": st.column_config.NumberColumn(format="₹ %.2f"),
                     })
    else:
        st.warning("No quantities entered. Go back and fill quantities.", icon="⚠️")

    # ── Totals ────────────────────────────────────────────────────────────────
    discount = grand_mrp - grand_land
    st.markdown(f"""
    <div class="card" style="max-width:440px;margin-top:12px;">
      <div class="card-title">💰 Totals</div>
      <div class="totals-row">
        <span>Gross MRP Value</span><span class="val">₹ {grand_mrp:,.2f}</span>
      </div>
      <div class="totals-row">
        <span>Distributor Discount</span><span class="val neg">− ₹ {discount:,.2f}</span>
      </div>
      <div class="totals-row net">
        <span>Net Payable (Distributor Landing)</span>
        <span class="val">₹ {grand_land:,.2f}</span>
      </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Download buttons ──────────────────────────────────────────────────────
    st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)
    dl1, dl2, dl3 = st.columns([1, 1, 2])

    # CSV download (always available)
    if line_rows:
        csv_buf = io.StringIO()
        pd.DataFrame(line_rows).to_csv(csv_buf, index=False)
        dl1.download_button(
            "⬇  Download CSV",
            data=csv_buf.getvalue().encode(),
            file_name=f"sintex_quotation_{date.today()}.csv",
            mime="text/csv",
            key="dl_csv",
        )

    # Excel download
    if line_rows:
        xls_buf = io.BytesIO()
        with pd.ExcelWriter(xls_buf, engine="openpyxl") as writer:
            # Sheet 1: Party Details
            party_data = []
            for k, v in st.session_state.bill_to.items():
                party_data.append({"Field": f"Bill To - {k}", "Value": v})
            for k, v in st.session_state.ship_to.items():
                party_data.append({"Field": f"Ship To - {k}", "Value": v})
            pd.DataFrame(party_data).to_excel(writer, sheet_name="Party Details", index=False)
            # Sheet 2: Line Items
            pd.DataFrame(line_rows).to_excel(writer, sheet_name="Quotation Lines", index=False)
            # Sheet 3: Summary
            summary = pd.DataFrame([
                {"Description": "Gross MRP Value",  "Amount (₹)": round(grand_mrp,  2)},
                {"Description": "Distributor Disc.", "Amount (₹)": round(discount,   2)},
                {"Description": "Net Payable",       "Amount (₹)": round(grand_land, 2)},
            ])
            summary.to_excel(writer, sheet_name="Summary", index=False)

        dl2.download_button(
            "⬇  Download Excel",
            data=xls_buf.getvalue(),
            file_name=f"sintex_quotation_{date.today()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_excel",
        )

    # PDF download
    if _HAS_REPORTLAB and line_rows:
        if st.session_state.pdf_bytes is None:
            with st.spinner("Generating PDF…"):
                try:
                    pdf_bytes = generate_pdf(
                        st.session_state.active_sheet,
                        sheet,
                        qty_map,
                        mrp_lookup,
                        st.session_state.bill_to,
                        st.session_state.ship_to,
                    )
                    st.session_state.pdf_bytes = pdf_bytes
                except Exception as e:
                    st.error(f"PDF generation failed: {e}")

        if st.session_state.pdf_bytes:
            dl3.download_button(
                "⬇  Download PDF",
                data=st.session_state.pdf_bytes,
                file_name=f"sintex_quotation_{date.today()}.pdf",
                mime="application/pdf",
                key="dl_pdf",
            )
    elif not _HAS_REPORTLAB:
        st.caption("📄 PDF download requires `reportlab`. Run: `pip install reportlab`")

    # ── Navigation ────────────────────────────────────────────────────────────
    st.markdown("<div style='height:14px'></div>", unsafe_allow_html=True)
    col_bk, col_nr, _ = st.columns([1, 1, 3])
    with col_bk:
        if st.button("◀  Back", key="back_step4"):
            st.session_state.pdf_bytes = None
            st.session_state.step = 4
            st.rerun()
    with col_nr:
        if st.button("🔄  New Quotation", key="new_quot"):
            for k in ["step","qty_map","bill_to","ship_to","pdf_bytes",
                      "image_bytes","ocr_numbers","ocr_done","qty_keys_seeded","zsd_search"]:
                if k in st.session_state:
                    del st.session_state[k]
            for k in list(st.session_state.keys()):
                if k.startswith("qty_"):
                    del st.session_state[k]
            st.rerun()