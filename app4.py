import io
import os
import re
import copy
import base64
import textwrap
from datetime import date
from difflib import SequenceMatcher

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

try:
    import requests as _requests
    _HAS_REQUESTS = True
except ImportError:
    _HAS_REQUESTS = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, HRFlowable)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    _HAS_REPORTLAB = True
except ImportError:
    _HAS_REPORTLAB = False

st.set_page_config(
    page_title="Sintex BAPL – Quotation Generator",
    page_icon="🔧",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@300;400;500;600;700&family=IBM+Plex+Mono:wght@400;500&display=swap');
:root {
  --navy:   #0A2342;
  --blue:   #1565C0;
  --sky:    #1E88E5;
  --teal:   #00796B;
  --gold:   #F9A825;
  --danger: #C62828;
  --border: #DEE3EC;
  --surface:#F4F6FA;
  --text:   #1A1F36;
  --muted:  #5A6880;
  --radius: 10px;
}
html, body, [class*="css"] {
  font-family: 'IBM Plex Sans', sans-serif !important;
  background: #F0F2F8 !important;
  color: var(--text);
}
#MainMenu, footer, header { visibility: hidden; }
.block-container { padding: 1.2rem 1.5rem 4rem !important; max-width: 1400px !important; }

.card {
  background: #fff;
  border: 1px solid var(--border);
  border-radius: var(--radius);
  padding: 20px 24px;
  margin-bottom: 18px;
  box-shadow: 0 1px 4px rgba(10,35,66,.07);
}
.card-title {
  font-size: 13px;
  font-weight: 700;
  text-transform: uppercase;
  letter-spacing: .7px;
  color: var(--navy);
  margin-bottom: 14px;
  display: flex;
  align-items: center;
  gap: 8px;
}
.section-anchor {
  display: flex;
  align-items: center;
  gap: 12px;
  background: linear-gradient(135deg, #0A2342 0%, #1565C0 100%);
  border-radius: 10px;
  padding: 14px 20px;
  margin: 24px 0 16px 0;
  box-shadow: 0 2px 10px rgba(10,35,66,.18);
}
.section-anchor h2 {
  font-size: 15px;
  font-weight: 700;
  color: #fff;
  margin: 0;
  letter-spacing: .3px;
}
.section-anchor .section-num {
  background: rgba(255,255,255,.2);
  color: #fff;
  border-radius: 50%;
  width: 28px; height: 28px;
  display: inline-flex;
  align-items: center;
  justify-content: center;
  font-size: 12px;
  font-weight: 800;
  flex-shrink: 0;
}

.app-header {
  background: linear-gradient(135deg, #0A2342 0%, #1565C0 100%);
  border-radius: 12px;
  padding: 18px 26px;
  display: flex; align-items: center; gap: 18px;
  margin-bottom: 22px;
  box-shadow: 0 4px 16px rgba(10,35,66,.25);
}
.app-header h1 { font-size: 20px; font-weight: 700; color: #fff; margin: 0; }
.app-header p  { font-size: 12px; color: rgba(255,255,255,.65); margin: 3px 0 0; }

/* OCR summary table — read-only */
.ocr-summary-table {
  width: 100%;
  border-collapse: collapse;
  font-size: 12px;
  margin-top: 8px;
}
.ocr-summary-table th {
  background: var(--navy);
  color: #fff;
  padding: 8px 12px;
  font-weight: 700;
  font-size: 11px;
  text-align: left;
  border: 1px solid rgba(255,255,255,.12);
  white-space: nowrap;
}
.ocr-summary-table th.num { text-align: right; }
.ocr-summary-table td {
  padding: 7px 12px;
  border: 1px solid var(--border);
  vertical-align: middle;
}
.ocr-summary-table td.num { text-align: right; font-family: 'IBM Plex Mono', monospace; }
.ocr-summary-table tr:nth-child(even) td { background: #F7F9FC; }
.ocr-summary-table tr:hover td { background: #EBF3FF; transition: background .12s; }
.ocr-summary-table .sku-mono { font-family: 'IBM Plex Mono', monospace; font-size: 10.5px; color: var(--blue); }
.ocr-summary-table .final-price { font-weight: 700; color: var(--navy); }
.ocr-summary-table .match-ok   { color: var(--teal);  font-weight: 700; font-size: 10px; }
.ocr-summary-table .match-warn { color: #E65100;      font-weight: 700; font-size: 10px; }
.ocr-summary-table .match-none { color: var(--danger);font-weight: 700; font-size: 10px; }
.ocr-summary-table tfoot td {
  background: #EBF3FF !important;
  font-weight: 700;
  color: var(--navy);
  border-top: 2px solid var(--navy);
}

.edit-section-header {
  display: flex;
  align-items: center;
  justify-content: space-between;
  padding: 10px 16px;
  border-radius: 8px;
  margin-bottom: 10px;
}
.edit-group-label {
  font-size: 11px;
  font-weight: 700;
  text-transform: uppercase;
  letter-spacing: .6px;
  padding: 4px 10px;
  border-radius: 4px;
  margin: 10px 0 6px 0;
  display: inline-block;
}
.edit-group-detected { background: #E8F5E9; color: var(--teal); }
.edit-group-undetected { background: #FFF3E0; color: #E65100; }

.stNumberInput > div > div > input {
  border-radius: 6px !important;
  border: 1.5px solid var(--border) !important;
  font-family: 'IBM Plex Mono', monospace !important;
  font-size: 13px !important;
  padding: 4px 6px !important;
  text-align: center !important;
}
.stButton > button {
  font-family: 'IBM Plex Sans', sans-serif !important;
  font-weight: 600 !important; font-size: 14px !important;
  border-radius: 8px !important; border: none !important;
  background: var(--navy) !important; color: #fff !important;
  padding: 10px 20px !important;
  box-shadow: 0 2px 8px rgba(10,35,66,.2) !important;
  transition: all .15s !important;
}
.stButton > button:hover { background: var(--sky) !important; transform: translateY(-1px) !important; }

.pill {
  display: inline-block; padding: 2px 10px;
  border-radius: 20px; font-size: 11px; font-weight: 600;
}
.pill-blue   { background: #EBF3FF; color: var(--blue); }
.pill-green  { background: #E8F5E9; color: var(--teal); }
.pill-gold   { background: #FFF8E1; color: #7B5B00; }

.totals-row {
  display: flex; justify-content: space-between; align-items: center;
  padding: 9px 0; border-bottom: 1px solid var(--border);
  font-size: 14px;
}
.totals-row:last-child { border: none; padding-top: 12px; }
.totals-row.net { font-weight: 700; font-size: 16px; color: var(--navy); }
.totals-row .val { font-family: 'IBM Plex Mono', monospace; font-weight: 600; }
.totals-row .val.neg { color: var(--danger); }

.party-box {
  border: 1px solid var(--border); border-radius: 8px;
  padding: 14px 16px; background: var(--surface);
}
.party-box h4 { margin: 0 0 10px; font-size: 12px; font-weight: 700;
  text-transform: uppercase; letter-spacing: .5px; color: var(--navy); }
.party-row { display: flex; gap: 6px; margin-bottom: 5px; font-size: 12.5px; }
.party-lbl { color: var(--muted); min-width: 80px; flex-shrink: 0; }
.party-val { font-weight: 500; color: var(--text); word-break: break-word; }

label { font-size:12px!important; font-weight:600!important; color:var(--muted)!important; }
</style>
""", unsafe_allow_html=True)

# ─── Paths ──────────────────────────────────────────────────────────────────
_HERE = os.path.dirname(os.path.abspath(__file__))
MRP_CSV_PATH  = os.path.join(_HERE, "MRP_State_chhattisghar.csv")
ZSD_CSV_PATH  = os.path.join(_HERE, "ZSD_CUST.csv")
SKU_XLSX_PATH = os.path.join(_HERE, "Sample form for Product list.xlsx")

# ─── Data loaders ────────────────────────────────────────────────────────────
@st.cache_data
def load_mrp():
    df = pd.read_csv(MRP_CSV_PATH)
    df["MRP_clean"] = (
        df["MRP(ZPR1-933)"]
        .astype(str).str.replace(",", "", regex=False)
        .pipe(pd.to_numeric, errors="coerce")
    )
    df["Distributor Landing"] = pd.to_numeric(
        df["Distributor Landing"].astype(str).str.replace(",", "", regex=False),
        errors="coerce",
    )
    lookup = {}
    for _, row in df.iterrows():
        mat = str(row["Material Number"]).strip()
        lookup[mat] = row.to_dict()
    return lookup

@st.cache_data
def load_zsd():
    df = pd.read_csv(ZSD_CSV_PATH, encoding="latin1")
    return df

@st.cache_data
def load_sku_sheets():
    wb  = load_workbook(SKU_XLSX_PATH, read_only=True)
    out = {}
    for sname in wb.sheetnames:
        ws   = wb[sname]
        data = [row for row in ws.iter_rows(values_only=True)
                if any(c is not None for c in row)]
        if len(data) < 3:
            continue
        col_ids    = [str(c).strip() if c is not None else "" for c in data[1][2:]]
        col_labels = [str(c).strip() if c is not None else "" for c in data[2][2:]]
        while col_ids and col_ids[-1] == "":
            col_ids.pop(); col_labels.pop()

        rows   = []
        section = "General"
        for raw in data[3:]:
            if not any(c is not None for c in raw):
                continue
            label = str(raw[0]).strip() if raw[0] is not None else ""
            if label.upper() in ("FITTINGS", "PIPES", "FITTING SCH 80") and all(
                c is None for c in raw[2:]
            ):
                section = label.title()
                continue
            if not label:
                continue
            skus = []
            for i in range(len(col_ids)):
                idx = i + 2
                val = raw[idx] if idx < len(raw) else None
                skus.append(str(val).strip() if val is not None else None)
            rows.append({"label": label, "section": section, "skus": skus})

        out[sname] = {
            "label":      sname,
            "col_ids":    col_ids,
            "col_labels": col_labels,
            "rows":       rows,
        }
    return out


@st.cache_data
def build_sku_catalogue():
    sheets = load_sku_sheets()
    cat = {}
    for sname, sheet in sheets.items():
        col_ids    = sheet["col_ids"]
        col_labels = sheet["col_labels"]
        for ri, row in enumerate(sheet["rows"]):
            for ci, sku in enumerate(row["skus"]):
                if not sku or sku in ("-", "None", ""):
                    continue
                cat[sku] = {
                    "sheet":      sname,
                    "row_label":  row["label"],
                    "section":    row["section"],
                    "col_id":     col_ids[ci] if ci < len(col_ids) else "",
                    "col_label":  col_labels[ci] if ci < len(col_labels) else "",
                    "ri":         ri,
                    "ci":         ci,
                }
    return cat


_OCR_FIXES = str.maketrans({
    'O': '0', 'I': '1', 'S': '5', 'B': '8', 'G': '6', 'Z': '2',
})

def _normalise_sku(raw: str) -> str:
    s = re.sub(r"[\s\-\._]", "", raw).upper()
    m = re.match(r'^([A-Z]+)(.*)$', s)
    if m:
        prefix, rest = m.group(1), m.group(2)
        rest_fixed = rest.translate(_OCR_FIXES)
        return prefix + rest_fixed
    return s.translate(_OCR_FIXES)


@st.cache_data
def _build_normalised_index(catalogue_keys: tuple):
    return {_normalise_sku(k): k for k in catalogue_keys}


def match_sku_to_catalogue(raw_sku: str, catalogue: dict,
                            norm_index: dict, threshold: float = 0.75):
    if not raw_sku:
        return None, 0.0, "none"

    norm_raw = _normalise_sku(raw_sku)
    raw_len  = len(norm_raw)

    if norm_raw in norm_index:
        return norm_index[norm_raw], 1.0, "exact"

    prefix_candidates = []
    for norm_cat, orig_cat in norm_index.items():
        if norm_cat.startswith(norm_raw) and raw_len >= 8:
            score = raw_len / max(len(norm_cat), 1)
            prefix_candidates.append((orig_cat, score))
    if prefix_candidates:
        best = max(prefix_candidates, key=lambda x: x[1])
        if best[1] >= 0.55:
            return best[0], min(0.97, best[1] + 0.3), "exact"

    if raw_len >= 8:
        sub_candidates = []
        for norm_cat, orig_cat in norm_index.items():
            if norm_raw in norm_cat:
                score = raw_len / max(len(norm_cat), 1)
                sub_candidates.append((orig_cat, score))
        if sub_candidates:
            best = max(sub_candidates, key=lambda x: x[1])
            if best[1] >= 0.50:
                return best[0], min(0.90, best[1] + 0.2), "fuzzy"

    has_digit = any(c.isdigit() for c in norm_raw)
    has_alpha = any(c.isalpha() for c in norm_raw)
    if raw_len < 7 or not (has_digit and has_alpha):
        return None, 0.0, "none"

    best_sku, best_score = None, 0.0
    for norm_cat, orig_cat in norm_index.items():
        score = SequenceMatcher(None, norm_raw, norm_cat).ratio()
        if score > best_score:
            best_score = score
            best_sku   = orig_cat

    if best_score >= threshold:
        return best_sku, best_score, "fuzzy"
    return None, best_score, "none"


_FORM_SIZES = ["15MM", "20MM", "25MM", "32MM", "40MM", "50MM"]

def _looks_like_product_code(token: str) -> bool:
    t = token.strip()
    if len(t) < 7:
        return False
    return (any(c.isdigit() for c in t) and
            any(c.isalpha() for c in t) and
            not t.isalpha())

def _word_x_center(polygon) -> float:
    if not polygon:
        return 0.0
    if isinstance(polygon[0], dict):
        xs = [p["x"] for p in polygon]
    else:
        xs = [polygon[i] for i in range(0, len(polygon), 2)]
    return (min(xs) + max(xs)) / 2.0

def _word_y_center(polygon) -> float:
    if not polygon:
        return 0.0
    if isinstance(polygon[0], dict):
        ys = [p["y"] for p in polygon]
    else:
        ys = [polygon[i] for i in range(1, len(polygon), 2)]
    return (min(ys) + max(ys)) / 2.0

def _extract_table_geometric(words_with_pos: list) -> list:
    if not words_with_pos:
        return []
    words_sorted = sorted(words_with_pos, key=lambda w: w["y"])
    row_gap  = 0.018
    rows     = []
    cur_row  = [words_sorted[0]]
    for w in words_sorted[1:]:
        if w["y"] - cur_row[-1]["y"] > row_gap:
            rows.append(cur_row)
            cur_row = [w]
        else:
            cur_row.append(w)
    rows.append(cur_row)

    header_row  = None
    col_x_map   = {}
    for row in rows:
        texts = [w["text"].upper().strip() for w in row]
        hits  = [s for s in _FORM_SIZES if s in texts]
        if len(hits) >= 2:
            header_row = row
            for w in row:
                t = w["text"].upper().strip()
                if t in _FORM_SIZES:
                    col_x_map[t] = w["x"]
            break

    if not col_x_map:
        return _extract_table_fallback(words_with_pos)

    size_labels_sorted = sorted(col_x_map.keys(), key=lambda s: col_x_map[s])
    col_x_sorted = [col_x_map[s] for s in size_labels_sorted]

    def _assign_size(x_val):
        best_size, best_dist = None, float("inf")
        for size in size_labels_sorted:
            d = abs(col_x_map[size] - x_val)
            if d < best_dist:
                best_dist = d
                best_size = size
        col_width = (max(col_x_sorted) - min(col_x_sorted)) / max(len(col_x_sorted) - 1, 1)
        if best_dist > col_width * 0.75:
            return None
        return best_size

    sku_x_max = min(col_x_sorted) - 0.01
    results   = []
    seen_keys = set()

    for row in rows:
        if row is header_row:
            continue
        row_words = sorted(row, key=lambda w: w["x"])
        sku_tokens = [w for w in row_words
                      if w["x"] <= sku_x_max and _looks_like_product_code(w["text"])]
        if not sku_tokens:
            continue
        raw_sku  = sku_tokens[0]["text"]
        raw_line = " ".join(w["text"] for w in row_words)
        qty_tokens = [w for w in row_words
                      if w["x"] > sku_x_max and re.match(r'^\d+$', w["text"].strip())]
        if not qty_tokens:
            key = (raw_sku.upper(), "")
            if key not in seen_keys:
                seen_keys.add(key)
                results.append({"raw_sku": raw_sku, "raw_size": "", "qty": 0, "raw_line": raw_line})
            continue
        for qt in qty_tokens:
            size = _assign_size(qt["x"])
            if size is None:
                continue
            try:
                qty_val = int(qt["text"].strip())
            except ValueError:
                continue
            key = (raw_sku.upper(), size)
            if key in seen_keys:
                continue
            seen_keys.add(key)
            results.append({"raw_sku": raw_sku, "raw_size": size, "qty": qty_val, "raw_line": raw_line})
    return results


def _extract_table_fallback(words_with_pos: list) -> list:
    if not words_with_pos:
        return []
    words_sorted = sorted(words_with_pos, key=lambda w: w["y"])
    row_gap  = 0.018
    rows     = []
    cur_row  = [words_sorted[0]]
    for w in words_sorted[1:]:
        if w["y"] - cur_row[-1]["y"] > row_gap:
            rows.append(cur_row)
            cur_row = [w]
        else:
            cur_row.append(w)
    rows.append(cur_row)
    results   = []
    seen_skus = set()
    for row in rows:
        texts    = [w["text"] for w in sorted(row, key=lambda x: x["x"])]
        raw_line = " ".join(texts)
        sku_toks = [t for t in texts if _looks_like_product_code(t)]
        if not sku_toks:
            continue
        raw_sku  = sku_toks[0]
        if raw_sku.upper() in seen_skus:
            continue
        seen_skus.add(raw_sku.upper())
        nums = [t for t in texts if re.match(r'^\d+$', t)]
        qty  = int(nums[-1]) if nums else 0
        results.append({"raw_sku": raw_sku, "raw_size": "", "qty": qty, "raw_line": raw_line})
    return results


def _ocr_document_intelligence(image_bytes: bytes, endpoint: str, key: str) -> list:
    import time
    url     = endpoint.rstrip("/") + "/formrecognizer/documentModels/prebuilt-read:analyze?api-version=2023-07-31"
    headers = {"Ocp-Apim-Subscription-Key": key, "Content-Type": "application/octet-stream"}
    resp    = _requests.post(url, headers=headers, data=image_bytes, timeout=30)
    resp.raise_for_status()
    op_url  = resp.headers["Operation-Location"]
    result  = {}
    for _ in range(20):
        time.sleep(1.5)
        r      = _requests.get(op_url, headers={"Ocp-Apim-Subscription-Key": key}, timeout=15)
        result = r.json()
        if result.get("status") == "succeeded":
            break
    analyze = result.get("analyzeResult", {})
    pages   = analyze.get("pages", [])
    words_with_pos = []
    for page in pages:
        pw = page.get("width",  1.0) or 1.0
        ph = page.get("height", 1.0) or 1.0
        for word in page.get("words", []):
            poly = word.get("polygon", [])
            text = word.get("content", "").strip()
            if not text:
                continue
            x = _word_x_center(poly) / pw
            y = _word_y_center(poly) / ph
            words_with_pos.append({"text": text, "x": x, "y": y})
    return _extract_table_geometric(words_with_pos)


def _ocr_computer_vision(image_bytes: bytes, endpoint: str, key: str) -> list:
    import time
    url     = endpoint.rstrip("/") + "/vision/v3.2/read/analyze"
    headers = {"Ocp-Apim-Subscription-Key": key, "Content-Type": "application/octet-stream"}
    resp    = _requests.post(url, headers=headers, data=image_bytes, timeout=30)
    resp.raise_for_status()
    op_url  = resp.headers["Operation-Location"]
    result  = {}
    for _ in range(20):
        time.sleep(1.5)
        r      = _requests.get(op_url, headers={"Ocp-Apim-Subscription-Key": key}, timeout=15)
        result = r.json()
        if result.get("status") == "succeeded":
            break
    analyze   = result.get("analyzeResult", {})
    read_res  = analyze.get("readResults", [])
    words_with_pos = []
    for page in read_res:
        pw = page.get("width",  1.0) or 1.0
        ph = page.get("height", 1.0) or 1.0
        for line in page.get("lines", []):
            for word in line.get("words", []):
                text = word.get("text", "").strip()
                if not text:
                    continue
                bbox = word.get("boundingBox", [])
                xs = [bbox[i] for i in range(0, len(bbox), 2)] if bbox else [0]
                ys = [bbox[i] for i in range(1, len(bbox), 2)] if bbox else [0]
                x  = ((min(xs) + max(xs)) / 2.0) / pw
                y  = ((min(ys) + max(ys)) / 2.0) / ph
                words_with_pos.append({"text": text, "x": x, "y": y})
    return _extract_table_geometric(words_with_pos)


def run_azure_ocr(image_bytes: bytes, azure_endpoint: str, azure_key: str) -> list:
    if not _HAS_REQUESTS:
        return []
    try:
        return _ocr_document_intelligence(image_bytes, azure_endpoint, azure_key)
    except Exception as di_err:
        di_status = getattr(getattr(di_err, "response", None), "status_code", None)
        if di_status in (401, 403, 404):
            return _ocr_computer_vision(image_bytes, azure_endpoint, azure_key)
        raise


def resolve_ocr_rows(raw_rows: list, catalogue: dict, mrp_lookup: dict) -> list:
    norm_index = _build_normalised_index(tuple(catalogue.keys()))
    resolved   = []
    seen_keys  = set()

    for r in raw_rows:
        raw_sku  = r["raw_sku"]
        raw_size = r.get("raw_size", "")
        qty      = r["qty"]

        matched_sku, confidence, match_type = match_sku_to_catalogue(
            raw_sku, catalogue, norm_index
        )

        if raw_size and matched_sku:
            size_norm = re.sub(r'[^0-9]', '', raw_size)
            base_norm = _normalise_sku(raw_sku)
            candidates_for_size = [
                (sku, entry) for sku, entry in catalogue.items()
                if _normalise_sku(sku).startswith(base_norm)
                   and re.sub(r'[^0-9]', '', entry.get("col_label", "")) == size_norm
            ]
            if candidates_for_size:
                matched_sku  = candidates_for_size[0][0]
                match_type   = "exact"
                confidence   = 1.0

        dedup_key = (matched_sku or raw_sku, raw_size)
        if dedup_key in seen_keys:
            continue
        seen_keys.add(dedup_key)

        info = {}
        mrp  = land = 0.0
        if matched_sku:
            info = mrp_lookup.get(matched_sku, {})
            mrp  = info.get("MRP_clean") or 0.0
            land = info.get("Distributor Landing") or mrp

        cat_entry = catalogue.get(matched_sku, {}) if matched_sku else {}
        resolved.append({
            "raw_sku":     raw_sku,
            "raw_size":    raw_size,
            "matched_sku": matched_sku or "—",
            "match_type":  match_type,
            "confidence":  round(confidence * 100, 1),
            "qty":         qty,
            "sheet":       cat_entry.get("sheet", "—"),
            "row_label":   cat_entry.get("row_label", "—"),
            "col_label":   cat_entry.get("col_label", raw_size),
            "mrp":         mrp,
            "landing":     land,
            "amount":      round(land * qty, 2),
            "ri":          cat_entry.get("ri"),
            "ci":          cat_entry.get("ci"),
        })
    return resolved


# ─── PDF generator ────────────────────────────────────────────────────────────
def generate_pdf_from_resolved(resolved_rows, mrp_lookup, bill_to, ship_to, dealer_info):
    buf    = io.BytesIO()
    doc    = SimpleDocTemplate(buf, pagesize=A4,
                               leftMargin=10*mm, rightMargin=10*mm,
                               topMargin=12*mm, bottomMargin=12*mm)
    styles = getSampleStyleSheet()
    navy   = colors.HexColor("#0A2342")
    sky    = colors.HexColor("#1565C0")

    title_style = ParagraphStyle("title", fontName="Helvetica-Bold",
                                 fontSize=14, textColor=colors.white, alignment=TA_CENTER)
    sub_style   = ParagraphStyle("sub", fontName="Helvetica", fontSize=8,
                                 textColor=colors.white, alignment=TA_CENTER)
    story = []

    hdr_data = [[
        Paragraph("SINTEX BAPL LIMITED", title_style),
        Paragraph("Kutesar Road, Raipur, Chhattisgarh – 492101<br/>GSTIN: 22AADCB1921F1ZE", sub_style),
    ]]
    hdr_table = Table(hdr_data, colWidths=["40%", "60%"])
    hdr_table.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,-1), navy),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0), (-1,-1), 8),
        ("BOTTOMPADDING", (0,0), (-1,-1), 8),
        ("LEFTPADDING",   (0,0), (-1,-1), 10),
    ]))
    story.append(hdr_table)
    story.append(Spacer(1, 4*mm))

    title_data = [[
        Paragraph("<b>QUOTATION</b>", ParagraphStyle("qt", fontName="Helvetica-Bold",
                  fontSize=13, textColor=navy)),
        Paragraph(f"<b>Date:</b> {date.today().strftime('%d-%m-%Y')}",
                  ParagraphStyle("qd", fontName="Helvetica", fontSize=8, alignment=TA_RIGHT)),
    ]]
    t = Table(title_data, colWidths=["60%", "40%"])
    t.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"MIDDLE"),
                           ("LINEBELOW",(0,0),(-1,-1),0.5,navy)]))
    story.append(t)
    story.append(Spacer(1, 3*mm))

    if any(dealer_info.values()):
        def party_lines(d, title):
            lines = [f"<b>{title}</b>"]
            for k, v in d.items():
                if v:
                    lines.append(f"<b>{k}:</b> {v}")
            return "<br/>".join(lines)
        dealer_data = [[
            Paragraph(party_lines(dealer_info, "DEALER / DISTRIBUTOR"),
                      ParagraphStyle("pt", fontName="Helvetica", fontSize=7.5, leading=11)),
        ]]
        dt = Table(dealer_data, colWidths=["100%"])
        dt.setStyle(TableStyle([
            ("BOX",          (0,0),(-1,-1), 0.5, navy),
            ("VALIGN",       (0,0),(-1,-1), "TOP"),
            ("TOPPADDING",   (0,0),(-1,-1), 5),
            ("BOTTOMPADDING",(0,0),(-1,-1), 5),
            ("LEFTPADDING",  (0,0),(-1,-1), 8),
            ("BACKGROUND",   (0,0),(-1,-1), colors.HexColor("#EBF3FF")),
        ]))
        story.append(dt)
        story.append(Spacer(1, 3*mm))

    def party_lines2(d, title):
        lines = [f"<b>{title}</b>"]
        for k, v in d.items():
            if v:
                lines.append(f"<b>{k}:</b> {v}")
        return "<br/>".join(lines)

    pt_data = [[
        Paragraph(party_lines2(bill_to, "BILL TO PARTY"),
                  ParagraphStyle("pt", fontName="Helvetica", fontSize=7.5, leading=11)),
        Paragraph(party_lines2(ship_to, "SHIP TO PARTY"),
                  ParagraphStyle("pt", fontName="Helvetica", fontSize=7.5, leading=11)),
    ]]
    pt = Table(pt_data, colWidths=["50%", "50%"])
    pt.setStyle(TableStyle([
        ("BOX",          (0,0),(-1,-1), 0.5, navy),
        ("INNERGRID",    (0,0),(-1,-1), 0.5, colors.HexColor("#C5D0E0")),
        ("VALIGN",       (0,0),(-1,-1), "TOP"),
        ("TOPPADDING",   (0,0),(-1,-1), 6),
        ("BOTTOMPADDING",(0,0),(-1,-1), 6),
        ("LEFTPADDING",  (0,0),(-1,-1), 8),
        ("BACKGROUND",   (0,0),(-1,-1), colors.HexColor("#F4F6FA")),
    ]))
    story.append(pt)
    story.append(Spacer(1, 5*mm))

    tbl_header = ["#", "Item / Description", "SKU Code", "Size", "MRP (₹)", "Qty", "Rate (₹)", "Amount (₹)"]
    tbl_rows   = [tbl_header]
    grand_total = 0.0
    for sno, r in enumerate(resolved_rows, 1):
        if r["qty"] <= 0 or r["matched_sku"] == "—":
            continue
        tbl_rows.append([
            str(sno),
            Paragraph(r["row_label"], ParagraphStyle("rl", fontName="Helvetica", fontSize=7)),
            Paragraph(f'<font name="Courier" size="6.5">{r["matched_sku"]}</font>',
                      ParagraphStyle("sk", fontName="Helvetica", fontSize=7)),
            r["col_label"],
            f"{r['mrp']:,.2f}",
            str(r["qty"]),
            f"{r['landing']:,.2f}",
            f"{r['amount']:,.2f}",
        ])
        grand_total += r["amount"]

    tbl_rows.append(["", "", "", "", "", "",
                     Paragraph("<b>GRAND TOTAL</b>",
                               ParagraphStyle("gt", fontName="Helvetica-Bold", fontSize=8)),
                     Paragraph(f"<b>₹ {grand_total:,.2f}</b>",
                               ParagraphStyle("gv", fontName="Helvetica-Bold", fontSize=8,
                                              alignment=TA_RIGHT))])

    col_widths = [8*mm, 55*mm, 38*mm, 16*mm, 18*mm, 12*mm, 18*mm, 22*mm]
    lt = Table(tbl_rows, colWidths=col_widths, repeatRows=1)
    n  = len(tbl_rows)
    lt.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,0),  navy),
        ("TEXTCOLOR",     (0,0), (-1,0),  colors.white),
        ("FONTNAME",      (0,0), (-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0,0), (-1,0),  7.5),
        ("ALIGN",         (0,0), (-1,-1), "CENTER"),
        ("ALIGN",         (1,1), (1,-1),  "LEFT"),
        ("ALIGN",         (2,1), (2,-1),  "LEFT"),
        ("FONTSIZE",      (0,1), (-1,-1), 7),
        ("ROWBACKGROUNDS",(0,1), (-1,-2), [colors.white, colors.HexColor("#F4F8FF")]),
        ("GRID",          (0,0), (-1,-1), 0.35, colors.HexColor("#C5D0E0")),
        ("TOPPADDING",    (0,0), (-1,-1), 3),
        ("BOTTOMPADDING", (0,0), (-1,-1), 3),
        ("BACKGROUND",    (0,n-1),(-1,n-1), colors.HexColor("#EBF3FF")),
        ("LINEABOVE",     (0,n-1),(-1,n-1), 1, navy),
    ]))
    story.append(lt)
    story.append(Spacer(1, 6*mm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=navy))
    story.append(Spacer(1, 2*mm))
    story.append(Paragraph(
        "This is a computer-generated quotation. Prices are subject to change without notice. "
        "Taxes applicable as per prevailing rates.",
        ParagraphStyle("footer", fontName="Helvetica", fontSize=7, textColor=colors.grey,
                       alignment=TA_CENTER)
    ))
    doc.build(story)
    return buf.getvalue()


# ─── Session-state defaults ───────────────────────────────────────────────────
DEFAULTS = {
    "image_bytes":       None,
    "ocr_raw_rows":      [],
    "ocr_resolved":      [],
    "ocr_done":          False,
    "qty_map":           {},
    "bill_to":           {},
    "ship_to":           {},
    "dealer_info":       {},
    "pdf_bytes":         None,
    "azure_endpoint":    "",
    "azure_key":         "",
    "_cam_recv_val":     "",
}
for k, v in DEFAULTS.items():
    if k not in st.session_state:
        st.session_state[k] = v

mrp_lookup  = load_mrp()
zsd_df      = load_zsd()
sku_sheets  = load_sku_sheets()
sku_cat     = build_sku_catalogue()

# ─── UI helpers ──────────────────────────────────────────────────────────────
def party_html(d, title):
    rows = "".join(
        f'<div class="party-row"><span class="party-lbl">{k}</span>'
        f'<span class="party-val">{v or "—"}</span></div>'
        for k, v in d.items()
    )
    return f'<div class="party-box"><h4>{title}</h4>{rows}</div>'

def section_header(num, icon, title):
    st.markdown(f"""
    <div class="section-anchor">
      <span class="section-num">{num}</span>
      <h2>{icon} {title}</h2>
    </div>
    """, unsafe_allow_html=True)


# ─── App header ──────────────────────────────────────────────────────────────
st.markdown("""
<div class="app-header">
  <div>
    <h1>🔧 Sintex BAPL – Quotation Generator</h1>
    <p>CPVC / UPVC Pipes &amp; Fittings · Chhattisgarh Price List · Single-page workflow</p>
  </div>
</div>
""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 1 — Dealer / Distributor + Party Details
# ═══════════════════════════════════════════════════════════════════════════════
section_header(1, "🏪", "Dealer / Distributor & Customer Details")

with st.expander("Expand to fill party details", expanded=True):
    st.markdown("#### 🏪 Dealer / Distributor Information")
    st.caption("Enter the details of the Dealer or Distributor raising this quotation.")

    da1, da2 = st.columns(2)
    dealer_name   = da1.text_input("Dealer / Distributor Name",  key="dealer_name")
    dealer_code   = da2.text_input("Dealer Code",                key="dealer_code")
    dealer_addr   = st.text_input("Address",                     key="dealer_address")
    da3, da4 = st.columns(2)
    dealer_phone  = da3.text_input("Phone",                      key="dealer_phone")
    dealer_gst    = da4.text_input("GST No.",                    key="dealer_gst")
    da5, da6 = st.columns(2)
    dealer_state  = da5.text_input("State",                      key="dealer_state")
    dealer_pan    = da6.text_input("PAN No.",                    key="dealer_pan")

    st.session_state.dealer_info = {
        "Name":    dealer_name,
        "Code":    dealer_code,
        "Address": dealer_addr,
        "Phone":   dealer_phone,
        "GST No.": dealer_gst,
        "State":   dealer_state,
        "PAN No.": dealer_pan,
    }

    st.markdown("---")
    st.markdown("#### 👤 Customer Information")
    cust_options = ["— Select or type below —"] + [
        f'{row["Customer Code"]} | {row["Customer Name"]}'
        for _, row in zsd_df.iterrows()
    ]
    sel = st.selectbox("🔍 Search Customer from Master", cust_options, key="zsd_search")

    def zsd_fill(prefix, row):
        addr = " ".join(filter(None, [
            str(row.get("Address 1","") or ""),
            str(row.get("Address 2","") or ""),
            str(row.get("Address 3","") or ""),
            str(row.get("City","")      or ""),
        ]))
        return {
            f"{prefix}_party_no":   str(row.get("Customer Code","") or ""),
            f"{prefix}_party_name": str(row.get("Customer Name","") or ""),
            f"{prefix}_address":    addr.strip(),
            f"{prefix}_phone":      str(row.get("Telephone","") or ""),
            f"{prefix}_mobile":     str(row.get("Mobile No.","") or ""),
            f"{prefix}_state_code": str(row.get("State Code","") or ""),
            f"{prefix}_state":      str(row.get("State Code Desc.","") or ""),
            f"{prefix}_gst":        str(row.get("GST Number","") or ""),
            f"{prefix}_pan":        str(row.get("PAN No.","") or ""),
        }

    if sel != cust_options[0]:
        code    = sel.split("|")[0].strip()
        matched = zsd_df[zsd_df["Customer Code"].astype(str) == code]
        if not matched.empty:
            row = matched.iloc[0]
            for k, v in zsd_fill("bill", row).items():
                if k not in st.session_state or not st.session_state[k]:
                    st.session_state[k] = v
            st.toast("Bill-to details filled from customer master.", icon="✅")

    st.markdown("#### 📋 Bill To Party")
    bc1, bc2 = st.columns(2)
    bill_party_no   = bc1.text_input("Bill to Party No.",   key="bill_party_no")
    bill_party_name = bc2.text_input("Bill to Party Name",  key="bill_party_name")
    bill_address    = st.text_input("Bill to Address",       key="bill_address")
    bc3, bc4 = st.columns(2)
    bill_phone      = bc3.text_input("Phone",                key="bill_phone")
    bill_mobile     = bc4.text_input("Mobile",               key="bill_mobile")
    bc5, bc6 = st.columns(2)
    bill_sc         = bc5.text_input("State Code",           key="bill_state_code")
    bill_state      = bc6.text_input("State",                key="bill_state")
    bc7, bc8 = st.columns(2)
    bill_gst        = bc7.text_input("GST No.",              key="bill_gst")
    bill_pan        = bc8.text_input("PAN No.",              key="bill_pan")

    same_as_bill = st.checkbox("Ship-to same as Bill-to", value=False, key="same_as_bill_chk")

    st.markdown("#### 🚚 Ship To Party")
    if same_as_bill:
        st.session_state["ship_party_no"]   = bill_party_no
        st.session_state["ship_party_name"] = bill_party_name
        st.session_state["ship_address"]    = bill_address
        st.session_state["ship_phone"]      = bill_phone
        st.session_state["ship_mobile"]     = bill_mobile
        st.session_state["ship_state_code"] = bill_sc
        st.session_state["ship_state"]      = bill_state
        st.session_state["ship_gst"]        = bill_gst
        st.session_state["ship_pan"]        = bill_pan

    sc1, sc2 = st.columns(2)
    ship_party_no   = sc1.text_input("Ship to Party No.",   key="ship_party_no")
    ship_party_name = sc2.text_input("Ship to Party Name",  key="ship_party_name")
    ship_address    = st.text_input("Ship to Address",       key="ship_address")
    sc3, sc4 = st.columns(2)
    ship_phone      = sc3.text_input("Phone ",               key="ship_phone")
    ship_mobile     = sc4.text_input("Mobile ",              key="ship_mobile")
    sc5, sc6 = st.columns(2)
    ship_sc         = sc5.text_input("State Code ",          key="ship_state_code")
    ship_state      = sc6.text_input("State ",               key="ship_state")
    sc7, sc8 = st.columns(2)
    ship_gst        = sc7.text_input("GST No. ",             key="ship_gst")
    ship_pan        = sc8.text_input("PAN No. ",             key="ship_pan")

    st.session_state.bill_to = {
        "Party No.":  bill_party_no,
        "Name":       bill_party_name,
        "Address":    bill_address,
        "Phone":      bill_phone,
        "Mobile":     bill_mobile,
        "State Code": bill_sc,
        "State":      bill_state,
        "GST No.":    bill_gst,
        "PAN No.":    bill_pan,
    }
    st.session_state.ship_to = {
        "Party No.":  ship_party_no,
        "Name":       ship_party_name,
        "Address":    ship_address,
        "Phone":      ship_phone,
        "Mobile":     ship_mobile,
        "State Code": ship_sc,
        "State":      ship_state,
        "GST No.":    ship_gst,
        "PAN No.":    ship_pan,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 2 — Capture Image & Run OCR
# ═══════════════════════════════════════════════════════════════════════════════
section_header(2, "📸", "Capture Handwritten Order Form")

st.markdown(
    "📸 **Photograph or upload** the handwritten order form. "
    "Azure OCR will detect SKU codes and quantities, match them to the product catalogue."
)

with st.expander("🔑 Azure OCR Credentials", expanded=not st.session_state.ocr_done):
    ep  = st.text_input("Azure Endpoint", value=st.session_state.azure_endpoint,
                        placeholder="https://YOUR_RESOURCE.cognitiveservices.azure.com",
                        key="az_ep_input")
    key = st.text_input("Azure Key", value=st.session_state.azure_key,
                        type="password", placeholder="32-character subscription key",
                        key="az_key_input")
    st.session_state.azure_endpoint = ep
    st.session_state.azure_key      = key

img_mode = st.radio("Image source",
                    ["📷  Camera (recommended)", "📁  Upload File"],
                    horizontal=True, label_visibility="collapsed",
                    key="img_mode_radio")

def _on_file_upload():
    uf = st.session_state.get("file_upload_input")
    if uf is not None:
        st.session_state.image_bytes  = uf.getvalue()
        st.session_state.ocr_done     = False
        st.session_state.ocr_raw_rows = []
        st.session_state.ocr_resolved = []
        st.session_state.pdf_bytes    = None

if img_mode == "📷  Camera (recommended)":
    import streamlit.components.v1 as components

    CAMERA_HTML = """
<style>
* { box-sizing:border-box; margin:0; padding:0; }
body { background:transparent; font-family:'IBM Plex Sans',sans-serif; }
#cam-wrap {
  width:100%; background:linear-gradient(160deg,#071829 0%,#0d2d56 100%);
  border-radius:14px; overflow:hidden; display:flex; flex-direction:column;
  align-items:center; box-shadow:0 8px 32px rgba(10,35,66,.4);
}
#video  { width:100%; max-height:62vh; object-fit:cover; display:block; }
#canvas { display:none; }
#preview{ width:100%; display:none; border-bottom:3px solid #1E88E5; }
.toolbar{ display:flex; gap:12px; padding:16px 20px 18px; width:100%;
  background:rgba(0,0,0,.3); justify-content:center; flex-wrap:wrap; }
.cam-btn{ display:inline-flex; align-items:center; justify-content:center; gap:8px;
  padding:12px 32px; border:none; border-radius:10px; font-size:14px; font-weight:700;
  cursor:pointer; transition:all .18s; letter-spacing:.3px; min-width:160px;
  font-family:'IBM Plex Sans',sans-serif; }
#btn-capture{ background:linear-gradient(135deg,#1565C0 0%,#1E88E5 100%); color:#fff;
  box-shadow:0 4px 16px rgba(21,101,192,.5); }
#btn-capture:hover{ transform:translateY(-2px); box-shadow:0 8px 24px rgba(21,101,192,.65); }
#btn-retake { background:rgba(255,255,255,.1); color:rgba(255,255,255,.9);
  border:1.5px solid rgba(255,255,255,.25); display:none; }
#btn-retake:hover{ background:rgba(255,255,255,.2); transform:translateY(-1px); }
.status-bar{ display:flex; align-items:center; gap:10px;
  background:rgba(0,0,0,.25); backdrop-filter:blur(8px);
  border-top:1px solid rgba(255,255,255,.07); padding:10px 20px; width:100%;
  font-size:12.5px; font-weight:500; color:rgba(255,255,255,.7); min-height:42px; }
.status-bar.success{ background:rgba(0,121,107,.3); color:#80CBC4; }
.status-bar.sending{ background:rgba(249,168,37,.12); color:#FFD54F; }
.status-bar.error  { background:rgba(198,40,40,.2);  color:#EF9A9A; }
.pulse-dot { width:8px; height:8px; border-radius:50%; background:currentColor;
  flex-shrink:0; animation:pulse 1.5s infinite ease-in-out; }
@keyframes pulse{ 0%,100%{opacity:1;transform:scale(1)} 50%{opacity:.3;transform:scale(.65)} }
.spinner{ width:16px; height:16px; flex-shrink:0; border:2.5px solid currentColor;
  border-top-color:transparent; border-radius:50%; animation:spin .65s linear infinite; }
@keyframes spin{ to{ transform:rotate(360deg); } }
.icon-static{ flex-shrink:0; font-size:15px; line-height:1; }
</style>
<div id="cam-wrap">
  <video id="video" autoplay playsinline muted></video>
  <canvas id="canvas"></canvas>
  <img id="preview" alt="captured photo"/>
  <div class="toolbar">
    <button class="cam-btn" id="btn-capture">📷&nbsp;&nbsp;Capture Photo</button>
    <button class="cam-btn" id="btn-retake">🔄&nbsp;&nbsp;Retake</button>
  </div>
  <div class="status-bar ready" id="status-bar">
    <span class="pulse-dot" id="status-icon"></span>
    <span id="status-txt">Starting camera…</span>
  </div>
</div>
<script>
(function(){
  const video=document.getElementById('video'),canvas=document.getElementById('canvas'),
        preview=document.getElementById('preview'),btnCap=document.getElementById('btn-capture'),
        btnRet=document.getElementById('btn-retake'),bar=document.getElementById('status-bar'),
        icon=document.getElementById('status-icon'),txt=document.getElementById('status-txt');
  function setStatus(m,mode){
    txt.textContent=m; bar.className='status-bar '+(mode||'ready');
    icon.className=''; icon.textContent=''; icon.style.cssText='';
    if(mode==='sending') icon.className='spinner';
    else if(mode==='success'){icon.className='icon-static';icon.textContent='✓';}
    else if(mode==='error') {icon.className='icon-static';icon.textContent='⚠';}
    else icon.className='pulse-dot';
  }
  async function startCamera(){
    try{
      const s=await navigator.mediaDevices.getUserMedia(
        {video:{facingMode:{ideal:'environment'},width:{ideal:4096},height:{ideal:3072}},audio:false});
      video.srcObject=s; await video.play();
      setStatus('Camera ready — frame the order form and tap Capture','ready');
    }catch(e){setStatus('Camera error: '+e.message,'error');}
  }
  function convolve3x3(data,w,h,kernel){
    const out=new Uint8ClampedArray(data.length);
    for(let y=1;y<h-1;y++) for(let x=1;x<w-1;x++){
      for(let c=0;c<3;c++){
        let sum=0;
        for(let ky=-1;ky<=1;ky++) for(let kx=-1;kx<=1;kx++)
          sum+=data[((y+ky)*w+(x+kx))*4+c]*kernel[(ky+1)*3+(kx+1)];
        out[(y*w+x)*4+c]=Math.max(0,Math.min(255,sum));
      }
      out[(y*w+x)*4+3]=255;
    }
    return out;
  }
  function sharpenAndEnhance(ctx,w,h){
    const imgData=ctx.getImageData(0,0,w,h),d=imgData.data;
    for(let i=0;i<d.length;i+=4)
      for(let c=0;c<3;c++){
        let v=Math.round(((d[i+c]/255-.5)*1.25+.5)*255);
        d[i+c]=Math.max(0,Math.min(255,v));
      }
    const kernel=[0,-1,0,-1,5,-1,0,-1,0],out=ctx.createImageData(w,h);
    out.data.set(convolve3x3(d,w,h,kernel)); ctx.putImageData(out,0,0);
  }
  btnCap.addEventListener('click',()=>{
    const w=video.videoWidth||1280,h=video.videoHeight||720;
    canvas.width=w; canvas.height=h;
    const ctx=canvas.getContext('2d');
    ctx.drawImage(video,0,0,w,h); sharpenAndEnhance(ctx,w,h);
    const b64=canvas.toDataURL('image/jpeg',0.97);
    preview.src=b64; video.style.display='none';
    preview.style.display='block'; btnCap.style.display='none'; btnRet.style.display='block';
    setStatus('Sending photo to app…','sending');
    window.parent.sessionStorage.setItem('sintex_cam_b64',b64);
    window.parent.postMessage({type:'SINTEX_CAM_CAPTURE',data:b64},'*');
    setTimeout(()=>setStatus('Photo captured — use Retake to redo, or run OCR below','success'),700);
  });
  btnRet.addEventListener('click',()=>{
    preview.style.display='none'; video.style.display='block';
    btnCap.style.display='block'; btnRet.style.display='none';
    window.parent.sessionStorage.removeItem('sintex_cam_b64');
    window.parent.postMessage({type:'SINTEX_CAM_RETAKE'},'*');
    setStatus('Camera ready — frame the order form and tap Capture','ready');
  });
  startCamera();
})();
</script>
"""
    components.html(CAMERA_HTML, height=580, scrolling=False)

    AUTO_BRIDGE = """
<script>
(function(){
  var attempts=0,MAX=120;
  function tryInject(){
    attempts++;
    if(attempts>MAX)return;
    var b64=window.parent.sessionStorage.getItem('sintex_cam_b64');
    if(!b64){setTimeout(tryInject,500);return;}
    var parent=window.parent.document,inputs=parent.querySelectorAll('input[type="text"]'),target=null;
    for(var i=inputs.length-1;i>=0;i--){
      var inp=inputs[i],st=window.parent.getComputedStyle(inp);
      if(parseFloat(st.height)<5||parseFloat(st.opacity)<0.1){target=inp;break;}
    }
    if(!target&&inputs.length)target=inputs[inputs.length-1];
    if(!target){setTimeout(tryInject,500);return;}
    var setter=Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype,'value').set;
    setter.call(target,b64);
    target.dispatchEvent(new Event('input',{bubbles:true}));
    target.dispatchEvent(new Event('change',{bubbles:true}));
    window.parent.sessionStorage.removeItem('sintex_cam_b64');
  }
  setTimeout(tryInject,800);
})();
</script>
"""
    components.html(AUTO_BRIDGE, height=0)

    cam_val = st.text_input(
        "sintex_cam_hidden_receiver",
        value=st.session_state.get("_cam_recv_val", ""),
        label_visibility="collapsed",
        key="sintex_cam_recv",
    )

    recv = st.session_state.get("sintex_cam_recv", "")
    if recv and recv.startswith("data:image"):
        import base64 as _b64
        _, encoded = recv.split(",", 1)
        raw = _b64.b64decode(encoded)
        if raw != st.session_state.image_bytes:
            st.session_state.image_bytes      = raw
            st.session_state.ocr_done         = False
            st.session_state.ocr_raw_rows     = []
            st.session_state.ocr_resolved     = []
            st.session_state["_cam_recv_val"] = recv
            st.session_state.pdf_bytes        = None
            st.rerun()

    components.html("""
<script>
(function(){
  var parent=window.parent.document,inputs=parent.querySelectorAll('input[type="text"]');
  if(inputs.length){
    var last=inputs[inputs.length-1];
    last.style.cssText+='height:1px!important;opacity:0!important;pointer-events:none!important;position:absolute!important;';
    last.setAttribute('data-sintex-cam','1');
    var wrap=last.closest('[data-testid="stTextInput"]');
    if(wrap)wrap.style.cssText+='height:0!important;overflow:hidden!important;margin:0!important;padding:0!important;';
  }
})();
</script>
""", height=0)

else:
    st.file_uploader("Upload image of order form", type=["jpg","jpeg","png"],
                     label_visibility="collapsed",
                     key="file_upload_input", on_change=_on_file_upload)

# Image preview
if st.session_state.image_bytes:
    with st.expander("📸 View Captured Image", expanded=False):
        st.image(st.session_state.image_bytes, use_container_width=True)

    ocr_col1, ocr_col2 = st.columns([1, 4])
    with ocr_col1:
        if st.button("🔍  Run OCR", key="run_ocr_btn"):
            _ep  = st.session_state.azure_endpoint.strip()
            _key = st.session_state.azure_key.strip()
            if not _ep or not _key:
                st.error("Please enter your **Azure Endpoint** and **Azure Key** above.", icon="🔑")
            else:
                with st.spinner("Sending image to Azure OCR… this may take 10–20 seconds"):
                    try:
                        raw_rows = run_azure_ocr(st.session_state.image_bytes, _ep, _key)
                        resolved = resolve_ocr_rows(raw_rows, sku_cat, mrp_lookup)
                        st.session_state.ocr_raw_rows = raw_rows
                        st.session_state.ocr_resolved = resolved
                        st.session_state.ocr_done     = True
                        st.session_state.pdf_bytes    = None
                        st.rerun()
                    except Exception as exc:
                        st.error(f"OCR failed: {exc}", icon="❌")
else:
    st.markdown("""
    <div style='background:#F4F6FA;border:2px dashed #DEE3EC;border-radius:10px;
                padding:40px;text-align:center;margin:14px 0;color:#5A6880;'>
      <div style='font-size:36px;margin-bottom:10px;'>🖼️</div>
      <div style='font-weight:600;font-size:14px;margin-bottom:4px;'>No image loaded</div>
      <div style='font-size:12px;'>Capture or upload an image above</div>
    </div>""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 3 — OCR Results: Summary Table (read-only) + Edit Section
# ═══════════════════════════════════════════════════════════════════════════════
section_header(3, "📋", "OCR Results & Order Review")

resolved = st.session_state.get("ocr_resolved", [])

if not resolved and not st.session_state.ocr_done:
    st.info("Run OCR above, or add items manually in the edit section below.", icon="ℹ️")

# ── 3A: READ-ONLY SUMMARY TABLE ──────────────────────────────────────────────
if resolved:
    n_matched = sum(1 for r in resolved if r["match_type"] != "none")
    n_nonzero = sum(1 for r in resolved if r["qty"] > 0)
    n_total   = len(resolved)

    st.success(
        f"✅ OCR detected **{n_total} rows** — **{n_matched}** matched to catalogue, "
        f"**{n_nonzero}** with non-zero quantity.",
        icon="🔍",
    )

    st.markdown("### 📊 Detected Items — Summary View")
    st.caption("Read-only overview of what the OCR found. Edit quantities and SKUs in the section below.")

    # Build HTML table
    grand_mrp_sum  = 0.0
    grand_land_sum = 0.0
    rows_html = ""
    for idx, r in enumerate(resolved):
        mt = r["match_type"]
        badge_cls  = "match-ok" if mt == "exact" else ("match-warn" if mt == "fuzzy" else "match-none")
        badge_icon = "✓ Exact" if mt == "exact" else ("~ Fuzzy" if mt == "fuzzy" else "✗ None")
        sku_disp   = r["matched_sku"] if r["matched_sku"] != "—" else "<em style='color:#aaa'>—</em>"
        mrp_v      = r["mrp"]
        land_v     = r["landing"]
        qty_v      = r["qty"]
        amount_v   = r["amount"]
        grand_mrp_sum  += mrp_v * qty_v
        grand_land_sum += amount_v

        # Highlight rows with zero qty in muted style
        row_style = "" if qty_v > 0 else "opacity:0.5;"

        rows_html += f"""
        <tr style="{row_style}">
          <td>{idx + 1}</td>
          <td style="font-weight:500;">{r['row_label']}</td>
          <td><span class="sku-mono">{sku_disp}</span></td>
          <td>{r.get('col_label') or r.get('raw_size','—')}</td>
          <td><span class="{badge_cls}">{badge_icon}</span></td>
          <td class="num">₹ {mrp_v:,.2f}</td>
          <td class="num">{qty_v}</td>
          <td class="num">₹ {land_v:,.2f}</td>
          <td class="num final-price">₹ {amount_v:,.2f}</td>
        </tr>
        """

    st.markdown(f"""
    <div style="overflow-x:auto;">
    <table class="ocr-summary-table">
      <thead>
        <tr>
          <th>#</th>
          <th>Product Name</th>
          <th>SKU Code</th>
          <th>Size</th>
          <th>Match</th>
          <th class="num">MRP (₹)</th>
          <th class="num">Qty</th>
          <th class="num">Rate (₹)</th>
          <th class="num">Final Price (₹)</th>
        </tr>
      </thead>
      <tbody>
        {rows_html}
      </tbody>
      <tfoot>
        <tr>
          <td colspan="5" style="text-align:right;font-weight:700;">TOTALS</td>
          <td class="num">₹ {grand_mrp_sum:,.2f}</td>
          <td class="num">{sum(r['qty'] for r in resolved)}</td>
          <td class="num">—</td>
          <td class="num final-price">₹ {grand_land_sum:,.2f}</td>
        </tr>
      </tfoot>
    </table>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

# ── 3B: EDIT SECTION ─────────────────────────────────────────────────────────
st.markdown("### ✏️ Edit Items")
st.caption(
    "Non-zero detected items appear first. Items with zero or undetected quantities appear below. "
    "Correct SKUs, sizes, and quantities as needed."
)

# Partition: detected (qty > 0 and matched) vs undetected/zero
detected_rows   = [r for r in resolved if r["qty"] > 0 and r["matched_sku"] != "—"]
undetected_rows = [r for r in resolved if not (r["qty"] > 0 and r["matched_sku"] != "—")]

# Also allow adding new manual rows
if "manual_rows" not in st.session_state:
    st.session_state.manual_rows = []

_ni_edit = _build_normalised_index(tuple(sku_cat.keys()))

def render_edit_row(r, idx, prefix):
    """Render one editable row. Returns updated r dict."""
    cols = st.columns([0.3, 1.8, 1.5, 1.0, 0.9, 1.1, 1.1, 0.9])

    cols[0].markdown(
        f'<div style="font-size:11px;color:#9BAECC;padding:8px 2px;text-align:center;">'
        f'{idx+1}</div>',
        unsafe_allow_html=True,
    )

    # Raw OCR reference
    cols[1].markdown(
        f'<div style="font-size:10px;color:#5A6880;padding:4px 2px;line-height:1.4;">'
        f'<b>Raw:</b> <code>{r["raw_sku"]}</code></div>',
        unsafe_allow_html=True,
    )

    new_sku = cols[2].text_input(
        "SKU", value=r["matched_sku"] if r["matched_sku"] != "—" else "",
        label_visibility="visible",
        key=f"{prefix}_sku_{idx}",
        placeholder="SKU code",
    )

    new_size = cols[3].text_input(
        "Size", value=r.get("raw_size", ""),
        label_visibility="visible",
        key=f"{prefix}_size_{idx}",
        placeholder="e.g. 15MM",
    )

    # Re-resolve if SKU or size changed
    changed = (new_sku and new_sku != r.get("matched_sku","—")) or (new_size != r.get("raw_size",""))
    if changed:
        probe_sku  = new_sku if new_sku else r["raw_sku"]
        re_sku, re_conf, re_type = match_sku_to_catalogue(probe_sku, sku_cat, _ni_edit)
        if re_sku:
            size_norm = re.sub(r"[^0-9]", "", new_size or r.get("raw_size",""))
            base_norm = _normalise_sku(probe_sku)
            size_hits = [
                (sku, ent) for sku, ent in sku_cat.items()
                if _normalise_sku(sku).startswith(base_norm)
                   and re.sub(r"[^0-9]", "", ent.get("col_label","")) == size_norm
            ]
            if size_hits:
                re_sku, cat_entry = size_hits[0]
                re_type = "exact"; re_conf = 1.0
            else:
                cat_entry = sku_cat[re_sku]
            info = mrp_lookup.get(re_sku, {})
            r = {**r,
                 "matched_sku": re_sku,
                 "raw_size":    new_size or r.get("raw_size",""),
                 "match_type":  re_type,
                 "confidence":  round(re_conf * 100, 1),
                 "sheet":       cat_entry.get("sheet","—"),
                 "row_label":   cat_entry.get("row_label","—"),
                 "col_label":   cat_entry.get("col_label","—"),
                 "mrp":         info.get("MRP_clean") or 0.0,
                 "landing":     info.get("Distributor Landing") or 0.0,
                 "ri":          cat_entry.get("ri"),
                 "ci":          cat_entry.get("ci"),
            }
        else:
            r = {**r, "matched_sku": new_sku or "—", "raw_size": new_size,
                 "match_type": "none", "confidence": 0.0,
                 "mrp": 0.0, "landing": 0.0, "ri": None, "ci": None}

    # MRP display
    cols[4].markdown(
        f'<div style="font-size:11px;font-family:monospace;padding:8px 2px;text-align:right;">'
        f'₹ {r["mrp"]:,.2f}</div>',
        unsafe_allow_html=True,
    )

    # Landing rate display
    cols[5].markdown(
        f'<div style="font-size:11px;font-family:monospace;padding:8px 2px;text-align:right;">'
        f'₹ {r["landing"]:,.2f}</div>',
        unsafe_allow_html=True,
    )

    new_qty = cols[6].number_input(
        "Qty", value=int(r["qty"]), min_value=0, step=1,
        label_visibility="visible",
        key=f"{prefix}_qty_{idx}",
    )

    amount = round(r["landing"] * new_qty, 2)
    cols[7].markdown(
        f'<div style="font-size:12px;font-family:monospace;font-weight:700;'
        f'color:#0A2342;padding:8px 2px;text-align:right;">₹ {amount:,.2f}</div>',
        unsafe_allow_html=True,
    )

    return {**r, "qty": new_qty, "amount": amount}


# Column headers for edit table
edit_hdr = st.columns([0.3, 1.8, 1.5, 1.0, 0.9, 1.1, 1.1, 0.9])
for col, lbl in zip(edit_hdr, ["#", "Raw OCR", "SKU Code ✏️", "Size ✏️", "MRP (₹)", "Rate (₹)", "Qty ✏️", "Amount (₹)"]):
    col.markdown(
        f"<div style='font-size:10.5px;font-weight:700;color:#0A2342;"
        f"background:#EBF3FF;border-radius:4px;padding:3px 8px;'>{lbl}</div>",
        unsafe_allow_html=True,
    )

st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)

updated_resolved = []

# — Detected (non-zero) rows first —
if detected_rows:
    st.markdown(
        '<span class="edit-group-label edit-group-detected">✅ Detected with Quantity</span>',
        unsafe_allow_html=True,
    )
    for idx, r in enumerate(detected_rows):
        updated_r = render_edit_row(r, idx, "det")
        updated_resolved.append(updated_r)
    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

# — Undetected / zero qty rows —
if undetected_rows:
    st.markdown(
        '<span class="edit-group-label edit-group-undetected">⚠️ Zero Qty / Unmatched — Review These</span>',
        unsafe_allow_html=True,
    )
    for idx, r in enumerate(undetected_rows):
        updated_r = render_edit_row(r, len(detected_rows) + idx, "und")
        updated_resolved.append(updated_r)
    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

# — Manual add rows —
if not resolved:
    st.markdown(
        '<span class="edit-group-label" style="background:#EBF3FF;color:#1565C0;">➕ Manual Entry</span>',
        unsafe_allow_html=True,
    )

manual_add_rows = []
for mi, mr in enumerate(st.session_state.manual_rows):
    mc1, mc2, mc3, mc4 = st.columns([2, 1, 1, 0.5])
    sku_val = mc1.text_input("SKU Code", value=mr.get("sku",""),
                             key=f"manual_sku_{mi}", label_visibility="visible")
    size_val = mc2.text_input("Size", value=mr.get("size",""),
                              key=f"manual_size_{mi}", label_visibility="visible",
                              placeholder="e.g. 15MM")
    qty_val  = mc3.number_input("Qty", value=int(mr.get("qty",0)), min_value=0, step=1,
                                key=f"manual_qty_{mi}", label_visibility="visible")
    manual_add_rows.append({"sku": sku_val, "size": size_val, "qty": qty_val})

st.session_state.manual_rows = manual_add_rows

add_col, apply_col, _ = st.columns([1, 1, 4])
with add_col:
    if st.button("➕ Add Row", key="add_manual_row"):
        st.session_state.manual_rows = manual_add_rows + [{"sku": "", "size": "", "qty": 0}]
        st.rerun()
with apply_col:
    if st.button("✅ Apply Manual Entries", key="apply_manual"):
        for mr in st.session_state.manual_rows:
            sku = mr["sku"].strip()
            qty = mr["qty"]
            if not sku:
                continue
            m_sku, conf, mtype = match_sku_to_catalogue(sku, sku_cat, _ni_edit)
            cat_entry = sku_cat.get(m_sku, {}) if m_sku else {}
            info      = mrp_lookup.get(m_sku, {}) if m_sku else {}
            mrp_v     = info.get("MRP_clean") or 0.0
            land_v    = info.get("Distributor Landing") or mrp_v
            updated_resolved.append({
                "raw_sku":     sku,
                "raw_size":    mr.get("size",""),
                "matched_sku": m_sku or "—",
                "match_type":  mtype,
                "confidence":  round(conf * 100, 1),
                "qty":         qty,
                "sheet":       cat_entry.get("sheet","—"),
                "row_label":   cat_entry.get("row_label","—"),
                "col_label":   cat_entry.get("col_label","—"),
                "mrp":         mrp_v,
                "landing":     land_v,
                "amount":      round(land_v * qty, 2),
                "ri":          cat_entry.get("ri"),
                "ci":          cat_entry.get("ci"),
            })
        st.session_state.ocr_resolved = updated_resolved
        st.session_state.manual_rows  = []
        st.success(f"✅ Applied.", icon="✅")
        st.rerun()

# Save updated resolved back
if updated_resolved or resolved:
    st.session_state.ocr_resolved = updated_resolved if updated_resolved else resolved

# ── Live totals ──────────────────────────────────────────────────────────────
resolved_now = st.session_state.ocr_resolved
grand_mrp_l  = sum(r["mrp"]     * r["qty"] for r in resolved_now if r["matched_sku"] != "—")
grand_land_l = sum(r["landing"] * r["qty"] for r in resolved_now if r["matched_sku"] != "—")
n_lines_l    = sum(1 for r in resolved_now if r["qty"] > 0 and r["matched_sku"] != "—")

st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)
st.markdown(f"""
<div class="card" style="max-width:440px;">
  <div class="card-title">📊 Live Summary</div>
  <div class="totals-row">
    <span>Active Line Items</span><span class="val">{n_lines_l}</span>
  </div>
  <div class="totals-row">
    <span>Gross MRP Value</span><span class="val">₹ {grand_mrp_l:,.2f}</span>
  </div>
  <div class="totals-row">
    <span>Distributor Landing</span><span class="val">₹ {grand_land_l:,.2f}</span>
  </div>
  <div class="totals-row net">
    <span>Net Payable (Landing)</span><span class="val">₹ {grand_land_l:,.2f}</span>
  </div>
</div>
""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 4 — Download
# ═══════════════════════════════════════════════════════════════════════════════
section_header(4, "⬇️", "Download Quotation")

resolved_rows_dl = [r for r in st.session_state.ocr_resolved
                    if r["qty"] > 0 and r["matched_sku"] != "—"]

if not resolved_rows_dl:
    st.warning("No valid line items with matched SKUs and non-zero quantities yet. Fill in items above.", icon="⚠️")
else:
    # Party preview
    pc1, pc2 = st.columns(2)
    with pc1:
        st.markdown(party_html(st.session_state.bill_to, "BILL TO PARTY"), unsafe_allow_html=True)
    with pc2:
        st.markdown(party_html(st.session_state.ship_to, "SHIP TO PARTY"), unsafe_allow_html=True)
    if any(st.session_state.dealer_info.values()):
        st.markdown(party_html(st.session_state.dealer_info, "DEALER / DISTRIBUTOR"),
                    unsafe_allow_html=True)

    st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)
    st.markdown("#### 📦 Final Order Lines")

    line_data   = []
    grand_mrp_d = grand_land_d = 0.0
    for r in resolved_rows_dl:
        grand_mrp_d  += r["mrp"]     * r["qty"]
        grand_land_d += r["landing"] * r["qty"]
        line_data.append({
            "Item":       r["row_label"],
            "SKU Code":   r["matched_sku"],
            "Size":       r["col_label"],
            "Sheet":      r["sheet"],
            "MRP (₹)":    round(r["mrp"],    2),
            "Qty":        r["qty"],
            "Rate (₹)":   round(r["landing"], 2),
            "Amount (₹)": round(r["amount"],  2),
        })

    df_lines = pd.DataFrame(line_data)
    st.dataframe(df_lines, use_container_width=True, hide_index=True,
                 column_config={
                     "MRP (₹)":    st.column_config.NumberColumn(format="₹ %.2f"),
                     "Rate (₹)":   st.column_config.NumberColumn(format="₹ %.2f"),
                     "Amount (₹)": st.column_config.NumberColumn(format="₹ %.2f"),
                 })

    discount_d = grand_mrp_d - grand_land_d
    st.markdown(f"""
    <div class="card" style="max-width:440px;margin-top:12px;">
      <div class="card-title">💰 Final Totals</div>
      <div class="totals-row">
        <span>Gross MRP Value</span><span class="val">₹ {grand_mrp_d:,.2f}</span>
      </div>
      <div class="totals-row">
        <span>Distributor Discount</span><span class="val neg">− ₹ {discount_d:,.2f}</span>
      </div>
      <div class="totals-row net">
        <span>Net Payable (Distributor Landing)</span>
        <span class="val">₹ {grand_land_d:,.2f}</span>
      </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)
    dl1, dl2, dl3, dl4 = st.columns([1, 1, 1, 2])

    # CSV
    csv_buf = io.StringIO()
    df_lines.to_csv(csv_buf, index=False)
    dl1.download_button(
        "⬇  Download CSV",
        data=csv_buf.getvalue().encode(),
        file_name=f"sintex_quotation_{date.today()}.csv",
        mime="text/csv",
        key="dl_csv",
    )

    # Excel
    xls_buf = io.BytesIO()
    with pd.ExcelWriter(xls_buf, engine="openpyxl") as writer:
        party_data = []
        for k, v in st.session_state.dealer_info.items():
            party_data.append({"Field": f"Dealer - {k}", "Value": v})
        for k, v in st.session_state.bill_to.items():
            party_data.append({"Field": f"Bill To - {k}", "Value": v})
        for k, v in st.session_state.ship_to.items():
            party_data.append({"Field": f"Ship To - {k}", "Value": v})
        pd.DataFrame(party_data).to_excel(writer, sheet_name="Party Details", index=False)
        df_lines.to_excel(writer, sheet_name="Quotation Lines", index=False)
        pd.DataFrame([
            {"Description": "Gross MRP Value",  "Amount (₹)": round(grand_mrp_d, 2)},
            {"Description": "Distributor Disc.", "Amount (₹)": round(discount_d,  2)},
            {"Description": "Net Payable",       "Amount (₹)": round(grand_land_d,2)},
        ]).to_excel(writer, sheet_name="Summary", index=False)

    dl2.download_button(
        "⬇  Download Excel",
        data=xls_buf.getvalue(),
        file_name=f"sintex_quotation_{date.today()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="dl_excel",
    )

    # PDF
    if _HAS_REPORTLAB:
        if st.button("🖨️  Generate PDF", key="gen_pdf_btn"):
            with st.spinner("Generating PDF…"):
                try:
                    pdf_bytes = generate_pdf_from_resolved(
                        resolved_rows_dl, mrp_lookup,
                        st.session_state.bill_to,
                        st.session_state.ship_to,
                        st.session_state.dealer_info,
                    )
                    st.session_state.pdf_bytes = pdf_bytes
                except Exception as e:
                    st.error(f"PDF generation failed: {e}")

        if st.session_state.pdf_bytes:
            dl3.download_button(
                "⬇  Download PDF",
                data=st.session_state.pdf_bytes,
                file_name=f"sintex_quotation_{date.today()}.pdf",
                mime="application/pdf",
                key="dl_pdf",
            )
    else:
        st.caption("📄 PDF download requires `reportlab`. Run: `pip install reportlab`")

    # New Quotation reset
    st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)
    if st.button("🔄  Start New Quotation", key="new_quot"):
        keys_to_clear = [
            "qty_map", "bill_to", "ship_to", "dealer_info", "pdf_bytes",
            "image_bytes", "ocr_raw_rows", "ocr_resolved", "ocr_done",
            "zsd_search", "manual_rows",
        ]
        for k in keys_to_clear:
            if k in st.session_state:
                del st.session_state[k]
        for k in list(st.session_state.keys()):
            if k.startswith(("qty_", "edit_", "manual_", "bill_", "ship_", "dealer_",
                              "det_", "und_", "az_", "file_")):
                del st.session_state[k]
        st.rerun()