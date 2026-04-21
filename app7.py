import io, os, re, copy, base64, json, time, requests
import pandas as pd
from openpyxl import load_workbook
import streamlit as st
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
import streamlit.components.v1 as components
from PIL import Image, ImageEnhance, ImageFilter, ExifTags

st.set_page_config(
    page_title="Sintex BAPL – Quotation Generator",
    page_icon="🔴", layout="centered", initial_sidebar_state="collapsed",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&family=JetBrains+Mono:wght@400;500&display=swap');
:root{--red:#C0211F;--dred:#8B1514;--lred:#FDEAEA;--lgray:#F7F7F7;--mgray:#E8E8E8;
  --text:#1A1A1A;--muted:#6B6B6B;--border:#DEDEDE;--green:#1E7E4A;
  --radius:12px;--shadow:0 2px 12px rgba(0,0,0,0.08);}
html,body,[class*="css"]{font-family:'Inter',-apple-system,sans-serif!important;
  color:var(--text);background:#F2F2F2;}
#MainMenu,footer,header{visibility:hidden;}
.block-container{padding:1rem 0.75rem 5rem!important;max-width:920px!important;margin:0 auto!important;}

.app-header{background:linear-gradient(135deg,var(--red),var(--dred));color:white;
  padding:18px 20px;border-radius:0;margin-bottom:24px;
  display:flex;align-items:center;gap:16px;box-shadow:0 4px 20px rgba(192,33,31,.35);}
.app-header-badge{background:rgba(255,255,255,.18);border-radius:10px;width:52px;height:52px;
  display:flex;align-items:center;justify-content:center;font-size:26px;flex-shrink:0;}
.app-header-text h1{font-size:17px;font-weight:800;margin:0;}
.app-header-text p{font-size:11.5px;margin:3px 0 0;opacity:.75;}

.step-navbar{display:flex;background:white;border-radius:var(--radius);
  box-shadow:var(--shadow);margin-bottom:20px;overflow:hidden;
  border:1px solid var(--border);}
.step-nav-item{flex:1;display:flex;flex-direction:column;align-items:center;
  padding:12px 6px 10px;cursor:default;position:relative;
  border-right:1px solid var(--border);transition:background .2s;}
.step-nav-item:last-child{border-right:none;}
.step-nav-item.locked{opacity:.45;cursor:not-allowed;}
.step-nav-item.active{background:var(--lred);}
.step-nav-item.done{background:#ECFDF5;cursor:pointer;}
.step-nav-item.active::after{content:'';position:absolute;bottom:0;left:0;right:0;
  height:3px;background:var(--red);}
.step-nav-item.done::after{content:'';position:absolute;bottom:0;left:0;right:0;
  height:3px;background:var(--green);}
.step-nav-dot{width:26px;height:26px;border-radius:50%;display:flex;align-items:center;
  justify-content:center;font-size:11px;font-weight:700;margin-bottom:4px;
  background:var(--mgray);color:var(--muted);}
.step-nav-item.active .step-nav-dot{background:var(--red);color:white;}
.step-nav-item.done .step-nav-dot{background:var(--green);color:white;}
.step-nav-label{font-size:10px;font-weight:600;color:var(--muted);text-align:center;}
.step-nav-item.active .step-nav-label{color:var(--red);}
.step-nav-item.done .step-nav-label{color:var(--green);}

.step-card{background:white;border-radius:var(--radius);box-shadow:var(--shadow);
  margin-bottom:20px;overflow:hidden;border:1px solid var(--border);}
.step-card-header{background:linear-gradient(90deg,var(--red),#D94644);padding:14px 18px;
  display:flex;align-items:center;gap:12px;}
.step-number{background:rgba(255,255,255,.25);border-radius:8px;width:30px;height:30px;
  min-width:30px;display:flex;align-items:center;justify-content:center;
  font-size:13px;font-weight:700;color:white;}
.step-number.done{background:#1E7E4A;}
.step-title{font-size:14px;font-weight:700;color:white;margin:0;}
.step-subtitle{font-size:11px;color:rgba(255,255,255,.75);margin:2px 0 0;}
.step-body{padding:18px;}

.stButton>button{width:100%;background:linear-gradient(135deg,var(--red),var(--dred))!important;
  color:white!important;border:none!important;border-radius:10px!important;padding:13px 20px!important;
  font-family:'Inter',sans-serif!important;font-size:14.5px!important;font-weight:600!important;
  box-shadow:0 4px 14px rgba(192,33,31,.28)!important;transition:all .2s!important;}
.stButton>button:hover{transform:translateY(-1px)!important;}
.btn-secondary>.stButton>button{background:white!important;color:var(--red)!important;
  border:2px solid var(--red)!important;box-shadow:none!important;margin-top:0!important;}
[data-testid="column"]{display:flex;flex-direction:column;justify-content:center;}
.btn-green>.stButton>button{background:linear-gradient(135deg,var(--green),#155d38)!important;
  color:white!important;border:none!important;
  box-shadow:0 4px 14px rgba(30,126,74,.28)!important;}
.btn-blue>.stButton>button{background:linear-gradient(135deg,#1d4ed8,#1e40af)!important;
  color:white!important;border:none!important;
  box-shadow:0 4px 14px rgba(29,78,216,.28)!important;}

.stTabs [data-baseweb="tab-list"]{gap:4px;background:var(--mgray);
  padding:4px;border-radius:10px;border:none!important;}
.stTabs [data-baseweb="tab"]{border-radius:8px!important;font-family:'Inter',sans-serif!important;
  font-weight:500!important;font-size:13px!important;color:var(--muted)!important;}
.stTabs [aria-selected="true"]{background:var(--red)!important;color:white!important;}

.info-box{background:#EEF6FF;border:1.5px solid #BFDBFE;border-radius:9px;
  padding:11px 15px;font-size:12.5px;color:#1E40AF;margin:12px 0;}
.warn-box{background:#FFFBEB;border:1.5px solid #FDE68A;border-radius:9px;
  padding:11px 15px;font-size:12.5px;color:#92400E;margin:12px 0;}
.success-box{background:#ECFDF5;border:1.5px solid #A7F3D0;border-radius:9px;
  padding:11px 15px;font-size:12.5px;color:#065F46;margin:12px 0;}
.error-box{background:#FEF2F2;border:1.5px solid #FECACA;border-radius:9px;
  padding:11px 15px;font-size:12.5px;color:#991B1B;margin:6px 0;}

.totals-box{background:white;border:1.5px solid var(--border);border-radius:var(--radius);
  padding:16px 18px;margin:16px 0;}
.total-row{display:flex;justify-content:space-between;padding:7px 0;font-size:13.5px;
  border-bottom:1px solid var(--mgray);}
.total-row:last-child{border:none;padding-top:10px;}
.total-row.grand .total-lbl{font-weight:700;color:var(--red);font-size:15px;}
.total-row.grand .total-val{font-weight:800;color:var(--red);font-size:15px;
  font-family:'JetBrains Mono',monospace;}
.total-lbl{color:var(--muted);font-weight:500;}
.total-val{font-family:'JetBrains Mono',monospace;font-weight:600;}
.total-val.neg{color:#c0392b;}

.ocr-wrap{overflow-x:auto;-webkit-overflow-scrolling:touch;margin:10px 0;}
.ocr-tbl{width:100%;border-collapse:collapse;font-family:'Inter',sans-serif;font-size:12px;}
.ocr-tbl th{background:var(--red);color:white;padding:8px 7px;text-align:center;
  font-size:10.5px;font-weight:600;white-space:nowrap;}
.ocr-tbl th.L{text-align:left;min-width:120px;}
.ocr-tbl td{padding:6px 7px;border-bottom:1px solid var(--mgray);vertical-align:middle;text-align:center;}
.ocr-tbl td.L{text-align:left;font-weight:500;font-size:11.5px;}
.ocr-tbl td.M{font-family:'JetBrains Mono',monospace;font-size:10px;color:#555;}
.ocr-tbl tr:nth-child(even) td{background:var(--lgray);}
.ocr-tbl .ok td{background:#ECFDF5!important;}
.ocr-tbl .no td{background:#FFFBEB!important;}

.raw-wrap{overflow-x:auto;-webkit-overflow-scrolling:touch;margin:8px 0;
  border:1.5px solid var(--border);border-radius:8px;}
.raw-tbl{border-collapse:collapse;font-size:11px;font-family:'JetBrains Mono',monospace;min-width:500px;}
.raw-tbl th{background:#2D2D2D;color:white;padding:6px 8px;white-space:nowrap;text-align:center;}
.raw-tbl td{padding:5px 8px;border:1px solid #eee;white-space:nowrap;}
.raw-tbl tr:nth-child(even) td{background:#fafafa;}
.raw-tbl .sku{background:#FFF3E0!important;font-weight:700;color:#8B1514;}
.raw-tbl .qty{color:#1E7E4A;font-weight:600;}
.raw-tbl .hdr td{background:#F0F0F0;font-weight:700;color:#333;font-family:'Inter',sans-serif;}

.party-section{background:var(--lgray);border-radius:10px;padding:14px;
  border:1px solid var(--border);margin-bottom:14px;}
.party-title{font-size:12px;font-weight:700;color:var(--red);text-transform:uppercase;
  letter-spacing:.6px;margin-bottom:10px;}
.fsl{font-size:11.5px;font-weight:700;text-transform:uppercase;letter-spacing:.8px;
  color:var(--muted);margin:14px 0 8px;padding-bottom:4px;border-bottom:1px solid var(--mgray);}
.dl-btn{display:flex;align-items:center;justify-content:center;gap:8px;width:100%;
  background:linear-gradient(135deg,var(--green),#155d38);color:white!important;
  text-align:center;padding:15px 20px;border-radius:10px;font-weight:700;font-size:14.5px;
  text-decoration:none!important;box-shadow:0 4px 18px rgba(30,126,74,.3);
  margin:8px 0;box-sizing:border-box;}

.pdf-preview-wrap{border:2px solid var(--border);border-radius:var(--radius);
  overflow:hidden;margin:16px 0;background:#525659;}

.state-inner-card{background:var(--lgray);border-radius:10px;padding:14px 16px;
  border:1px solid var(--border);margin-bottom:16px;}
.state-inner-title{font-size:11.5px;font-weight:700;color:var(--red);text-transform:uppercase;
  letter-spacing:.6px;margin-bottom:10px;}

.capture-fullwidth [data-testid="stCameraInput"]{
  width:100vw!important;
  max-width:100%!important;
  margin-left:0!important;
  margin-right:0!important;
}
.capture-fullwidth [data-testid="stCameraInput"] video,
.capture-fullwidth [data-testid="stCameraInput"] img{
  width:100%!important;
  max-height:72vh!important;
  min-height:300px!important;
  object-fit:cover!important;
  display:block!important;
}
.capture-fullwidth [data-testid="stCameraInputButton"]{
  width:100%!important;
  height:56px!important;
  border-radius:0!important;
  font-size:16px!important;
  font-weight:700!important;
  background:linear-gradient(135deg,#C0211F,#8B1514)!important;
  color:white!important;
  border:none!important;
  box-shadow:none!important;
  letter-spacing:.5px!important;
}

.rotate-toolbar{display:flex;align-items:center;justify-content:center;gap:10px;
  background:#1a1a1a;padding:10px 14px;border-radius:0 0 0 0;}
.rotate-btn-wrap .stButton>button{
  background:#2a2a2a!important;
  color:white!important;
  border:1.5px solid #444!important;
  border-radius:8px!important;
  padding:8px 16px!important;
  font-size:13px!important;
  font-weight:600!important;
  box-shadow:none!important;
  width:auto!important;
}
.rotate-btn-wrap .stButton>button:hover{background:#C0211F!important;border-color:#C0211F!important;}

.img-quality-badge{display:inline-flex;align-items:center;gap:6px;
  background:rgba(30,126,74,.12);border:1.5px solid rgba(30,126,74,.3);
  color:#1E7E4A;font-size:11px;font-weight:700;
  padding:4px 10px;border-radius:20px;margin-left:8px;}

@keyframes sintex-pulse{
  0%,100%{opacity:1;box-shadow:0 0 6px #C0211F;}
  50%{opacity:.25;box-shadow:0 0 2px #C0211F;}
}

@media(max-width:600px){
  .block-container{padding:.75rem 0 5rem!important;}
  .step-body{padding:14px 12px;}
  .step-nav-label{font-size:9px;}
  .capture-fullwidth [data-testid="stCameraInput"] video,
  .capture-fullwidth [data-testid="stCameraInput"] img{
    max-height:60vh!important;
    min-height:250px!important;
  }
}
</style>
<script>
(function() {
  function fixHeader() {
    const header = document.querySelector('.app-header');
    if (!header) { setTimeout(fixHeader, 100); return; }
    const existing = document.getElementById('sintex-sticky-header');
    if (existing) return;
    const portal = document.createElement('div');
    portal.id = 'sintex-sticky-header';
    portal.style.cssText = [
      'position:fixed','top:0','left:0','right:0','z-index:9999',
      'background:linear-gradient(135deg,#C0211F,#8B1514)',
      'display:flex','align-items:center','gap:16px',
      'padding:18px 20px',
      'box-shadow:0 4px 20px rgba(192,33,31,.35)',
    ].join(';');
    portal.innerHTML = header.innerHTML;
    document.body.appendChild(portal);
    header.style.visibility = 'hidden';
    header.style.height = header.offsetHeight + 'px';
    const ph = portal.offsetHeight;
    document.body.style.paddingTop = ph + 'px';
    const main = document.querySelector('.block-container');
    if(main) main.style.paddingTop = '1rem';
  }
  if(document.readyState === 'loading') {
    document.addEventListener('DOMContentLoaded', fixHeader);
  } else {
    fixHeader();
  }
})();
</script>
""", unsafe_allow_html=True)

XLSX_PATH      = os.path.join(os.path.dirname(__file__), "Sample form for Product list.xlsx")
MRP_PATH       = os.path.join(os.path.dirname(__file__), "MRP_State_chhattisghar.csv")
CUST_PATH      = os.path.join(os.path.dirname(__file__), "ZSD_CUST.csv")
AZURE_ENDPOINT = os.environ.get("AZURE_OCR_ENDPOINT", "")
AZURE_KEY      = os.environ.get("AZURE_OCR_KEY", "")

SIZE_ALIASES: dict[str, str] = {}
for _mm in ["15","20","25","32","40","50","63","75","90","110"]:
    SIZE_ALIASES[_mm + "MM"]      = _mm + "MM"
    SIZE_ALIASES[_mm + " MM"]     = _mm + "MM"
    SIZE_ALIASES[_mm]             = _mm + "MM"
SIZE_ALIASES.update({
    '1/2"':"15MM",'1/2':"15MM",'½"':"15MM",
    '3/4"':"20MM",'3/4':"20MM",'¾"':"20MM",
    '1"'  :"25MM",
    '1.25"':"32MM",'1-1/4"':"32MM",'1¼"':"32MM",
    '1.5"':"40MM",'1-1/2"':"40MM",'1½"':"40MM",
    '2"'  :"50MM",
})

ALL_INDIA_STATES = [
    "Andhra Pradesh", "Arunachal Pradesh", "Assam", "Bihar", "Chhattisgarh",
    "Goa", "Gujarat", "Haryana", "Himachal Pradesh", "Jharkhand", "Karnataka",
    "Kerala", "Madhya Pradesh", "Maharashtra", "Manipur", "Meghalaya", "Mizoram",
    "Nagaland", "Odisha", "Punjab", "Rajasthan", "Sikkim", "Tamil Nadu",
    "Telangana", "Tripura", "Uttar Pradesh", "Uttarakhand", "West Bengal",
    "Andaman and Nicobar Islands", "Chandigarh", "Dadra and Nagar Haveli and Daman and Diu",
    "Delhi", "Jammu and Kashmir", "Ladakh", "Lakshadweep", "Puducherry",
]


def fix_exif_orientation(img: Image.Image) -> Image.Image:
    try:
        exif_data = img._getexif()
        if exif_data is None:
            return img
        orientation_key = next(
            (k for k, v in ExifTags.TAGS.items() if v == "Orientation"), None
        )
        if orientation_key is None or orientation_key not in exif_data:
            return img
        orientation = exif_data[orientation_key]
        rotation_map = {3: 180, 6: 270, 8: 90}
        if orientation in rotation_map:
            img = img.rotate(rotation_map[orientation], expand=True)
    except Exception:
        pass
    return img


def enhance_image_for_ocr(raw_bytes: bytes, rotation_degrees: int = 0) -> bytes:
    img = Image.open(io.BytesIO(raw_bytes))
    img = fix_exif_orientation(img)

    if img.mode not in ("RGB", "L"):
        img = img.convert("RGB")

    max_dim = 3000
    w, h = img.size
    if max(w, h) > max_dim:
        ratio = max_dim / max(w, h)
        img = img.resize((int(w * ratio), int(h * ratio)), Image.LANCZOS)
    elif max(w, h) < 1200:
        ratio = 1200 / max(w, h)
        img = img.resize((int(w * ratio), int(h * ratio)), Image.LANCZOS)

    if rotation_degrees != 0:
        img = img.rotate(-rotation_degrees, expand=True)

    enhancer_contrast = ImageEnhance.Contrast(img)
    img = enhancer_contrast.enhance(1.35)

    enhancer_sharpness = ImageEnhance.Sharpness(img)
    img = enhancer_sharpness.enhance(2.2)

    enhancer_brightness = ImageEnhance.Brightness(img)
    img = enhancer_brightness.enhance(1.08)

    out = io.BytesIO()
    img.save(out, format="JPEG", quality=97, optimize=True, subsampling=0)
    return out.getvalue()


def get_image_dimensions(raw_bytes: bytes) -> tuple[int, int]:
    try:
        img = Image.open(io.BytesIO(raw_bytes))
        return img.size
    except Exception:
        return (0, 0)


@st.cache_data(show_spinner=False)
def load_sku_sheets() -> dict:
    wb = load_workbook(XLSX_PATH, read_only=True)
    result = {}
    for sn in wb.sheetnames:
        ws  = wb[sn]
        all_rows = list(ws.iter_rows(values_only=True))
        r1  = list(all_rows[1]) if len(all_rows) > 1 else []
        r2  = list(all_rows[2]) if len(all_rows) > 2 else []
        cis = [ci for ci,v in enumerate(r1) if ci>=2 and v is not None]
        if not cis:
            cis = [ci for ci,v in enumerate(r2) if ci>=2 and v is not None]
        od  = {ci: str(r1[ci]).strip() if ci<len(r1) and r1[ci] else "" for ci in cis}
        inch= {ci: str(r2[ci]).strip() if ci<len(r2) and r2[ci] else "" for ci in cis}
        rows, sec = [], "PIPES"
        for raw in all_rows[3:]:
            row   = list(raw)
            if not any(v for v in row if v is not None): continue
            label = str(row[0]).strip() if row[0] else ""
            if not label or label == "\xa0": continue
            up = label.upper()
            if up in {"FITTINGS","PIPES","FITTING SCH 80","FITTING SCH80"}:
                sec = up; rows.append({"sh":True,"label":label}); continue
            if (len(row)>1 and (row[1] is None or str(row[1]).strip()=="\xa0")
                    and not any(str(v).strip().startswith("CP") for v in row[2:] if v)):
                sec = label; rows.append({"sh":True,"label":label}); continue
            cells = {}
            for ci in cis:
                if ci<len(row) and row[ci] is not None:
                    v = str(row[ci]).strip()
                    if v and v not in("-","\xa0") and v.startswith("CP"):
                        cells[ci] = v
            rows.append({"sh":False,"label":label,"sec":sec,"cells":cells})
        result[sn] = {"od":od,"inch":inch,"cis":cis,"rows":rows}
    return result


@st.cache_data(show_spinner=False)
def build_sku_master(sku_sheets: dict) -> dict:
    master = {}
    for sn, sd in sku_sheets.items():
        sec = "PIPES"
        for rd in sd["rows"]:
            if rd.get("sh"):
                sec = rd["label"]; continue
            for ci, sku in rd["cells"].items():
                master[sku] = {
                    "sheet": sn, "product": rd["label"], "section": sec,
                    "od_size": sd["od"].get(ci,""), "inch_size": sd["inch"].get(ci,""),
                }
    return master


@st.cache_data(show_spinner=False)
def build_prefix_index(sku_master: dict) -> dict:
    bps, bp = {}, {}
    for sku, info in sku_master.items():
        sz = info.get("od_size","").upper().replace(" ","")
        for l in range(4, len(sku)+1):
            p = sku[:l]
            bp.setdefault(p, []).append(sku)
            if sz:
                bps[(p, sz)] = sku
    return {"bps": bps, "bp": bp}


@st.cache_data(show_spinner=False)
def load_mrp_data_for_state(state_name: str) -> dict:
    df = pd.read_csv(MRP_PATH, encoding="latin-1")
    df["_m"] = pd.to_numeric(
        df["MRP(ZPR1-933)"].astype(str).str.replace(",","").str.strip(), errors="coerce"
    ).fillna(0)
    df["_d"] = pd.to_numeric(
        df["Distributor Landing"].astype(str).str.replace(",","").str.strip(), errors="coerce"
    ).fillna(0)
    if "State Name" in df.columns and state_name:
        state_df = df[df["State Name"].str.strip().str.lower() == state_name.strip().lower()]
        if not state_df.empty:
            df = state_df
    r = {}
    for _, row in df.iterrows():
        mat = str(row["Material Number"]).strip()
        r[mat] = {
            "mrp": float(row["_m"]),
            "distributor_landing": float(row["_d"]),
            "description": str(row["Material Description"]).strip(),
        }
    return r


@st.cache_data(show_spinner=False)
def load_state_names_from_csv() -> list[str]:
    try:
        df = pd.read_csv(MRP_PATH, encoding="latin-1")
        if "State Name" in df.columns:
            states = (
                df["State Name"]
                .dropna()
                .str.strip()
                .unique()
                .tolist()
            )
            states = [s for s in states if s]
            if states:
                return sorted(states)
    except Exception:
        pass
    return ALL_INDIA_STATES


@st.cache_data(show_spinner=False)
def load_customers() -> list:
    df  = pd.read_csv(CUST_PATH, encoding="latin-1")
    out = []
    for _, r in df.iterrows():
        name = str(r.get("Customer Name","")).strip()
        code = str(r.get("Customer Code","")).strip()
        if name and name != "nan":
            out.append({
                "code": code, "name": name,
                "address": " ".join(filter(None,[
                    str(r.get("Address 1","") or "").strip(),
                    str(r.get("Address 2","") or "").strip(),
                    str(r.get("City","") or "").strip(),
                    str(r.get("State Code Desc.","") or "").strip()])),
                "phone":  str(r.get("Telephone","")  or "").strip(),
                "mobile": str(r.get("Mobile No.","") or "").strip(),
                "state_code": str(r.get("State Code","") or "").strip(),
                "state": str(r.get("State Code Desc.","") or "").strip(),
                "gst": str(r.get("GST Number","") or "").strip(),
                "pan": str(r.get("PAN No.","") or "").strip(),
                "display": f"{code} – {name}",
            })
    return out


def _poly_bbox(poly):
    if not poly:
        return 0, 0, 0, 0
    if isinstance(poly[0], dict):
        xs, ys = [p["x"] for p in poly], [p["y"] for p in poly]
    else:
        xs, ys = poly[0::2], poly[1::2]
    return min(xs), min(ys), max(xs)-min(xs), max(ys)-min(ys)


def _words_v3v4(data: dict) -> list[dict]:
    out = []
    ar  = data.get("analyzeResult", data)
    for page in ar.get("pages", []):
        pw = page.get("width", 1) or 1
        ph = page.get("height", 1) or 1
        for w in page.get("words", []):
            x, y, ww, hh = _poly_bbox(w.get("polygon", w.get("boundingBox",[])))
            out.append({"text": w.get("content", w.get("text","")),
                        "x":x,"y":y,"w":ww,"h":hh,"cx":x+ww/2,"cy":y+hh/2,
                        "pw":pw,"ph":ph})
    return out


def _words_v2(data: dict) -> list[dict]:
    out = []
    for page in data.get("analyzeResult",{}).get("readResults",[]):
        pw = page.get("width",1) or 1
        ph = page.get("height",1) or 1
        for line in page.get("lines",[]):
            for w in line.get("words",[]):
                x,y,ww,hh = _poly_bbox(w.get("boundingBox",[]))
                out.append({"text":w.get("text",""),
                            "x":x,"y":y,"w":ww,"h":hh,"cx":x+ww/2,"cy":y+hh/2,
                            "pw":pw,"ph":ph})
    return out


def run_azure_ocr(img: bytes, endpoint: str, key: str) -> list[dict]:
    ep  = endpoint.rstrip("/")
    hdr = {"Ocp-Apim-Subscription-Key": key, "Content-Type": "application/octet-stream"}

    def post(url):
        return requests.post(url, headers=hdr, data=img, timeout=60)

    def poll_async(resp) -> dict:
        if resp.status_code != 202:
            resp.raise_for_status()
            return resp.json()
        op = resp.headers.get("Operation-Location") or resp.headers.get("operation-location","")
        if not op:
            raise ValueError("No Operation-Location header")
        for _ in range(30):
            time.sleep(2)
            p = requests.get(op, headers={"Ocp-Apim-Subscription-Key": key}, timeout=30)
            p.raise_for_status()
            d = p.json()
            s = d.get("status","")
            if s == "succeeded": return d
            if s == "failed":    raise ValueError("Azure failed: " + json.dumps(d.get("error",{})))
        raise TimeoutError("Azure OCR timed out")

    r = post(f"{ep}/documentintelligence/documentModels/prebuilt-read:analyze?api-version=2024-02-29-preview")
    if r.status_code not in (200, 202):
        r = post(f"{ep}/formrecognizer/documentModels/prebuilt-read:analyze?api-version=2023-07-31")
    if r.status_code not in (200, 202):
        r = post(f"{ep}/formrecognizer/v2.1/layout/analyze")
        r.raise_for_status()
        op = r.headers.get("Operation-Location","")
        if not op: raise ValueError("No Operation-Location in v2 response")
        for _ in range(20):
            time.sleep(2)
            p = requests.get(op, headers={"Ocp-Apim-Subscription-Key": key}, timeout=30)
            p.raise_for_status()
            d = p.json()
            if d.get("status") == "succeeded": return _words_v2(d)
            if d.get("status") == "failed":    raise ValueError("Azure v2 failed")
        raise TimeoutError("Azure v2 timed out")

    return _words_v3v4(poll_async(r))


def reconstruct_table(words: list[dict]) -> list[list[str]]:
    if not words:
        return []

    heights = sorted(w["h"] for w in words if w["h"] > 0)
    med_h   = heights[len(heights)//2] if heights else 8
    row_tol = max(med_h * 0.55, 2)
    pw      = words[0]["pw"] or 1
    merge_x = pw * 0.016
    col_tol = pw * 0.025

    by_y = sorted(words, key=lambda w: w["cy"])
    raw_rows: list[list[dict]] = []
    cur: list[dict] = []
    cy_avg = None
    for w in by_y:
        if cy_avg is None or abs(w["cy"] - cy_avg) <= row_tol:
            cur.append(w)
            cy_avg = (cy_avg + w["cy"]) / 2 if cy_avg else w["cy"]
        else:
            raw_rows.append(sorted(cur, key=lambda ww: ww["cx"]))
            cur    = [w]
            cy_avg = w["cy"]
    if cur:
        raw_rows.append(sorted(cur, key=lambda ww: ww["cx"]))

    cell_rows: list[list[dict]] = []
    for row in raw_rows:
        cells, buf, buf_cx, buf_rx = [], row[0]["text"], row[0]["cx"], row[0]["x"]+row[0]["w"]
        for w in row[1:]:
            if w["x"] - buf_rx <= merge_x:
                buf    += " " + w["text"]
                buf_rx  = w["x"] + w["w"]
                buf_cx  = (buf_cx + w["cx"]) / 2
            else:
                cells.append({"text": buf.strip(), "cx": buf_cx})
                buf, buf_cx, buf_rx = w["text"], w["cx"], w["x"]+w["w"]
        cells.append({"text": buf.strip(), "cx": buf_cx})
        cell_rows.append(cells)

    all_cx  = sorted(c["cx"] for row in cell_rows for c in row)
    col_cxs: list[float] = []
    for cx in all_cx:
        placed = False
        for i, cc in enumerate(col_cxs):
            if abs(cx - cc) <= col_tol:
                col_cxs[i] = (cc + cx) / 2
                placed = True; break
        if not placed:
            col_cxs.append(cx)
    col_cxs.sort()
    nc = len(col_cxs)

    def nearest(cx):
        return min(range(nc), key=lambda i: abs(col_cxs[i] - cx))

    grid: list[list[str]] = []
    for row in cell_rows:
        arr = [""] * nc
        for cell in row:
            ci = nearest(cell["cx"])
            arr[ci] = (arr[ci] + " " + cell["text"]).strip() if arr[ci] else cell["text"]
        grid.append(arr)

    return grid


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", s.strip().upper())

def _size_of(cell: str) -> str | None:
    n = _norm(cell).replace(" ","").rstrip(".")
    v = SIZE_ALIASES.get(n)
    if not v:
        v = SIZE_ALIASES.get(re.sub(r'[^0-9]','',n) + "MM") if re.sub(r'[^0-9]','',n) else None
    return v

def _is_sku_prefix(cell: str) -> bool:
    return bool(re.match(r'^CP[A-Z0-9]{3,}', cell.strip().upper()))

def _clean_ocr_num(cell: str) -> int | None:
    c = cell.strip().replace(",","")
    c = c.replace("O","0").replace("o","0")
    c = c.replace("l","1").replace("I","1").replace("|","1")
    c = c.replace("S","5").replace("s","5").replace("G","6")
    m = re.search(r'\d+', c)
    if m:
        v = int(m.group())
        return v if 1 <= v <= 9999 else None
    return None

def _is_qty(cell: str) -> bool:
    return _clean_ocr_num(cell) is not None


def analyze_table(grid: list[list[str]]) -> dict:
    if not grid:
        return {"rows":[], "meta":{}, "grid":grid}

    nc = max(len(r) for r in grid)

    best_hdr, best_szc = 0, {}
    for ri, row in enumerate(grid[:15]):
        szc = {ci: _size_of(cell) for ci,cell in enumerate(row) if _size_of(cell)}
        if len(szc) > len(best_szc):
            best_szc, best_hdr = szc, ri

    if not best_szc:
        for ri, row in enumerate(grid[:15]):
            szc = {}
            for ci, cell in enumerate(row):
                n = _norm(cell).replace(" ","")
                if re.fullmatch(r'\d{2,3}', n) and (n+"MM") in SIZE_ALIASES:
                    szc[ci] = n+"MM"
            if len(szc) > len(best_szc):
                best_szc, best_hdr = szc, ri

    size_cols  = best_szc
    header_row = best_hdr

    sku_votes: dict[int,int] = {}
    for row in grid[header_row+1:]:
        for ci, cell in enumerate(row):
            if _is_sku_prefix(cell):
                sku_votes[ci] = sku_votes.get(ci,0) + 1
    sku_col = max(sku_votes, key=sku_votes.get) if sku_votes else None

    skip = set(size_cols) | ({sku_col} if sku_col is not None else set())
    name_col = next((c for c in range(nc) if c not in skip), None)

    rows = []
    for ri in range(header_row+1, len(grid)):
        row       = grid[ri]
        sku_pfx   = row[sku_col].strip().upper() if sku_col is not None and sku_col < len(row) else ""
        product   = row[name_col].strip()        if name_col is not None and name_col < len(row) else ""

        if not sku_pfx and not product:
            continue
        if _norm(product) in {"SKU NAME","PRODUCT","DESCRIPTION","ITEM","S NO","SR NO",
                               "SKU CODE","SIZE","CATEGORY","SECTION"}:
            continue
        if not sku_pfx and not any(_is_qty(row[ci]) for ci in size_cols if ci < len(row)):
            continue

        sizes: dict[str,int] = {}
        for ci, sz_label in size_cols.items():
            if ci < len(row):
                q = _clean_ocr_num(row[ci])
                if q:
                    sizes[sz_label] = q

        if not sizes:
            continue

        rows.append({"product": product, "sku_prefix": sku_pfx, "sizes": sizes, "ri": ri})

    return {
        "rows": rows,
        "meta": {"header_row": header_row, "sku_col": sku_col,
                 "name_col": name_col, "size_cols": size_cols},
        "grid": grid,
    }


_OCR_FIXES = str.maketrans({
    "O":"0","o":"0","Q":"0",
    "I":"1","l":"1","|":"1",
    "Z":"2","z":"2",
    "S":"5","s":"5",
    "G":"6","b":"6",
    "B":"8",
})

def _fix_sku(raw: str) -> str:
    s = raw.strip().upper()
    return s[:2] + s[2:].translate(_OCR_FIXES) if len(s) > 2 else s


def match_sku(prefix: str, size: str, pidx: dict, master: dict) -> str | None:
    if not prefix:
        return None

    sz   = size.upper().replace(" ","")
    pfx  = _fix_sku(prefix)
    bps  = pidx["bps"]
    bp   = pidx["bp"]

    if pfx in master and master[pfx].get("od_size","").upper().replace(" ","") == sz:
        return pfx

    for l in range(min(len(pfx),20), 3, -1):
        key = (pfx[:l], sz)
        if key in bps:
            return bps[key]

    for l in range(min(len(pfx),20), 3, -1):
        for sku in bp.get(pfx[:l], []):
            if master[sku].get("od_size","").upper().replace(" ","") == sz:
                return sku

    best_sku, best_score = None, 0
    for sku, info in master.items():
        if info.get("od_size","").upper().replace(" ","") != sz:
            continue
        min_l = min(len(pfx), len(sku))
        if min_l < 4:
            continue
        score = sum(a==b for a,b in zip(pfx[:min_l], sku[:min_l]))
        if score >= max(4, min_l - 2) and score > best_score:
            best_score, best_sku = score, sku

    return best_sku


def build_quantities(extracted_rows: list[dict], pidx: dict, master: dict):
    line_items: list[dict] = []
    quantities: dict[str,int] = {}
    log: list[dict] = []
    for row in extracted_rows:
        for sz, qty in row["sizes"].items():
            full = match_sku(row["sku_prefix"], sz, pidx, master)
            if full:
                quantities[full] = quantities.get(full, 0) + qty
                line_items.append({"sku": full, "qty": qty})
            log.append({
                "product":  row["product"],
                "prefix":   row["sku_prefix"],
                "size":     sz,
                "qty":      qty,
                "full_sku": full or "—",
                "status":   "matched" if full else "unmatched",
            })
    return quantities, log, line_items


def build_pdf(quantities: dict, mrp_data: dict, bill_to: dict,
              ship_to: dict, sku_master: dict, line_items: list = None) -> bytes:
    buf = io.BytesIO()
    PAGE = landscape(A4)
    LEFT_M = RIGHT_M = 12 * mm
    TOP_M = BOT_M = 12 * mm
    USABLE_W = PAGE[0] - LEFT_M - RIGHT_M

    doc = SimpleDocTemplate(buf, pagesize=PAGE,
                            leftMargin=LEFT_M, rightMargin=RIGHT_M,
                            topMargin=TOP_M, bottomMargin=BOT_M)

    BLACK   = colors.HexColor("#000000")
    DKGRAY  = colors.HexColor("#1A1A1A")
    MDGRAY  = colors.HexColor("#4A4A4A")
    GRAY    = colors.HexColor("#6B6B6B")
    LGRAY   = colors.HexColor("#F0F0F0")
    MLGRAY  = colors.HexColor("#D0D0D0")
    XLGRAY  = colors.HexColor("#E8E8E8")
    WHITE   = colors.white
    HDRFILL = colors.HexColor("#1A1A1A")
    ALTFILL = colors.HexColor("#F5F5F5")
    ACCLINE = colors.HexColor("#555555")

    sty = getSampleStyleSheet()
    def ps(n, **kw): return ParagraphStyle(n, parent=sty["Normal"], **kw)

    story = []

    hdr_left = ps("hl", fontName="Times-Bold", fontSize=9, textColor=WHITE, leading=14)
    hdr_ctr  = ps("hc", fontName="Times-Bold", fontSize=20, textColor=WHITE, alignment=TA_CENTER)
    hdr_rt   = ps("hr", fontName="Times-Roman", fontSize=8, textColor=WHITE,
                  alignment=TA_RIGHT, leading=13)

    selected_state = st.session_state.get("selected_state", "Chhattisgarh")

    hdr = Table([[
        Paragraph("Sintex BAPL Limited<br/>"
                  "<font size='7.5'>Kutesar Road, Raipur, CG 492101</font><br/>"
                  "<font size='7.5'>GSTIN: 22AADCB1921F1ZE</font>", hdr_left),
        Paragraph("SALES QUOTATION", hdr_ctr),
        Paragraph(f"State: <b>{selected_state}</b><br/>"
                  "<font size='7'>CPVC / UPVC Pipes &amp; Fittings</font>", hdr_rt),
    ]], colWidths=[USABLE_W*0.30, USABLE_W*0.40, USABLE_W*0.30])
    hdr.setStyle(TableStyle([
        ("BACKGROUND",   (0,0), (-1,-1), HDRFILL),
        ("TEXTCOLOR",    (0,0), (-1,-1), WHITE),
        ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",   (0,0), (-1,-1), 12),
        ("BOTTOMPADDING",(0,0), (-1,-1), 12),
        ("LEFTPADDING",  (0,0), (-1,-1), 14),
        ("RIGHTPADDING", (0,0), (-1,-1), 14),
        ("LINEBELOW",    (0,0), (-1,-1), 2, ACCLINE),
    ]))
    story += [hdr, Spacer(1, 4*mm)]

    plbl = ps("plbl", fontName="Times-Bold", fontSize=7.5, textColor=WHITE)
    pval = ps("pval", fontName="Times-Roman", fontSize=7.5, textColor=BLACK, leading=11)

    def party_cell(d):
        ph_raw  = d.get('phone','').strip()
        mb_raw  = d.get('mobile','').strip()
        ph_disp = f"+91 {ph_raw}" if ph_raw else ""
        mb_disp = f"+91 {mb_raw}" if mb_raw else ""
        lines = [
            f"<b>Party No.:</b> {d.get('party_no','')}",
            f"<b>Name &amp; Address:</b> {d.get('name','').replace(chr(10),' ')}",
            f"<b>Phone:</b> {ph_disp}   <b>Mobile:</b> {mb_disp}",
            f"<b>State:</b> {d.get('state_code','')} – {d.get('state','')}",
            f"<b>GST:</b> {d.get('gst','')}   <b>PAN:</b> {d.get('pan','')}",
        ]
        return Paragraph("<br/>".join(lines), pval)

    HALF = USABLE_W / 2
    pt = Table([
        [Paragraph("BILL TO PARTY", plbl), Paragraph("SHIP TO PARTY", plbl)],
        [party_cell(bill_to),              party_cell(ship_to)],
    ], colWidths=[HALF, HALF])
    pt.setStyle(TableStyle([
        ("BACKGROUND",   (0,0), (-1,0), MDGRAY),
        ("TEXTCOLOR",    (0,0), (-1,0), WHITE),
        ("BACKGROUND",   (0,1), (-1,1), XLGRAY),
        ("BOX",          (0,0), (-1,-1), 0.8, MLGRAY),
        ("INNERGRID",    (0,0), (-1,-1), 0.5, MLGRAY),
        ("VALIGN",       (0,0), (-1,-1), "TOP"),
        ("TOPPADDING",   (0,0), (-1,-1), 7),
        ("BOTTOMPADDING",(0,0), (-1,-1), 7),
        ("LEFTPADDING",  (0,0), (-1,-1), 10),
        ("RIGHTPADDING", (0,0), (-1,-1), 10),
    ]))
    story += [pt, Spacer(1, 5*mm)]

    sch  = ps("sch",  fontName="Times-Bold",   fontSize=7,   textColor=WHITE, alignment=TA_CENTER)
    cell = ps("c",    fontName="Times-Roman",  fontSize=7.5, textColor=DKGRAY)
    celc = ps("cc",   fontName="Times-Roman",  fontSize=7.5, textColor=DKGRAY, alignment=TA_CENTER)
    celr = ps("cr",   fontName="Times-Roman",  fontSize=7.5, textColor=DKGRAY, alignment=TA_RIGHT)
    celb = ps("cb",   fontName="Times-Bold",   fontSize=7.5, textColor=BLACK,  alignment=TA_RIGHT)
    skuc = ps("sk",   fontName="Courier",      fontSize=6.5, textColor=MDGRAY)
    qtys_ps = ps("qp",fontName="Times-Bold",   fontSize=8,   textColor=BLACK,  alignment=TA_CENTER)

    heads = ["S.No", "Product", "SKU Code", "OD", "Inch", "MRP (₹)", "Qty", "Total MRP (₹)", "Dist. Landing (₹)", "Taxable (₹)"]
    cw_mm = [8, 68, 48, 14, 12, 22, 10, 26, 26, 26]
    cw = [x * mm for x in cw_mm]
    scale = USABLE_W / sum(cw)
    cw = [x * scale for x in cw]

    trows = [[Paragraph(h, sch) for h in heads]]
    ln = gm = gd = gt = 0

    if line_items:
        items_iter = line_items
    else:
        items_iter = [{"sku": s, "qty": q} for s, q in quantities.items() if q > 0]

    for item in items_iter:
        sku = item["sku"]; qty = item["qty"]
        if qty <= 0:
            continue
        info  = sku_master.get(sku, {})
        mi    = mrp_data.get(sku, {})
        mrp   = mi.get("mrp", 0.)
        dist  = mi.get("distributor_landing", 0.)
        tot   = round(mrp * qty, 2)
        tax   = round(dist * qty, 2)
        ln += 1; gm += tot; gd += (mrp - dist) * qty; gt += tax
        trows.append([
            Paragraph(str(ln),                    celc),
            Paragraph(info.get("product",""),     cell),
            Paragraph(sku,                        skuc),
            Paragraph(info.get("od_size",""),     celc),
            Paragraph(info.get("inch_size",""),   celc),
            Paragraph(f"{mrp:,.2f}",              celr),
            Paragraph(str(qty),                   qtys_ps),
            Paragraph(f"{tot:,.2f}",              celb),
            Paragraph(f"{dist:,.2f}",             celr),
            Paragraph(f"{tax:,.2f}",              celb),
        ])
    if ln == 0:
        trows.append([Paragraph("No items", cell)] + [""] * 9)

    t2 = Table(trows, colWidths=cw, repeatRows=1)
    ts = [
        ("BACKGROUND",    (0,0),  (-1,0),  HDRFILL),
        ("TEXTCOLOR",     (0,0),  (-1,0),  WHITE),
        ("BOX",           (0,0),  (-1,-1), 0.8,  DKGRAY),
        ("INNERGRID",     (0,1),  (-1,-1), 0.3,  MLGRAY),
        ("LINEBELOW",     (0,0),  (-1,0),  1.5,  ACCLINE),
        ("VALIGN",        (0,0),  (-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0),  (-1,-1), 4),
        ("BOTTOMPADDING", (0,0),  (-1,-1), 4),
        ("LEFTPADDING",   (0,0),  (-1,-1), 4),
        ("RIGHTPADDING",  (0,0),  (-1,-1), 4),
    ]
    for ri in range(1, len(trows)):
        if ri % 2 == 0:
            ts.append(("BACKGROUND", (0, ri), (-1, ri), ALTFILL))
    t2.setStyle(TableStyle(ts))
    story += [t2, Spacer(1, 0)]

    lbl_style = ps("tlbl", fontName="Times-Roman", fontSize=8,  textColor=GRAY,   alignment=TA_RIGHT)
    lbl_bold  = ps("tbb",  fontName="Times-Bold",  fontSize=8,  textColor=BLACK,  alignment=TA_RIGHT)
    val_style = ps("tval", fontName="Times-Bold",  fontSize=8,  textColor=BLACK,  alignment=TA_RIGHT)
    val_emph  = ps("tvr",  fontName="Times-Bold",  fontSize=9,  textColor=BLACK,  alignment=TA_RIGHT)
    val_neg   = ps("tvn",  fontName="Times-Roman", fontSize=8,  textColor=MDGRAY, alignment=TA_RIGHT)

    LABEL_W = 55 * mm
    VALUE_W = 38 * mm
    FILL_W  = USABLE_W - LABEL_W - VALUE_W

    tot_rows = [
        ["", Paragraph("Gross Total (MRP):", lbl_style),
              Paragraph(f"₹  {gm:,.2f}", val_style)],
        ["", Paragraph("Less Distributor Discount:", lbl_style),
              Paragraph(f"– ₹  {gd:,.2f}", val_neg)],
        ["", Paragraph("Net Taxable Value:", lbl_bold),
              Paragraph(f"₹  {gt:,.2f}", val_emph)],
    ]
    tot_tbl = Table(tot_rows, colWidths=[FILL_W, LABEL_W, VALUE_W])
    tot_tbl.setStyle(TableStyle([
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0), (-1,-1), 5),
        ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ("LEFTPADDING",   (0,0), (-1,-1), 6),
        ("RIGHTPADDING",  (0,0), (-1,-1), 6),
        ("BOX",           (1,0), (2,-1),  0.8, MLGRAY),
        ("INNERGRID",     (1,0), (2,-1),  0.4, XLGRAY),
        ("BACKGROUND",    (1,2), (2,2),   XLGRAY),
        ("LINEABOVE",     (1,2), (2,2),   1.5, DKGRAY),
        ("LINEBELOW",     (1,2), (2,2),   1.5, DKGRAY),
        ("ALIGN",         (1,0), (2,-1),  "RIGHT"),
    ]))
    story += [tot_tbl, Spacer(1, 5*mm)]

    ft = Table([[
        Paragraph("<i>Computer-generated quotation. Subject to change without notice.</i>",
                  ps("ft", fontName="Times-Italic", fontSize=7, textColor=GRAY)),
        Paragraph("<b>Authorised Signatory</b>",
                  ps("sg", fontName="Times-Bold", fontSize=8, textColor=BLACK,
                     alignment=TA_RIGHT)),
    ]], colWidths=[USABLE_W * 0.65, USABLE_W * 0.35])
    ft.setStyle(TableStyle([
        ("LINEABOVE",     (0,0), (-1,0), 0.8, MLGRAY),
        ("VALIGN",        (0,0), (-1,0), "MIDDLE"),
        ("TOPPADDING",    (0,0), (-1,0), 6),
        ("BOTTOMPADDING", (0,0), (-1,0), 6),
    ]))
    story.append(ft)

    doc.build(story)
    buf.seek(0)
    return buf.read()


def _ss(k, v):
    if k not in st.session_state: st.session_state[k] = v

for k, v in [
    ("step", 1), ("image_bytes", None), ("ocr_done", False),
    ("quantities", {}), ("match_log", []), ("ocr_grid", []),
    ("ocr_meta", {}), ("ocr_extracted", []), ("line_items", []),
    ("bill_to", {}), ("ship_to", {}), ("party_confirmed", False), ("pdf_bytes", None),
    ("azure_endpoint", AZURE_ENDPOINT), ("azure_key", AZURE_KEY),
    ("image_submitted", False),
    ("editor_open", False),
    ("raw_image_bytes", None),
    ("ocr_reviewed", False),
    ("upload_key", 0),
    ("selected_state", "Chhattisgarh"),
    ("state_confirmed", False),
    ("cam_preview_bytes", None),
    ("image_rotation", 0),
    ("capture_mode", "camera"),
]:
    _ss(k, v)

sku_sheets   = load_sku_sheets()
sku_master   = build_sku_master(sku_sheets)
prefix_index = build_prefix_index(sku_master)
customers    = load_customers()

mrp_data = load_mrp_data_for_state(st.session_state.selected_state)
available_states = load_state_names_from_csv()

@st.cache_data(show_spinner=False)
def build_product_size_map(sku_master_frozen: str) -> dict:
    master = json.loads(sku_master_frozen)
    pmap: dict[str, dict[str, str]] = {}
    for sku, info in master.items():
        prod = info.get("product", "").strip()
        sz   = info.get("od_size", "").strip()
        if prod and sz:
            pmap.setdefault(prod, {})[sz] = sku
    return pmap


def validate_party(d: dict, prefix: str) -> list[str]:
    errors = []
    phone  = d.get("phone", "").strip()
    mobile = d.get("mobile", "").strip()
    pno    = d.get("party_no", "").strip()

    if phone and not re.fullmatch(r'\d{10}', phone):
        errors.append(f"{prefix} Phone must be exactly 10 digits (no spaces/symbols).")
    if mobile and not re.fullmatch(r'\d{10}', mobile):
        errors.append(f"{prefix} Mobile must be exactly 10 digits (no spaces/symbols).")
    if pno and not re.fullmatch(r'\d+', pno):
        errors.append(f"{prefix} Party No. must contain digits only.")
    return errors


def is_party_complete(d: dict) -> bool:
    return bool(
        d.get("party_no","").strip() and
        d.get("name","").strip() and
        d.get("phone","").strip() and
        d.get("mobile","").strip() and
        d.get("state_code","").strip() and
        d.get("state","").strip() and
        d.get("gst","").strip() and
        d.get("pan","").strip()
    )


st.markdown("""
<div class="app-header">
  <div class="app-header-badge">🔴</div>
  <div class="app-header-text">
    <h1>Sintex BAPL — Quotation Generator</h1>
    <p>CPVC / UPVC Pipes &amp; Fittings · Chhattisgarh · Any Form Layout</p>
  </div>
</div>""", unsafe_allow_html=True)


step_unlocked = [
    True,
    st.session_state.image_submitted,
    st.session_state.ocr_reviewed,
    st.session_state.party_confirmed,
]

cur_active = 1
if st.session_state.party_confirmed:
    cur_active = 4
elif st.session_state.ocr_reviewed:
    cur_active = 3
elif st.session_state.image_submitted:
    cur_active = 2

nav_html = '<div class="step-navbar">'
for i, lbl in enumerate(["📷 Capture", "⚡ Edit/Review Prices", "👤 User Information", "📄 Download and Send Quotation"], 1):
    is_active = (i == cur_active)
    is_done   = step_unlocked[i-1] and (i < cur_active)
    is_locked = not step_unlocked[i-1]
    cls = "locked" if is_locked else ("active" if is_active else ("done" if is_done else ""))
    dot_txt = "✓" if is_done else str(i)
    nav_html += f"""
    <div class="step-nav-item {cls}">
      <div class="step-nav-dot">{dot_txt}</div>
      <div class="step-nav-label">{lbl}</div>
    </div>"""
nav_html += "</div>"
st.markdown(nav_html, unsafe_allow_html=True)


def render_rotated_preview(raw_bytes: bytes, rotation: int) -> str:
    try:
        img = Image.open(io.BytesIO(raw_bytes))
        img = fix_exif_orientation(img)
        if img.mode not in ("RGB", "L"):
            img = img.convert("RGB")
        max_dim = 1200
        w, h = img.size
        if max(w, h) > max_dim:
            ratio = max_dim / max(w, h)
            img = img.resize((int(w * ratio), int(h * ratio)), Image.LANCZOS)
        if rotation != 0:
            img = img.rotate(-rotation, expand=True)
        preview_buf = io.BytesIO()
        img.save(preview_buf, format="JPEG", quality=85)
        return base64.b64encode(preview_buf.getvalue()).decode()
    except Exception:
        return base64.b64encode(raw_bytes).decode()


def render_step1():
    done = st.session_state.image_submitted and st.session_state.ocr_done

    st.markdown(f"""
    <div class="step-card"><div class="step-card-header">
      <div class="step-number {'done' if done else ''}">{'✓' if done else '1'}</div>
      <div><div class="step-title">Step 1 — Capture Order Form</div>
      <div class="step-subtitle">Select your state, then upload or photograph any Sintex order form</div></div>
    </div><div class="step-body">""", unsafe_allow_html=True)

    try:
        current_idx = available_states.index(st.session_state.selected_state)
    except ValueError:
        current_idx = 0

    st.markdown('<div class="state-inner-card">'
                '<div class="state-inner-title">🗺️ Select State / Region</div>',
                unsafe_allow_html=True)

    chosen = st.selectbox(
        "State",
        options=available_states,
        index=current_idx,
        key="state_selectbox",
        label_visibility="collapsed",
        help="Select the state for which this quotation is being generated.",
    )

    if chosen != st.session_state.selected_state:
        st.session_state.selected_state = chosen
        load_mrp_data_for_state.clear()
        if not st.session_state.image_submitted:
            st.rerun()
        else:
            st.markdown(
                '<div class="warn-box">⚠️ State changed. '
                'The new pricing will apply when you regenerate the PDF.</div>',
                unsafe_allow_html=True,
            )

    st.markdown(
        f'<div class="info-box" style="margin-top:8px;margin-bottom:0;">📍 Selected: '
        f'<b>{st.session_state.selected_state}</b> — '
        f'MRP &amp; distributor landing prices loaded for this state.</div>',
        unsafe_allow_html=True,
    )
    st.markdown('</div>', unsafe_allow_html=True)

    with st.expander("🔧 Azure OCR Settings", expanded=not st.session_state.azure_key):
        st.markdown('<div class="info-box">Endpoint: <code>https://&lt;resource&gt;.cognitiveservices.azure.com</code><br/>'
                    'Auto-tries Document Intelligence v4 → v3 → Form Recognizer v2.1</div>',
                    unsafe_allow_html=True)
        st.session_state.azure_endpoint = st.text_input(
            "Azure Endpoint", value=st.session_state.azure_endpoint,
            placeholder="https://YOUR-RESOURCE.cognitiveservices.azure.com")
        st.session_state.azure_key = st.text_input(
            "Azure API Key", value=st.session_state.azure_key,
            type="password", placeholder="••••••••••••••••••••••••••••••••")

    st.markdown("""<div class="info-box">
    📋 <b>Any form layout supported.</b> The OCR engine extracts bounding-box positions of every
    word, then <b>spatially reconstructs the table</b> — it finds which column has SKU prefixes,
    which have size headers (15MM/20MM/…), and reads quantities from the right cells.</div>""",
    unsafe_allow_html=True)

    if not st.session_state.image_submitted:
        upload_key_suffix = st.session_state.upload_key

        st.markdown("""
        <style>
        [data-testid="stFileUploaderDropzone"]{
          border:2.5px dashed #C0211F!important;
          border-radius:10px!important;
          background:rgba(192,33,31,0.04)!important;
        }
        [data-testid="stFileUploaderDropzone"]:hover{
          background:rgba(192,33,31,0.10)!important;
        }
        [data-testid="stFileUploaderDropzoneInstructions"] svg{color:#C0211F!important;}
        [data-testid="stFileUploaderDropzoneInstructions"] span{color:#C0211F!important;font-weight:600!important;}
        [data-testid="stFileUploaderDropzoneInstructions"] small{color:#888!important;}
        button[data-testid="baseButton-secondary"]{
          background:linear-gradient(135deg,#C0211F,#8B1514)!important;
          color:white!important;border:none!important;
          border-radius:10px!important;font-weight:700!important;
        }
        </style>
        """, unsafe_allow_html=True)

        st.markdown("""
        <style>
        @keyframes sintex-pulse{
          0%,100%{opacity:1;}
          50%{opacity:.3;}
        }
        .sintex-capture-wrap{
          border-radius:12px;
          overflow:hidden;
          border:2px solid #2a2a2a;
          margin-left:-0.75rem;
          margin-right:-0.75rem;
        }
        .sintex-capture-wrap [data-baseweb="tab-list"]{
          background:#1a1a1a!important;
          border-radius:0!important;
          padding:0!important;
          gap:0!important;
          border-bottom:none!important;
        }
        .sintex-capture-wrap [data-baseweb="tab"]{
          background:transparent!important;
          color:#888!important;
          border:none!important;
          border-bottom:3px solid transparent!important;
          border-radius:0!important;
          padding:14px 24px!important;
          font-size:14px!important;
          font-weight:600!important;
          font-family:'Inter',sans-serif!important;
          flex:1!important;
          justify-content:center!important;
        }
        .sintex-capture-wrap [aria-selected="true"]{
          color:#C0211F!important;
          border-bottom:3px solid #C0211F!important;
          background:#1f0a0a!important;
        }
        .sintex-capture-wrap [data-baseweb="tab-panel"]{
          padding:0!important;
          border-top:none!important;
          background:#000!important;
        }

        .sintex-capture-wrap [data-testid="stCameraInput"]{
          width:100%!important;
          display:block!important;
          background:#000!important;
        }
        .sintex-capture-wrap [data-testid="stCameraInput"] video{
          width:100%!important;
          height:62vw!important;
          max-height:520px!important;
          min-height:280px!important;
          object-fit:cover!important;
          display:block!important;
          background:#000!important;
        }
        .sintex-capture-wrap [data-testid="stCameraInput"] img{
          width:100%!important;
          max-height:520px!important;
          min-height:280px!important;
          object-fit:contain!important;
          display:block!important;
          background:#000!important;
        }
        .sintex-capture-wrap [data-testid="stCameraInputButton"]{
          width:100%!important;
          height:58px!important;
          border-radius:0!important;
          font-size:16px!important;
          font-weight:700!important;
          background:linear-gradient(135deg,#C0211F,#8B1514)!important;
          color:white!important;
          border:none!important;
          box-shadow:none!important;
          letter-spacing:.5px!important;
        }

        .cam-overlay-bar{
          background:#1a1a1a;
          display:flex;
          align-items:center;
          justify-content:space-between;
          padding:8px 16px;
          border-bottom:1px solid #2a2a2a;
        }
        .cam-rec-dot{
          display:inline-flex;
          align-items:center;
          gap:7px;
          background:rgba(192,33,31,0.85);
          color:white;
          font-size:11px;
          font-weight:800;
          letter-spacing:1.4px;
          padding:4px 12px;
          border-radius:5px;
          font-family:'Inter',sans-serif;
        }
        .cam-rec-dot span.dot{
          width:9px;height:9px;border-radius:50%;
          background:white;display:inline-block;
          animation:sintex-pulse 1.2s ease-in-out infinite;
        }
        .cam-tip{
          color:#888;font-size:11px;font-family:'Inter',sans-serif;
        }

        .upload-panel-inner{
          background:#111;
          padding:20px 16px;
          min-height:200px;
          display:flex;
          flex-direction:column;
          justify-content:center;
        }
        .upload-panel-inner [data-testid="stFileUploaderDropzone"]{
          border:2px dashed #C0211F!important;
          background:rgba(192,33,31,0.06)!important;
          border-radius:10px!important;
        }

        @media(max-width:600px){
          .sintex-capture-wrap{
            margin-left:-0.5rem;
            margin-right:-0.5rem;
          }
          .sintex-capture-wrap [data-testid="stCameraInput"] video{
            height:72vw!important;
            max-height:420px!important;
          }
          .sintex-capture-wrap [data-testid="stCameraInput"] img{
            max-height:420px!important;
          }
          .sintex-capture-wrap [data-testid="stCameraInputButton"]{
            height:52px!important;
            font-size:15px!important;
          }
        }
        </style>
        """, unsafe_allow_html=True)

        # ── Navbar for capture mode ──────────────────────────────────────────
        if "capture_mode" not in st.session_state:
            st.session_state.capture_mode = "camera"

        st.markdown("""
        <style>
        .capture-mode-nav{
          display:flex;background:#1a1a1a;border-radius:12px 12px 0 0;
          overflow:hidden;border:2px solid #2a2a2a;border-bottom:none;margin-bottom:0;
        }
        .capture-mode-nav button{
          flex:1;padding:14px 10px;border:none;background:transparent;
          color:#777;font-size:14px;font-weight:700;font-family:'Inter',sans-serif;
          cursor:pointer;border-bottom:3px solid transparent;transition:all .2s;
          letter-spacing:.3px;
        }
        .capture-mode-nav button.active{
          color:#C0211F;border-bottom:3px solid #C0211F;background:#1f0a0a;
        }
        .capture-mode-nav button:hover:not(.active){color:#ccc;background:#222;}
        </style>
        """, unsafe_allow_html=True)

        col_cam_nav, col_up_nav = st.columns(2)
        with col_cam_nav:
            cam_active = "active" if st.session_state.capture_mode == "camera" else ""
            if st.button("📷  Camera", key="nav_camera",
                         use_container_width=True):
                st.session_state.capture_mode = "camera"
                st.rerun()
        with col_up_nav:
            up_active = "active" if st.session_state.capture_mode == "upload" else ""
            if st.button("📁  Upload File", key="nav_upload",
                         use_container_width=True):
                st.session_state.capture_mode = "upload"
                st.rerun()

        # highlight active nav button via JS
        st.markdown(f"""
        <script>
        (function(){{
          var mode = "{st.session_state.capture_mode}";
          var btns = window.parent.document.querySelectorAll(
            'button[data-testid="baseButton-secondary"]');
          // handled via rerun + CSS below
        }})();
        </script>
        <style>
        /* style the active nav btn */
        div[data-testid="column"]:nth-child({"1" if st.session_state.capture_mode == "camera" else "2"})
          button[kind="secondary"],
        div[data-testid="column"]:nth-child({"1" if st.session_state.capture_mode == "camera" else "2"})
          .stButton>button{{
          background:#1f0a0a!important;
          color:#C0211F!important;
          border-bottom:3px solid #C0211F!important;
          border-radius:10px 10px 0 0!important;
        }}
        </style>
        """, unsafe_allow_html=True)

        img_bytes = None

        # ── CAMERA MODE ──────────────────────────────────────────────────────
        if st.session_state.capture_mode == "camera":

            st.markdown("""
            <style>
            .sintex-native-cam-wrap{
              background:#000;border:2px solid #2a2a2a;
              border-radius:0 0 12px 12px;overflow:hidden;
            }
            .cam-instruction-bar{
              background:#111;padding:12px 16px;
              border-bottom:1px solid #222;
            }
            .cam-instruction-bar p{
              color:#aaa;font-size:12.5px;font-family:'Inter',sans-serif;
              margin:0;line-height:1.5;
            }
            .cam-instruction-bar b{color:#fff;}
            .native-cam-btn-wrap{
              padding:20px 16px;display:flex;flex-direction:column;
              align-items:center;gap:12px;background:#0a0a0a;
            }
            /* Style the file_uploader used as camera trigger */
            .native-cam-btn-wrap [data-testid="stFileUploaderDropzone"]{
              display:none!important;
            }
            .native-cam-btn-wrap label[data-testid="stFileUploaderDropzoneInput"]{
              display:none!important;
            }
            </style>
            <div class="sintex-native-cam-wrap">
              <div class="cam-instruction-bar">
                <p>📷 <b>Native Camera Capture</b> — tap the button below to open your
                device camera directly. The photo will appear here for review.</p>
              </div>
            </div>
            """, unsafe_allow_html=True)

            # Native camera trigger using HTML input with capture="environment"
            # We use a Streamlit component to inject a real <input capture> button
            cam_html = f"""
            <style>
            .sintex-cam-outer{{
              background:#0a0a0a;
              border:2px solid #2a2a2a;
              border-top:none;
              border-radius:0 0 12px 12px;
              padding:24px 16px 20px;
              display:flex;flex-direction:column;align-items:center;gap:16px;
            }}
            .sintex-cam-btn{{
              display:inline-flex;align-items:center;justify-content:center;gap:10px;
              background:linear-gradient(135deg,#C0211F,#8B1514);
              color:white;border:none;border-radius:12px;
              padding:16px 32px;font-size:16px;font-weight:700;
              font-family:'Inter',sans-serif;cursor:pointer;
              box-shadow:0 4px 20px rgba(192,33,31,.45);
              letter-spacing:.4px;width:100%;max-width:340px;
              transition:transform .15s,box-shadow .15s;
            }}
            .sintex-cam-btn:hover{{transform:translateY(-2px);box-shadow:0 8px 28px rgba(192,33,31,.5);}}
            .sintex-cam-btn:active{{transform:translateY(0);}}
            .cam-tip{{color:#555;font-size:11.5px;font-family:'Inter',sans-serif;text-align:center;}}
            #sintex-cam-input{{display:none;}}
            #sintex-preview-section{{display:none;width:100%;margin-top:8px;}}
            #sintex-preview-img{{
              width:100%;max-height:480px;object-fit:contain;
              display:block;border-radius:8px;background:#111;
            }}
            .preview-bar{{
              background:#1a1a1a;padding:8px 14px;border-radius:8px 8px 0 0;
              display:flex;align-items:center;justify-content:space-between;
              margin-bottom:0;
            }}
            .preview-label{{color:#aaa;font-size:12px;font-weight:600;font-family:'Inter',sans-serif;}}
            .captured-badge{{background:#1E7E4A;color:white;font-size:10px;font-weight:700;
              padding:3px 10px;border-radius:4px;letter-spacing:1px;}}
            .sintex-use-btn{{
              display:inline-flex;align-items:center;justify-content:center;gap:8px;
              background:linear-gradient(135deg,#1E7E4A,#155d38);
              color:white;border:none;border-radius:10px;
              padding:14px 24px;font-size:15px;font-weight:700;
              font-family:'Inter',sans-serif;cursor:pointer;
              box-shadow:0 4px 16px rgba(30,126,74,.4);
              width:100%;max-width:340px;margin-top:8px;
              transition:transform .15s;
            }}
            .sintex-use-btn:hover{{transform:translateY(-1px);}}
            .sintex-retake-btn{{
              display:inline-flex;align-items:center;justify-content:center;gap:8px;
              background:transparent;color:#888;
              border:1.5px solid #333;border-radius:10px;
              padding:10px 20px;font-size:13px;font-weight:600;
              font-family:'Inter',sans-serif;cursor:pointer;
              width:100%;max-width:340px;
              transition:all .15s;
            }}
            .sintex-retake-btn:hover{{border-color:#C0211F;color:#C0211F;}}
            #sintex-b64-output{{display:none;}}
            #sintex-dim-output{{display:none;}}
            </style>

            <div class="sintex-cam-outer">
              <input type="file" id="sintex-cam-input"
                     accept="image/*" capture="environment" />

              <div id="sintex-shoot-section" style="display:flex;flex-direction:column;
                align-items:center;gap:12px;width:100%;">
                <button class="sintex-cam-btn" onclick="document.getElementById('sintex-cam-input').click()">
                  📷 &nbsp; Open Camera & Take Photo
                </button>
                <p class="cam-tip">Opens your device's native camera app.<br/>
                  Point at the order form and take the photo.</p>
              </div>

              <div id="sintex-preview-section">
                <div class="preview-bar">
                  <span class="preview-label">📸 Preview</span>
                  <span class="captured-badge">✓ CAPTURED</span>
                </div>
                <img id="sintex-preview-img" src="" alt="preview"/>
                <div style="display:flex;flex-direction:column;align-items:center;
                  gap:8px;margin-top:10px;width:100%;">
                  <button class="sintex-use-btn" id="sintex-use-btn"
                    onclick="submitPhotoToStreamlit()">
                    ✅ &nbsp; Use This Photo
                  </button>
                  <button class="sintex-retake-btn"
                    onclick="retakePhoto()">
                    🔄 &nbsp; Retake
                  </button>
                </div>
              </div>

              <textarea id="sintex-b64-output" rows="1"></textarea>
              <textarea id="sintex-dim-output" rows="1"></textarea>
            </div>

            <script>
            (function(){{
              var input  = document.getElementById('sintex-cam-input');
              var prevSec= document.getElementById('sintex-preview-section');
              var shootSec=document.getElementById('sintex-shoot-section');
              var prevImg= document.getElementById('sintex-preview-img');
              var b64Out = document.getElementById('sintex-b64-output');
              var dimOut = document.getElementById('sintex-dim-output');
              var capturedB64 = null;

              input.addEventListener('change', function(){{
                var file = input.files[0];
                if (!file) return;
                var reader = new FileReader();
                reader.onload = function(e){{
                  capturedB64 = e.target.result; // data:image/...;base64,...
                  prevImg.src = capturedB64;
                  shootSec.style.display = 'none';
                  prevSec.style.display  = 'block';
                  b64Out.value = capturedB64;
                  dimOut.value = file.size + '|' + file.name;
                }};
                reader.readAsDataURL(file);
              }});

              window.retakePhoto = function(){{
                input.value = '';
                capturedB64 = null;
                prevImg.src = '';
                prevSec.style.display  = 'none';
                shootSec.style.display = 'flex';
                b64Out.value = '';
                dimOut.value = '';
              }};

              window.submitPhotoToStreamlit = function(){{
                if (!capturedB64) return;
                // Post to parent streamlit via query param trick
                var b64data = capturedB64.split(',')[1];
                window.parent.postMessage({{
                  type: 'sintex_photo',
                  b64: b64data,
                  mime: capturedB64.split(';')[0].split(':')[1]
                }}, '*');
              }};
            }})();
            </script>
            """

            # Use Streamlit components to render the camera HTML
            # We capture the photo via a hidden file_uploader acting as the bridge
            import streamlit.components.v1 as stc

            # The actual native camera capture widget
            components.html(cam_html, height=420, scrolling=False)

            # Bridge: hidden file uploader that we trigger programmatically via JS message
            # Since postMessage can't directly call st.session_state, we use a workaround:
            # render a real st.camera_input or file_uploader hidden, but styled as the
            # native button — OR use query_params as a relay.
            # Best approach for Streamlit: use a file_uploader with accept image/* 
            # but render our own button via HTML that clicks the hidden input.

            st.markdown("""
            <style>
            /* Hide the default streamlit file uploader UI, show only via our custom button */
            .sintex-hidden-cam-upload [data-testid="stFileUploaderDropzone"]{
              opacity:0!important;height:1px!important;overflow:hidden!important;
              position:absolute!important;pointer-events:none!important;
            }
            .sintex-hidden-cam-upload [data-testid="stFileUploaderDropzone"] *{{
              display:none!important;
            }}
            .sintex-hidden-cam-upload{{
              height:0;overflow:hidden;margin:0;padding:0;
            }}
            </style>
            """, unsafe_allow_html=True)

            # Hidden uploader that the user won't interact with directly —
            # but we do need a real Streamlit uploader to receive the file.
            # The HTML component above handles camera → we need to relay.
            # SOLUTION: Use a visible st.file_uploader with capture attribute injected via JS.

            st.markdown('<div style="display:none" id="sintex-cam-uploader-wrap">',
                        unsafe_allow_html=True)
            cam_capture_file = st.file_uploader(
                "camera_bridge",
                type=["jpg","jpeg","png","webp","heic","heif"],
                key=f"native_cam_{upload_key_suffix}",
                label_visibility="collapsed",
            )
            st.markdown('</div>', unsafe_allow_html=True)

            # Inject JS to add capture="environment" to the hidden input and wire our button
            components.html(f"""
            <script>
            (function waitForInput(){{
              // Find the file input inside the Streamlit uploader in the parent frame
              var iframes = window.parent.document.querySelectorAll('iframe');
              var inputs  = window.parent.document.querySelectorAll(
                'input[type="file"][accept*="image"]');

              if(inputs.length === 0){{
                setTimeout(waitForInput, 200); return;
              }}

              // Add capture attribute to the most recently added file input
              var lastInput = inputs[inputs.length - 1];
              lastInput.setAttribute('capture', 'environment');
              lastInput.setAttribute('accept', 'image/*');

              // Listen for our custom button click from the component above
              window.addEventListener('message', function(ev){{
                if(ev.data && ev.data.type === 'sintex_photo'){{
                  // Convert base64 to File and dispatch change event
                  var b64 = ev.data.b64;
                  var mime= ev.data.mime || 'image/jpeg';
                  var byteStr = atob(b64);
                  var arr = new Uint8Array(byteStr.length);
                  for(var i=0;i<byteStr.length;i++) arr[i]=byteStr.charCodeAt(i);
                  var blob = new Blob([arr], {{type: mime}});
                  var file = new File([blob], 'capture.jpg', {{type: mime}});
                  var dt   = new DataTransfer();
                  dt.items.add(file);
                  lastInput.files = dt.files;
                  lastInput.dispatchEvent(new Event('change', {{bubbles:true}}));
                }}
              }}, false);

              // Also wire the native open-camera button to this input directly
              // so clicking our custom HTML button opens the native camera
              var camBtns = document.querySelectorAll('.sintex-cam-btn');
              if(camBtns.length){{
                camBtns[0].addEventListener('click', function(e){{
                  e.preventDefault();
                  lastInput.click();
                }});
              }}
            }})();
            </script>
            """, height=0, scrolling=False)

            raw_cam = cam_capture_file.getvalue() if cam_capture_file is not None else None

            if raw_cam is not None:
                rotation = st.session_state.image_rotation
                b64_prev = render_rotated_preview(raw_cam, rotation)
                w, h = get_image_dimensions(raw_cam)

                st.markdown(f"""
                <div style="background:#000;margin-top:16px;border-radius:10px;overflow:hidden;
                  border:1.5px solid #333;">
                  <div style="background:#1a1a1a;padding:8px 16px;display:flex;
                    align-items:center;justify-content:space-between;">
                    <span style="color:#888;font-size:12px;font-weight:600;
                      font-family:'Inter',sans-serif;">
                      📸 Captured Image &nbsp;
                      <span style="color:#555;font-size:10px;">{w}×{h}px</span>
                    </span>
                    <span style="background:#1E7E4A;color:white;font-size:10px;font-weight:700;
                      padding:3px 10px;border-radius:4px;letter-spacing:1px;">✓ CAPTURED</span>
                  </div>
                  <img src="data:image/jpeg;base64,{b64_prev}"
                    style="width:100%;max-height:520px;min-height:200px;
                    object-fit:contain;display:block;background:#000;"/>
                </div>
                """, unsafe_allow_html=True)

                st.markdown("""
                <div style="background:#1a1a1a;padding:10px 14px;border-radius:8px;
                  margin-top:8px;text-align:center;">
                  <span style="color:#aaa;font-size:12px;font-weight:600;
                    font-family:'Inter',sans-serif;">🔄 Rotate if needed:</span>
                </div>
                """, unsafe_allow_html=True)

                rc1, rc2, rc3, rc4 = st.columns(4)
                with rc1:
                    st.markdown('<div class="rotate-btn-wrap">', unsafe_allow_html=True)
                    if st.button("↺ 90° L", key="rot_ccw"):
                        st.session_state.image_rotation = (st.session_state.image_rotation - 90) % 360
                        st.rerun()
                    st.markdown('</div>', unsafe_allow_html=True)
                with rc2:
                    st.markdown('<div class="rotate-btn-wrap">', unsafe_allow_html=True)
                    if st.button("↻ 90° R", key="rot_cw"):
                        st.session_state.image_rotation = (st.session_state.image_rotation + 90) % 360
                        st.rerun()
                    st.markdown('</div>', unsafe_allow_html=True)
                with rc3:
                    st.markdown('<div class="rotate-btn-wrap">', unsafe_allow_html=True)
                    if st.button("↕ 180°", key="rot_180"):
                        st.session_state.image_rotation = (st.session_state.image_rotation + 180) % 360
                        st.rerun()
                    st.markdown('</div>', unsafe_allow_html=True)
                with rc4:
                    st.markdown('<div class="rotate-btn-wrap">', unsafe_allow_html=True)
                    if st.button("⟲ Reset", key="rot_reset"):
                        st.session_state.image_rotation = 0
                        st.rerun()
                    st.markdown('</div>', unsafe_allow_html=True)

                st.markdown("""
                <div style="background:#0d1a0d;padding:6px 14px;text-align:center;
                  margin-top:4px;border-radius:6px;">
                  <span style="color:#4ade80;font-size:11px;font-family:'Inter',sans-serif;">
                    ✓ Make sure form text is upright before submitting
                  </span>
                </div>
                """, unsafe_allow_html=True)

                st.markdown('<div class="btn-green" style="margin-top:14px;">',
                            unsafe_allow_html=True)
                if st.button("✅  Submit & Process Image", key="submit_cam_photo"):
                    img_bytes = raw_cam
                st.markdown('</div>', unsafe_allow_html=True)

        # ── UPLOAD MODE ──────────────────────────────────────────────────────
        else:
            st.markdown("""
            <style>
            .sintex-upload-outer{
              background:#111;border:2px solid #2a2a2a;
              border-radius:0 0 12px 12px;padding:20px 16px;
            }
            .sintex-upload-outer [data-testid="stFileUploaderDropzone"]{
              border:2px dashed #C0211F!important;
              background:rgba(192,33,31,0.06)!important;
              border-radius:10px!important;
            }
            .sintex-upload-outer [data-testid="stFileUploaderDropzoneInstructions"] span{
              color:#C0211F!important;font-weight:700!important;font-size:14px!important;
            }
            .sintex-upload-outer [data-testid="stFileUploaderDropzoneInstructions"] small{
              color:#888!important;font-size:12px!important;
            }
            button[data-testid="baseButton-secondary"]{
              background:linear-gradient(135deg,#C0211F,#8B1514)!important;
              color:white!important;border:none!important;
              border-radius:8px!important;font-weight:700!important;
            }
            </style>
            <div class="sintex-upload-outer">
            """, unsafe_allow_html=True)

            up_file = st.file_uploader(
                "📁 Choose image from device memory",
                type=["jpg", "jpeg", "png", "webp"],
                key=f"up_{upload_key_suffix}",
                label_visibility="visible",
            )

            if up_file is not None:
                raw = up_file.read()
                rotation = st.session_state.image_rotation
                b64_up = render_rotated_preview(raw, rotation)
                w, h = get_image_dimensions(raw)

                st.markdown(f"""
                <div style="margin-top:14px;border:1.5px solid #333;border-radius:8px;
                  overflow:hidden;">
                  <div style="background:#1a1a1a;padding:6px 12px;display:flex;
                    align-items:center;justify-content:space-between;">
                    <span style="color:#aaa;font-size:11px;">📎 {up_file.name}</span>
                    <span style="color:#555;font-size:10px;">{w}×{h}px</span>
                  </div>
                  <img src="data:image/jpeg;base64,{b64_up}"
                    style="width:100%;max-height:400px;object-fit:contain;
                    display:block;background:#000;"/>
                </div>
                """, unsafe_allow_html=True)

                st.markdown("""
                <div style="background:#1a1a1a;padding:10px 14px;border-radius:8px;
                  margin-top:10px;text-align:center;">
                  <span style="color:#aaa;font-size:12px;font-weight:600;
                    font-family:'Inter',sans-serif;">🔄 Rotate if needed:</span>
                </div>
                """, unsafe_allow_html=True)

                ur1, ur2, ur3, ur4 = st.columns(4)
                with ur1:
                    st.markdown('<div class="rotate-btn-wrap">', unsafe_allow_html=True)
                    if st.button("↺ 90° L", key="urot_ccw"):
                        st.session_state.image_rotation = (st.session_state.image_rotation - 90) % 360
                        st.rerun()
                    st.markdown('</div>', unsafe_allow_html=True)
                with ur2:
                    st.markdown('<div class="rotate-btn-wrap">', unsafe_allow_html=True)
                    if st.button("↻ 90° R", key="urot_cw"):
                        st.session_state.image_rotation = (st.session_state.image_rotation + 90) % 360
                        st.rerun()
                    st.markdown('</div>', unsafe_allow_html=True)
                with ur3:
                    st.markdown('<div class="rotate-btn-wrap">', unsafe_allow_html=True)
                    if st.button("↕ 180°", key="urot_180"):
                        st.session_state.image_rotation = (st.session_state.image_rotation + 180) % 360
                        st.rerun()
                    st.markdown('</div>', unsafe_allow_html=True)
                with ur4:
                    st.markdown('<div class="rotate-btn-wrap">', unsafe_allow_html=True)
                    if st.button("⟲ Reset", key="urot_reset"):
                        st.session_state.image_rotation = 0
                        st.rerun()
                    st.markdown('</div>', unsafe_allow_html=True)

                st.markdown('<div class="btn-green" style="margin-top:12px;">',
                            unsafe_allow_html=True)
                if st.button("✅  Submit Uploaded Image", key="submit_upload"):
                    img_bytes = raw
                st.markdown('</div>', unsafe_allow_html=True)

            st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('</div>', unsafe_allow_html=True)

        if img_bytes is not None:
            for k in ["quantities","match_log","ocr_grid","ocr_meta","ocr_extracted","line_items"]:
                st.session_state[k] = [] if isinstance(st.session_state[k], list) else {}
            st.session_state.image_bytes     = img_bytes
            st.session_state.raw_image_bytes = img_bytes
            st.session_state.ocr_done        = False
            st.session_state.ocr_reviewed    = False
            st.session_state.party_confirmed = False
            st.session_state.pdf_bytes       = None
            st.session_state.image_submitted = True
            st.rerun()

    else:
        if st.session_state.ocr_done:
            st.markdown('<div class="success-box">✅ Image submitted &amp; OCR completed</div>',
                        unsafe_allow_html=True)
        else:
            raw_bytes = st.session_state.raw_image_bytes
            if raw_bytes:
                rotation = st.session_state.image_rotation
                b64_sub = render_rotated_preview(raw_bytes, rotation)
                w, h = get_image_dimensions(raw_bytes)
                st.markdown(f"""
                <div style="border:1.5px solid var(--border);border-radius:10px;overflow:hidden;
                  margin-bottom:14px;">
                  <div style="background:#1a1a1a;padding:7px 14px;display:flex;
                    align-items:center;justify-content:space-between;">
                    <span style="color:#aaa;font-size:12px;font-weight:600;">📷 Submitted Image</span>
                    <span style="color:#555;font-size:10px;">{w}×{h}px</span>
                  </div>
                  <img src="data:image/jpeg;base64,{b64_sub}"
                    style="width:100%;max-height:300px;object-fit:contain;display:block;background:#000;"/>
                </div>
                """, unsafe_allow_html=True)

                st.markdown("""
                <div style="background:#f8f8f8;padding:10px 14px;border:1px solid var(--border);
                  border-radius:8px;margin-bottom:8px;">
                  <span style="font-size:12px;font-weight:600;color:#444;">🔄 Adjust rotation if needed:</span>
                </div>
                """, unsafe_allow_html=True)

                sr1, sr2, sr3, sr4 = st.columns(4)
                with sr1:
                    st.markdown('<div class="rotate-btn-wrap">', unsafe_allow_html=True)
                    if st.button("↺ 90° L", key="srot_ccw"):
                        st.session_state.image_rotation = (st.session_state.image_rotation - 90) % 360
                        st.rerun()
                    st.markdown('</div>', unsafe_allow_html=True)
                with sr2:
                    st.markdown('<div class="rotate-btn-wrap">', unsafe_allow_html=True)
                    if st.button("↻ 90° R", key="srot_cw"):
                        st.session_state.image_rotation = (st.session_state.image_rotation + 90) % 360
                        st.rerun()
                    st.markdown('</div>', unsafe_allow_html=True)
                with sr3:
                    st.markdown('<div class="rotate-btn-wrap">', unsafe_allow_html=True)
                    if st.button("↕ 180°", key="srot_180"):
                        st.session_state.image_rotation = (st.session_state.image_rotation + 180) % 360
                        st.rerun()
                    st.markdown('</div>', unsafe_allow_html=True)
                with sr4:
                    st.markdown('<div class="rotate-btn-wrap">', unsafe_allow_html=True)
                    if st.button("⟲ Reset", key="srot_reset"):
                        st.session_state.image_rotation = 0
                        st.rerun()
                    st.markdown('</div>', unsafe_allow_html=True)

            st.markdown('<div class="success-box">✅ Image submitted — proceed to Step 2 for OCR</div>',
                        unsafe_allow_html=True)

    st.markdown("</div></div>", unsafe_allow_html=True)


def render_step2():
    if not step_unlocked[1]:
        return

    done = st.session_state.ocr_done
    st.markdown(f"""
    <div class="step-card"><div class="step-card-header">
      <div class="step-number {'done' if done else ''}">{'✓' if done else '2'}</div>
      <div><div class="step-title">Step 2 — OCR · Spatial Reconstruction · SKU Matching</div>
      <div class="step-subtitle">Detects table structure from bounding boxes — works with any form layout</div></div>
    </div><div class="step-body">""", unsafe_allow_html=True)

    if not done:
        c1, c2 = st.columns(2)
        with c1:
            run = st.button("⚡  Run OCR & Match", key="run_ocr",
                            disabled=not (st.session_state.azure_key and st.session_state.azure_endpoint and st.session_state.image_bytes))
        with c2:
            skip = st.button("✏️  Enter SKUs Manually", key="skip_ocr")

        if not st.session_state.image_bytes:
            st.markdown('<div class="error-box">❌ No image captured — go back to Step 1 and capture or upload an image first.</div>',
                        unsafe_allow_html=True)
        if not st.session_state.azure_endpoint:
            st.markdown('<div class="warn-box">⚠️ Azure Endpoint not set — configure in Step 1 → Azure OCR Settings.</div>',
                        unsafe_allow_html=True)
        if not st.session_state.azure_key:
            st.markdown('<div class="warn-box">⚠️ Azure API Key not set — configure in Step 1 → Azure OCR Settings.</div>',
                        unsafe_allow_html=True)

        if run:
            ph = st.empty()
            try:
                ph.info("⏳ 1/4 — Enhancing image quality for OCR…")
                enhanced_bytes = enhance_image_for_ocr(
                    st.session_state.image_bytes,
                    st.session_state.get("image_rotation", 0)
                )

                ph.info("⏳ 2/4 — Sending enhanced image to Azure Document Intelligence…")
                words = run_azure_ocr(enhanced_bytes,
                                      st.session_state.azure_endpoint,
                                      st.session_state.azure_key)

                ph.info(f"🔲 3/4 — Reconstructing table from {len(words)} OCR words…")
                grid = reconstruct_table(words)

                ph.info(f"🔍 4/4 — Analysing structure &amp; matching SKUs ({len(grid)} rows)…")
                res  = analyze_table(grid)
                qtys, log, line_items = build_quantities(res["rows"], prefix_index, sku_master)

                st.session_state.ocr_grid      = grid
                st.session_state.ocr_meta      = res["meta"]
                st.session_state.ocr_extracted = res["rows"]
                st.session_state.quantities    = qtys
                st.session_state.match_log     = log
                st.session_state.line_items    = line_items
                st.session_state.ocr_done      = True

                nm = sum(1 for m in log if m["status"] == "matched")
                nu = sum(1 for m in log if m["status"] == "unmatched")
                if nm == 0:
                    ph.warning(f"⚠️ Table reconstructed ({len(grid)} rows, {len(res['rows'])} data rows) "
                               f"but 0 SKUs matched. Check Raw OCR Table & Match Log tabs below.")
                else:
                    ph.success(f"✅ Done — {nm} detections matched, {nu} unmatched. Review below.")
                time.sleep(0.8); st.rerun()

            except requests.exceptions.HTTPError as e:
                ph.empty()
                status = e.response.status_code if e.response else "?"
                try:    body = e.response.json().get("error", {}).get("message", "")
                except: body = (e.response.text or "")[:300]
                st.error(f"❌ Azure HTTP {status}: {body or str(e)}")
            except requests.exceptions.ConnectionError:
                ph.empty(); st.error("❌ Cannot connect to Azure — check endpoint URL.")
            except TimeoutError as e:
                ph.empty(); st.error(f"❌ Timeout: {e}")
            except Exception as e:
                ph.empty(); st.error(f"❌ {type(e).__name__}: {e}")

        if skip:
            st.session_state.quantities = {}
            st.session_state.match_log  = []
            st.session_state.ocr_done   = True
            st.rerun()

    if done:
        grid = st.session_state.ocr_grid
        meta = st.session_state.ocr_meta
        log  = st.session_state.match_log
        qtys = st.session_state.quantities
        nm = sum(1 for m in log if m["status"] == "matched")
        nu = sum(1 for m in log if m["status"] == "unmatched")

        if nm: st.markdown(f'<div class="success-box">✅ {nm} SKU-size detections matched → {nm} line items in document</div>',
                           unsafe_allow_html=True)
        if nu: st.markdown(f'<div class="warn-box">⚠️ {nu} pair(s) unmatched — check Match Log tab</div>',
                           unsafe_allow_html=True)
        if not nm and not nu:
            st.markdown('<div class="warn-box">No OCR data — use Edit tab to add SKUs manually.</div>',
                        unsafe_allow_html=True)

        tab_edit, tab1, tab2, tab3 = st.tabs(["✏️ Edit / Add", "📋 Matched Items", "🔲 Raw OCR Table", "🔍 Match Log"])

        with tab_edit:
            st.caption("Add or correct items by selecting Product Name and Size.")

            _master_json = json.dumps(sku_master)
            pmap = build_product_size_map(_master_json)
            product_names = sorted(pmap.keys())

            st.markdown('<div class="fsl">Add Item by Product &amp; Size</div>', unsafe_allow_html=True)
            col_prod, col_sz, col_qty, col_add = st.columns([4, 2, 1, 1])

            with col_prod:
                sel_product = st.selectbox(
                    "Product Name",
                    options=["— select product —"] + product_names,
                    key="add_product_select",
                    label_visibility="collapsed",
                )

            if sel_product and sel_product != "— select product —":
                avail_sizes = sorted(pmap[sel_product].keys(),
                                     key=lambda s: int(re.sub(r'\D','',s) or 0))
            else:
                avail_sizes = []

            with col_sz:
                sel_size = st.selectbox(
                    "Size",
                    options=["— size —"] + avail_sizes,
                    key="add_size_select",
                    label_visibility="collapsed",
                    disabled=(not avail_sizes),
                )

            with col_qty:
                add_qty = st.number_input("Qty", min_value=1, value=1, step=1,
                                          key="add_qty_input",
                                          label_visibility="collapsed")

            resolved_sku = None
            if (sel_product and sel_product != "— select product —"
                    and sel_size and sel_size != "— size —"):
                resolved_sku = pmap[sel_product].get(sel_size)

            with col_add:
                st.markdown("<br/>", unsafe_allow_html=True)
                if st.button("➕", key="add_by_product", disabled=(resolved_sku is None)):
                    updated = copy.deepcopy(st.session_state.quantities)
                    updated[resolved_sku] = updated.get(resolved_sku, 0) + add_qty
                    new_items = copy.deepcopy(st.session_state.line_items)
                    new_items.append({"sku": resolved_sku, "qty": add_qty})
                    st.session_state.quantities = updated
                    st.session_state.line_items = new_items
                    info_added = sku_master.get(resolved_sku, {})
                    st.success(f"Added: {info_added.get('product',resolved_sku)} "
                               f"({sel_size}) × {add_qty} → SKU: {resolved_sku}")
                    st.rerun()

            if resolved_sku:
                st.markdown(
                    f'<div class="info-box" style="margin-top:6px;padding:8px 12px;">'
                    f'🔗 Resolved SKU: <code><b>{resolved_sku}</b></code></div>',
                    unsafe_allow_html=True,
                )
            elif sel_product != "— select product —" and sel_size != "— size —" and avail_sizes:
                st.markdown(
                    '<div class="warn-box" style="margin-top:6px;">⚠️ No SKU found for this combination.</div>',
                    unsafe_allow_html=True,
                )

            with st.expander("🔑 Advanced: Add by SKU Code directly"):
                ca, cb, cc = st.columns([3, 1, 1])
                with ca: si = st.text_input("SKU Code", key="mski", placeholder="e.g. CPF11BV00000015")
                with cb: qi = st.number_input("Qty", min_value=0, value=1, step=1, key="mqi")
                with cc:
                    st.markdown("<br/>", unsafe_allow_html=True)
                    if st.button("➕ Add SKU", key="add_sku"):
                        st_clean = si.strip().upper()
                        if st_clean in sku_master:
                            updated = copy.deepcopy(st.session_state.quantities)
                            updated[st_clean] = updated.get(st_clean, 0) + qi
                            new_items = copy.deepcopy(st.session_state.line_items)
                            new_items.append({"sku": st_clean, "qty": qi})
                            st.session_state.quantities = updated
                            st.session_state.line_items = new_items
                            st.success(f"Added {st_clean} × {qi}"); st.rerun()
                        elif st_clean:
                            st.error(f"'{st_clean}' not in master.")

            updated = copy.deepcopy(st.session_state.quantities)
            items = {s: q for s, q in updated.items() if q > 0}
            if items:
                st.markdown('<div class="fsl">Edit Quantities</div>', unsafe_allow_html=True)
                il = list(items.items())
                for i in range(0, len(il), 3):
                    chunk = il[i:i+3]; rcols = st.columns(len(chunk))
                    for col, (sku, qty) in zip(rcols, chunk):
                        info = sku_master.get(sku, {})
                        lbl  = f"{info.get('product', sku)[:20]}\n{info.get('od_size','')}"
                        with col:
                            nq = st.number_input(lbl, min_value=0, value=int(qty),
                                                 step=1, key=f"eq_{sku}")
                            updated[sku] = nq
            st.session_state.quantities = updated

        with tab1:
            line_items = st.session_state.line_items
            if line_items:
                h = """<div class="ocr-wrap"><table class="ocr-tbl">
                <thead><tr><th style="text-align:center;width:40px">#</th>
                <th class="L">Product</th><th class="L">SKU</th>
                <th>OD</th><th>Inch</th><th>Qty</th>
                <th>MRP(₹)</th><th>Dist.Landing(₹)</th></tr></thead><tbody>"""
                for idx, item in enumerate(line_items, 1):
                    sku  = item["sku"]; qty = item["qty"]
                    info = sku_master.get(sku, {}); mi = mrp_data.get(sku, {})
                    h += (f'<tr><td style="color:#999;font-size:10px">{idx}</td>'
                          f'<td class="L">{info.get("product", "?")}</td>'
                          f'<td class="M">{sku}</td><td>{info.get("od_size","")}</td>'
                          f'<td>{info.get("inch_size","")}</td>'
                          f'<td><b style="color:#C0211F">{qty}</b></td>'
                          f'<td>₹{mi.get("mrp",0):,.2f}</td>'
                          f'<td>₹{mi.get("distributor_landing",0):,.2f}</td></tr>')
                h += "</tbody></table></div>"
                st.markdown(h, unsafe_allow_html=True)
            else:
                st.markdown('<div class="warn-box">No matched items. Use Edit / Add tab to add manually.</div>',
                            unsafe_allow_html=True)

        with tab2:
            st.caption("Spatially reconstructed table from OCR bounding boxes. "
                       "🟠 = detected SKU column · 🟢 = quantity cells · "
                       "Bold row = detected header.")
            if grid:
                skuc = meta.get("sku_col")
                hrow = meta.get("header_row", 0)

                h = '<div class="raw-wrap"><table class="raw-tbl"><thead><tr>'
                h += "".join(f"<th>C{i}</th>" for i in range(len(grid[0])))
                h += "</tr></thead><tbody>"
                for ri, row in enumerate(grid):
                    h += '<tr class="hdr">' if ri == hrow else "<tr>"
                    for ci, cell in enumerate(row):
                        css = ""
                        if ci == skuc: css = 'class="sku"'
                        h += f"<td {css}>{cell or '·'}</td>"
                    h += "</tr>"
                h += "</tbody></table></div>"
                st.markdown(h, unsafe_allow_html=True)
                st.markdown(
                    f'<div class="info-box">📊 {len(grid)} rows · {len(grid[0]) if grid else 0} cols · '
                    f'Header row: {hrow} · SKU col: {skuc}</div>', unsafe_allow_html=True)
            else:
                st.info("Run OCR first.")

        with tab3:
            st.caption("Every SKU prefix + size → full SKU match attempt.")
            if log:
                h = """<div class="ocr-wrap"><table class="ocr-tbl">
                <thead><tr><th class="L">Product</th><th class="L">Prefix (OCR)</th>
                <th>Size</th><th>Qty</th><th class="L">Full SKU</th><th>Status</th>
                </tr></thead><tbody>"""
                for m in log:
                    cls = "ok" if m["status"] == "matched" else "no"
                    ico = "✅" if m["status"] == "matched" else "⚠️"
                    h += (f'<tr class="{cls}"><td class="L">{m["product"]}</td>'
                          f'<td class="M">{m["prefix"]}</td><td>{m["size"]}</td>'
                          f'<td><b>{m["qty"]}</b></td><td class="M">{m["full_sku"]}</td>'
                          f'<td>{ico}</td></tr>')
                h += "</tbody></table></div>"
                st.markdown(h, unsafe_allow_html=True)
            else:
                st.info("Run OCR first.")

        iord  = {s: q for s, q in st.session_state.quantities.items() if q > 0}
        gmrp  = sum(mrp_data.get(s, {}).get("mrp", 0) * q        for s, q in iord.items())
        gdist = sum(mrp_data.get(s, {}).get("distributor_landing", 0) * q for s, q in iord.items())
        disc  = gmrp - gdist
        st.markdown(f"""
        <div class="totals-box">
          <div class="total-row"><span class="total-lbl">Line items</span><span class="total-val">{len(st.session_state.line_items) if st.session_state.line_items else len(iord)}</span></div>
          <div class="total-row"><span class="total-lbl">Gross MRP</span><span class="total-val">₹ {gmrp:,.2f}</span></div>
          <div class="total-row"><span class="total-lbl">Distributor Discount</span><span class="total-val neg">− ₹ {disc:,.2f}</span></div>
          <div class="total-row grand"><span class="total-lbl">Net Distributor Landing</span><span class="total-val">₹ {gdist:,.2f}</span></div>
        </div>""", unsafe_allow_html=True)

        if st.button("Proceed to fill Customer Details", key="goto3"):
            st.session_state.ocr_reviewed = True
            st.rerun()

    st.markdown("</div></div>", unsafe_allow_html=True)


def render_step3():
    if not step_unlocked[2]:
        return

    done = st.session_state.party_confirmed
    st.markdown(f"""
    <div class="step-card"><div class="step-card-header">
      <div class="step-number {'done' if done else ''}">{'✓' if done else '3'}</div>
      <div><div class="step-title">Step 3 — Dealer / Distributor Details</div>
      <div class="step-subtitle">Bill To &amp; Ship To party information</div></div>
    </div><div class="step-body">""", unsafe_allow_html=True)

    def party_form(pfx, title, emoji):
        st.markdown(f'<div class="party-section"><div class="party-title">{emoji} {title}</div>',
                    unsafe_allow_html=True)

        c1, c2 = st.columns(2)
        with c1: pno = st.text_input("Party No.", key=f"{pfx}_pno",
                                      placeholder="Digits only")
        with c2: gst = st.text_input("GST No.",   key=f"{pfx}_gst")

        na = st.text_area("Name & Address", key=f"{pfx}_na", height=80)

        c3, c4 = st.columns(2)
        with c3:
            st.markdown("**Phone** (+91)")
            ph = st.text_input("+91 Phone", key=f"{pfx}_ph",
                               label_visibility="collapsed",
                               placeholder="10-digit number")
        with c4:
            st.markdown("**Mobile** (+91)")
            mb = st.text_input("+91 Mobile", key=f"{pfx}_mb",
                               label_visibility="collapsed",
                               placeholder="10-digit number")

        c5, c6 = st.columns(2)
        with c5:
            st_ = st.text_input("State", key=f"{pfx}_st")
        with c6:
            pan = st.text_input("PAN No.", key=f"{pfx}_pan")

        st.markdown("</div>", unsafe_allow_html=True)

        return {"party_no": pno, "name": na, "phone": ph, "mobile": mb,
                "state": st_, "gst": gst, "pan": pan}

    bill_to = party_form("bill", "Dealer/Distributor Details", "🏢")
    ship_to = party_form("ship", "Customer Details", "🚚")

    all_errors = validate_party(bill_to, "Bill To") + validate_party(ship_to, "Ship To")

    if all_errors:
        for e in all_errors:
            st.markdown(f'<div class="error-box">⚠️ {e}</div>', unsafe_allow_html=True)

    def do_confirm(b, s):
        st.session_state.bill_to         = b
        st.session_state.ship_to         = s
        st.session_state.party_confirmed = True
        with st.spinner("Generating PDF…"):
            try:
                pdf = build_pdf(st.session_state.quantities, mrp_data,
                                b, s, sku_master,
                                st.session_state.get("line_items", []))
                st.session_state.pdf_bytes = pdf
                st.rerun()
            except Exception as e:
                st.error(f"PDF error: {e}")

    if is_party_complete(bill_to) and is_party_complete(ship_to) and not all_errors:
        st.markdown('<div class="success-box">✅ All fields filled — generating PDF automatically…</div>',
                    unsafe_allow_html=True)
        if not st.session_state.party_confirmed:
            do_confirm(bill_to, ship_to)

    btn_label = "Generate Quotation"
    if not all_errors:
        if st.button(btn_label, key="confirm"):
            do_confirm(bill_to, ship_to)
    else:
        st.markdown(
            "<div style='font-size:12px;color:#92400E;margin-top:4px;'>"
            "Fix the validation errors above, then confirm.</div>", unsafe_allow_html=True)
        if st.button("⚡ Skip Validation & Generate PDF", key="confirm_force"):
            do_confirm(bill_to, ship_to)

    st.markdown("</div></div>", unsafe_allow_html=True)


def render_step4():
    if not step_unlocked[3]:
        return

    st.markdown("""
    <div class="step-card"><div class="step-card-header">
      <div class="step-number done">✓</div>
      <div><div class="step-title">Step 4 — Preview &amp; Download Quotation</div>
      <div class="step-subtitle">Review your Sales Quotation PDF before downloading</div></div>
    </div><div class="step-body">""", unsafe_allow_html=True)

    line_items = st.session_state.get("line_items", [])
    if line_items:
        items_for_total = line_items
        gmrp  = sum(mrp_data.get(it["sku"], {}).get("mrp", 0) * it["qty"] for it in items_for_total)
        gdist = sum(mrp_data.get(it["sku"], {}).get("distributor_landing", 0) * it["qty"] for it in items_for_total)
        n_lines = len(items_for_total)
    else:
        iord  = {s: q for s, q in st.session_state.quantities.items() if q > 0}
        gmrp  = sum(mrp_data.get(s, {}).get("mrp", 0) * q        for s, q in iord.items())
        gdist = sum(mrp_data.get(s, {}).get("distributor_landing", 0) * q for s, q in iord.items())
        n_lines = len(iord)
    disc = gmrp - gdist

    st.markdown(f"""
    <div class="success-box">✅ Quotation ready — <b>{n_lines} line items</b></div>
    <div class="totals-box">
      <div class="total-row"><span class="total-lbl">Total Line Items</span><span class="total-val">{n_lines}</span></div>
      <div class="total-row"><span class="total-lbl">Gross MRP</span><span class="total-val">₹ {gmrp:,.2f}</span></div>
      <div class="total-row"><span class="total-lbl">Distributor Discount</span><span class="total-val neg">− ₹ {disc:,.2f}</span></div>
      <div class="total-row grand"><span class="total-lbl">Net Taxable Value</span><span class="total-val">₹ {gdist:,.2f}</span></div>
    </div>""", unsafe_allow_html=True)

    if st.session_state.pdf_bytes:
        st.markdown("""
        <div style="margin:20px 0 8px;">
          <div style="font-size:13px;font-weight:700;color:#1A1A1A;margin-bottom:4px;">
            📄 Document Preview
          </div>
          <div style="font-size:11.5px;color:#6B6B6B;margin-bottom:10px;">
            Review the quotation below. Use the viewer controls to navigate pages.
          </div>
        </div>""", unsafe_allow_html=True)

        b64 = base64.b64encode(st.session_state.pdf_bytes).decode()
        pdf_display = f"""
        <div class="pdf-preview-wrap">
          <iframe
            src="data:application/pdf;base64,{b64}#toolbar=1&navpanes=1&scrollbar=1&view=FitH"
            width="100%"
            height="620px"
            style="border:none;display:block;"
            type="application/pdf"
          >
            <p style="color:white;padding:20px;text-align:center;">
              Your browser does not support embedded PDF preview.<br/>
              Please use the download button below.
            </p>
          </iframe>
        </div>"""
        st.markdown(pdf_display, unsafe_allow_html=True)
        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

        col_dl, col_wa, col_new = st.columns(3)

        with col_dl:
            st.markdown(
                f'<a class="dl-btn" href="data:application/pdf;base64,{b64}" '
                f'download="sintex_quotation.pdf" '
                f'style="background:linear-gradient(135deg,#C0211F,#8B1514);">'
                f'📥 &nbsp; Download Quotation'
                f'</a>',
                unsafe_allow_html=True,
            )

        with col_wa:
            wa_text = "Please find attached the Sintex BAPL Sales Quotation."
            wa_url  = f"https://wa.me/?text={requests.utils.quote(wa_text)}"
            st.markdown(
                f'<a class="dl-btn" href="{wa_url}" target="_blank" rel="noopener noreferrer" '
                f'style="background:linear-gradient(135deg,#C0211F,#8B1514);">'
                f'📲 &nbsp; Send Quotation'
                f'</a>',
                unsafe_allow_html=True,
            )

        with col_new:
            st.markdown("<div style='margin-top:6px;'>", unsafe_allow_html=True)
            if st.button("🔄  New Quotation", key="newq"):
                    for k in ["image_bytes", "pdf_bytes"]: st.session_state[k] = None
                    for k in ["quantities", "bill_to", "ship_to", "ocr_meta"]: st.session_state[k] = {}
                    for k in ["match_log", "ocr_grid", "ocr_extracted", "line_items"]: st.session_state[k] = []
                    st.session_state["ocr_done"]        = False
                    st.session_state["ocr_reviewed"]    = False
                    st.session_state["party_confirmed"] = False
                    st.session_state["image_submitted"] = False
                    st.session_state["upload_key"]     += 1
                    st.session_state["image_rotation"]  = 0
                    st.session_state["capture_mode"]    = "camera"
                    st.rerun()
            st.markdown("</div>", unsafe_allow_html=True)

    else:
        if st.button("🔄  New Quotation", key="newq_no_pdf"):
            for k in ["image_bytes", "pdf_bytes"]: st.session_state[k] = None
            for k in ["quantities", "bill_to", "ship_to", "ocr_meta"]: st.session_state[k] = {}
            for k in ["match_log", "ocr_grid", "ocr_extracted", "line_items"]: st.session_state[k] = []
            st.session_state["ocr_done"]        = False
            st.session_state["ocr_reviewed"]    = False
            st.session_state["party_confirmed"] = False
            st.session_state["image_submitted"] = False
            st.session_state["upload_key"]     += 1
            st.session_state["image_rotation"]  = 0
            st.session_state["capture_mode"]    = "camera"
            st.rerun()

    st.markdown("</div></div>", unsafe_allow_html=True)

render_step1()
render_step2()
render_step3()
render_step4()

st.markdown("""
<div style="text-align:center;padding:32px 0 8px;font-size:11px;color:#AAA;">
  Sintex BAPL Limited &nbsp;·&nbsp; CPVC / UPVC Quotation System &nbsp;·&nbsp; Chhattisgarh
</div>""", unsafe_allow_html=True)