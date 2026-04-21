import io
import os
import re
import copy
import base64
import textwrap
from datetime import date
from difflib import SequenceMatcher

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

try:
    import requests as _requests
    _HAS_REQUESTS = True
except ImportError:
    _HAS_REQUESTS = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, HRFlowable)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    _HAS_REPORTLAB = True
except ImportError:
    _HAS_REPORTLAB = False

st.set_page_config(
    page_title="Sintex BAPL – Quotation Generator",
    page_icon="🔧",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@300;400;500;600;700&family=IBM+Plex+Mono:wght@400;500&display=swap');
:root {
  --navy:   #0A2342;
  --blue:   #1565C0;
  --sky:    #1E88E5;
  --teal:   #00796B;
  --gold:   #F9A825;
  --danger: #C62828;
  --border: #DEE3EC;
  --surface:#F4F6FA;
  --text:   #1A1F36;
  --muted:  #5A6880;
  --radius: 10px;
}
html, body, [class*="css"] {
  font-family: 'IBM Plex Sans', sans-serif !important;
  background: #F0F2F8 !important;
  color: var(--text);
}
#MainMenu, footer, header { visibility: hidden; }
.block-container { padding: 1.2rem 1.5rem 4rem !important; max-width: 1400px !important; }

.card {
  background: #fff;
  border: 1px solid var(--border);
  border-radius: var(--radius);
  padding: 20px 24px;
  margin-bottom: 18px;
  box-shadow: 0 1px 4px rgba(10,35,66,.07);
}
.card-title {
  font-size: 13px;
  font-weight: 700;
  text-transform: uppercase;
  letter-spacing: .7px;
  color: var(--navy);
  margin-bottom: 14px;
  display: flex;
  align-items: center;
  gap: 8px;
}
.step-badge {
  display: inline-flex;
  align-items: center;
  justify-content: center;
  background: var(--navy);
  color: #fff;
  border-radius: 50%;
  width: 24px; height: 24px;
  font-size: 11px;
  font-weight: 700;
  flex-shrink: 0;
}
.step-badge.done { background: var(--teal); }

.app-header {
  background: linear-gradient(135deg, #0A2342 0%, #1565C0 100%);
  border-radius: 12px;
  padding: 18px 26px;
  display: flex; align-items: center; gap: 18px;
  margin-bottom: 22px;
  box-shadow: 0 4px 16px rgba(10,35,66,.25);
}
.app-header h1 { font-size: 20px; font-weight: 700; color: #fff; margin: 0; }
.app-header p  { font-size: 12px; color: rgba(255,255,255,.65); margin: 3px 0 0; }

.sku-table { width: 100%; border-collapse: collapse; font-size: 11.5px; }
.sku-table th {
  background: var(--navy); color: #fff;
  padding: 6px 8px; text-align: center;
  font-weight: 600; font-size: 10.5px;
  border: 1px solid rgba(255,255,255,.15);
  white-space: nowrap;
}
.sku-table th.row-hdr { text-align: left; min-width: 180px; }
.sku-table tr:nth-child(even) td { background: #F7F9FC; }
.sku-table td {
  padding: 5px 8px; border: 1px solid var(--border);
  text-align: center; vertical-align: middle;
}
.sku-table td.row-label {
  text-align: left; font-weight: 500;
  color: var(--text); white-space: nowrap;
}
.sku-table td.section-hdr {
  background: #E8EDF6; font-weight: 700;
  color: var(--navy); font-size: 11px;
  text-align: left; padding: 4px 8px;
}
.sku-code { font-family: 'IBM Plex Mono', monospace; font-size: 10px; color: var(--blue); }

.stNumberInput > div > div > input {
  border-radius: 6px !important;
  border: 1.5px solid var(--border) !important;
  font-family: 'IBM Plex Mono', monospace !important;
  font-size: 13px !important;
  padding: 4px 6px !important;
  text-align: center !important;
}

.stButton > button {
  font-family: 'IBM Plex Sans', sans-serif !important;
  font-weight: 600 !important; font-size: 14px !important;
  border-radius: 8px !important; border: none !important;
  background: var(--navy) !important; color: #fff !important;
  padding: 10px 20px !important;
  box-shadow: 0 2px 8px rgba(10,35,66,.2) !important;
  transition: all .15s !important;
}
.stButton > button:hover { background: var(--sky) !important; transform: translateY(-1px) !important; }

.stTabs [data-baseweb="tab-list"] {
  background: var(--surface); padding: 4px; border-radius: 10px;
  border: 1px solid var(--border); gap: 4px;
}
.stTabs [data-baseweb="tab"] {
  border-radius: 8px !important; font-weight: 600 !important;
  font-size: 13px !important; color: var(--muted) !important;
}
.stTabs [aria-selected="true"] {
  background: var(--navy) !important; color: #fff !important;
}

.pill {
  display: inline-block; padding: 2px 10px;
  border-radius: 20px; font-size: 11px; font-weight: 600;
}
.pill-blue   { background: #EBF3FF; color: var(--blue); }
.pill-green  { background: #E8F5E9; color: var(--teal); }
.pill-gold   { background: #FFF8E1; color: #7B5B00; }

.totals-row {
  display: flex; justify-content: space-between; align-items: center;
  padding: 9px 0; border-bottom: 1px solid var(--border);
  font-size: 14px;
}
.totals-row:last-child { border: none; padding-top: 12px; }
.totals-row.net { font-weight: 700; font-size: 16px; color: var(--navy); }
.totals-row .val { font-family: 'IBM Plex Mono', monospace; font-weight: 600; }
.totals-row .val.neg { color: var(--danger); }

.party-box {
  border: 1px solid var(--border); border-radius: 8px;
  padding: 14px 16px; background: var(--surface);
}
.party-box h4 { margin: 0 0 10px; font-size: 12px; font-weight: 700;
  text-transform: uppercase; letter-spacing: .5px; color: var(--navy); }
.party-row { display: flex; gap: 6px; margin-bottom: 5px; font-size: 12.5px; }
.party-lbl { color: var(--muted); min-width: 80px; flex-shrink: 0; }
.party-val { font-weight: 500; color: var(--text); word-break: break-word; }

.ocr-table { width:100%; border-collapse:collapse; font-size:12px; margin-top:10px; }
.ocr-table th { background:var(--navy); color:#fff; padding:6px 10px;
  font-weight:600; font-size:11px; text-align:left; border:1px solid rgba(255,255,255,.15); }
.ocr-table td { padding:5px 10px; border:1px solid var(--border); vertical-align:middle; }
.ocr-table tr:nth-child(even) td { background:#F7F9FC; }
.match-ok   { color:var(--teal);  font-weight:700; }
.match-warn { color:#E65100;      font-weight:700; }
.match-none { color:var(--danger);font-weight:700; }

input[data-sintex-cam="1"] { opacity:0!important; height:0!important; pointer-events:none!important; }
label { font-size:12px!important; font-weight:600!important; color:var(--muted)!important; }
</style>
""", unsafe_allow_html=True)

# ─── Paths ──────────────────────────────────────────────────────────────────
_HERE = os.path.dirname(os.path.abspath(__file__))
MRP_CSV_PATH  = os.path.join(_HERE, "MRP_State_chhattisghar.csv")
ZSD_CSV_PATH  = os.path.join(_HERE, "ZSD_CUST.csv")
SKU_XLSX_PATH = os.path.join(_HERE, "Sample form for Product list.xlsx")

# ─── Data loaders ────────────────────────────────────────────────────────────
@st.cache_data
def load_mrp():
    df = pd.read_csv(MRP_CSV_PATH)
    df["MRP_clean"] = (
        df["MRP(ZPR1-933)"]
        .astype(str).str.replace(",", "", regex=False)
        .pipe(pd.to_numeric, errors="coerce")
    )
    df["Distributor Landing"] = pd.to_numeric(
        df["Distributor Landing"].astype(str).str.replace(",", "", regex=False),
        errors="coerce",
    )
    lookup = {}
    for _, row in df.iterrows():
        mat = str(row["Material Number"]).strip()
        lookup[mat] = row.to_dict()
    return lookup

@st.cache_data
def load_zsd():
    df = pd.read_csv(ZSD_CSV_PATH, encoding="latin1")
    return df

@st.cache_data
def load_sku_sheets():
    wb  = load_workbook(SKU_XLSX_PATH, read_only=True)
    out = {}
    for sname in wb.sheetnames:
        ws   = wb[sname]
        data = [row for row in ws.iter_rows(values_only=True)
                if any(c is not None for c in row)]
        if len(data) < 3:
            continue
        col_ids    = [str(c).strip() if c is not None else "" for c in data[1][2:]]
        col_labels = [str(c).strip() if c is not None else "" for c in data[2][2:]]
        while col_ids and col_ids[-1] == "":
            col_ids.pop(); col_labels.pop()

        rows   = []
        section = "General"
        for raw in data[3:]:
            if not any(c is not None for c in raw):
                continue
            label = str(raw[0]).strip() if raw[0] is not None else ""
            if label.upper() in ("FITTINGS", "PIPES", "FITTING SCH 80") and all(
                c is None for c in raw[2:]
            ):
                section = label.title()
                continue
            if not label:
                continue
            skus = []
            for i in range(len(col_ids)):
                idx = i + 2
                val = raw[idx] if idx < len(raw) else None
                skus.append(str(val).strip() if val is not None else None)
            rows.append({"label": label, "section": section, "skus": skus})

        out[sname] = {
            "label":      sname,
            "col_ids":    col_ids,
            "col_labels": col_labels,
            "rows":       rows,
        }
    return out


# ─── Build a flat SKU catalogue from all sheets ───────────────────────────
@st.cache_data
def build_sku_catalogue():
    """Returns {sku_code: {"sheet":…, "row_label":…, "section":…, "col_id":…, "col_label":…, "ri":…, "ci":…}}"""
    sheets = load_sku_sheets()
    cat = {}
    for sname, sheet in sheets.items():
        col_ids    = sheet["col_ids"]
        col_labels = sheet["col_labels"]
        for ri, row in enumerate(sheet["rows"]):
            for ci, sku in enumerate(row["skus"]):
                if not sku or sku in ("-", "None", ""):
                    continue
                cat[sku] = {
                    "sheet":      sname,
                    "row_label":  row["label"],
                    "section":    row["section"],
                    "col_id":     col_ids[ci] if ci < len(col_ids) else "",
                    "col_label":  col_labels[ci] if ci < len(col_labels) else "",
                    "ri":         ri,
                    "ci":         ci,
                }
    return cat


# ─── SKU normalise + match ───────────────────────────────────────────────────

# Characters that OCR commonly confuses in product codes
_OCR_FIXES = str.maketrans({
    'O': '0',   # letter O → digit 0
    'I': '1',   # letter I → digit 1  (in numeric positions)
    'S': '5',   # letter S → digit 5  (in numeric positions)
    'B': '8',   # letter B → digit 8  (in numeric positions)
    'G': '6',   # letter G → digit 6
    'Z': '2',   # letter Z → digit 2
})

def _normalise_sku(raw: str) -> str:
    """
    Normalise a raw token for comparison:
      1. Strip whitespace, hyphens, dots, underscores
      2. Uppercase
      3. Apply OCR character-substitutions on the NUMERIC suffix portion only
         (we preserve the alphabetic prefix so CPF stays CPF)
    """
    s = re.sub(r"[\s\-\._]", "", raw).upper()
    # Find where the numeric portion begins (after the leading alpha prefix)
    # e.g. CPF11BV000  →  prefix='CPF', rest='11BV000'
    m = re.match(r'^([A-Z]+)(.*)$', s)
    if m:
        prefix, rest = m.group(1), m.group(2)
        # Apply OCR fixes only on the rest (mixed alphanum section)
        rest_fixed = rest.translate(_OCR_FIXES)
        return prefix + rest_fixed
    return s.translate(_OCR_FIXES)


@st.cache_data
def _build_normalised_index(catalogue_keys: tuple):
    """
    Pre-compute normalised versions of every catalogue SKU for fast lookup.
    Returns {normalised_sku: original_sku}
    """
    return {_normalise_sku(k): k for k in catalogue_keys}


def match_sku_to_catalogue(raw_sku: str, catalogue: dict,
                            norm_index: dict, threshold: float = 0.75):
    """
    Multi-strategy SKU matching:
      1. Exact normalised match (after OCR fix)
      2. Prefix-anchored match  — OCR often drops trailing chars
      3. Substring match        — raw token is contained in a catalogue SKU
      4. SequenceMatcher fuzzy  — catch remaining near-matches
    Returns (best_sku, confidence_0_to_1, 'exact'|'fuzzy'|'none').
    """
    if not raw_sku:
        return None, 0.0, "none"

    norm_raw = _normalise_sku(raw_sku)
    raw_len  = len(norm_raw)

    # ── 1. Exact after normalisation ─────────────────────────────────────────
    if norm_raw in norm_index:
        return norm_index[norm_raw], 1.0, "exact"

    # ── 2. Prefix anchor: catalogue SKU starts with the raw token ────────────
    #    e.g. raw = "CPF11BV000"  →  catalogue = "CPF11BV00000015"
    prefix_candidates = []
    for norm_cat, orig_cat in norm_index.items():
        if norm_cat.startswith(norm_raw) and raw_len >= 8:
            # Score = raw_len / cat_len  (longer match = higher score)
            score = raw_len / max(len(norm_cat), 1)
            prefix_candidates.append((orig_cat, score))
    if prefix_candidates:
        best = max(prefix_candidates, key=lambda x: x[1])
        # Only accept if prefix covers ≥55 % of the catalogue SKU
        if best[1] >= 0.55:
            return best[0], min(0.97, best[1] + 0.3), "exact"

    # ── 3. Substring: raw token fully contained in catalogue SKU ─────────────
    if raw_len >= 8:
        sub_candidates = []
        for norm_cat, orig_cat in norm_index.items():
            if norm_raw in norm_cat:
                score = raw_len / max(len(norm_cat), 1)
                sub_candidates.append((orig_cat, score))
        if sub_candidates:
            best = max(sub_candidates, key=lambda x: x[1])
            if best[1] >= 0.50:
                return best[0], min(0.90, best[1] + 0.2), "fuzzy"

    # ── 4. SequenceMatcher fuzzy ─────────────────────────────────────────────
    #    Only worthwhile for tokens that look like product codes (≥7 chars,
    #    mixed alpha+digits).  Skip pure words.
    has_digit = any(c.isdigit() for c in norm_raw)
    has_alpha = any(c.isalpha() for c in norm_raw)
    if raw_len < 7 or not (has_digit and has_alpha):
        return None, 0.0, "none"

    best_sku, best_score = None, 0.0
    for norm_cat, orig_cat in norm_index.items():
        score = SequenceMatcher(None, norm_raw, norm_cat).ratio()
        if score > best_score:
            best_score = score
            best_sku   = orig_cat

    if best_score >= threshold:
        return best_sku, best_score, "fuzzy"
    return None, best_score, "none"


# ─── OCR helpers ─────────────────────────────────────────────────────────────

_PRODUCT_CODE_RE = re.compile(
    r'\b([A-Z]{2,}[0-9]{2,}[A-Z0-9]{2,})\b',
    re.IGNORECASE,
)

def _looks_like_product_code(token: str) -> bool:
    t = token.strip()
    if len(t) < 7:
        return False
    return (any(c.isdigit() for c in t) and
            any(c.isalpha() for c in t) and
            not t.isalpha())


def _word_x_center(polygon) -> float:
    """Return the horizontal centre of a word given its polygon (flat list or list of {x,y})."""
    if not polygon:
        return 0.0
    if isinstance(polygon[0], dict):
        xs = [p["x"] for p in polygon]
    else:
        # flat list [x0,y0,x1,y1,x2,y2,x3,y3]
        xs = [polygon[i] for i in range(0, len(polygon), 2)]
    return (min(xs) + max(xs)) / 2.0


def _word_y_center(polygon) -> float:
    if not polygon:
        return 0.0
    if isinstance(polygon[0], dict):
        ys = [p["y"] for p in polygon]
    else:
        ys = [polygon[i] for i in range(1, len(polygon), 2)]
    return (min(ys) + max(ys)) / 2.0


def _cluster_columns(x_values: list, gap_fraction: float = 0.04, page_width: float = 1.0) -> list:
    """
    Given a sorted list of x-centres, cluster them into column bands.
    Returns list of (x_min, x_max) band tuples.
    """
    if not x_values:
        return []
    gap = gap_fraction * page_width
    bands = []
    band_start = band_end = x_values[0]
    for x in x_values[1:]:
        if x - band_end > gap:
            bands.append((band_start, band_end))
            band_start = band_end = x
        else:
            band_end = x
    bands.append((band_start, band_end))
    return bands


# Known Sintex size labels in the order they appear on the form (left→right)
_FORM_SIZES = ["15MM", "20MM", "25MM", "32MM", "40MM", "50MM"]


def _extract_table_geometric(words_with_pos: list) -> list:
    """
    Core table parser.  words_with_pos is a list of dicts:
        {"text": str, "x": float, "y": float}
    where x,y are normalised 0-1 coordinates.

    Algorithm:
    1. Identify rows by clustering words with similar y-coordinates.
    2. Within each row, find the SKU-code token (mixed alpha+digit, ≥7 chars).
    3. Find the size-column header row (contains "15MM", "20MM", …).
    4. For each data row that has a SKU, assign quantity tokens to size columns
       based on their x-position relative to the header x-positions.
    5. Emit one {"raw_sku", "raw_size", "qty", "raw_line"} dict per
       (SKU, non-zero quantity, size).
    """
    if not words_with_pos:
        return []

    # ── Step 1: cluster words into rows by y-coordinate ─────────────────────
    # Sort by y
    words_sorted = sorted(words_with_pos, key=lambda w: w["y"])
    row_gap  = 0.018   # ~1.8 % of page height between rows
    rows     = []      # list of list-of-word-dicts
    cur_row  = [words_sorted[0]]
    for w in words_sorted[1:]:
        if w["y"] - cur_row[-1]["y"] > row_gap:
            rows.append(cur_row)
            cur_row = [w]
        else:
            cur_row.append(w)
    rows.append(cur_row)

    # ── Step 2: find the size-header row ────────────────────────────────────
    header_row  = None
    col_x_map   = {}   # size_label → x_centre
    for row in rows:
        texts = [w["text"].upper().strip() for w in row]
        hits  = [s for s in _FORM_SIZES if s in texts]
        if len(hits) >= 2:   # at least two size labels on the same row
            header_row = row
            for w in row:
                t = w["text"].upper().strip()
                if t in _FORM_SIZES:
                    col_x_map[t] = w["x"]
            break

    # Fallback: no geometric header found — use plain-text line extraction
    if not col_x_map:
        return _extract_table_fallback(words_with_pos)

    # Sort size labels by their x position
    size_labels_sorted = sorted(col_x_map.keys(), key=lambda s: col_x_map[s])

    # Build column bands: midpoints between adjacent column centres
    col_x_sorted = [col_x_map[s] for s in size_labels_sorted]
    def _assign_size(x_val):
        """Return the size label whose column centre is closest to x_val."""
        best_size, best_dist = None, float("inf")
        for size in size_labels_sorted:
            d = abs(col_x_map[size] - x_val)
            if d < best_dist:
                best_dist = d
                best_size = size
        # Only accept if within half the typical column width
        col_width = (max(col_x_sorted) - min(col_x_sorted)) / max(len(col_x_sorted) - 1, 1)
        if best_dist > col_width * 0.75:
            return None
        return best_size

    # ── Step 3: find the SKU-code column x-range ────────────────────────────
    # SKU codes appear to the LEFT of the first size column
    sku_x_max = min(col_x_sorted) - 0.01

    # ── Step 4: parse each data row ─────────────────────────────────────────
    results   = []
    seen_keys = set()   # (raw_sku_upper, size) already emitted

    for row in rows:
        if row is header_row:
            continue

        row_words = sorted(row, key=lambda w: w["x"])

        # Find SKU token(s) in the left zone
        sku_tokens = [w for w in row_words
                      if w["x"] <= sku_x_max
                      and _looks_like_product_code(w["text"])]
        if not sku_tokens:
            continue

        raw_sku   = sku_tokens[0]["text"]  # take the leftmost product-code token
        raw_line  = " ".join(w["text"] for w in row_words)

        # Find quantity tokens to the RIGHT of the SKU zone
        qty_tokens = [w for w in row_words
                      if w["x"] > sku_x_max
                      and re.match(r'^\d+$', w["text"].strip())]

        if not qty_tokens:
            # No quantities on this row — still emit with qty=0 so user can fix
            key = (raw_sku.upper(), "")
            if key not in seen_keys:
                seen_keys.add(key)
                results.append({
                    "raw_sku":  raw_sku,
                    "raw_size": "",
                    "qty":      0,
                    "raw_line": raw_line,
                })
            continue

        # Assign each qty token to a size column
        for qt in qty_tokens:
            size = _assign_size(qt["x"])
            if size is None:
                continue
            try:
                qty_val = int(qt["text"].strip())
            except ValueError:
                continue
            key = (raw_sku.upper(), size)
            if key in seen_keys:
                continue
            seen_keys.add(key)
            results.append({
                "raw_sku":  raw_sku,
                "raw_size": size,
                "qty":      qty_val,
                "raw_line": raw_line,
            })

    return results


def _extract_table_fallback(words_with_pos: list) -> list:
    """
    Fallback when no size-header row is detected.
    Groups words into text lines and looks for lines that contain a
    product-code token plus one or more integers.
    """
    if not words_with_pos:
        return []

    words_sorted = sorted(words_with_pos, key=lambda w: w["y"])
    row_gap  = 0.018
    rows     = []
    cur_row  = [words_sorted[0]]
    for w in words_sorted[1:]:
        if w["y"] - cur_row[-1]["y"] > row_gap:
            rows.append(cur_row)
            cur_row = [w]
        else:
            cur_row.append(w)
    rows.append(cur_row)

    results   = []
    seen_skus = set()

    for row in rows:
        texts    = [w["text"] for w in sorted(row, key=lambda x: x["x"])]
        raw_line = " ".join(texts)
        sku_toks = [t for t in texts if _looks_like_product_code(t)]
        if not sku_toks:
            continue
        raw_sku  = sku_toks[0]
        if raw_sku.upper() in seen_skus:
            continue
        seen_skus.add(raw_sku.upper())
        nums = [t for t in texts if re.match(r'^\d+$', t)]
        qty  = int(nums[-1]) if nums else 0
        results.append({"raw_sku": raw_sku, "raw_size": "", "qty": qty, "raw_line": raw_line})

    return results


# ── API callers — return raw words with positions ─────────────────────────────

def _ocr_document_intelligence(image_bytes: bytes, endpoint: str, key: str) -> list:
    """
    Call Azure Document Intelligence (prebuilt-read) and return
    list of {"raw_sku", "raw_size", "qty", "raw_line"}.
    """
    import time
    url     = endpoint.rstrip("/") + "/formrecognizer/documentModels/prebuilt-read:analyze?api-version=2023-07-31"
    headers = {"Ocp-Apim-Subscription-Key": key, "Content-Type": "application/octet-stream"}
    resp    = _requests.post(url, headers=headers, data=image_bytes, timeout=30)
    resp.raise_for_status()
    op_url  = resp.headers["Operation-Location"]
    result  = {}
    for _ in range(20):
        time.sleep(1.5)
        r      = _requests.get(op_url, headers={"Ocp-Apim-Subscription-Key": key}, timeout=15)
        result = r.json()
        if result.get("status") == "succeeded":
            break

    analyze = result.get("analyzeResult", {})
    pages   = analyze.get("pages", [])

    words_with_pos = []
    for page in pages:
        pw = page.get("width",  1.0) or 1.0
        ph = page.get("height", 1.0) or 1.0
        for word in page.get("words", []):
            poly = word.get("polygon", [])
            text = word.get("content", "").strip()
            if not text:
                continue
            x = _word_x_center(poly) / pw
            y = _word_y_center(poly) / ph
            words_with_pos.append({"text": text, "x": x, "y": y})

    return _extract_table_geometric(words_with_pos)


def _ocr_computer_vision(image_bytes: bytes, endpoint: str, key: str) -> list:
    """
    Fallback: Azure Computer Vision Read API (v3.2).
    Words here also have bounding boxes.
    """
    import time
    url     = endpoint.rstrip("/") + "/vision/v3.2/read/analyze"
    headers = {"Ocp-Apim-Subscription-Key": key, "Content-Type": "application/octet-stream"}
    resp    = _requests.post(url, headers=headers, data=image_bytes, timeout=30)
    resp.raise_for_status()
    op_url  = resp.headers["Operation-Location"]
    result  = {}
    for _ in range(20):
        time.sleep(1.5)
        r      = _requests.get(op_url, headers={"Ocp-Apim-Subscription-Key": key}, timeout=15)
        result = r.json()
        if result.get("status") == "succeeded":
            break

    analyze   = result.get("analyzeResult", {})
    read_res  = analyze.get("readResults", [])
    words_with_pos = []
    for page in read_res:
        pw = page.get("width",  1.0) or 1.0
        ph = page.get("height", 1.0) or 1.0
        for line in page.get("lines", []):
            for word in line.get("words", []):
                text = word.get("text", "").strip()
                if not text:
                    continue
                bbox = word.get("boundingBox", [])
                # boundingBox is [x0,y0,x1,y1,x2,y2,x3,y3]
                xs = [bbox[i] for i in range(0, len(bbox), 2)] if bbox else [0]
                ys = [bbox[i] for i in range(1, len(bbox), 2)] if bbox else [0]
                x  = ((min(xs) + max(xs)) / 2.0) / pw
                y  = ((min(ys) + max(ys)) / 2.0) / ph
                words_with_pos.append({"text": text, "x": x, "y": y})

    return _extract_table_geometric(words_with_pos)


def run_azure_ocr(image_bytes: bytes, azure_endpoint: str, azure_key: str) -> list:
    """Returns list of {"raw_sku", "raw_size", "qty", "raw_line"}"""
    if not _HAS_REQUESTS:
        return []
    try:
        return _ocr_document_intelligence(image_bytes, azure_endpoint, azure_key)
    except Exception as di_err:
        di_status = getattr(getattr(di_err, "response", None), "status_code", None)
        if di_status in (401, 403, 404):
            return _ocr_computer_vision(image_bytes, azure_endpoint, azure_key)
        raise


def resolve_ocr_rows(raw_rows: list, catalogue: dict, mrp_lookup: dict) -> list:
    """
    For each raw OCR row {"raw_sku", "raw_size", "qty", "raw_line"},
    match the SKU base code to the catalogue, then find the specific
    catalogue entry for the requested size, and pull pricing.

    One input row → one output resolved dict.
    Rows with qty == 0 are kept so the user can still review/correct them.
    """
    norm_index = _build_normalised_index(tuple(catalogue.keys()))
    resolved   = []
    seen_keys  = set()   # (matched_sku, raw_size) to de-dup

    for r in raw_rows:
        raw_sku  = r["raw_sku"]
        raw_size = r.get("raw_size", "")
        qty      = r["qty"]

        # ── Match base SKU code → find all catalogue entries for that base ──
        # The catalogue key includes the size suffix (e.g. CPF11BEL9001515 for ½")
        # We match the OCR token (e.g. CPF11BEL90) which is the prefix/base code.
        # Then if raw_size is known, pick the specific size variant; otherwise
        # take the best-scored match overall.

        # Step A: get the best overall match
        matched_sku, confidence, match_type = match_sku_to_catalogue(
            raw_sku, catalogue, norm_index
        )

        # Step B: if we know the size, try to find a better/specific size match
        if raw_size and matched_sku:
            # normalise the size for comparison: "15MM" → "15", "½"" → "15" etc.
            size_norm = re.sub(r'[^0-9]', '', raw_size)   # keep digits only
            # Among all catalogue SKUs whose normalised code starts with the
            # same base as our match, find the one whose col_label matches size.
            base_norm = _normalise_sku(raw_sku)
            candidates_for_size = [
                (sku, entry) for sku, entry in catalogue.items()
                if _normalise_sku(sku).startswith(base_norm)
                   and re.sub(r'[^0-9]', '', entry.get("col_label", "")) == size_norm
            ]
            if candidates_for_size:
                matched_sku  = candidates_for_size[0][0]
                match_type   = "exact"
                confidence   = 1.0

        dedup_key = (matched_sku or raw_sku, raw_size)
        if dedup_key in seen_keys:
            continue
        seen_keys.add(dedup_key)

        info = {}
        mrp  = land = 0.0
        if matched_sku:
            info = mrp_lookup.get(matched_sku, {})
            mrp  = info.get("MRP_clean") or 0.0
            land = info.get("Distributor Landing") or mrp

        cat_entry = catalogue.get(matched_sku, {}) if matched_sku else {}
        resolved.append({
            "raw_sku":     raw_sku,
            "raw_size":    raw_size,
            "matched_sku": matched_sku or "—",
            "match_type":  match_type,
            "confidence":  round(confidence * 100, 1),
            "qty":         qty,
            "sheet":       cat_entry.get("sheet", "—"),
            "row_label":   cat_entry.get("row_label", "—"),
            "col_label":   cat_entry.get("col_label", raw_size),
            "mrp":         mrp,
            "landing":     land,
            "amount":      round(land * qty, 2),
            "ri":          cat_entry.get("ri"),
            "ci":          cat_entry.get("ci"),
        })
    return resolved


# ─── PDF generator ────────────────────────────────────────────────────────────
def generate_pdf_from_resolved(resolved_rows, mrp_lookup, bill_to, ship_to, dealer_info):
    buf    = io.BytesIO()
    doc    = SimpleDocTemplate(buf, pagesize=A4,
                               leftMargin=10*mm, rightMargin=10*mm,
                               topMargin=12*mm, bottomMargin=12*mm)
    styles = getSampleStyleSheet()
    navy   = colors.HexColor("#0A2342")
    sky    = colors.HexColor("#1565C0")

    title_style = ParagraphStyle("title", fontName="Helvetica-Bold",
                                 fontSize=14, textColor=colors.white, alignment=TA_CENTER)
    sub_style   = ParagraphStyle("sub", fontName="Helvetica", fontSize=8,
                                 textColor=colors.white, alignment=TA_CENTER)

    story = []

    hdr_data = [[
        Paragraph("SINTEX BAPL LIMITED", title_style),
        Paragraph("Kutesar Road, Raipur, Chhattisgarh – 492101<br/>GSTIN: 22AADCB1921F1ZE", sub_style),
    ]]
    hdr_table = Table(hdr_data, colWidths=["40%", "60%"])
    hdr_table.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,-1), navy),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0), (-1,-1), 8),
        ("BOTTOMPADDING", (0,0), (-1,-1), 8),
        ("LEFTPADDING",   (0,0), (-1,-1), 10),
    ]))
    story.append(hdr_table)
    story.append(Spacer(1, 4*mm))

    title_data = [[
        Paragraph("<b>QUOTATION</b>", ParagraphStyle("qt", fontName="Helvetica-Bold",
                  fontSize=13, textColor=navy)),
        Paragraph(f"<b>Date:</b> {date.today().strftime('%d-%m-%Y')}",
                  ParagraphStyle("qd", fontName="Helvetica", fontSize=8, alignment=TA_RIGHT)),
    ]]
    t = Table(title_data, colWidths=["60%", "40%"])
    t.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"MIDDLE"),
                           ("LINEBELOW",(0,0),(-1,-1),0.5,navy)]))
    story.append(t)
    story.append(Spacer(1, 3*mm))

    # Dealer / Distributor block
    if any(dealer_info.values()):
        def party_lines(d, title):
            lines = [f"<b>{title}</b>"]
            for k, v in d.items():
                if v:
                    lines.append(f"<b>{k}:</b> {v}")
            return "<br/>".join(lines)

        dealer_data = [[
            Paragraph(party_lines(dealer_info, "DEALER / DISTRIBUTOR"),
                      ParagraphStyle("pt", fontName="Helvetica", fontSize=7.5, leading=11)),
        ]]
        dt = Table(dealer_data, colWidths=["100%"])
        dt.setStyle(TableStyle([
            ("BOX",          (0,0),(-1,-1), 0.5, navy),
            ("VALIGN",       (0,0),(-1,-1), "TOP"),
            ("TOPPADDING",   (0,0),(-1,-1), 5),
            ("BOTTOMPADDING",(0,0),(-1,-1), 5),
            ("LEFTPADDING",  (0,0),(-1,-1), 8),
            ("BACKGROUND",   (0,0),(-1,-1), colors.HexColor("#EBF3FF")),
        ]))
        story.append(dt)
        story.append(Spacer(1, 3*mm))

    def party_lines2(d, title):
        lines = [f"<b>{title}</b>"]
        for k, v in d.items():
            if v:
                lines.append(f"<b>{k}:</b> {v}")
        return "<br/>".join(lines)

    pt_data = [[
        Paragraph(party_lines2(bill_to, "BILL TO PARTY"),
                  ParagraphStyle("pt", fontName="Helvetica", fontSize=7.5, leading=11)),
        Paragraph(party_lines2(ship_to, "SHIP TO PARTY"),
                  ParagraphStyle("pt", fontName="Helvetica", fontSize=7.5, leading=11)),
    ]]
    pt = Table(pt_data, colWidths=["50%", "50%"])
    pt.setStyle(TableStyle([
        ("BOX",          (0,0),(-1,-1), 0.5, navy),
        ("INNERGRID",    (0,0),(-1,-1), 0.5, colors.HexColor("#C5D0E0")),
        ("VALIGN",       (0,0),(-1,-1), "TOP"),
        ("TOPPADDING",   (0,0),(-1,-1), 6),
        ("BOTTOMPADDING",(0,0),(-1,-1), 6),
        ("LEFTPADDING",  (0,0),(-1,-1), 8),
        ("BACKGROUND",   (0,0),(-1,-1), colors.HexColor("#F4F6FA")),
    ]))
    story.append(pt)
    story.append(Spacer(1, 5*mm))

    tbl_header = ["#", "Item / Description", "SKU Code", "Size", "MRP (₹)", "Qty", "Rate (₹)", "Amount (₹)"]
    tbl_rows   = [tbl_header]
    grand_total = 0.0
    for sno, r in enumerate(resolved_rows, 1):
        if r["qty"] <= 0 or r["matched_sku"] == "—":
            continue
        tbl_rows.append([
            str(sno),
            Paragraph(r["row_label"], ParagraphStyle("rl", fontName="Helvetica", fontSize=7)),
            Paragraph(f'<font name="Courier" size="6.5">{r["matched_sku"]}</font>',
                      ParagraphStyle("sk", fontName="Helvetica", fontSize=7)),
            r["col_label"],
            f"{r['mrp']:,.2f}",
            str(r["qty"]),
            f"{r['landing']:,.2f}",
            f"{r['amount']:,.2f}",
        ])
        grand_total += r["amount"]

    tbl_rows.append(["", "", "", "", "", "",
                     Paragraph("<b>GRAND TOTAL</b>",
                               ParagraphStyle("gt", fontName="Helvetica-Bold", fontSize=8)),
                     Paragraph(f"<b>₹ {grand_total:,.2f}</b>",
                               ParagraphStyle("gv", fontName="Helvetica-Bold", fontSize=8,
                                              alignment=TA_RIGHT))])

    col_widths = [8*mm, 55*mm, 38*mm, 16*mm, 18*mm, 12*mm, 18*mm, 22*mm]
    lt = Table(tbl_rows, colWidths=col_widths, repeatRows=1)
    n  = len(tbl_rows)
    lt.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,0),  navy),
        ("TEXTCOLOR",     (0,0), (-1,0),  colors.white),
        ("FONTNAME",      (0,0), (-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0,0), (-1,0),  7.5),
        ("ALIGN",         (0,0), (-1,-1), "CENTER"),
        ("ALIGN",         (1,1), (1,-1),  "LEFT"),
        ("ALIGN",         (2,1), (2,-1),  "LEFT"),
        ("FONTSIZE",      (0,1), (-1,-1), 7),
        ("ROWBACKGROUNDS",(0,1), (-1,-2), [colors.white, colors.HexColor("#F4F8FF")]),
        ("GRID",          (0,0), (-1,-1), 0.35, colors.HexColor("#C5D0E0")),
        ("TOPPADDING",    (0,0), (-1,-1), 3),
        ("BOTTOMPADDING", (0,0), (-1,-1), 3),
        ("BACKGROUND",    (0,n-1),(-1,n-1), colors.HexColor("#EBF3FF")),
        ("LINEABOVE",     (0,n-1),(-1,n-1), 1, navy),
    ]))
    story.append(lt)
    story.append(Spacer(1, 6*mm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=navy))
    story.append(Spacer(1, 2*mm))
    story.append(Paragraph(
        "This is a computer-generated quotation. Prices are subject to change without notice. "
        "Taxes applicable as per prevailing rates.",
        ParagraphStyle("footer", fontName="Helvetica", fontSize=7, textColor=colors.grey,
                       alignment=TA_CENTER)
    ))
    doc.build(story)
    return buf.getvalue()


# ─── Session-state defaults ───────────────────────────────────────────────────
DEFAULTS = {
    "step":              1,
    "image_bytes":       None,
    "ocr_raw_rows":      [],     # list of {"raw_sku","qty","raw_line"}
    "ocr_resolved":      [],     # list of resolved dicts
    "ocr_done":          False,
    "qty_map":           {},     # {(ri,ci): qty} built from resolved
    "bill_to":           {},
    "ship_to":           {},
    "dealer_info":       {},
    "pdf_bytes":         None,
    "azure_endpoint":    "",
    "azure_key":         "",
    "_cam_recv_val":     "",
}
for k, v in DEFAULTS.items():
    if k not in st.session_state:
        st.session_state[k] = v

mrp_lookup  = load_mrp()
zsd_df      = load_zsd()
sku_sheets  = load_sku_sheets()
sku_cat     = build_sku_catalogue()


# ─── UI helpers ──────────────────────────────────────────────────────────────
def badge(n, done=False):
    cls = "step-badge done" if done else "step-badge"
    return f'<span class="{cls}">{n}</span>'

def party_html(d, title):
    rows = "".join(
        f'<div class="party-row"><span class="party-lbl">{k}</span>'
        f'<span class="party-val">{v or "—"}</span></div>'
        for k, v in d.items()
    )
    return f'<div class="party-box"><h4>{title}</h4>{rows}</div>'


# ─── App header ──────────────────────────────────────────────────────────────
st.markdown("""
<div class="app-header">
  <div>
    <h1>🔧 Sintex BAPL – Quotation Generator</h1>
    <p>CPVC / UPVC Pipes &amp; Fittings · Chhattisgarh Price List</p>
  </div>
</div>
""", unsafe_allow_html=True)

step = st.session_state.step
cols_steps = st.columns(4)
STEP_LABELS = ["1  Party Details", "2  Capture & OCR", "3  Review & Edit", "4  Download"]
for i, (col, lbl) in enumerate(zip(cols_steps, STEP_LABELS), 1):
    done   = step > i
    active = step == i
    bg = "#0A2342" if active else ("#00796B" if done else "#DEE3EC")
    tc = "#fff" if (active or done) else "#5A6880"
    col.markdown(f"""
    <div style="background:{bg};color:{tc};border-radius:8px;
                padding:8px 12px;text-align:center;font-size:12px;font-weight:600;">
      {'✓ ' if done else ''}{lbl}
    </div>""", unsafe_allow_html=True)

st.markdown("<div style='height:18px'></div>", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 1 — Dealer / Distributor + Party Details
# ═══════════════════════════════════════════════════════════════════════════════
if step == 1:
    st.markdown(f"""
    <div class="card">
      <div class="card-title">{badge(1)} Dealer / Distributor & Customer Details</div>
    </div>""", unsafe_allow_html=True)

    # ── Dealer / Distributor section ─────────────────────────────────────────
    st.markdown("### 🏪 Dealer / Distributor Information")
    st.caption("Enter the details of the Dealer or Distributor raising this quotation.")

    da1, da2 = st.columns(2)
    dealer_name   = da1.text_input("Dealer / Distributor Name",  key="dealer_name")
    dealer_code   = da2.text_input("Dealer Code",                key="dealer_code")
    dealer_addr   = st.text_input("Address",                     key="dealer_address")
    da3, da4 = st.columns(2)
    dealer_phone  = da3.text_input("Phone",                      key="dealer_phone")
    dealer_gst    = da4.text_input("GST No.",                    key="dealer_gst")
    da5, da6 = st.columns(2)
    dealer_state  = da5.text_input("State",                      key="dealer_state")
    dealer_pan    = da6.text_input("PAN No.",                    key="dealer_pan")

    st.session_state.dealer_info = {
        "Name":    dealer_name,
        "Code":    dealer_code,
        "Address": dealer_addr,
        "Phone":   dealer_phone,
        "GST No.": dealer_gst,
        "State":   dealer_state,
        "PAN No.": dealer_pan,
    }

    st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)
    st.markdown("---")

    # ── Customer lookup ──────────────────────────────────────────────────────
    st.markdown("### 👤 Customer Information")
    cust_options = ["— Select or type below —"] + [
        f'{row["Customer Code"]} | {row["Customer Name"]}'
        for _, row in zsd_df.iterrows()
    ]
    sel = st.selectbox("🔍 Search Customer from Master", cust_options, key="zsd_search")

    def zsd_fill(prefix, row):
        addr = " ".join(filter(None, [
            str(row.get("Address 1","") or ""),
            str(row.get("Address 2","") or ""),
            str(row.get("Address 3","") or ""),
            str(row.get("City","")      or ""),
        ]))
        return {
            f"{prefix}_party_no":   str(row.get("Customer Code","") or ""),
            f"{prefix}_party_name": str(row.get("Customer Name","") or ""),
            f"{prefix}_address":    addr.strip(),
            f"{prefix}_phone":      str(row.get("Telephone","") or ""),
            f"{prefix}_mobile":     str(row.get("Mobile No.","") or ""),
            f"{prefix}_state_code": str(row.get("State Code","") or ""),
            f"{prefix}_state":      str(row.get("State Code Desc.","") or ""),
            f"{prefix}_gst":        str(row.get("GST Number","") or ""),
            f"{prefix}_pan":        str(row.get("PAN No.","") or ""),
        }

    if sel != cust_options[0]:
        code    = sel.split("|")[0].strip()
        matched = zsd_df[zsd_df["Customer Code"].astype(str) == code]
        if not matched.empty:
            row = matched.iloc[0]
            for k, v in zsd_fill("bill", row).items():
                if k not in st.session_state or not st.session_state[k]:
                    st.session_state[k] = v
            st.toast("Bill-to details filled from customer master.", icon="✅")

    st.markdown("#### 📋 Bill To Party")
    bc1, bc2 = st.columns(2)
    bill_party_no   = bc1.text_input("Bill to Party No.",   key="bill_party_no")
    bill_party_name = bc2.text_input("Bill to Party Name",  key="bill_party_name")
    bill_address    = st.text_input("Bill to Address",       key="bill_address")
    bc3, bc4 = st.columns(2)
    bill_phone      = bc3.text_input("Phone",                key="bill_phone")
    bill_mobile     = bc4.text_input("Mobile",               key="bill_mobile")
    bc5, bc6 = st.columns(2)
    bill_sc         = bc5.text_input("State Code",           key="bill_state_code")
    bill_state      = bc6.text_input("State",                key="bill_state")
    bc7, bc8 = st.columns(2)
    bill_gst        = bc7.text_input("GST No.",              key="bill_gst")
    bill_pan        = bc8.text_input("PAN No.",              key="bill_pan")

    st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)
    same_as_bill = st.checkbox("Ship-to same as Bill-to", value=False, key="same_as_bill_chk")

    st.markdown("#### 🚚 Ship To Party")
    if same_as_bill:
        st.session_state["ship_party_no"]   = bill_party_no
        st.session_state["ship_party_name"] = bill_party_name
        st.session_state["ship_address"]    = bill_address
        st.session_state["ship_phone"]      = bill_phone
        st.session_state["ship_mobile"]     = bill_mobile
        st.session_state["ship_state_code"] = bill_sc
        st.session_state["ship_state"]      = bill_state
        st.session_state["ship_gst"]        = bill_gst
        st.session_state["ship_pan"]        = bill_pan

    sc1, sc2 = st.columns(2)
    ship_party_no   = sc1.text_input("Ship to Party No.",   key="ship_party_no")
    ship_party_name = sc2.text_input("Ship to Party Name",  key="ship_party_name")
    ship_address    = st.text_input("Ship to Address",       key="ship_address")
    sc3, sc4 = st.columns(2)
    ship_phone      = sc3.text_input("Phone ",               key="ship_phone")
    ship_mobile     = sc4.text_input("Mobile ",              key="ship_mobile")
    sc5, sc6 = st.columns(2)
    ship_sc         = sc5.text_input("State Code ",          key="ship_state_code")
    ship_state      = sc6.text_input("State ",               key="ship_state")
    sc7, sc8 = st.columns(2)
    ship_gst        = sc7.text_input("GST No. ",             key="ship_gst")
    ship_pan        = sc8.text_input("PAN No. ",             key="ship_pan")

    st.session_state.bill_to = {
        "Party No.":  bill_party_no,
        "Name":       bill_party_name,
        "Address":    bill_address,
        "Phone":      bill_phone,
        "Mobile":     bill_mobile,
        "State Code": bill_sc,
        "State":      bill_state,
        "GST No.":    bill_gst,
        "PAN No.":    bill_pan,
    }
    st.session_state.ship_to = {
        "Party No.":  ship_party_no,
        "Name":       ship_party_name,
        "Address":    ship_address,
        "Phone":      ship_phone,
        "Mobile":     ship_mobile,
        "State Code": ship_sc,
        "State":      ship_state,
        "GST No.":    ship_gst,
        "PAN No.":    ship_pan,
    }

    with st.expander("👁 Preview All Party Details"):
        pcols = st.columns(2)
        with pcols[0]:
            st.markdown(party_html(st.session_state.bill_to, "BILL TO PARTY"),
                        unsafe_allow_html=True)
        with pcols[1]:
            st.markdown(party_html(st.session_state.ship_to, "SHIP TO PARTY"),
                        unsafe_allow_html=True)

    st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)
    if st.button("▶  Next: Capture & OCR", key="go_step2"):
        st.session_state.step = 2
        st.rerun()


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 2 — Capture Image & Run OCR
# ═══════════════════════════════════════════════════════════════════════════════
elif step == 2:
    st.markdown(f"""
    <div class="card">
      <div class="card-title">{badge(2)} Capture Handwritten Order Form</div>
    </div>""", unsafe_allow_html=True)

    st.markdown(
        "📸 **Photograph or upload** the handwritten order form. "
        "Azure OCR will detect SKU codes and quantities, match them to the product catalogue, "
        "and pre-fill the review table. You can also skip and fill manually.",
    )

    with st.expander("🔑 Azure OCR Credentials (required to run OCR)", expanded=True):
        ep  = st.text_input("Azure Endpoint", value=st.session_state.azure_endpoint,
                            placeholder="https://YOUR_RESOURCE.cognitiveservices.azure.com",
                            key="az_ep_input")
        key = st.text_input("Azure Key", value=st.session_state.azure_key,
                            type="password", placeholder="32-character subscription key",
                            key="az_key_input")
        st.session_state.azure_endpoint = ep
        st.session_state.azure_key      = key

    img_mode = st.radio("Image source",
                        ["📷  Camera (recommended)", "📁  Upload File"],
                        horizontal=True, label_visibility="collapsed",
                        key="img_mode_radio")

    def _on_file_upload():
        uf = st.session_state.get("file_upload_input")
        if uf is not None:
            st.session_state.image_bytes  = uf.getvalue()
            st.session_state.ocr_done     = False
            st.session_state.ocr_raw_rows = []
            st.session_state.ocr_resolved = []

    if img_mode == "📷  Camera (recommended)":
        import streamlit.components.v1 as components

        CAMERA_HTML = """
<style>
* { box-sizing:border-box; margin:0; padding:0; }
body { background:transparent; font-family:'IBM Plex Sans',sans-serif; }
#cam-wrap {
  width:100%; background:linear-gradient(160deg,#071829 0%,#0d2d56 100%);
  border-radius:14px; overflow:hidden; display:flex; flex-direction:column;
  align-items:center; box-shadow:0 8px 32px rgba(10,35,66,.4);
}
#video  { width:100%; max-height:62vh; object-fit:cover; display:block; }
#canvas { display:none; }
#preview{ width:100%; display:none; border-bottom:3px solid #1E88E5; }
.toolbar{ display:flex; gap:12px; padding:16px 20px 18px; width:100%;
  background:rgba(0,0,0,.3); justify-content:center; flex-wrap:wrap; }
.cam-btn{ display:inline-flex; align-items:center; justify-content:center; gap:8px;
  padding:12px 32px; border:none; border-radius:10px; font-size:14px; font-weight:700;
  cursor:pointer; transition:all .18s; letter-spacing:.3px; min-width:160px;
  font-family:'IBM Plex Sans',sans-serif; }
#btn-capture{ background:linear-gradient(135deg,#1565C0 0%,#1E88E5 100%); color:#fff;
  box-shadow:0 4px 16px rgba(21,101,192,.5); }
#btn-capture:hover{ transform:translateY(-2px); box-shadow:0 8px 24px rgba(21,101,192,.65); }
#btn-retake { background:rgba(255,255,255,.1); color:rgba(255,255,255,.9);
  border:1.5px solid rgba(255,255,255,.25); display:none; }
#btn-retake:hover{ background:rgba(255,255,255,.2); transform:translateY(-1px); }
.status-bar{ display:flex; align-items:center; gap:10px;
  background:rgba(0,0,0,.25); backdrop-filter:blur(8px);
  border-top:1px solid rgba(255,255,255,.07); padding:10px 20px; width:100%;
  font-size:12.5px; font-weight:500; color:rgba(255,255,255,.7); min-height:42px; }
.status-bar.success{ background:rgba(0,121,107,.3); color:#80CBC4; }
.status-bar.sending{ background:rgba(249,168,37,.12); color:#FFD54F; }
.status-bar.error  { background:rgba(198,40,40,.2);  color:#EF9A9A; }
.pulse-dot { width:8px; height:8px; border-radius:50%; background:currentColor;
  flex-shrink:0; animation:pulse 1.5s infinite ease-in-out; }
@keyframes pulse{ 0%,100%{opacity:1;transform:scale(1)} 50%{opacity:.3;transform:scale(.65)} }
.spinner{ width:16px; height:16px; flex-shrink:0; border:2.5px solid currentColor;
  border-top-color:transparent; border-radius:50%; animation:spin .65s linear infinite; }
@keyframes spin{ to{ transform:rotate(360deg); } }
.icon-static{ flex-shrink:0; font-size:15px; line-height:1; }
</style>
<div id="cam-wrap">
  <video id="video" autoplay playsinline muted></video>
  <canvas id="canvas"></canvas>
  <img id="preview" alt="captured photo"/>
  <div class="toolbar">
    <button class="cam-btn" id="btn-capture">📷&nbsp;&nbsp;Capture Photo</button>
    <button class="cam-btn" id="btn-retake">🔄&nbsp;&nbsp;Retake</button>
  </div>
  <div class="status-bar ready" id="status-bar">
    <span class="pulse-dot" id="status-icon"></span>
    <span id="status-txt">Starting camera…</span>
  </div>
</div>
<script>
(function(){
  const video=document.getElementById('video'),canvas=document.getElementById('canvas'),
        preview=document.getElementById('preview'),btnCap=document.getElementById('btn-capture'),
        btnRet=document.getElementById('btn-retake'),bar=document.getElementById('status-bar'),
        icon=document.getElementById('status-icon'),txt=document.getElementById('status-txt');
  function setStatus(m,mode){
    txt.textContent=m; bar.className='status-bar '+(mode||'ready');
    icon.className=''; icon.textContent=''; icon.style.cssText='';
    if(mode==='sending') icon.className='spinner';
    else if(mode==='success'){icon.className='icon-static';icon.textContent='✓';}
    else if(mode==='error') {icon.className='icon-static';icon.textContent='⚠';}
    else icon.className='pulse-dot';
  }
  async function startCamera(){
    try{
      const s=await navigator.mediaDevices.getUserMedia(
        {video:{facingMode:{ideal:'environment'},width:{ideal:4096},height:{ideal:3072}},audio:false});
      video.srcObject=s; await video.play();
      setStatus('Camera ready — frame the order form and tap Capture','ready');
    }catch(e){setStatus('Camera error: '+e.message,'error');}
  }
  function convolve3x3(data,w,h,kernel){
    const out=new Uint8ClampedArray(data.length);
    for(let y=1;y<h-1;y++) for(let x=1;x<w-1;x++){
      for(let c=0;c<3;c++){
        let sum=0;
        for(let ky=-1;ky<=1;ky++) for(let kx=-1;kx<=1;kx++)
          sum+=data[((y+ky)*w+(x+kx))*4+c]*kernel[(ky+1)*3+(kx+1)];
        out[(y*w+x)*4+c]=Math.max(0,Math.min(255,sum));
      }
      out[(y*w+x)*4+3]=255;
    }
    return out;
  }
  function sharpenAndEnhance(ctx,w,h){
    const imgData=ctx.getImageData(0,0,w,h),d=imgData.data;
    for(let i=0;i<d.length;i+=4)
      for(let c=0;c<3;c++){
        let v=Math.round(((d[i+c]/255-.5)*1.25+.5)*255);
        d[i+c]=Math.max(0,Math.min(255,v));
      }
    const kernel=[0,-1,0,-1,5,-1,0,-1,0],out=ctx.createImageData(w,h);
    out.data.set(convolve3x3(d,w,h,kernel)); ctx.putImageData(out,0,0);
  }
  btnCap.addEventListener('click',()=>{
    const w=video.videoWidth||1280,h=video.videoHeight||720;
    canvas.width=w; canvas.height=h;
    const ctx=canvas.getContext('2d');
    ctx.drawImage(video,0,0,w,h); sharpenAndEnhance(ctx,w,h);
    const b64=canvas.toDataURL('image/jpeg',0.97);
    preview.src=b64; video.style.display='none';
    preview.style.display='block'; btnCap.style.display='none'; btnRet.style.display='block';
    setStatus('Sending photo to app…','sending');
    window.parent.sessionStorage.setItem('sintex_cam_b64',b64);
    window.parent.postMessage({type:'SINTEX_CAM_CAPTURE',data:b64},'*');
    setTimeout(()=>setStatus('Photo captured — use Retake to redo, or proceed to OCR','success'),700);
  });
  btnRet.addEventListener('click',()=>{
    preview.style.display='none'; video.style.display='block';
    btnCap.style.display='block'; btnRet.style.display='none';
    window.parent.sessionStorage.removeItem('sintex_cam_b64');
    window.parent.postMessage({type:'SINTEX_CAM_RETAKE'},'*');
    setStatus('Camera ready — frame the order form and tap Capture','ready');
  });
  startCamera();
})();
</script>
"""
        components.html(CAMERA_HTML, height=580, scrolling=False)

        AUTO_BRIDGE = """
<script>
(function(){
  var attempts=0,MAX=120;
  function tryInject(){
    attempts++;
    if(attempts>MAX)return;
    var b64=window.parent.sessionStorage.getItem('sintex_cam_b64');
    if(!b64){setTimeout(tryInject,500);return;}
    var parent=window.parent.document,inputs=parent.querySelectorAll('input[type="text"]'),target=null;
    for(var i=inputs.length-1;i>=0;i--){
      var inp=inputs[i],st=window.parent.getComputedStyle(inp);
      if(parseFloat(st.height)<5||parseFloat(st.opacity)<0.1){target=inp;break;}
    }
    if(!target&&inputs.length)target=inputs[inputs.length-1];
    if(!target){setTimeout(tryInject,500);return;}
    var setter=Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype,'value').set;
    setter.call(target,b64);
    target.dispatchEvent(new Event('input',{bubbles:true}));
    target.dispatchEvent(new Event('change',{bubbles:true}));
    window.parent.sessionStorage.removeItem('sintex_cam_b64');
  }
  setTimeout(tryInject,800);
})();
</script>
"""
        components.html(AUTO_BRIDGE, height=0)

        cam_val = st.text_input(
            "sintex_cam_hidden_receiver",
            value=st.session_state.get("_cam_recv_val", ""),
            label_visibility="collapsed",
            key="sintex_cam_recv",
        )

        recv = st.session_state.get("sintex_cam_recv", "")
        if recv and recv.startswith("data:image"):
            import base64 as _b64
            _, encoded = recv.split(",", 1)
            raw = _b64.b64decode(encoded)
            if raw != st.session_state.image_bytes:
                st.session_state.image_bytes      = raw
                st.session_state.ocr_done         = False
                st.session_state.ocr_raw_rows     = []
                st.session_state.ocr_resolved     = []
                st.session_state["_cam_recv_val"] = recv
                st.rerun()

        components.html("""
<script>
(function(){
  var parent=window.parent.document,inputs=parent.querySelectorAll('input[type="text"]');
  if(inputs.length){
    var last=inputs[inputs.length-1];
    last.style.cssText+='height:1px!important;opacity:0!important;pointer-events:none!important;position:absolute!important;';
    last.setAttribute('data-sintex-cam','1');
    var wrap=last.closest('[data-testid="stTextInput"]');
    if(wrap)wrap.style.cssText+='height:0!important;overflow:hidden!important;margin:0!important;padding:0!important;';
  }
})();
</script>
""", height=0)

    else:
        st.file_uploader("Upload image of order form", type=["jpg","jpeg","png"],
                         label_visibility="collapsed",
                         key="file_upload_input", on_change=_on_file_upload)

    # ── Image preview ────────────────────────────────────────────────────────
    if st.session_state.image_bytes:
        st.markdown("""
        <div style='background:#fff;border:1px solid #DEE3EC;border-radius:10px;
                    padding:12px;margin:14px 0 6px;'>
          <div style='font-size:11px;font-weight:700;text-transform:uppercase;
                      letter-spacing:.5px;color:#0A2342;margin-bottom:10px;'>
            📸 Captured Image
          </div>""", unsafe_allow_html=True)
        st.image(st.session_state.image_bytes, use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)
    else:
        st.markdown("""
        <div style='background:#F4F6FA;border:2px dashed #DEE3EC;border-radius:10px;
                    padding:40px;text-align:center;margin:14px 0;color:#5A6880;'>
          <div style='font-size:36px;margin-bottom:10px;'>🖼️</div>
          <div style='font-weight:600;font-size:14px;margin-bottom:4px;'>No image loaded</div>
          <div style='font-size:12px;'>Capture or upload an image above</div>
        </div>""", unsafe_allow_html=True)

    # ── OCR result summary ───────────────────────────────────────────────────
    if st.session_state.ocr_done:
        resolved = st.session_state.ocr_resolved
        n_matched = sum(1 for r in resolved if r["match_type"] != "none")
        n_total   = len(resolved)
        if n_total > 0:
            st.success(
                f"✅ OCR complete — **{n_total} rows** detected, "
                f"**{n_matched} SKUs** matched to catalogue. "
                "Review and edit in the next step.",
                icon="🔍",
            )
        else:
            st.warning("OCR ran but found no recognisable SKU rows. You can fill manually in Step 3.",
                       icon="⚠️")

    st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)
    btn_cols = st.columns([1, 1, 1, 3])

    with btn_cols[0]:
        if st.button("◀  Back", key="back_step1"):
            st.session_state.step = 1
            st.rerun()

    with btn_cols[1]:
        if st.button("🔍  Run OCR", key="run_ocr_btn"):
            if not st.session_state.image_bytes:
                st.error("Please upload or capture an image first.", icon="🖼️")
            else:
                _ep  = st.session_state.azure_endpoint.strip()
                _key = st.session_state.azure_key.strip()
                if not _ep or not _key:
                    st.error("Please enter your **Azure Endpoint** and **Azure Key** above.", icon="🔑")
                else:
                    with st.spinner("Sending image to Azure OCR… this may take 10–20 seconds"):
                        try:
                            raw_rows = run_azure_ocr(st.session_state.image_bytes, _ep, _key)
                            resolved = resolve_ocr_rows(raw_rows, sku_cat, mrp_lookup)
                            st.session_state.ocr_raw_rows = raw_rows
                            st.session_state.ocr_resolved = resolved
                            st.session_state.ocr_done     = True
                            st.rerun()
                        except Exception as exc:
                            st.error(f"OCR failed: {exc}", icon="❌")

    with btn_cols[2]:
        if st.button("▶  Next: Review & Edit", key="go_step3"):
            st.session_state.step = 3
            st.rerun()


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 3 — Review OCR Table + Manual Editing
# ═══════════════════════════════════════════════════════════════════════════════
elif step == 3:
    st.markdown(f"""
    <div class="card">
      <div class="card-title">{badge(3, step > 3)} Review Detected SKUs & Quantities</div>
    </div>""", unsafe_allow_html=True)

    resolved = st.session_state.get("ocr_resolved", [])

    if resolved:
        st.info(
            "The table below shows OCR-detected rows matched to the product catalogue. "
            "**Edit quantities and correct any mismatched SKUs before proceeding.**",
            icon="📋",
        )

        # Legend
        st.markdown("""
        <div style="display:flex;gap:16px;margin-bottom:10px;flex-wrap:wrap;">
          <span><span class="match-ok">●</span> Exact match</span>
          <span><span class="match-warn">●</span> Fuzzy match (verify)</span>
          <span><span class="match-none">●</span> No match found</span>
        </div>
        """, unsafe_allow_html=True)

        # ── Editable review table ────────────────────────────────────────────
        # Columns: Raw OCR | OCR Size | Matched SKU | Item | Conf% | MRP | Rate | Qty
        HDR_WIDTHS = [1.4, 0.9, 1.5, 1.8, 0.8, 1.1, 1.1, 0.9]
        HDR_LABELS = ["Raw OCR", "OCR Size", "Matched SKU", "Item / Size", "Conf%",
                      "MRP (₹)", "Rate (₹)", "Qty"]
        hdr = st.columns(HDR_WIDTHS)
        for col, lbl in zip(hdr, HDR_LABELS):
            col.markdown(
                f"<div style='font-size:11px;font-weight:700;color:#0A2342;"
                f"background:#EBF3FF;border-radius:4px;padding:3px 6px;text-align:center;'>"
                f"{lbl}</div>",
                unsafe_allow_html=True,
            )

        updated_resolved = []
        _ni_step3 = _build_normalised_index(tuple(sku_cat.keys()))

        for idx, r in enumerate(resolved):
            cols = st.columns(HDR_WIDTHS)

            # Col 0: Raw OCR token (read-only display)
            cols[0].markdown(
                f'<div style="font-family:monospace;font-size:10.5px;padding:6px 3px;'
                f'color:#5A6880;word-break:break-all;">{r["raw_sku"]}<br/>'
                f'<span style="color:#9BAECC;font-size:9px;">{r.get("raw_size","")}</span></div>',
                unsafe_allow_html=True,
            )

            # Col 1: OCR-detected size (editable)
            new_size = cols[1].text_input(
                label=f"size_{idx}",
                value=r.get("raw_size", ""),
                label_visibility="collapsed",
                key=f"edit_size_{idx}",
                placeholder="e.g. 15MM",
            )

            # Col 2: Matched SKU (editable — user can override)
            new_sku = cols[2].text_input(
                label=f"sku_{idx}",
                value=r["matched_sku"] if r["matched_sku"] != "—" else "",
                label_visibility="collapsed",
                key=f"edit_sku_{idx}",
            )

            # Re-resolve whenever SKU or size is changed
            changed = (new_sku and new_sku != r["matched_sku"]) or (new_size != r.get("raw_size",""))
            if changed:
                probe_sku  = new_sku if new_sku else r["raw_sku"]
                re_sku, re_conf, re_type = match_sku_to_catalogue(probe_sku, sku_cat, _ni_step3)
                if re_sku:
                    # Try to narrow to the exact size variant
                    size_norm = re.sub(r"[^0-9]", "", new_size or r.get("raw_size",""))
                    base_norm = _normalise_sku(probe_sku)
                    size_hits = [
                        (sku, ent) for sku, ent in sku_cat.items()
                        if _normalise_sku(sku).startswith(base_norm)
                           and re.sub(r"[^0-9]", "", ent.get("col_label","")) == size_norm
                    ]
                    if size_hits:
                        re_sku, cat_entry = size_hits[0]
                        re_type = "exact"; re_conf = 1.0
                    else:
                        cat_entry = sku_cat[re_sku]
                    info = mrp_lookup.get(re_sku, {})
                    r = {**r,
                         "matched_sku": re_sku,
                         "raw_size":    new_size or r.get("raw_size",""),
                         "match_type":  re_type,
                         "confidence":  round(re_conf * 100, 1),
                         "sheet":       cat_entry.get("sheet","—"),
                         "row_label":   cat_entry.get("row_label","—"),
                         "col_label":   cat_entry.get("col_label","—"),
                         "mrp":         info.get("MRP_clean") or 0.0,
                         "landing":     info.get("Distributor Landing") or 0.0,
                         "ri":          cat_entry.get("ri"),
                         "ci":          cat_entry.get("ci"),
                    }
                else:
                    r = {**r, "matched_sku": new_sku or "—", "raw_size": new_size,
                         "match_type": "none", "confidence": 0.0,
                         "mrp": 0.0, "landing": 0.0, "ri": None, "ci": None}

            # Col 3: Item / size description (read-only)
            item_lbl = r["row_label"]
            size_lbl = r.get("col_label") or r.get("raw_size","")
            mt = r["match_type"]
            badge_class = "match-ok" if mt == "exact" else ("match-warn" if mt == "fuzzy" else "match-none")
            badge_icon  = "✓" if mt == "exact" else ("~" if mt == "fuzzy" else "✗")
            cols[3].markdown(
                f'<div style="font-size:11px;padding:4px 3px;line-height:1.4;">'
                f'<b>{item_lbl[:30]}</b><br/>'
                f'<span style="font-size:10px;color:#5A6880;">{size_lbl}</span>'
                f'</div>',
                unsafe_allow_html=True,
            )

            # Col 4: Confidence badge
            cols[4].markdown(
                f'<div class="{badge_class}" style="padding:6px 3px;font-size:11px;text-align:center;">'
                f'{badge_icon} {r["confidence"]:.0f}%</div>',
                unsafe_allow_html=True,
            )

            # Col 5: MRP
            cols[5].markdown(
                f'<div style="font-size:11px;padding:6px 3px;font-family:monospace;text-align:right;">'
                f'₹ {r["mrp"]:,.2f}</div>',
                unsafe_allow_html=True,
            )

            # Col 6: Landing rate
            cols[6].markdown(
                f'<div style="font-size:11px;padding:6px 3px;font-family:monospace;text-align:right;">'
                f'₹ {r["landing"]:,.2f}</div>',
                unsafe_allow_html=True,
            )

            # Col 7: Qty (editable number input)
            new_qty = cols[7].number_input(
                label=f"qty_{idx}",
                value=int(r["qty"]),
                min_value=0,
                step=1,
                label_visibility="collapsed",
                key=f"edit_qty_{idx}",
            )
            r = {**r, "qty": new_qty, "amount": round(r["landing"] * new_qty, 2)}
            updated_resolved.append(r)

        st.session_state.ocr_resolved = updated_resolved

    else:
        # No OCR — allow manual SKU entry
        st.warning("No OCR results available. Use the form below to add items manually.", icon="📝")

        st.markdown("#### ➕ Add Items Manually")
        st.caption("Type a SKU code — it will be looked up in the catalogue automatically.")

        if "manual_rows" not in st.session_state:
            st.session_state.manual_rows = [{"sku": "", "qty": 0}]

        new_manual = []
        for mi, mr in enumerate(st.session_state.manual_rows):
            mc1, mc2, mc3 = st.columns([3, 1, 0.5])
            sku_val = mc1.text_input(f"SKU Code", value=mr["sku"],
                                     key=f"manual_sku_{mi}", label_visibility="visible")
            qty_val = mc2.number_input("Qty", value=int(mr["qty"]), min_value=0, step=1,
                                       key=f"manual_qty_{mi}", label_visibility="visible")
            new_manual.append({"sku": sku_val, "qty": qty_val})

        if st.button("➕ Add Row"):
            st.session_state.manual_rows = new_manual + [{"sku": "", "qty": 0}]
            st.rerun()
        else:
            st.session_state.manual_rows = new_manual

        if st.button("✅ Apply Manual Entries", key="apply_manual"):
            manual_resolved = []
            for mr in st.session_state.manual_rows:
                sku = mr["sku"].strip()
                qty = mr["qty"]
                if not sku:
                    continue
                _ni2 = _build_normalised_index(tuple(sku_cat.keys()))
                m_sku, conf, mtype = match_sku_to_catalogue(sku, sku_cat, _ni2)
                cat_entry = sku_cat.get(m_sku, {}) if m_sku else {}
                info      = mrp_lookup.get(m_sku, {}) if m_sku else {}
                mrp_v     = info.get("MRP_clean") or 0.0
                land_v    = info.get("Distributor Landing") or mrp_v
                manual_resolved.append({
                    "raw_sku":     sku,
                    "matched_sku": m_sku or "—",
                    "match_type":  mtype,
                    "confidence":  round(conf * 100, 1),
                    "qty":         qty,
                    "sheet":       cat_entry.get("sheet","—"),
                    "row_label":   cat_entry.get("row_label","—"),
                    "col_label":   cat_entry.get("col_label","—"),
                    "mrp":         mrp_v,
                    "landing":     land_v,
                    "amount":      round(land_v * qty, 2),
                    "ri":          cat_entry.get("ri"),
                    "ci":          cat_entry.get("ci"),
                })
            st.session_state.ocr_resolved = manual_resolved
            st.success(f"✅ {len(manual_resolved)} item(s) applied.", icon="✅")
            st.rerun()

    # ── Live totals ──────────────────────────────────────────────────────────
    resolved_now = st.session_state.ocr_resolved
    grand_mrp  = sum(r["mrp"]     * r["qty"] for r in resolved_now if r["matched_sku"] != "—")
    grand_land = sum(r["landing"] * r["qty"] for r in resolved_now if r["matched_sku"] != "—")
    n_lines    = sum(1 for r in resolved_now if r["qty"] > 0 and r["matched_sku"] != "—")

    st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)
    st.markdown(f"""
    <div class="card" style="max-width:420px;">
      <div class="card-title">📊 Live Summary</div>
      <div class="totals-row">
        <span>Active Line Items</span><span class="val">{n_lines}</span>
      </div>
      <div class="totals-row">
        <span>Gross MRP Value</span><span class="val">₹ {grand_mrp:,.2f}</span>
      </div>
      <div class="totals-row">
        <span>Distributor Landing</span><span class="val">₹ {grand_land:,.2f}</span>
      </div>
      <div class="totals-row net">
        <span>Net Payable (Landing)</span><span class="val">₹ {grand_land:,.2f}</span>
      </div>
    </div>
    """, unsafe_allow_html=True)

    col_b1, col_b2, _ = st.columns([1, 1, 3])
    with col_b1:
        if st.button("◀  Back", key="back_step2"):
            st.session_state.step = 2
            st.rerun()
    with col_b2:
        if st.button("▶  Next: Download", key="go_step4"):
            # Build qty_map from resolved for downstream compatibility
            qty_map = {}
            for r in st.session_state.ocr_resolved:
                if r["ri"] is not None and r["ci"] is not None and r["qty"] > 0:
                    qty_map[(r["ri"], r["ci"])] = r["qty"]
            st.session_state.qty_map = qty_map
            st.session_state.step = 4
            st.rerun()


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 4 — Download
# ═══════════════════════════════════════════════════════════════════════════════
elif step == 4:
    st.markdown(f"""
    <div class="card">
      <div class="card-title">{badge(4)} Review & Download Quotation</div>
    </div>""", unsafe_allow_html=True)

    resolved_rows = [r for r in st.session_state.ocr_resolved
                     if r["qty"] > 0 and r["matched_sku"] != "—"]

    # Party summary
    pc1, pc2 = st.columns(2)
    with pc1:
        st.markdown(party_html(st.session_state.bill_to, "BILL TO PARTY"), unsafe_allow_html=True)
    with pc2:
        st.markdown(party_html(st.session_state.ship_to, "SHIP TO PARTY"), unsafe_allow_html=True)

    if any(st.session_state.dealer_info.values()):
        st.markdown(party_html(st.session_state.dealer_info, "DEALER / DISTRIBUTOR"),
                    unsafe_allow_html=True)

    st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)
    st.markdown("#### 📦 Ordered Items")

    if resolved_rows:
        line_data = []
        grand_mrp = grand_land = 0.0
        for r in resolved_rows:
            grand_mrp  += r["mrp"]     * r["qty"]
            grand_land += r["landing"] * r["qty"]
            line_data.append({
                "Item":        r["row_label"],
                "SKU Code":    r["matched_sku"],
                "Size":        r["col_label"],
                "Sheet":       r["sheet"],
                "MRP (₹)":     round(r["mrp"],     2),
                "Qty":         r["qty"],
                "Rate (₹)":    round(r["landing"],  2),
                "Amount (₹)":  round(r["amount"],   2),
            })

        df_lines = pd.DataFrame(line_data)
        st.dataframe(df_lines, use_container_width=True, hide_index=True,
                     column_config={
                         "MRP (₹)":    st.column_config.NumberColumn(format="₹ %.2f"),
                         "Rate (₹)":   st.column_config.NumberColumn(format="₹ %.2f"),
                         "Amount (₹)": st.column_config.NumberColumn(format="₹ %.2f"),
                     })

        discount = grand_mrp - grand_land
        st.markdown(f"""
        <div class="card" style="max-width:440px;margin-top:12px;">
          <div class="card-title">💰 Totals</div>
          <div class="totals-row">
            <span>Gross MRP Value</span><span class="val">₹ {grand_mrp:,.2f}</span>
          </div>
          <div class="totals-row">
            <span>Distributor Discount</span><span class="val neg">− ₹ {discount:,.2f}</span>
          </div>
          <div class="totals-row net">
            <span>Net Payable (Distributor Landing)</span>
            <span class="val">₹ {grand_land:,.2f}</span>
          </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)
        dl1, dl2, dl3 = st.columns([1, 1, 2])

        # CSV
        csv_buf = io.StringIO()
        df_lines.to_csv(csv_buf, index=False)
        dl1.download_button(
            "⬇  Download CSV",
            data=csv_buf.getvalue().encode(),
            file_name=f"sintex_quotation_{date.today()}.csv",
            mime="text/csv",
            key="dl_csv",
        )

        # Excel
        xls_buf = io.BytesIO()
        with pd.ExcelWriter(xls_buf, engine="openpyxl") as writer:
            party_data = []
            for k, v in st.session_state.dealer_info.items():
                party_data.append({"Field": f"Dealer - {k}", "Value": v})
            for k, v in st.session_state.bill_to.items():
                party_data.append({"Field": f"Bill To - {k}", "Value": v})
            for k, v in st.session_state.ship_to.items():
                party_data.append({"Field": f"Ship To - {k}", "Value": v})
            pd.DataFrame(party_data).to_excel(writer, sheet_name="Party Details", index=False)
            df_lines.to_excel(writer, sheet_name="Quotation Lines", index=False)
            pd.DataFrame([
                {"Description": "Gross MRP Value",   "Amount (₹)": round(grand_mrp,  2)},
                {"Description": "Distributor Disc.",  "Amount (₹)": round(discount,   2)},
                {"Description": "Net Payable",        "Amount (₹)": round(grand_land, 2)},
            ]).to_excel(writer, sheet_name="Summary", index=False)

        dl2.download_button(
            "⬇  Download Excel",
            data=xls_buf.getvalue(),
            file_name=f"sintex_quotation_{date.today()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="dl_excel",
        )

        # PDF
        if _HAS_REPORTLAB:
            if st.session_state.pdf_bytes is None:
                with st.spinner("Generating PDF…"):
                    try:
                        pdf_bytes = generate_pdf_from_resolved(
                            resolved_rows, mrp_lookup,
                            st.session_state.bill_to,
                            st.session_state.ship_to,
                            st.session_state.dealer_info,
                        )
                        st.session_state.pdf_bytes = pdf_bytes
                    except Exception as e:
                        st.error(f"PDF generation failed: {e}")

            if st.session_state.pdf_bytes:
                dl3.download_button(
                    "⬇  Download PDF",
                    data=st.session_state.pdf_bytes,
                    file_name=f"sintex_quotation_{date.today()}.pdf",
                    mime="application/pdf",
                    key="dl_pdf",
                )
        else:
            st.caption("📄 PDF download requires `reportlab`. Run: `pip install reportlab`")

    else:
        st.warning("No valid line items found. Go back and check the OCR review step.", icon="⚠️")

    st.markdown("<div style='height:14px'></div>", unsafe_allow_html=True)
    col_bk, col_nr, _ = st.columns([1, 1, 3])
    with col_bk:
        if st.button("◀  Back", key="back_step3"):
            st.session_state.pdf_bytes = None
            st.session_state.step = 3
            st.rerun()
    with col_nr:
        if st.button("🔄  New Quotation", key="new_quot"):
            keys_to_clear = [
                "step", "qty_map", "bill_to", "ship_to", "dealer_info", "pdf_bytes",
                "image_bytes", "ocr_raw_rows", "ocr_resolved", "ocr_done",
                "zsd_search", "manual_rows",
            ]
            for k in keys_to_clear:
                if k in st.session_state:
                    del st.session_state[k]
            for k in list(st.session_state.keys()):
                if k.startswith(("qty_", "edit_", "manual_", "bill_", "ship_", "dealer_")):
                    del st.session_state[k]
            st.rerun()

# import io
# import os
# import re
# import base64
# import time
# from datetime import date

# import pandas as pd
# import streamlit as st
# from openpyxl import load_workbook

# try:
#     import requests as _requests
#     _HAS_REQUESTS = True
# except ImportError:
#     _HAS_REQUESTS = False

# try:
#     from reportlab.lib.pagesizes import A4
#     from reportlab.lib import colors
#     from reportlab.lib.units import mm
#     from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
#                                     Paragraph, Spacer, HRFlowable)
#     from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
#     from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
#     _HAS_REPORTLAB = True
# except ImportError:
#     _HAS_REPORTLAB = False

# st.set_page_config(
#     page_title="Sintex BAPL – Quotation Generator",
#     page_icon="🔧",
#     layout="wide",
#     initial_sidebar_state="collapsed",
# )

# st.markdown("""
# <style>
# @import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@300;400;500;600;700&family=IBM+Plex+Mono:wght@400;500&display=swap');
# :root {
#   --navy:   #0A2342;
#   --blue:   #1565C0;
#   --sky:    #1E88E5;
#   --teal:   #00796B;
#   --gold:   #F9A825;
#   --danger: #C62828;
#   --border: #DEE3EC;
#   --surface:#F4F6FA;
#   --text:   #1A1F36;
#   --muted:  #5A6880;
#   --radius: 10px;
# }
# html, body, [class*="css"] {
#   font-family: 'IBM Plex Sans', sans-serif !important;
#   background: #F0F2F8 !important;
#   color: var(--text);
# }
# #MainMenu, footer, header { visibility: hidden; }
# .block-container { padding: 1.2rem 1.5rem 4rem !important; max-width: 1400px !important; }
# .card {
#   background: #fff; border: 1px solid var(--border);
#   border-radius: var(--radius); padding: 20px 24px;
#   margin-bottom: 18px; box-shadow: 0 1px 4px rgba(10,35,66,.07);
# }
# .card-title {
#   font-size: 13px; font-weight: 700; text-transform: uppercase;
#   letter-spacing: .7px; color: var(--navy); margin-bottom: 14px;
#   display: flex; align-items: center; gap: 8px;
# }
# .step-badge {
#   display: inline-flex; align-items: center; justify-content: center;
#   background: var(--navy); color: #fff; border-radius: 50%;
#   width: 24px; height: 24px; font-size: 11px; font-weight: 700; flex-shrink: 0;
# }
# .step-badge.done { background: var(--teal); }
# .app-header {
#   background: linear-gradient(135deg, #0A2342 0%, #1565C0 100%);
#   border-radius: 12px; padding: 18px 26px;
#   display: flex; align-items: center; gap: 18px;
#   margin-bottom: 22px; box-shadow: 0 4px 16px rgba(10,35,66,.25);
# }
# .app-header h1 { font-size: 20px; font-weight: 700; color: #fff; margin: 0; }
# .app-header p  { font-size: 12px; color: rgba(255,255,255,.65); margin: 3px 0 0; }
# .stButton > button {
#   font-family: 'IBM Plex Sans', sans-serif !important;
#   font-weight: 600 !important; font-size: 14px !important;
#   border-radius: 8px !important; border: none !important;
#   background: var(--navy) !important; color: #fff !important;
#   padding: 10px 20px !important;
#   box-shadow: 0 2px 8px rgba(10,35,66,.2) !important;
#   transition: all .15s !important;
# }
# .stButton > button:hover { background: var(--sky) !important; transform: translateY(-1px) !important; }
# .totals-row {
#   display: flex; justify-content: space-between; align-items: center;
#   padding: 9px 0; border-bottom: 1px solid var(--border); font-size: 14px;
# }
# .totals-row:last-child { border: none; padding-top: 12px; }
# .totals-row.net { font-weight: 700; font-size: 16px; color: var(--navy); }
# .totals-row .val { font-family: 'IBM Plex Mono', monospace; font-weight: 600; }
# .totals-row .val.neg { color: var(--danger); }
# .party-box {
#   border: 1px solid var(--border); border-radius: 8px;
#   padding: 14px 16px; background: var(--surface);
# }
# .party-box h4 { margin: 0 0 10px; font-size: 12px; font-weight: 700;
#   text-transform: uppercase; letter-spacing: .5px; color: var(--navy); }
# .party-row { display: flex; gap: 6px; margin-bottom: 5px; font-size: 12.5px; }
# .party-lbl { color: var(--muted); min-width: 80px; flex-shrink: 0; }
# .party-val { font-weight: 500; color: var(--text); word-break: break-word; }
# label { font-size:12px!important; font-weight:600!important; color:var(--muted)!important; }
# </style>
# """, unsafe_allow_html=True)

# # ─── Paths ───────────────────────────────────────────────────────────────────
# _HERE = os.path.dirname(os.path.abspath(__file__))
# MRP_CSV_PATH  = os.path.join(_HERE, "MRP_State_chhattisghar.csv")
# ZSD_CSV_PATH  = os.path.join(_HERE, "ZSD_CUST.csv")
# SKU_XLSX_PATH = os.path.join(_HERE, "Sample form for Product list.xlsx")

# # ─── Data loaders ─────────────────────────────────────────────────────────────
# @st.cache_data
# def load_mrp():
#     """Returns dict: {material_number_upper: {MRP_clean, Distributor_Landing, description}}"""
#     df = pd.read_csv(MRP_CSV_PATH)
#     def _clean_num(val):
#         try:
#             return float(str(val).replace(",", "").strip())
#         except Exception:
#             return 0.0
#     lookup = {}
#     for _, row in df.iterrows():
#         mat = str(row["Material Number"]).strip().upper()
#         lookup[mat] = {
#             "mrp":         _clean_num(row.get("MRP(ZPR1-933)", 0)),
#             "landing":     _clean_num(row.get("Distributor Landing", 0)),
#             "description": str(row.get("Material Description", "")).strip(),
#         }
#     return lookup

# @st.cache_data
# def load_zsd():
#     df = pd.read_csv(ZSD_CSV_PATH, encoding="latin1")
#     return df

# @st.cache_data
# def build_sku_grid():
#     """
#     Parse the Excel SKU catalogue and build a nested dict:
#       grid[sheet_name][row_label][base_sku][size_mm] = full_sku_code
#     Also returns a flat reverse lookup:
#       full_sku_upper → {sheet, row_label, base_sku, size_mm, size_label}
#     And a list of (base_sku, row_label, sheet, sizes_list) for display.
#     """
#     wb = load_workbook(SKU_XLSX_PATH, read_only=True)

#     # full_sku → info
#     full_sku_lookup: dict = {}
#     # base_sku_upper → list of {size_mm, full_sku, size_label, row_label, sheet}
#     base_to_variants: dict = {}
#     # For display: list of sheets with their rows
#     sheets_info: list = []

#     for sname in wb.sheetnames:
#         ws = wb[sname]
#         rows_data = [r for r in ws.iter_rows(values_only=True)
#                      if any(c is not None for c in r)]
#         if len(rows_data) < 3:
#             continue

#         # Row 0: sheet title / ignored
#         # Row 1: column headers (Pipes/Fittings label, SKU Code, size labels like 15MM/20MM...)
#         # Row 2: sub-labels (inch equivalents like ½", ¾"...)
#         # Row 3+: data rows

#         header_row  = rows_data[1]   # e.g. ('Pipes', 'SKU Code', '15MM', '20MM', ...)
#         sublabel_row = rows_data[2]  # e.g. (None, None, '½"', '¾"', ...)

#         # Build column index: col_idx → (size_mm_str, size_label_str)
#         col_info = {}
#         for ci, val in enumerate(header_row[2:], start=2):
#             if val is not None:
#                 size_mm = str(val).strip()   # e.g. "15MM", "20X15"
#                 sub     = rows_data[2][ci] if ci < len(rows_data[2]) else None
#                 size_lbl = str(sub).strip() if sub else size_mm
#                 col_info[ci] = (size_mm, size_lbl)

#         sheet_rows = []
#         for raw in rows_data[3:]:
#             if not any(c is not None for c in raw):
#                 continue
#             row_label = str(raw[0]).strip() if raw[0] is not None else ""
#             base_sku  = str(raw[1]).strip() if raw[1] is not None else ""
#             if not row_label:
#                 continue
#             # Section separator row (FITTINGS, PIPES header)
#             if base_sku in ("", "\xa0") and all(
#                 (raw[ci] if ci < len(raw) else None) is None
#                 for ci in col_info
#             ):
#                 continue

#             sizes_in_row = []
#             for ci, (size_mm, size_lbl) in col_info.items():
#                 full_sku = raw[ci] if ci < len(raw) else None
#                 if full_sku is None or str(full_sku).strip() in ("", "-", "None"):
#                     continue
#                 full_sku = str(full_sku).strip()
#                 full_sku_up = full_sku.upper()

#                 full_sku_lookup[full_sku_up] = {
#                     "sheet":      sname,
#                     "row_label":  row_label,
#                     "base_sku":   base_sku.upper(),
#                     "size_mm":    size_mm,
#                     "size_label": size_lbl,
#                 }
#                 base_up = base_sku.upper()
#                 if base_up not in base_to_variants:
#                     base_to_variants[base_up] = []
#                 base_to_variants[base_up].append({
#                     "size_mm":    size_mm,
#                     "size_label": size_lbl,
#                     "full_sku":   full_sku_up,
#                     "row_label":  row_label,
#                     "sheet":      sname,
#                     "ci":         ci,
#                 })
#                 sizes_in_row.append(size_mm)

#             if row_label and sizes_in_row:
#                 sheet_rows.append({
#                     "row_label": row_label,
#                     "base_sku":  base_sku.upper(),
#                     "sizes":     sizes_in_row,
#                 })

#         sheets_info.append({
#             "name":     sname,
#             "col_info": col_info,
#             "rows":     sheet_rows,
#         })

#     return full_sku_lookup, base_to_variants, sheets_info


# # ─── OCR helpers ──────────────────────────────────────────────────────────────

# def _word_centre(polygon):
#     """Return (x, y) centre given a polygon (flat list or list of {x,y})."""
#     if not polygon:
#         return 0.0, 0.0
#     if isinstance(polygon[0], dict):
#         xs = [p["x"] for p in polygon]
#         ys = [p["y"] for p in polygon]
#     else:
#         xs = [polygon[i] for i in range(0, len(polygon), 2)]
#         ys = [polygon[i] for i in range(1, len(polygon), 2)]
#     return (min(xs)+max(xs))/2.0, (min(ys)+max(ys))/2.0


# def _cluster_rows(words: list, page_height: float, gap_frac: float = 0.012) -> list:
#     """Group word dicts (with 'x','y' normalised) into horizontal rows."""
#     if not words:
#         return []
#     words = sorted(words, key=lambda w: w["y"])
#     gap = gap_frac * page_height
#     rows, cur = [], [words[0]]
#     for w in words[1:]:
#         if w["y"] - cur[-1]["y"] > gap:
#             rows.append(cur)
#             cur = [w]
#         else:
#             cur.append(w)
#     rows.append(cur)
#     return rows


# def _is_base_sku(token: str) -> bool:
#     """Looks like a printed Sintex SKU code – alphanumeric, ≥6 chars."""
#     t = token.strip().upper()
#     return (len(t) >= 6
#             and any(c.isdigit() for c in t)
#             and any(c.isalpha() for c in t))


# def _is_size_header(token: str) -> bool:
#     t = token.strip().upper()
#     return bool(re.match(r'^\d{2}(MM|X\d+|x\d+)$', t))


# def _closest_column(x: float, col_centres: dict) -> str | None:
#     """Return the size_mm key whose x-centre is closest to x."""
#     if not col_centres:
#         return None
#     best, best_d = None, float("inf")
#     for size, cx in col_centres.items():
#         d = abs(cx - x)
#         if d < best_d:
#             best_d = d
#             best = size
#     col_width = max(col_centres.values()) - min(col_centres.values())
#     if col_width > 0 and best_d > col_width * 0.7:
#         return None
#     return best


# # ─── Azure Document Intelligence OCR ─────────────────────────────────────────

# def _ocr_azure(image_bytes: bytes, endpoint: str, key: str) -> list:
#     """
#     Call Azure Document Intelligence (prebuilt-read) and return a flat list of:
#       {"text": str, "x": float, "y": float}   (normalised 0-1 coordinates)
#     """
#     url = endpoint.rstrip("/") + "/formrecognizer/documentModels/prebuilt-read:analyze?api-version=2023-07-31"
#     hdrs = {"Ocp-Apim-Subscription-Key": key, "Content-Type": "application/octet-stream"}
#     resp = _requests.post(url, headers=hdrs, data=image_bytes, timeout=30)
#     resp.raise_for_status()
#     op_url = resp.headers["Operation-Location"]
#     result = {}
#     for _ in range(30):
#         time.sleep(1.5)
#         r = _requests.get(op_url, headers={"Ocp-Apim-Subscription-Key": key}, timeout=15)
#         result = r.json()
#         if result.get("status") == "succeeded":
#             break

#     words_with_pos = []
#     for page in result.get("analyzeResult", {}).get("pages", []):
#         pw = page.get("width", 1.0) or 1.0
#         ph = page.get("height", 1.0) or 1.0
#         for word in page.get("words", []):
#             text = word.get("content", "").strip()
#             if not text:
#                 continue
#             poly = word.get("polygon", [])
#             x, y = _word_centre(poly)
#             words_with_pos.append({"text": text, "x": x / pw, "y": y / ph})
#     return words_with_pos


# def _ocr_azure_cv(image_bytes: bytes, endpoint: str, key: str) -> list:
#     """Fallback: Azure Computer Vision Read API v3.2."""
#     url = endpoint.rstrip("/") + "/vision/v3.2/read/analyze"
#     hdrs = {"Ocp-Apim-Subscription-Key": key, "Content-Type": "application/octet-stream"}
#     resp = _requests.post(url, headers=hdrs, data=image_bytes, timeout=30)
#     resp.raise_for_status()
#     op_url = resp.headers["Operation-Location"]
#     result = {}
#     for _ in range(30):
#         time.sleep(1.5)
#         r = _requests.get(op_url, headers={"Ocp-Apim-Subscription-Key": key}, timeout=15)
#         result = r.json()
#         if result.get("status") == "succeeded":
#             break

#     words_with_pos = []
#     for page in result.get("analyzeResult", {}).get("readResults", []):
#         pw = page.get("width", 1.0) or 1.0
#         ph = page.get("height", 1.0) or 1.0
#         for line in page.get("lines", []):
#             for word in line.get("words", []):
#                 text = word.get("text", "").strip()
#                 if not text:
#                     continue
#                 bb = word.get("boundingBox", [])
#                 xs = [bb[i] for i in range(0, len(bb), 2)] if bb else [0]
#                 ys = [bb[i] for i in range(1, len(bb), 2)] if bb else [0]
#                 x = ((min(xs)+max(xs))/2.0) / pw
#                 y = ((min(ys)+max(ys))/2.0) / ph
#                 words_with_pos.append({"text": text, "x": x, "y": y})
#     return words_with_pos


# def run_azure_ocr(image_bytes: bytes, endpoint: str, key: str) -> list:
#     """Returns flat list of {"text", "x", "y"}."""
#     try:
#         return _ocr_azure(image_bytes, endpoint, key)
#     except Exception as e:
#         code = getattr(getattr(e, "response", None), "status_code", None)
#         if code in (401, 403, 404):
#             return _ocr_azure_cv(image_bytes, endpoint, key)
#         raise


# # ─── Core: parse OCR word positions → (base_sku, size_mm, qty) rows ──────────

# def parse_order_from_words(
#     words: list,
#     base_to_variants: dict,
#     mrp_lookup: dict,
# ) -> list:
#     """
#     Given a flat list of {"text", "x", "y"} from OCR (normalised 0-1 coords),
#     reconstruct the order form table.

#     Logic:
#     1.  Find the HEADER ROW: the row containing "SKU" or size labels (15MM, 20MM…).
#         Extract the x-centre for each size column.
#     2.  For each subsequent row:
#         a. Find the SKU-CODE token (printed, in the left zone before size columns).
#            This is the BASE SKU.
#         b. Find NUMERIC tokens to the right → each maps to the nearest size column.
#         c. Emit one result per (base_sku, size, qty) where qty > 0.

#     Returns list of dicts with keys:
#       base_sku, size_mm, qty, full_sku, row_label, sheet, mrp, landing, amount
#     """
#     if not words:
#         return []

#     # ── Step 1: Cluster into rows by y-coordinate ───────────────────────────
#     text_rows = _cluster_rows(words, page_height=1.0, gap_frac=0.015)

#     # ── Step 2: Find the header row and calibrate column x-positions ─────────
#     size_col_x: dict = {}   # size_mm → x_centre
#     header_row_idx = None

#     for ri, row in enumerate(text_rows):
#         tokens = [w["text"].strip().upper() for w in row]
#         size_hits = [t for t in tokens if _is_size_header(t)]
#         # Also check for "15MM" / "20MM" written as separate tokens
#         if len(size_hits) >= 2:
#             for w in row:
#                 t = w["text"].strip().upper()
#                 if _is_size_header(t):
#                     size_col_x[t] = w["x"]
#             header_row_idx = ri
#             break

#     # If no size header detected, try a fallback: look for repeated patterns
#     # of "MM" or numeric column headers
#     if not size_col_x:
#         # Try to infer from any row that has ≥3 numeric-looking values with similar y
#         for ri, row in enumerate(text_rows):
#             # Look for row with 3+ purely numeric tokens spread across x-axis
#             nums = [w for w in row if re.match(r'^\d{2,4}$', w["text"].strip())]
#             alpha = [w for w in row if re.match(r'^[A-Z]{2,}$', w["text"].strip().upper())]
#             if len(nums) >= 3 and len(alpha) == 0:
#                 # Treat this as a size-label row: assign dummy size labels
#                 nums_sorted = sorted(nums, key=lambda w: w["x"])
#                 default_sizes = ["15MM", "20MM", "25MM", "32MM", "40MM", "50MM"]
#                 for j, nw in enumerate(nums_sorted):
#                     if j < len(default_sizes):
#                         size_col_x[default_sizes[j]] = nw["x"]
#                 header_row_idx = ri
#                 break

#     # ── Step 3: Determine the x boundary between SKU zone and value zone ─────
#     # SKU codes appear to the LEFT of the first size column
#     if size_col_x:
#         sku_x_boundary = min(size_col_x.values()) - 0.02
#     else:
#         sku_x_boundary = 0.35  # fallback: left 35% of page

#     # ── Step 4: Build a lookup set of known base SKU codes (normalised) ──────
#     known_bases_upper = set(base_to_variants.keys())

#     # ── Step 5: Parse each data row ──────────────────────────────────────────
#     results = []
#     seen = set()  # (base_sku, size_mm) already emitted

#     data_rows = text_rows[header_row_idx + 1:] if header_row_idx is not None else text_rows

#     for row in data_rows:
#         row_sorted = sorted(row, key=lambda w: w["x"])

#         # a) Find SKU/base-code token in the left zone
#         left_tokens = [w for w in row_sorted if w["x"] <= sku_x_boundary]

#         base_sku = None
#         for w in left_tokens:
#             t = w["text"].strip().upper()
#             if t in known_bases_upper:
#                 base_sku = t
#                 break
#             # Fuzzy: partial prefix match
#             if _is_base_sku(t):
#                 for kb in known_bases_upper:
#                     if kb.startswith(t[:6]) or t.startswith(kb[:6]):
#                         base_sku = kb
#                         break
#             if base_sku:
#                 break

#         if not base_sku:
#             continue

#         # b) Find numeric tokens in the right zone (size columns)
#         right_tokens = [w for w in row_sorted if w["x"] > sku_x_boundary]
#         qty_tokens   = [w for w in right_tokens if re.match(r'^\d+$', w["text"].strip())]

#         if not qty_tokens:
#             continue

#         variants = base_to_variants.get(base_sku, [])
#         variant_size_map = {v["size_mm"]: v for v in variants}

#         for qt in qty_tokens:
#             qty_val = int(qt["text"].strip())
#             if qty_val == 0:
#                 continue

#             # c) Map this qty token to the nearest size column
#             if size_col_x:
#                 size_mm = _closest_column(qt["x"], size_col_x)
#             else:
#                 # If no header calibration, assign sizes in left-to-right order
#                 qty_sorted_x = sorted(qty_tokens, key=lambda w: w["x"])
#                 col_idx = qty_sorted_x.index(qt)
#                 default_sizes = ["15MM", "20MM", "25MM", "32MM", "40MM", "50MM"]
#                 size_mm = default_sizes[col_idx] if col_idx < len(default_sizes) else None

#             if size_mm is None:
#                 continue

#             # d) Find the full SKU for this (base_sku, size_mm)
#             variant = variant_size_map.get(size_mm)
#             if variant is None:
#                 # Try case-insensitive match
#                 for k, v in variant_size_map.items():
#                     if k.upper() == size_mm.upper():
#                         variant = v
#                         break

#             dedup_key = (base_sku, size_mm)
#             if dedup_key in seen:
#                 continue
#             seen.add(dedup_key)

#             full_sku = variant["full_sku"] if variant else None
#             row_label = variant["row_label"] if variant else base_sku
#             sheet     = variant["sheet"]     if variant else "—"
#             size_lbl  = variant["size_label"] if variant else size_mm

#             mrp_data  = mrp_lookup.get(full_sku, {}) if full_sku else {}
#             mrp_val   = mrp_data.get("mrp",     0.0)
#             land_val  = mrp_data.get("landing", 0.0)
#             desc      = mrp_data.get("description", row_label)

#             results.append({
#                 "base_sku":   base_sku,
#                 "size_mm":    size_mm,
#                 "size_label": size_lbl,
#                 "full_sku":   full_sku or "—",
#                 "row_label":  row_label,
#                 "description": desc,
#                 "sheet":      sheet,
#                 "mrp":        mrp_val,
#                 "landing":    land_val,
#                 "qty":        qty_val,
#                 "amount":     round(land_val * qty_val, 2),
#                 "matched":    full_sku is not None and mrp_val > 0,
#             })

#     return results


# # ─── PDF generator ────────────────────────────────────────────────────────────
# def generate_pdf(order_lines: list, bill_to: dict, ship_to: dict, dealer_info: dict) -> bytes:
#     if not _HAS_REPORTLAB:
#         raise RuntimeError("reportlab not installed")

#     buf    = io.BytesIO()
#     doc    = SimpleDocTemplate(buf, pagesize=A4,
#                                leftMargin=10*mm, rightMargin=10*mm,
#                                topMargin=12*mm, bottomMargin=12*mm)
#     navy  = colors.HexColor("#0A2342")
#     sky   = colors.HexColor("#1565C0")
#     story = []

#     def _para(text, **kw):
#         s = ParagraphStyle("x", **kw)
#         return Paragraph(text, s)

#     # Header
#     hdr = Table([[
#         _para("SINTEX BAPL LIMITED", fontName="Helvetica-Bold", fontSize=14,
#               textColor=colors.white, alignment=TA_CENTER),
#         _para("Kutesar Road, Raipur, Chhattisgarh – 492101<br/>GSTIN: 22AADCB1921F1ZE",
#               fontName="Helvetica", fontSize=8, textColor=colors.white, alignment=TA_CENTER),
#     ]], colWidths=["40%", "60%"])
#     hdr.setStyle(TableStyle([
#         ("BACKGROUND",   (0,0),(-1,-1), navy),
#         ("VALIGN",       (0,0),(-1,-1), "MIDDLE"),
#         ("TOPPADDING",   (0,0),(-1,-1), 8),
#         ("BOTTOMPADDING",(0,0),(-1,-1), 8),
#         ("LEFTPADDING",  (0,0),(-1,-1), 10),
#     ]))
#     story.append(hdr)
#     story.append(Spacer(1, 4*mm))

#     # Title
#     qt = Table([[
#         _para("<b>QUOTATION</b>", fontName="Helvetica-Bold", fontSize=13, textColor=navy),
#         _para(f"<b>Date:</b> {date.today().strftime('%d-%m-%Y')}",
#               fontName="Helvetica", fontSize=8, alignment=TA_RIGHT),
#     ]], colWidths=["60%","40%"])
#     qt.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"MIDDLE"),
#                              ("LINEBELOW",(0,0),(-1,-1),0.5,navy)]))
#     story.append(qt)
#     story.append(Spacer(1, 3*mm))

#     def party_block(d, title):
#         lines = [f"<b>{title}</b>"] + [f"<b>{k}:</b> {v}" for k,v in d.items() if v]
#         return _para("<br/>".join(lines), fontName="Helvetica", fontSize=7.5, leading=11)

#     if any(dealer_info.values()):
#         dt = Table([[party_block(dealer_info, "DEALER / DISTRIBUTOR")]], colWidths=["100%"])
#         dt.setStyle(TableStyle([
#             ("BOX",(0,0),(-1,-1),0.5,navy), ("VALIGN",(0,0),(-1,-1),"TOP"),
#             ("TOPPADDING",(0,0),(-1,-1),5), ("BOTTOMPADDING",(0,0),(-1,-1),5),
#             ("LEFTPADDING",(0,0),(-1,-1),8),
#             ("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#EBF3FF")),
#         ]))
#         story.append(dt)
#         story.append(Spacer(1, 3*mm))

#     pt = Table([[party_block(bill_to,"BILL TO PARTY"), party_block(ship_to,"SHIP TO PARTY")]],
#                colWidths=["50%","50%"])
#     pt.setStyle(TableStyle([
#         ("BOX",(0,0),(-1,-1),0.5,navy),
#         ("INNERGRID",(0,0),(-1,-1),0.5,colors.HexColor("#C5D0E0")),
#         ("VALIGN",(0,0),(-1,-1),"TOP"),
#         ("TOPPADDING",(0,0),(-1,-1),6), ("BOTTOMPADDING",(0,0),(-1,-1),6),
#         ("LEFTPADDING",(0,0),(-1,-1),8),
#         ("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#F4F6FA")),
#     ]))
#     story.append(pt)
#     story.append(Spacer(1, 5*mm))

#     # Line items table
#     hdr_row = ["#","SKU Code","Description","Size","MRP (₹)","Qty","Rate (₹)","Amount (₹)"]
#     tbl_rows = [hdr_row]
#     grand_mrp = grand_land = 0.0
#     for sno, r in enumerate(order_lines, 1):
#         if r["qty"] <= 0 or r["full_sku"] == "—":
#             continue
#         tbl_rows.append([
#             str(sno),
#             _para(f'<font name="Courier" size="6.5">{r["full_sku"]}</font>',
#                   fontName="Helvetica", fontSize=7),
#             _para(r["description"][:60], fontName="Helvetica", fontSize=7),
#             r["size_label"],
#             f"{r['mrp']:,.2f}",
#             str(r["qty"]),
#             f"{r['landing']:,.2f}",
#             f"{r['amount']:,.2f}",
#         ])
#         grand_mrp  += r["mrp"]     * r["qty"]
#         grand_land += r["landing"] * r["qty"]

#     n = len(tbl_rows)
#     tbl_rows.append(["","","","","","",
#                      _para("<b>GRAND TOTAL</b>", fontName="Helvetica-Bold", fontSize=8),
#                      _para(f"<b>₹ {grand_land:,.2f}</b>",
#                            fontName="Helvetica-Bold", fontSize=8, alignment=TA_RIGHT)])

#     lt = Table(tbl_rows,
#                colWidths=[8*mm, 36*mm, 55*mm, 16*mm, 18*mm, 12*mm, 18*mm, 22*mm],
#                repeatRows=1)
#     lt.setStyle(TableStyle([
#         ("BACKGROUND",     (0,0),(-1,0),  navy),
#         ("TEXTCOLOR",      (0,0),(-1,0),  colors.white),
#         ("FONTNAME",       (0,0),(-1,0),  "Helvetica-Bold"),
#         ("FONTSIZE",       (0,0),(-1,0),  7.5),
#         ("ALIGN",          (0,0),(-1,-1), "CENTER"),
#         ("ALIGN",          (2,1),(2,-1),  "LEFT"),
#         ("FONTSIZE",       (0,1),(-1,-1), 7),
#         ("ROWBACKGROUNDS", (0,1),(-1,-2), [colors.white, colors.HexColor("#F4F8FF")]),
#         ("GRID",           (0,0),(-1,-1), 0.35, colors.HexColor("#C5D0E0")),
#         ("TOPPADDING",     (0,0),(-1,-1), 3),
#         ("BOTTOMPADDING",  (0,0),(-1,-1), 3),
#         ("BACKGROUND",     (0,n-1),(-1,n-1), colors.HexColor("#EBF3FF")),
#         ("LINEABOVE",      (0,n-1),(-1,n-1), 1, navy),
#     ]))
#     story.append(lt)
#     story.append(Spacer(1, 6*mm))
#     story.append(HRFlowable(width="100%", thickness=0.5, color=navy))
#     story.append(Spacer(1, 2*mm))
#     story.append(_para(
#         "This is a computer-generated quotation. Prices are subject to change without notice. "
#         "Taxes applicable as per prevailing rates.",
#         fontName="Helvetica", fontSize=7, textColor=colors.grey, alignment=TA_CENTER,
#     ))
#     doc.build(story)
#     return buf.getvalue()


# # ─── Session-state defaults ───────────────────────────────────────────────────
# DEFAULTS = {
#     "step":           1,
#     "image_bytes":    None,
#     "ocr_words":      [],    # raw word list from Azure OCR
#     "order_lines":    [],    # parsed + resolved order lines
#     "ocr_done":       False,
#     "bill_to":        {},
#     "ship_to":        {},
#     "dealer_info":    {},
#     "pdf_bytes":      None,
#     "azure_endpoint": "",
#     "azure_key":      "",
#     "_cam_recv_val":  "",
# }
# for k, v in DEFAULTS.items():
#     if k not in st.session_state:
#         st.session_state[k] = v

# mrp_lookup              = load_mrp()
# zsd_df                  = load_zsd()
# full_sku_lookup, base_to_variants, sheets_info = build_sku_grid()


# # ─── UI helpers ──────────────────────────────────────────────────────────────
# def badge(n, done=False):
#     cls = "step-badge done" if done else "step-badge"
#     return f'<span class="{cls}">{n}</span>'

# def party_html(d, title):
#     rows = "".join(
#         f'<div class="party-row"><span class="party-lbl">{k}</span>'
#         f'<span class="party-val">{v or "—"}</span></div>'
#         for k, v in d.items()
#     )
#     return f'<div class="party-box"><h4>{title}</h4>{rows}</div>'


# # ─── App header ──────────────────────────────────────────────────────────────
# st.markdown("""
# <div class="app-header">
#   <div>
#     <h1>🔧 Sintex BAPL – Quotation Generator</h1>
#     <p>CPVC / UPVC Pipes &amp; Fittings · Chhattisgarh Price List</p>
#   </div>
# </div>
# """, unsafe_allow_html=True)

# step = st.session_state.step
# cols_steps = st.columns(4)
# STEP_LABELS = ["1  Party Details", "2  Capture & OCR", "3  Review & Edit", "4  Download"]
# for i, (col, lbl) in enumerate(zip(cols_steps, STEP_LABELS), 1):
#     done   = step > i
#     active = step == i
#     bg = "#0A2342" if active else ("#00796B" if done else "#DEE3EC")
#     tc = "#fff" if (active or done) else "#5A6880"
#     col.markdown(f"""
#     <div style="background:{bg};color:{tc};border-radius:8px;
#                 padding:8px 12px;text-align:center;font-size:12px;font-weight:600;">
#       {'✓ ' if done else ''}{lbl}
#     </div>""", unsafe_allow_html=True)

# st.markdown("<div style='height:18px'></div>", unsafe_allow_html=True)


# # ═══════════════════════════════════════════════════════════════════════════════
# # STEP 1 — Party Details
# # ═══════════════════════════════════════════════════════════════════════════════
# if step == 1:
#     st.markdown(f"""
#     <div class="card">
#       <div class="card-title">{badge(1)} Dealer / Distributor & Customer Details</div>
#     </div>""", unsafe_allow_html=True)

#     st.markdown("### 🏪 Dealer / Distributor Information")
#     da1, da2 = st.columns(2)
#     dealer_name  = da1.text_input("Dealer / Distributor Name", key="dealer_name")
#     dealer_code  = da2.text_input("Dealer Code",               key="dealer_code")
#     dealer_addr  = st.text_input("Address",                    key="dealer_address")
#     da3, da4 = st.columns(2)
#     dealer_phone = da3.text_input("Phone",   key="dealer_phone")
#     dealer_gst   = da4.text_input("GST No.", key="dealer_gst")
#     da5, da6 = st.columns(2)
#     dealer_state = da5.text_input("State",   key="dealer_state")
#     dealer_pan   = da6.text_input("PAN No.", key="dealer_pan")

#     st.session_state.dealer_info = {
#         "Name":    dealer_name, "Code":    dealer_code,
#         "Address": dealer_addr, "Phone":   dealer_phone,
#         "GST No.": dealer_gst,  "State":   dealer_state, "PAN No.": dealer_pan,
#     }

#     st.markdown("---")
#     st.markdown("### 👤 Customer Information")

#     cust_options = ["— Select or type below —"] + [
#         f'{row["Customer Code"]} | {row["Customer Name"]}'
#         for _, row in zsd_df.iterrows()
#     ]
#     sel = st.selectbox("🔍 Search Customer from Master", cust_options, key="zsd_search")

#     def zsd_fill(prefix, row):
#         addr = " ".join(filter(None, [
#             str(row.get("Address 1","") or ""), str(row.get("Address 2","") or ""),
#             str(row.get("Address 3","") or ""), str(row.get("City","") or ""),
#         ]))
#         return {
#             f"{prefix}_party_no":   str(row.get("Customer Code","") or ""),
#             f"{prefix}_party_name": str(row.get("Customer Name","")  or ""),
#             f"{prefix}_address":    addr.strip(),
#             f"{prefix}_phone":      str(row.get("Telephone","")     or ""),
#             f"{prefix}_mobile":     str(row.get("Mobile No.","")    or ""),
#             f"{prefix}_state_code": str(row.get("State Code","")    or ""),
#             f"{prefix}_state":      str(row.get("State Code Desc.","") or ""),
#             f"{prefix}_gst":        str(row.get("GST Number","")    or ""),
#             f"{prefix}_pan":        str(row.get("PAN No.","")       or ""),
#         }

#     if sel != cust_options[0]:
#         code = sel.split("|")[0].strip()
#         matched = zsd_df[zsd_df["Customer Code"].astype(str) == code]
#         if not matched.empty:
#             row = matched.iloc[0]
#             for k, v in zsd_fill("bill", row).items():
#                 if k not in st.session_state or not st.session_state[k]:
#                     st.session_state[k] = v
#             st.toast("Bill-to details filled from customer master.", icon="✅")

#     st.markdown("#### 📋 Bill To Party")
#     bc1, bc2 = st.columns(2)
#     bill_party_no   = bc1.text_input("Bill to Party No.",  key="bill_party_no")
#     bill_party_name = bc2.text_input("Bill to Party Name", key="bill_party_name")
#     bill_address    = st.text_input("Bill to Address",     key="bill_address")
#     bc3, bc4 = st.columns(2)
#     bill_phone  = bc3.text_input("Phone",  key="bill_phone")
#     bill_mobile = bc4.text_input("Mobile", key="bill_mobile")
#     bc5, bc6 = st.columns(2)
#     bill_sc    = bc5.text_input("State Code", key="bill_state_code")
#     bill_state = bc6.text_input("State",      key="bill_state")
#     bc7, bc8 = st.columns(2)
#     bill_gst = bc7.text_input("GST No.", key="bill_gst")
#     bill_pan = bc8.text_input("PAN No.", key="bill_pan")

#     same_as_bill = st.checkbox("Ship-to same as Bill-to", value=False, key="same_as_bill_chk")

#     st.markdown("#### 🚚 Ship To Party")
#     if same_as_bill:
#         for k, v in [("ship_party_no", bill_party_no), ("ship_party_name", bill_party_name),
#                      ("ship_address", bill_address),   ("ship_phone", bill_phone),
#                      ("ship_mobile", bill_mobile),     ("ship_state_code", bill_sc),
#                      ("ship_state", bill_state),       ("ship_gst", bill_gst),
#                      ("ship_pan", bill_pan)]:
#             st.session_state[k] = v

#     sc1, sc2 = st.columns(2)
#     ship_party_no   = sc1.text_input("Ship to Party No.",  key="ship_party_no")
#     ship_party_name = sc2.text_input("Ship to Party Name", key="ship_party_name")
#     ship_address    = st.text_input("Ship to Address",     key="ship_address")
#     sc3, sc4 = st.columns(2)
#     ship_phone  = sc3.text_input("Phone ",  key="ship_phone")
#     ship_mobile = sc4.text_input("Mobile ", key="ship_mobile")
#     sc5, sc6 = st.columns(2)
#     ship_sc    = sc5.text_input("State Code ", key="ship_state_code")
#     ship_state = sc6.text_input("State ",      key="ship_state")
#     sc7, sc8 = st.columns(2)
#     ship_gst = sc7.text_input("GST No. ", key="ship_gst")
#     ship_pan = sc8.text_input("PAN No. ", key="ship_pan")

#     st.session_state.bill_to = {
#         "Party No.": bill_party_no,  "Name":       bill_party_name,
#         "Address":   bill_address,   "Phone":      bill_phone,
#         "Mobile":    bill_mobile,    "State Code": bill_sc,
#         "State":     bill_state,     "GST No.":    bill_gst,    "PAN No.": bill_pan,
#     }
#     st.session_state.ship_to = {
#         "Party No.": ship_party_no,  "Name":       ship_party_name,
#         "Address":   ship_address,   "Phone":      ship_phone,
#         "Mobile":    ship_mobile,    "State Code": ship_sc,
#         "State":     ship_state,     "GST No.":    ship_gst,    "PAN No.": ship_pan,
#     }

#     st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)
#     if st.button("▶  Next: Capture & OCR", key="go_step2"):
#         st.session_state.step = 2
#         st.rerun()


# # ═══════════════════════════════════════════════════════════════════════════════
# # STEP 2 — Capture Image & Run OCR
# # ═══════════════════════════════════════════════════════════════════════════════
# elif step == 2:
#     st.markdown(f"""
#     <div class="card">
#       <div class="card-title">{badge(2)} Capture Order Form & Run OCR</div>
#     </div>""", unsafe_allow_html=True)

#     st.markdown(
#         "📸 **Photograph or upload** the handwritten order form. "
#         "The system will detect the printed SKU codes and handwritten quantities, "
#         "then automatically calculate pricing from the Chhattisgarh price list."
#     )

#     with st.expander("🔑 Azure OCR Credentials (required)", expanded=True):
#         ep  = st.text_input("Azure Endpoint", value=st.session_state.azure_endpoint,
#                             placeholder="https://YOUR_RESOURCE.cognitiveservices.azure.com",
#                             key="az_ep_input")
#         key = st.text_input("Azure Key", value=st.session_state.azure_key,
#                             type="password", placeholder="32-character subscription key",
#                             key="az_key_input")
#         st.session_state.azure_endpoint = ep
#         st.session_state.azure_key      = key

#     img_mode = st.radio("Image source",
#                         ["📷  Camera (recommended)", "📁  Upload File"],
#                         horizontal=True, label_visibility="collapsed", key="img_mode_radio")

#     def _on_file_upload():
#         uf = st.session_state.get("file_upload_input")
#         if uf is not None:
#             st.session_state.image_bytes = uf.getvalue()
#             st.session_state.ocr_done    = False
#             st.session_state.ocr_words   = []
#             st.session_state.order_lines = []

#     if img_mode == "📷  Camera (recommended)":
#         import streamlit.components.v1 as components

#         CAMERA_HTML = """
# <style>
# * { box-sizing:border-box; margin:0; padding:0; }
# body { background:transparent; font-family:'IBM Plex Sans',sans-serif; }
# #cam-wrap {
#   width:100%; background:linear-gradient(160deg,#071829 0%,#0d2d56 100%);
#   border-radius:14px; overflow:hidden; display:flex; flex-direction:column;
#   align-items:center; box-shadow:0 8px 32px rgba(10,35,66,.4);
# }
# #video  { width:100%; max-height:62vh; object-fit:cover; display:block; }
# #canvas { display:none; }
# #preview{ width:100%; display:none; border-bottom:3px solid #1E88E5; }
# .toolbar{ display:flex; gap:12px; padding:16px 20px 18px; width:100%;
#   background:rgba(0,0,0,.3); justify-content:center; flex-wrap:wrap; }
# .cam-btn{ display:inline-flex; align-items:center; justify-content:center; gap:8px;
#   padding:12px 32px; border:none; border-radius:10px; font-size:14px; font-weight:700;
#   cursor:pointer; transition:all .18s; letter-spacing:.3px; min-width:160px;
#   font-family:'IBM Plex Sans',sans-serif; }
# #btn-capture{ background:linear-gradient(135deg,#1565C0 0%,#1E88E5 100%); color:#fff;
#   box-shadow:0 4px 16px rgba(21,101,192,.5); }
# #btn-capture:hover{ transform:translateY(-2px); }
# #btn-retake { background:rgba(255,255,255,.1); color:rgba(255,255,255,.9);
#   border:1.5px solid rgba(255,255,255,.25); display:none; }
# .status-bar{ display:flex; align-items:center; gap:10px;
#   background:rgba(0,0,0,.25); padding:10px 20px; width:100%;
#   font-size:12.5px; font-weight:500; color:rgba(255,255,255,.7); min-height:42px; }
# .pulse-dot { width:8px; height:8px; border-radius:50%; background:currentColor;
#   flex-shrink:0; animation:pulse 1.5s infinite ease-in-out; }
# @keyframes pulse{ 0%,100%{opacity:1} 50%{opacity:.3} }
# .spinner{ width:16px; height:16px; flex-shrink:0; border:2.5px solid currentColor;
#   border-top-color:transparent; border-radius:50%; animation:spin .65s linear infinite; }
# @keyframes spin{ to{ transform:rotate(360deg); } }
# </style>
# <div id="cam-wrap">
#   <video id="video" autoplay playsinline muted></video>
#   <canvas id="canvas"></canvas>
#   <img id="preview" alt="captured photo"/>
#   <div class="toolbar">
#     <button class="cam-btn" id="btn-capture">📷&nbsp;&nbsp;Capture Photo</button>
#     <button class="cam-btn" id="btn-retake">🔄&nbsp;&nbsp;Retake</button>
#   </div>
#   <div class="status-bar" id="status-bar">
#     <span class="pulse-dot"></span>
#     <span id="status-txt">Starting camera…</span>
#   </div>
# </div>
# <script>
# (function(){
#   const video=document.getElementById('video'),canvas=document.getElementById('canvas'),
#         preview=document.getElementById('preview'),btnCap=document.getElementById('btn-capture'),
#         btnRet=document.getElementById('btn-retake'),txt=document.getElementById('status-txt');
#   async function startCamera(){
#     try{
#       const s=await navigator.mediaDevices.getUserMedia(
#         {video:{facingMode:{ideal:'environment'},width:{ideal:4096},height:{ideal:3072}},audio:false});
#       video.srcObject=s; await video.play();
#       txt.textContent='Camera ready — frame the order form and tap Capture';
#     }catch(e){txt.textContent='Camera error: '+e.message;}
#   }
#   btnCap.addEventListener('click',()=>{
#     const w=video.videoWidth||1280,h=video.videoHeight||720;
#     canvas.width=w; canvas.height=h;
#     canvas.getContext('2d').drawImage(video,0,0,w,h);
#     const b64=canvas.toDataURL('image/jpeg',0.97);
#     preview.src=b64; video.style.display='none';
#     preview.style.display='block'; btnCap.style.display='none'; btnRet.style.display='block';
#     txt.textContent='Photo captured — proceed to OCR';
#     window.parent.sessionStorage.setItem('sintex_cam_b64',b64);
#     window.parent.postMessage({type:'SINTEX_CAM_CAPTURE',data:b64},'*');
#   });
#   btnRet.addEventListener('click',()=>{
#     preview.style.display='none'; video.style.display='block';
#     btnCap.style.display='block'; btnRet.style.display='none';
#     window.parent.sessionStorage.removeItem('sintex_cam_b64');
#     txt.textContent='Camera ready — frame the order form and tap Capture';
#   });
#   startCamera();
# })();
# </script>
# """
#         components.html(CAMERA_HTML, height=580, scrolling=False)

#         AUTO_BRIDGE = """
# <script>
# (function(){
#   var attempts=0;
#   function tryInject(){
#     if(++attempts>120) return;
#     var b64=window.parent.sessionStorage.getItem('sintex_cam_b64');
#     if(!b64){setTimeout(tryInject,500);return;}
#     var inputs=window.parent.document.querySelectorAll('input[type="text"]');
#     var target=null;
#     for(var i=inputs.length-1;i>=0;i--){
#       var st=window.parent.getComputedStyle(inputs[i]);
#       if(parseFloat(st.height)<5||parseFloat(st.opacity)<0.1){target=inputs[i];break;}
#     }
#     if(!target&&inputs.length) target=inputs[inputs.length-1];
#     if(!target){setTimeout(tryInject,500);return;}
#     var setter=Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype,'value').set;
#     setter.call(target,b64);
#     target.dispatchEvent(new Event('input',{bubbles:true}));
#     target.dispatchEvent(new Event('change',{bubbles:true}));
#     window.parent.sessionStorage.removeItem('sintex_cam_b64');
#   }
#   setTimeout(tryInject,800);
# })();
# </script>
# """
#         components.html(AUTO_BRIDGE, height=0)

#         cam_val = st.text_input(
#             "sintex_cam_hidden",
#             value=st.session_state.get("_cam_recv_val", ""),
#             label_visibility="collapsed",
#             key="sintex_cam_recv",
#         )
#         recv = st.session_state.get("sintex_cam_recv", "")
#         if recv and recv.startswith("data:image"):
#             _, encoded = recv.split(",", 1)
#             raw = base64.b64decode(encoded)
#             if raw != st.session_state.image_bytes:
#                 st.session_state.image_bytes  = raw
#                 st.session_state.ocr_done     = False
#                 st.session_state.ocr_words    = []
#                 st.session_state.order_lines  = []
#                 st.session_state["_cam_recv_val"] = recv
#                 st.rerun()

#         components.html("""
# <script>
# (function(){
#   var inputs=window.parent.document.querySelectorAll('input[type="text"]');
#   if(inputs.length){
#     var last=inputs[inputs.length-1];
#     last.style.cssText+='height:1px!important;opacity:0!important;pointer-events:none!important;position:absolute!important;';
#     var wrap=last.closest('[data-testid="stTextInput"]');
#     if(wrap)wrap.style.cssText+='height:0!important;overflow:hidden!important;margin:0!important;padding:0!important;';
#   }
# })();
# </script>
# """, height=0)

#     else:
#         st.file_uploader("Upload image of order form", type=["jpg","jpeg","png"],
#                          label_visibility="collapsed",
#                          key="file_upload_input", on_change=_on_file_upload)

#     # Image preview
#     if st.session_state.image_bytes:
#         st.markdown("""<div style='background:#fff;border:1px solid #DEE3EC;border-radius:10px;
#                     padding:12px;margin:14px 0 6px;'>
#           <div style='font-size:11px;font-weight:700;text-transform:uppercase;
#                       letter-spacing:.5px;color:#0A2342;margin-bottom:10px;'>
#             📸 Captured Image</div>""", unsafe_allow_html=True)
#         st.image(st.session_state.image_bytes, use_container_width=True)
#         st.markdown("</div>", unsafe_allow_html=True)
#     else:
#         st.markdown("""<div style='background:#F4F6FA;border:2px dashed #DEE3EC;border-radius:10px;
#                     padding:40px;text-align:center;margin:14px 0;color:#5A6880;'>
#           <div style='font-size:36px;margin-bottom:10px;'>🖼️</div>
#           <div style='font-weight:600;font-size:14px;margin-bottom:4px;'>No image loaded</div>
#           <div style='font-size:12px;'>Capture or upload an image above</div>
#         </div>""", unsafe_allow_html=True)

#     # OCR status
#     if st.session_state.ocr_done:
#         n = len(st.session_state.order_lines)
#         matched = sum(1 for r in st.session_state.order_lines if r["matched"])
#         if n > 0:
#             st.success(
#                 f"✅ OCR complete — **{n} order line(s)** detected, "
#                 f"**{matched} matched** to catalogue with pricing. "
#                 "Review and edit in the next step.", icon="🔍"
#             )
#         else:
#             st.warning(
#                 "OCR ran but could not parse any order lines. "
#                 "Check that the image clearly shows the form with printed SKU codes "
#                 "and handwritten quantities. You can also add items manually in Step 3.",
#                 icon="⚠️"
#             )

#     # ── How OCR works explanation ─────────────────────────────────────────────
#     with st.expander("ℹ️ How does OCR work here?"):
#         st.markdown("""
# **The form parsing logic:**

# 1. Azure OCR scans the image and returns all text with x/y coordinates.
# 2. The system finds the **header row** containing size labels (`15MM`, `20MM`, `25MM`, `32MM`, `40MM`, `50MM`) and records the x-position of each column.
# 3. For each product row, it identifies the **printed base SKU code** (e.g. `CPF11BV000` for Ball Valve) in the left zone.
# 4. It then reads **handwritten numbers** to the right of the SKU, assigning each number to the nearest size column by x-position.
# 5. Using the SKU catalogue (Excel file), it maps `base_sku + size` → **full SKU** (e.g. `CPF11BV00000015`).
# 6. Pricing is looked up from the Chhattisgarh MRP price list CSV.

# **For best results:** Take a well-lit, flat photo with the full form visible. Ensure the printed SKU codes and handwritten numbers are both clearly readable.
#         """)

#     st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)
#     btn_cols = st.columns([1, 1, 1, 3])

#     with btn_cols[0]:
#         if st.button("◀  Back", key="back_step1"):
#             st.session_state.step = 1
#             st.rerun()

#     with btn_cols[1]:
#         if st.button("🔍  Run OCR", key="run_ocr_btn"):
#             if not st.session_state.image_bytes:
#                 st.error("Please upload or capture an image first.", icon="🖼️")
#             else:
#                 _ep  = st.session_state.azure_endpoint.strip()
#                 _key = st.session_state.azure_key.strip()
#                 if not _ep or not _key:
#                     st.error("Please enter your **Azure Endpoint** and **Azure Key** above.", icon="🔑")
#                 else:
#                     with st.spinner("Sending image to Azure OCR… 10–20 seconds"):
#                         try:
#                             words = run_azure_ocr(
#                                 st.session_state.image_bytes, _ep, _key
#                             )
#                             lines = parse_order_from_words(
#                                 words, base_to_variants, mrp_lookup
#                             )
#                             st.session_state.ocr_words   = words
#                             st.session_state.order_lines = lines
#                             st.session_state.ocr_done    = True
#                             st.rerun()
#                         except Exception as exc:
#                             st.error(f"OCR failed: {exc}", icon="❌")

#     with btn_cols[2]:
#         if st.button("▶  Next: Review & Edit", key="go_step3"):
#             st.session_state.step = 3
#             st.rerun()


# # ═══════════════════════════════════════════════════════════════════════════════
# # STEP 3 — Review & Edit Order Lines
# # ═══════════════════════════════════════════════════════════════════════════════
# elif step == 3:
#     st.markdown(f"""
#     <div class="card">
#       <div class="card-title">{badge(3, step > 3)} Review & Edit Order Lines</div>
#     </div>""", unsafe_allow_html=True)

#     order_lines = st.session_state.get("order_lines", [])

#     tabs = st.tabs(["📋 OCR Results / Edit", "➕ Add Items Manually"])

#     # ── Tab 1: Edit OCR results ────────────────────────────────────────────────
#     with tabs[0]:
#         if order_lines:
#             st.info(
#                 "Each row = one (product, size) combination detected from the form. "
#                 "Edit quantities or delete rows as needed.", icon="📋"
#             )

#             HDR = ["Full SKU Code", "Product", "Size", "MRP (₹)", "Landing (₹)", "Qty", "Amount (₹)", "Del"]
#             WIDTHS = [1.8, 2.5, 0.9, 1.1, 1.1, 0.7, 1.2, 0.3]
#             hdr_cols = st.columns(WIDTHS)
#             for col, h in zip(hdr_cols, HDR):
#                 col.markdown(
#                     f"<div style='font-size:11px;font-weight:700;color:#0A2342;"
#                     f"background:#EBF3FF;border-radius:4px;padding:3px 6px;text-align:center;'>"
#                     f"{h}</div>", unsafe_allow_html=True
#                 )

#             rows_to_keep = []
#             delete_indices = set()

#             for idx, r in enumerate(order_lines):
#                 cols = st.columns(WIDTHS)

#                 # Full SKU (editable)
#                 new_sku = cols[0].text_input(
#                     f"sku_{idx}", value=r["full_sku"] if r["full_sku"] != "—" else "",
#                     label_visibility="collapsed", key=f"sku_{idx}"
#                 )

#                 # Product name (display)
#                 cols[1].markdown(
#                     f'<div style="font-size:11px;padding:6px 3px;line-height:1.4;">'
#                     f'<b>{r["row_label"][:28]}</b></div>', unsafe_allow_html=True
#                 )

#                 # Size (display)
#                 cols[2].markdown(
#                     f'<div style="font-size:11px;padding:6px 3px;text-align:center;">'
#                     f'{r["size_label"]}</div>', unsafe_allow_html=True
#                 )

#                 # Re-resolve if SKU changed
#                 full_sku_key = new_sku.strip().upper() if new_sku.strip() else r["full_sku"]
#                 mrp_data  = mrp_lookup.get(full_sku_key, {})
#                 mrp_val   = mrp_data.get("mrp",     r["mrp"])
#                 land_val  = mrp_data.get("landing", r["landing"])

#                 cols[3].markdown(
#                     f'<div style="font-size:11px;padding:6px 3px;font-family:monospace;text-align:right;">'
#                     f'₹{mrp_val:,.2f}</div>', unsafe_allow_html=True
#                 )
#                 cols[4].markdown(
#                     f'<div style="font-size:11px;padding:6px 3px;font-family:monospace;text-align:right;">'
#                     f'₹{land_val:,.2f}</div>', unsafe_allow_html=True
#                 )

#                 # Qty (editable)
#                 new_qty = cols[5].number_input(
#                     f"qty_{idx}", value=int(r["qty"]), min_value=0, step=1,
#                     label_visibility="collapsed", key=f"qty_{idx}"
#                 )

#                 amt = round(land_val * new_qty, 2)
#                 cols[6].markdown(
#                     f'<div style="font-size:11px;padding:6px 3px;font-family:monospace;text-align:right;'
#                     f'font-weight:600;">₹{amt:,.2f}</div>', unsafe_allow_html=True
#                 )

#                 # Delete checkbox
#                 del_it = cols[7].checkbox("", key=f"del_{idx}", label_visibility="collapsed")
#                 if del_it:
#                     delete_indices.add(idx)

#                 rows_to_keep.append({**r,
#                     "full_sku": full_sku_key,
#                     "mrp":      mrp_val,
#                     "landing":  land_val,
#                     "qty":      new_qty,
#                     "amount":   amt,
#                     "matched":  mrp_val > 0,
#                 })

#             # Apply deletions
#             updated = [r for i, r in enumerate(rows_to_keep) if i not in delete_indices]
#             st.session_state.order_lines = updated

#             if delete_indices:
#                 if st.button("🗑 Remove selected rows", key="del_rows"):
#                     st.rerun()

#         else:
#             st.info("No OCR results yet. Use the 'Add Items Manually' tab or go back to run OCR.", icon="📋")

#     # ── Tab 2: Manual entry ────────────────────────────────────────────────────
#     with tabs[1]:
#         st.markdown("**Add items by selecting from the catalogue:**")

#         # Sheet selector
#         sheet_names = [s["name"] for s in sheets_info]
#         sel_sheet = st.selectbox("Product Sheet", sheet_names, key="man_sheet")
#         sheet_obj = next((s for s in sheets_info if s["name"] == sel_sheet), None)

#         if sheet_obj:
#             # Product row selector
#             prod_rows = [r for r in sheet_obj["rows"] if r["base_sku"]]
#             prod_labels = [f'{r["row_label"]} ({r["base_sku"]})' for r in prod_rows]
#             if prod_labels:
#                 sel_prod_idx = st.selectbox("Product", range(len(prod_labels)),
#                                             format_func=lambda i: prod_labels[i],
#                                             key="man_prod")
#                 sel_prod = prod_rows[sel_prod_idx]
#                 base_sku = sel_prod["base_sku"]

#                 # Size selector
#                 variants = base_to_variants.get(base_sku, [])
#                 if variants:
#                     size_labels = [f'{v["size_mm"]} ({v["size_label"]}) — {v["full_sku"]}' for v in variants]
#                     sel_size_idx = st.selectbox("Size", range(len(size_labels)),
#                                                 format_func=lambda i: size_labels[i],
#                                                 key="man_size")
#                     sel_var = variants[sel_size_idx]
#                     full_sku = sel_var["full_sku"]
#                     mrp_data = mrp_lookup.get(full_sku, {})
#                     mrp_v    = mrp_data.get("mrp",     0.0)
#                     land_v   = mrp_data.get("landing", 0.0)

#                     st.markdown(
#                         f"**SKU:** `{full_sku}` | **MRP:** ₹{mrp_v:,.2f} | **Landing:** ₹{land_v:,.2f}"
#                     )
#                     man_qty = st.number_input("Quantity", min_value=1, value=1, step=1, key="man_qty")

#                     if st.button("➕ Add to Order", key="add_manual"):
#                         # Check for duplicate
#                         existing = [r for r in st.session_state.order_lines
#                                     if r["full_sku"] == full_sku]
#                         if existing:
#                             for r in st.session_state.order_lines:
#                                 if r["full_sku"] == full_sku:
#                                     r["qty"]    += man_qty
#                                     r["amount"] = round(r["landing"] * r["qty"], 2)
#                             st.success(f"Updated qty for {full_sku}.", icon="✅")
#                         else:
#                             st.session_state.order_lines.append({
#                                 "base_sku":    base_sku,
#                                 "size_mm":     sel_var["size_mm"],
#                                 "size_label":  sel_var["size_label"],
#                                 "full_sku":    full_sku,
#                                 "row_label":   sel_prod["row_label"],
#                                 "description": mrp_data.get("description", sel_prod["row_label"]),
#                                 "sheet":       sel_var["sheet"],
#                                 "mrp":         mrp_v,
#                                 "landing":     land_v,
#                                 "qty":         man_qty,
#                                 "amount":      round(land_v * man_qty, 2),
#                                 "matched":     mrp_v > 0,
#                             })
#                             st.success(f"Added {full_sku} × {man_qty}.", icon="✅")
#                         st.rerun()
#                 else:
#                     st.warning("No size variants found for this product.", icon="⚠️")

#     # ── Live totals ──────────────────────────────────────────────────────────
#     active_lines = [r for r in st.session_state.order_lines
#                     if r["qty"] > 0 and r["full_sku"] != "—"]
#     grand_mrp  = sum(r["mrp"]     * r["qty"] for r in active_lines)
#     grand_land = sum(r["landing"] * r["qty"] for r in active_lines)

#     st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)
#     st.markdown(f"""
#     <div class="card" style="max-width:420px;">
#       <div class="card-title">📊 Live Summary</div>
#       <div class="totals-row">
#         <span>Active Line Items</span><span class="val">{len(active_lines)}</span>
#       </div>
#       <div class="totals-row">
#         <span>Gross MRP Value</span><span class="val">₹ {grand_mrp:,.2f}</span>
#       </div>
#       <div class="totals-row">
#         <span>Distributor Landing</span><span class="val">₹ {grand_land:,.2f}</span>
#       </div>
#       <div class="totals-row net">
#         <span>Net Payable (Landing)</span><span class="val">₹ {grand_land:,.2f}</span>
#       </div>
#     </div>
#     """, unsafe_allow_html=True)

#     col_b1, col_b2, _ = st.columns([1, 1, 3])
#     with col_b1:
#         if st.button("◀  Back", key="back_step2"):
#             st.session_state.step = 2
#             st.rerun()
#     with col_b2:
#         if st.button("▶  Next: Download", key="go_step4"):
#             st.session_state.step = 4
#             st.rerun()


# # ═══════════════════════════════════════════════════════════════════════════════
# # STEP 4 — Download
# # ═══════════════════════════════════════════════════════════════════════════════
# elif step == 4:
#     st.markdown(f"""
#     <div class="card">
#       <div class="card-title">{badge(4)} Review & Download Quotation</div>
#     </div>""", unsafe_allow_html=True)

#     active_lines = [r for r in st.session_state.order_lines
#                     if r["qty"] > 0 and r["full_sku"] != "—"]

#     # Party summary
#     pc1, pc2 = st.columns(2)
#     with pc1:
#         st.markdown(party_html(st.session_state.bill_to, "BILL TO PARTY"), unsafe_allow_html=True)
#     with pc2:
#         st.markdown(party_html(st.session_state.ship_to, "SHIP TO PARTY"), unsafe_allow_html=True)
#     if any(st.session_state.dealer_info.values()):
#         st.markdown(party_html(st.session_state.dealer_info, "DEALER / DISTRIBUTOR"),
#                     unsafe_allow_html=True)

#     st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)
#     st.markdown("#### 📦 Ordered Items")

#     if active_lines:
#         grand_mrp = grand_land = 0.0
#         line_data = []
#         for r in active_lines:
#             grand_mrp  += r["mrp"]     * r["qty"]
#             grand_land += r["landing"] * r["qty"]
#             line_data.append({
#                 "Full SKU":      r["full_sku"],
#                 "Product":       r["row_label"],
#                 "Description":   r["description"][:50],
#                 "Size":          r["size_label"],
#                 "Sheet":         r["sheet"],
#                 "MRP (₹)":       round(r["mrp"],     2),
#                 "Qty":           r["qty"],
#                 "Rate (₹)":      round(r["landing"],  2),
#                 "Amount (₹)":    round(r["amount"],   2),
#             })

#         df_lines = pd.DataFrame(line_data)
#         st.dataframe(df_lines, use_container_width=True, hide_index=True,
#                      column_config={
#                          "MRP (₹)":    st.column_config.NumberColumn(format="₹ %.2f"),
#                          "Rate (₹)":   st.column_config.NumberColumn(format="₹ %.2f"),
#                          "Amount (₹)": st.column_config.NumberColumn(format="₹ %.2f"),
#                      })

#         discount = grand_mrp - grand_land
#         st.markdown(f"""
#         <div class="card" style="max-width:440px;margin-top:12px;">
#           <div class="card-title">💰 Totals</div>
#           <div class="totals-row">
#             <span>Gross MRP Value</span><span class="val">₹ {grand_mrp:,.2f}</span>
#           </div>
#           <div class="totals-row">
#             <span>Distributor Discount</span>
#             <span class="val neg">− ₹ {discount:,.2f}</span>
#           </div>
#           <div class="totals-row net">
#             <span>Net Payable (Distributor Landing)</span>
#             <span class="val">₹ {grand_land:,.2f}</span>
#           </div>
#         </div>
#         """, unsafe_allow_html=True)

#         st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)
#         dl1, dl2, dl3 = st.columns([1, 1, 1])

#         # CSV download
#         csv_buf = io.StringIO()
#         df_lines.to_csv(csv_buf, index=False)
#         dl1.download_button(
#             "⬇  Download CSV",
#             data=csv_buf.getvalue().encode(),
#             file_name=f"sintex_quotation_{date.today()}.csv",
#             mime="text/csv", key="dl_csv",
#         )

#         # Excel download
#         xls_buf = io.BytesIO()
#         with pd.ExcelWriter(xls_buf, engine="openpyxl") as writer:
#             party_rows = []
#             for k, v in st.session_state.dealer_info.items():
#                 party_rows.append({"Field": f"Dealer - {k}", "Value": v})
#             for k, v in st.session_state.bill_to.items():
#                 party_rows.append({"Field": f"Bill To - {k}", "Value": v})
#             for k, v in st.session_state.ship_to.items():
#                 party_rows.append({"Field": f"Ship To - {k}", "Value": v})
#             pd.DataFrame(party_rows).to_excel(writer, sheet_name="Party Details", index=False)
#             df_lines.to_excel(writer, sheet_name="Quotation Lines", index=False)
#             pd.DataFrame([
#                 {"Description": "Gross MRP Value",   "Amount (₹)": round(grand_mrp,  2)},
#                 {"Description": "Distributor Disc.",  "Amount (₹)": round(discount,   2)},
#                 {"Description": "Net Payable",        "Amount (₹)": round(grand_land, 2)},
#             ]).to_excel(writer, sheet_name="Summary", index=False)

#         dl2.download_button(
#             "⬇  Download Excel",
#             data=xls_buf.getvalue(),
#             file_name=f"sintex_quotation_{date.today()}.xlsx",
#             mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#             key="dl_excel",
#         )

#         # PDF download
#         if _HAS_REPORTLAB:
#             if st.session_state.pdf_bytes is None:
#                 with st.spinner("Generating PDF…"):
#                     try:
#                         st.session_state.pdf_bytes = generate_pdf(
#                             active_lines,
#                             st.session_state.bill_to,
#                             st.session_state.ship_to,
#                             st.session_state.dealer_info,
#                         )
#                     except Exception as e:
#                         st.error(f"PDF generation failed: {e}")
#             if st.session_state.pdf_bytes:
#                 dl3.download_button(
#                     "⬇  Download PDF",
#                     data=st.session_state.pdf_bytes,
#                     file_name=f"sintex_quotation_{date.today()}.pdf",
#                     mime="application/pdf", key="dl_pdf",
#                 )
#         else:
#             st.caption("📄 PDF requires `reportlab`. Run: `pip install reportlab`")

#     else:
#         st.warning("No valid line items found. Go back and check the OCR / manual entry step.", icon="⚠️")

#     st.markdown("<div style='height:14px'></div>", unsafe_allow_html=True)
#     col_bk, col_nr, _ = st.columns([1, 1, 3])
#     with col_bk:
#         if st.button("◀  Back", key="back_step3"):
#             st.session_state.pdf_bytes = None
#             st.session_state.step = 3
#             st.rerun()
#     with col_nr:
#         if st.button("🔄  New Quotation", key="new_quot"):
#             keys_to_clear = [
#                 "step", "bill_to", "ship_to", "dealer_info", "pdf_bytes",
#                 "image_bytes", "ocr_words", "order_lines", "ocr_done", "zsd_search",
#             ]
#             for k in keys_to_clear:
#                 st.session_state.pop(k, None)
#             for k in list(st.session_state.keys()):
#                 if k.startswith(("qty_", "sku_", "del_", "man_", "bill_", "ship_", "dealer_")):
#                     st.session_state.pop(k, None)
#             st.rerun()