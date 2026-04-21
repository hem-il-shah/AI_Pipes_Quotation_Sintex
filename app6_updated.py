"""
Sintex BAPL Limited – Pipes Quotation Generator
Single-page Streamlit app — mobile-first, red/white Sintex brand palette

OCR Strategy (spatial table reconstruction — works with ANY form layout):
1. Azure Document Intelligence extracts all words + bounding box coordinates
2. Words are clustered into rows (by Y center) and columns (by X center) → 2-D grid
3. Header row detected → size columns (15MM/20MM/etc) and SKU-code column identified
4. For each data row: sku_prefix + size → look up full SKU in MRP master
5. Multi-strategy matching: exact prefix, prefix+size index, fuzzy char-score
"""

import io, os, re, copy, base64, json, time, requests
import pandas as pd
from openpyxl import load_workbook
import streamlit as st
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

# ─── Page Config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Sintex BAPL – Quotation Generator",
    page_icon="🔴", layout="centered", initial_sidebar_state="collapsed",
)

# ─── CSS ──────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&family=JetBrains+Mono:wght@400;500&display=swap');
:root{--red:#C0211F;--dred:#8B1514;--lred:#FDEAEA;--lgray:#F7F7F7;--mgray:#E8E8E8;
  --text:#1A1A1A;--muted:#6B6B6B;--border:#DEDEDE;--green:#1E7E4A;
  --radius:12px;--shadow:0 2px 12px rgba(0,0,0,0.08);}
html,body,[class*="css"]{font-family:'Inter',-apple-system,sans-serif!important;
  color:var(--text);background:#F2F2F2;}
#MainMenu,footer,header{visibility:hidden;}
.block-container{padding:1rem 0.75rem 5rem!important;max-width:920px!important;margin:0 auto!important;}

.app-header{background:linear-gradient(135deg,var(--red),var(--dred));color:white;
  padding:18px 20px;border-radius:var(--radius);margin-bottom:24px;
  display:flex;align-items:center;gap:16px;box-shadow:0 4px 20px rgba(192,33,31,.35);}
.app-header-badge{background:rgba(255,255,255,.18);border-radius:10px;width:52px;height:52px;
  display:flex;align-items:center;justify-content:center;font-size:26px;flex-shrink:0;}
.app-header-text h1{font-size:17px;font-weight:800;margin:0;}
.app-header-text p{font-size:11.5px;margin:3px 0 0;opacity:.75;}

/* ── TOP NAV TAB BAR ── */
.step-navbar{display:flex;background:white;border-radius:var(--radius);
  box-shadow:var(--shadow);margin-bottom:20px;overflow:hidden;
  border:1px solid var(--border);}
.step-nav-item{flex:1;display:flex;flex-direction:column;align-items:center;
  padding:12px 6px 10px;cursor:default;position:relative;
  border-right:1px solid var(--border);transition:background .2s;}
.step-nav-item:last-child{border-right:none;}
.step-nav-item.locked{opacity:.45;cursor:not-allowed;}
.step-nav-item.active{background:var(--lred);}
.step-nav-item.done{background:#ECFDF5;cursor:pointer;}
.step-nav-item.active::after{content:'';position:absolute;bottom:0;left:0;right:0;
  height:3px;background:var(--red);}
.step-nav-item.done::after{content:'';position:absolute;bottom:0;left:0;right:0;
  height:3px;background:var(--green);}
.step-nav-dot{width:26px;height:26px;border-radius:50%;display:flex;align-items:center;
  justify-content:center;font-size:11px;font-weight:700;margin-bottom:4px;
  background:var(--mgray);color:var(--muted);}
.step-nav-item.active .step-nav-dot{background:var(--red);color:white;}
.step-nav-item.done .step-nav-dot{background:var(--green);color:white;}
.step-nav-label{font-size:10px;font-weight:600;color:var(--muted);text-align:center;}
.step-nav-item.active .step-nav-label{color:var(--red);}
.step-nav-item.done .step-nav-label{color:var(--green);}

.step-card{background:white;border-radius:var(--radius);box-shadow:var(--shadow);
  margin-bottom:20px;overflow:hidden;border:1px solid var(--border);}
.step-card-header{background:linear-gradient(90deg,var(--red),#D94644);padding:14px 18px;
  display:flex;align-items:center;gap:12px;}
.step-number{background:rgba(255,255,255,.25);border-radius:8px;width:30px;height:30px;
  min-width:30px;display:flex;align-items:center;justify-content:center;
  font-size:13px;font-weight:700;color:white;}
.step-number.done{background:#1E7E4A;}
.step-title{font-size:14px;font-weight:700;color:white;margin:0;}
.step-subtitle{font-size:11px;color:rgba(255,255,255,.75);margin:2px 0 0;}
.step-body{padding:18px;}

.stButton>button{width:100%;background:linear-gradient(135deg,var(--red),var(--dred))!important;
  color:white!important;border:none!important;border-radius:10px!important;padding:13px 20px!important;
  font-family:'Inter',sans-serif!important;font-size:14.5px!important;font-weight:600!important;
  box-shadow:0 4px 14px rgba(192,33,31,.28)!important;transition:all .2s!important;}
.stButton>button:hover{transform:translateY(-1px)!important;}
.btn-secondary>.stButton>button{background:white!important;color:var(--red)!important;
  border:2px solid var(--red)!important;box-shadow:none!important;}

.stTabs [data-baseweb="tab-list"]{gap:4px;background:var(--mgray);
  padding:4px;border-radius:10px;border:none!important;}
.stTabs [data-baseweb="tab"]{border-radius:8px!important;font-family:'Inter',sans-serif!important;
  font-weight:500!important;font-size:13px!important;color:var(--muted)!important;}
.stTabs [aria-selected="true"]{background:var(--red)!important;color:white!important;}

.info-box{background:#EEF6FF;border:1.5px solid #BFDBFE;border-radius:9px;
  padding:11px 15px;font-size:12.5px;color:#1E40AF;margin:12px 0;}
.warn-box{background:#FFFBEB;border:1.5px solid #FDE68A;border-radius:9px;
  padding:11px 15px;font-size:12.5px;color:#92400E;margin:12px 0;}
.success-box{background:#ECFDF5;border:1.5px solid #A7F3D0;border-radius:9px;
  padding:11px 15px;font-size:12.5px;color:#065F46;margin:12px 0;}

.totals-box{background:white;border:1.5px solid var(--border);border-radius:var(--radius);
  padding:16px 18px;margin:16px 0;}
.total-row{display:flex;justify-content:space-between;padding:7px 0;font-size:13.5px;
  border-bottom:1px solid var(--mgray);}
.total-row:last-child{border:none;padding-top:10px;}
.total-row.grand .total-lbl{font-weight:700;color:var(--red);font-size:15px;}
.total-row.grand .total-val{font-weight:800;color:var(--red);font-size:15px;
  font-family:'JetBrains Mono',monospace;}
.total-lbl{color:var(--muted);font-weight:500;}
.total-val{font-family:'JetBrains Mono',monospace;font-weight:600;}
.total-val.neg{color:#c0392b;}

.ocr-wrap{overflow-x:auto;-webkit-overflow-scrolling:touch;margin:10px 0;}
.ocr-tbl{width:100%;border-collapse:collapse;font-family:'Inter',sans-serif;font-size:12px;}
.ocr-tbl th{background:var(--red);color:white;padding:8px 7px;text-align:center;
  font-size:10.5px;font-weight:600;white-space:nowrap;}
.ocr-tbl th.L{text-align:left;min-width:120px;}
.ocr-tbl td{padding:6px 7px;border-bottom:1px solid var(--mgray);vertical-align:middle;text-align:center;}
.ocr-tbl td.L{text-align:left;font-weight:500;font-size:11.5px;}
.ocr-tbl td.M{font-family:'JetBrains Mono',monospace;font-size:10px;color:#555;}
.ocr-tbl tr:nth-child(even) td{background:var(--lgray);}
.ocr-tbl .ok td{background:#ECFDF5!important;}
.ocr-tbl .no td{background:#FFFBEB!important;}

.raw-wrap{overflow-x:auto;-webkit-overflow-scrolling:touch;margin:8px 0;
  border:1.5px solid var(--border);border-radius:8px;}
.raw-tbl{border-collapse:collapse;font-size:11px;font-family:'JetBrains Mono',monospace;min-width:500px;}
.raw-tbl th{background:#2D2D2D;color:white;padding:6px 8px;white-space:nowrap;text-align:center;}
.raw-tbl td{padding:5px 8px;border:1px solid #eee;white-space:nowrap;}
.raw-tbl tr:nth-child(even) td{background:#fafafa;}
.raw-tbl .sku{background:#FFF3E0!important;font-weight:700;color:#8B1514;}
.raw-tbl .qty{color:#1E7E4A;font-weight:600;}
.raw-tbl .hdr td{background:#F0F0F0;font-weight:700;color:#333;font-family:'Inter',sans-serif;}

.party-section{background:var(--lgray);border-radius:10px;padding:14px;
  border:1px solid var(--border);margin-bottom:14px;}
.party-title{font-size:12px;font-weight:700;color:var(--red);text-transform:uppercase;
  letter-spacing:.6px;margin-bottom:10px;}
.fsl{font-size:11.5px;font-weight:700;text-transform:uppercase;letter-spacing:.8px;
  color:var(--muted);margin:14px 0 8px;padding-bottom:4px;border-bottom:1px solid var(--mgray);}
.dl-btn{display:flex;align-items:center;justify-content:center;gap:8px;width:100%;
  background:linear-gradient(135deg,var(--green),#155d38);color:white!important;
  text-align:center;padding:15px 20px;border-radius:10px;font-weight:700;font-size:14.5px;
  text-decoration:none!important;box-shadow:0 4px 18px rgba(30,126,74,.3);
  margin:8px 0;box-sizing:border-box;}

/* PDF Preview embed */
.pdf-preview-wrap{border:2px solid var(--border);border-radius:var(--radius);
  overflow:hidden;margin:16px 0;background:#525659;}

@media(max-width:600px){
  .block-container{padding:.75rem .5rem 5rem!important;}
  .step-body{padding:14px 12px;}
  .step-nav-label{font-size:9px;}
}
</style>
""", unsafe_allow_html=True)

# ─── Paths ────────────────────────────────────────────────────────────────────
XLSX_PATH      = os.path.join(os.path.dirname(__file__), "Sample form for Product list.xlsx")
MRP_PATH       = os.path.join(os.path.dirname(__file__), "MRP_State_chhattisghar.csv")
CUST_PATH      = os.path.join(os.path.dirname(__file__), "ZSD_CUST.csv")
AZURE_ENDPOINT = os.environ.get("AZURE_OCR_ENDPOINT", "")
AZURE_KEY      = os.environ.get("AZURE_OCR_KEY", "")

# ─── Canonical size labels ────────────────────────────────────────────────────
SIZE_ALIASES: dict[str, str] = {}
for _mm in ["15","20","25","32","40","50","63","75","90","110"]:
    SIZE_ALIASES[_mm + "MM"]      = _mm + "MM"
    SIZE_ALIASES[_mm + " MM"]     = _mm + "MM"
    SIZE_ALIASES[_mm]             = _mm + "MM"
SIZE_ALIASES.update({
    '1/2"':"15MM",'1/2':"15MM",'½"':"15MM",
    '3/4"':"20MM",'3/4':"20MM",'¾"':"20MM",
    '1"'  :"25MM",
    '1.25"':"32MM",'1-1/4"':"32MM",'1¼"':"32MM",
    '1.5"':"40MM",'1-1/2"':"40MM",'1½"':"40MM",
    '2"'  :"50MM",
})

# ═══════════════════════════════════════════════════════════════════════════════
# DATA LOADING
# ═══════════════════════════════════════════════════════════════════════════════

@st.cache_data(show_spinner=False)
def load_sku_sheets() -> dict:
    wb = load_workbook(XLSX_PATH, read_only=True)
    result = {}
    for sn in wb.sheetnames:
        ws  = wb[sn]
        all_rows = list(ws.iter_rows(values_only=True))
        r1  = list(all_rows[1]) if len(all_rows) > 1 else []
        r2  = list(all_rows[2]) if len(all_rows) > 2 else []
        cis = [ci for ci,v in enumerate(r1) if ci>=2 and v is not None]
        if not cis:
            cis = [ci for ci,v in enumerate(r2) if ci>=2 and v is not None]
        od  = {ci: str(r1[ci]).strip() if ci<len(r1) and r1[ci] else "" for ci in cis}
        inch= {ci: str(r2[ci]).strip() if ci<len(r2) and r2[ci] else "" for ci in cis}
        rows, sec = [], "PIPES"
        for raw in all_rows[3:]:
            row   = list(raw)
            if not any(v for v in row if v is not None): continue
            label = str(row[0]).strip() if row[0] else ""
            if not label or label == "\xa0": continue
            up = label.upper()
            if up in {"FITTINGS","PIPES","FITTING SCH 80","FITTING SCH80"}:
                sec = up; rows.append({"sh":True,"label":label}); continue
            if (len(row)>1 and (row[1] is None or str(row[1]).strip()=="\xa0")
                    and not any(str(v).strip().startswith("CP") for v in row[2:] if v)):
                sec = label; rows.append({"sh":True,"label":label}); continue
            cells = {}
            for ci in cis:
                if ci<len(row) and row[ci] is not None:
                    v = str(row[ci]).strip()
                    if v and v not in("-","\xa0") and v.startswith("CP"):
                        cells[ci] = v
            rows.append({"sh":False,"label":label,"sec":sec,"cells":cells})
        result[sn] = {"od":od,"inch":inch,"cis":cis,"rows":rows}
    return result


@st.cache_data(show_spinner=False)
def build_sku_master(sku_sheets: dict) -> dict:
    master = {}
    for sn, sd in sku_sheets.items():
        sec = "PIPES"
        for rd in sd["rows"]:
            if rd.get("sh"):
                sec = rd["label"]; continue
            for ci, sku in rd["cells"].items():
                master[sku] = {
                    "sheet": sn, "product": rd["label"], "section": sec,
                    "od_size": sd["od"].get(ci,""), "inch_size": sd["inch"].get(ci,""),
                }
    return master


@st.cache_data(show_spinner=False)
def build_prefix_index(sku_master: dict) -> dict:
    """
    Two indexes for fast matching:
    by_prefix_size: (prefix, size_norm) → full_sku
    by_prefix:      prefix → [full_sku, ...]
    Built for every prefix length from 4 to len(sku).
    """
    bps, bp = {}, {}
    for sku, info in sku_master.items():
        sz = info.get("od_size","").upper().replace(" ","")
        for l in range(4, len(sku)+1):
            p = sku[:l]
            bp.setdefault(p, []).append(sku)
            if sz:
                bps[(p, sz)] = sku   # last-write OK; same prefix+size → same SKU
    return {"bps": bps, "bp": bp}


@st.cache_data(show_spinner=False)
def load_mrp_data() -> dict:
    df = pd.read_csv(MRP_PATH, encoding="latin-1")
    df["_m"] = pd.to_numeric(df["MRP(ZPR1-933)"].astype(str).str.replace(",","").str.strip(), errors="coerce").fillna(0)
    df["_d"] = pd.to_numeric(df["Distributor Landing"].astype(str).str.replace(",","").str.strip(), errors="coerce").fillna(0)
    r = {}
    for _, row in df.iterrows():
        mat = str(row["Material Number"]).strip()
        r[mat] = {"mrp": float(row["_m"]), "distributor_landing": float(row["_d"]),
                  "description": str(row["Material Description"]).strip()}
    return r


@st.cache_data(show_spinner=False)
def load_customers() -> list:
    df  = pd.read_csv(CUST_PATH, encoding="latin-1")
    out = []
    for _, r in df.iterrows():
        name = str(r.get("Customer Name","")).strip()
        code = str(r.get("Customer Code","")).strip()
        if name and name != "nan":
            out.append({
                "code": code, "name": name,
                "address": " ".join(filter(None,[
                    str(r.get("Address 1","") or "").strip(),
                    str(r.get("Address 2","") or "").strip(),
                    str(r.get("City","") or "").strip(),
                    str(r.get("State Code Desc.","") or "").strip()])),
                "phone":  str(r.get("Telephone","")  or "").strip(),
                "mobile": str(r.get("Mobile No.","") or "").strip(),
                "state_code": str(r.get("State Code","") or "").strip(),
                "state": str(r.get("State Code Desc.","") or "").strip(),
                "gst": str(r.get("GST Number","") or "").strip(),
                "pan": str(r.get("PAN No.","") or "").strip(),
                "display": f"{code} – {name}",
            })
    return out


# ═══════════════════════════════════════════════════════════════════════════════
# AZURE OCR — word extraction
# ═══════════════════════════════════════════════════════════════════════════════

def _poly_bbox(poly):
    """(x, y, w, h) from flat list or list-of-dicts polygon."""
    if not poly:
        return 0, 0, 0, 0
    if isinstance(poly[0], dict):
        xs, ys = [p["x"] for p in poly], [p["y"] for p in poly]
    else:
        xs, ys = poly[0::2], poly[1::2]
    return min(xs), min(ys), max(xs)-min(xs), max(ys)-min(ys)


def _words_v3v4(data: dict) -> list[dict]:
    out = []
    ar  = data.get("analyzeResult", data)
    for page in ar.get("pages", []):
        pw = page.get("width", 1) or 1
        ph = page.get("height", 1) or 1
        for w in page.get("words", []):
            x, y, ww, hh = _poly_bbox(w.get("polygon", w.get("boundingBox",[])))
            out.append({"text": w.get("content", w.get("text","")),
                        "x":x,"y":y,"w":ww,"h":hh,"cx":x+ww/2,"cy":y+hh/2,
                        "pw":pw,"ph":ph})
    return out


def _words_v2(data: dict) -> list[dict]:
    out = []
    for page in data.get("analyzeResult",{}).get("readResults",[]):
        pw = page.get("width",1) or 1
        ph = page.get("height",1) or 1
        for line in page.get("lines",[]):
            for w in line.get("words",[]):
                x,y,ww,hh = _poly_bbox(w.get("boundingBox",[]))
                out.append({"text":w.get("text",""),
                            "x":x,"y":y,"w":ww,"h":hh,"cx":x+ww/2,"cy":y+hh/2,
                            "pw":pw,"ph":ph})
    return out


def run_azure_ocr(img: bytes, endpoint: str, key: str) -> list[dict]:
    ep  = endpoint.rstrip("/")
    hdr = {"Ocp-Apim-Subscription-Key": key, "Content-Type": "application/octet-stream"}

    def post(url):
        return requests.post(url, headers=hdr, data=img, timeout=60)

    def poll_async(resp) -> dict:
        if resp.status_code != 202:
            resp.raise_for_status()
            return resp.json()
        op = resp.headers.get("Operation-Location") or resp.headers.get("operation-location","")
        if not op:
            raise ValueError("No Operation-Location header")
        for _ in range(30):
            time.sleep(2)
            p = requests.get(op, headers={"Ocp-Apim-Subscription-Key": key}, timeout=30)
            p.raise_for_status()
            d = p.json()
            s = d.get("status","")
            if s == "succeeded": return d
            if s == "failed":    raise ValueError("Azure failed: " + json.dumps(d.get("error",{})))
        raise TimeoutError("Azure OCR timed out")

    # v4 Document Intelligence
    r = post(f"{ep}/documentintelligence/documentModels/prebuilt-read:analyze?api-version=2024-02-29-preview")
    if r.status_code not in (200, 202):
        # v3 Form Recognizer
        r = post(f"{ep}/formrecognizer/documentModels/prebuilt-read:analyze?api-version=2023-07-31")
    if r.status_code not in (200, 202):
        # v2.1 layout (always async)
        r = post(f"{ep}/formrecognizer/v2.1/layout/analyze")
        r.raise_for_status()
        op = r.headers.get("Operation-Location","")
        if not op: raise ValueError("No Operation-Location in v2 response")
        for _ in range(20):
            time.sleep(2)
            p = requests.get(op, headers={"Ocp-Apim-Subscription-Key": key}, timeout=30)
            p.raise_for_status()
            d = p.json()
            if d.get("status") == "succeeded": return _words_v2(d)
            if d.get("status") == "failed":    raise ValueError("Azure v2 failed")
        raise TimeoutError("Azure v2 timed out")

    return _words_v3v4(poll_async(r))


# ═══════════════════════════════════════════════════════════════════════════════
# SPATIAL TABLE RECONSTRUCTION
# ═══════════════════════════════════════════════════════════════════════════════

def reconstruct_table(words: list[dict]) -> list[list[str]]:
    """
    Cluster OCR words into a 2-D grid purely from bounding-box positions.

    Row clustering : words within (median_height × 0.55) of each other share a row.
    Column grid    : collect all cell-center X values, cluster within (page_width × 0.025).
    Cell merging   : within a row, words within (page_width × 0.016) of each other merge.
    """
    if not words:
        return []

    heights = sorted(w["h"] for w in words if w["h"] > 0)
    med_h   = heights[len(heights)//2] if heights else 8
    row_tol = max(med_h * 0.55, 2)
    pw      = words[0]["pw"] or 1
    merge_x = pw * 0.016
    col_tol = pw * 0.025

    # ── cluster into rows ──────────────────────────────────────────────────
    by_y = sorted(words, key=lambda w: w["cy"])
    raw_rows: list[list[dict]] = []
    cur: list[dict] = []
    cy_avg = None
    for w in by_y:
        if cy_avg is None or abs(w["cy"] - cy_avg) <= row_tol:
            cur.append(w)
            cy_avg = (cy_avg + w["cy"]) / 2 if cy_avg else w["cy"]
        else:
            raw_rows.append(sorted(cur, key=lambda ww: ww["cx"]))
            cur    = [w]
            cy_avg = w["cy"]
    if cur:
        raw_rows.append(sorted(cur, key=lambda ww: ww["cx"]))

    # ── merge nearby words within each row into cells ──────────────────────
    cell_rows: list[list[dict]] = []   # cell = {text, cx}
    for row in raw_rows:
        cells, buf, buf_cx, buf_rx = [], row[0]["text"], row[0]["cx"], row[0]["x"]+row[0]["w"]
        for w in row[1:]:
            if w["x"] - buf_rx <= merge_x:
                buf    += " " + w["text"]
                buf_rx  = w["x"] + w["w"]
                buf_cx  = (buf_cx + w["cx"]) / 2
            else:
                cells.append({"text": buf.strip(), "cx": buf_cx})
                buf, buf_cx, buf_rx = w["text"], w["cx"], w["x"]+w["w"]
        cells.append({"text": buf.strip(), "cx": buf_cx})
        cell_rows.append(cells)

    # ── build global column grid ───────────────────────────────────────────
    all_cx  = sorted(c["cx"] for row in cell_rows for c in row)
    col_cxs: list[float] = []
    for cx in all_cx:
        placed = False
        for i, cc in enumerate(col_cxs):
            if abs(cx - cc) <= col_tol:
                col_cxs[i] = (cc + cx) / 2
                placed = True; break
        if not placed:
            col_cxs.append(cx)
    col_cxs.sort()
    nc = len(col_cxs)

    def nearest(cx):
        return min(range(nc), key=lambda i: abs(col_cxs[i] - cx))

    # ── assign cells to columns ────────────────────────────────────────────
    grid: list[list[str]] = []
    for row in cell_rows:
        arr = [""] * nc
        for cell in row:
            ci = nearest(cell["cx"])
            arr[ci] = (arr[ci] + " " + cell["text"]).strip() if arr[ci] else cell["text"]
        grid.append(arr)

    return grid


# ═══════════════════════════════════════════════════════════════════════════════
# TABLE ANALYSIS — detect structure and extract rows
# ═══════════════════════════════════════════════════════════════════════════════

def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", s.strip().upper())

def _size_of(cell: str) -> str | None:
    n = _norm(cell).replace(" ","").rstrip(".")
    v = SIZE_ALIASES.get(n)
    if not v:
        # strip trailing non-digits in case OCR added noise: "20MM." "20Mm"
        v = SIZE_ALIASES.get(re.sub(r'[^0-9]','',n) + "MM") if re.sub(r'[^0-9]','',n) else None
    return v

def _is_sku_prefix(cell: str) -> bool:
    return bool(re.match(r'^CP[A-Z0-9]{3,}', cell.strip().upper()))

def _clean_ocr_num(cell: str) -> int | None:
    """Parse a quantity from potentially noisy OCR (O→0, l→1, S→5, etc.)."""
    c = cell.strip().replace(",","")
    c = c.replace("O","0").replace("o","0")
    c = c.replace("l","1").replace("I","1").replace("|","1")
    c = c.replace("S","5").replace("s","5").replace("G","6")
    m = re.search(r'\d+', c)
    if m:
        v = int(m.group())
        return v if 1 <= v <= 9999 else None
    return None

def _is_qty(cell: str) -> bool:
    return _clean_ocr_num(cell) is not None


def analyze_table(grid: list[list[str]]) -> dict:
    """
    Identify columns: SKU column, size columns, product-name column.
    Returns extracted data rows and meta info.
    """
    if not grid:
        return {"rows":[], "meta":{}, "grid":grid}

    nc = max(len(r) for r in grid)

    # ── Find header row: row with the most recognisable size-header cells ──
    best_hdr, best_szc = 0, {}
    for ri, row in enumerate(grid[:15]):          # header is in first 15 rows
        szc = {ci: _size_of(cell) for ci,cell in enumerate(row) if _size_of(cell)}
        if len(szc) > len(best_szc):
            best_szc, best_hdr = szc, ri

    # If still empty, try rows where numeric-only cells align
    if not best_szc:
        for ri, row in enumerate(grid[:15]):
            szc = {}
            for ci, cell in enumerate(row):
                n = _norm(cell).replace(" ","")
                if re.fullmatch(r'\d{2,3}', n) and (n+"MM") in SIZE_ALIASES:
                    szc[ci] = n+"MM"
            if len(szc) > len(best_szc):
                best_szc, best_hdr = szc, ri

    size_cols  = best_szc      # {col_idx: "15MM", ...}
    header_row = best_hdr

    # ── Find SKU column ────────────────────────────────────────────────────
    sku_votes: dict[int,int] = {}
    for row in grid[header_row+1:]:
        for ci, cell in enumerate(row):
            if _is_sku_prefix(cell):
                sku_votes[ci] = sku_votes.get(ci,0) + 1
    sku_col = max(sku_votes, key=sku_votes.get) if sku_votes else None

    # ── Find name column ───────────────────────────────────────────────────
    skip = set(size_cols) | ({sku_col} if sku_col is not None else set())
    name_col = next((c for c in range(nc) if c not in skip), None)

    # ── Extract data rows ──────────────────────────────────────────────────
    rows = []
    for ri in range(header_row+1, len(grid)):
        row       = grid[ri]
        sku_pfx   = row[sku_col].strip().upper() if sku_col is not None and sku_col < len(row) else ""
        product   = row[name_col].strip()        if name_col is not None and name_col < len(row) else ""

        if not sku_pfx and not product:
            continue
        # Skip obvious section/sub-header rows
        if _norm(product) in {"SKU NAME","PRODUCT","DESCRIPTION","ITEM","S NO","SR NO",
                               "SKU CODE","SIZE","CATEGORY","SECTION"}:
            continue
        # If sku_pfx looks like a section label (no CP prefix) and no qty → skip
        if not sku_pfx and not any(_is_qty(row[ci]) for ci in size_cols if ci < len(row)):
            continue

        sizes: dict[str,int] = {}
        for ci, sz_label in size_cols.items():
            if ci < len(row):
                q = _clean_ocr_num(row[ci])
                if q:
                    sizes[sz_label] = q

        if not sizes:
            continue   # row has no quantities — ignore

        rows.append({"product": product, "sku_prefix": sku_pfx, "sizes": sizes, "ri": ri})

    return {
        "rows": rows,
        "meta": {"header_row": header_row, "sku_col": sku_col,
                 "name_col": name_col, "size_cols": size_cols},
        "grid": grid,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# SKU MATCHING
# ═══════════════════════════════════════════════════════════════════════════════

_OCR_FIXES = str.maketrans({
    "O":"0","o":"0","Q":"0",
    "I":"1","l":"1","|":"1",
    "Z":"2","z":"2",
    "S":"5","s":"5",
    "G":"6","b":"6",
    "B":"8",
})

def _fix_sku(raw: str) -> str:
    s = raw.strip().upper()
    # Only fix characters after the first 2 (CP is always letters)
    return s[:2] + s[2:].translate(_OCR_FIXES) if len(s) > 2 else s


def match_sku(prefix: str, size: str, pidx: dict, master: dict) -> str | None:
    """
    Find full SKU for prefix + size.
    Tries multiple strategies in descending confidence order.
    """
    if not prefix:
        return None

    sz   = size.upper().replace(" ","")
    pfx  = _fix_sku(prefix)
    bps  = pidx["bps"]
    bp   = pidx["bp"]

    # Strategy 1: exact full key in master with matching size
    if pfx in master and master[pfx].get("od_size","").upper().replace(" ","") == sz:
        return pfx

    # Strategy 2: prefix+size index (longest prefix first)
    for l in range(min(len(pfx),20), 3, -1):
        key = (pfx[:l], sz)
        if key in bps:
            return bps[key]

    # Strategy 3: among SKUs sharing prefix, filter by size
    for l in range(min(len(pfx),20), 3, -1):
        for sku in bp.get(pfx[:l], []):
            if master[sku].get("od_size","").upper().replace(" ","") == sz:
                return sku

    # Strategy 4: fuzzy character match — allow up to 2 mismatches
    best_sku, best_score = None, 0
    for sku, info in master.items():
        if info.get("od_size","").upper().replace(" ","") != sz:
            continue
        min_l = min(len(pfx), len(sku))
        if min_l < 4:
            continue
        score = sum(a==b for a,b in zip(pfx[:min_l], sku[:min_l]))
        # Require at least (min_l - 2) chars to match
        if score >= max(4, min_l - 2) and score > best_score:
            best_score, best_sku = score, sku

    return best_sku


def build_quantities(extracted_rows: list[dict], pidx: dict, master: dict):
    # line_items: list of individual detections, NO grouping — each detection is separate
    line_items: list[dict] = []
    # quantities dict: sku→total qty, used only for totals display in UI
    quantities: dict[str,int] = {}
    log: list[dict] = []
    for row in extracted_rows:
        for sz, qty in row["sizes"].items():
            full = match_sku(row["sku_prefix"], sz, pidx, master)
            if full:
                quantities[full] = quantities.get(full, 0) + qty
                line_items.append({"sku": full, "qty": qty})
            log.append({
                "product":  row["product"],
                "prefix":   row["sku_prefix"],
                "size":     sz,
                "qty":      qty,
                "full_sku": full or "—",
                "status":   "matched" if full else "unmatched",
            })
    return quantities, log, line_items


# ═══════════════════════════════════════════════════════════════════════════════
# PDF GENERATION
# CHANGES: (1) Full-width tables, formal layout, fixed totals table
#          (4) Removed "Sheet" column
# ═══════════════════════════════════════════════════════════════════════════════

def build_pdf(quantities: dict, mrp_data: dict, bill_to: dict,
              ship_to: dict, sku_master: dict, line_items: list = None) -> bytes:
    buf = io.BytesIO()
    # Use full landscape A4 with narrow margins for maximum table width
    PAGE = landscape(A4)
    LEFT_M = RIGHT_M = 12 * mm
    TOP_M = BOT_M = 12 * mm
    USABLE_W = PAGE[0] - LEFT_M - RIGHT_M  # full usable width

    doc = SimpleDocTemplate(buf, pagesize=PAGE,
                            leftMargin=LEFT_M, rightMargin=RIGHT_M,
                            topMargin=TOP_M, bottomMargin=BOT_M)

    # ── Colour palette ──────────────────────────────────────────────────────
    RED    = colors.HexColor("#C0211F")
    DRED   = colors.HexColor("#8B1514")
    BLACK  = colors.HexColor("#1A1A1A")
    DGRAY  = colors.HexColor("#404040")
    MGRAY  = colors.HexColor("#AAAAAA")
    LGRAY  = colors.HexColor("#F5F5F5")
    MLGRAY = colors.HexColor("#DEDEDE")
    WHITE  = colors.white
    GREEN  = colors.HexColor("#1E7E4A")

    sty = getSampleStyleSheet()
    def ps(n, **kw): return ParagraphStyle(n, parent=sty["Normal"], **kw)

    # ── Header ──────────────────────────────────────────────────────────────
    story = []

    hdr_left = ps("hl", fontName="Helvetica-Bold", fontSize=9, textColor=WHITE, leading=14)
    hdr_ctr  = ps("hc", fontName="Helvetica-Bold", fontSize=18, textColor=WHITE,
                  alignment=TA_CENTER)
    hdr_rt   = ps("hr", fontName="Helvetica", fontSize=8, textColor=WHITE,
                  alignment=TA_RIGHT, leading=13)

    hdr = Table([[
        Paragraph("Sintex BAPL Limited<br/>"
                  "<font size='7.5'>Kutesar Road, Raipur, CG 492101</font><br/>"
                  "<font size='7.5'>GSTIN: 22AADCB1921F1ZE</font>", hdr_left),
        Paragraph("SALES QUOTATION", hdr_ctr),
        Paragraph("State: <b>Chhattisgarh</b><br/>"
                  "<font size='7'>CPVC / UPVC Pipes &amp; Fittings</font>", hdr_rt),
    ]], colWidths=[USABLE_W*0.30, USABLE_W*0.40, USABLE_W*0.30])
    hdr.setStyle(TableStyle([
        ("BACKGROUND",  (0,0), (-1,-1), DRED),
        ("TEXTCOLOR",   (0,0), (-1,-1), WHITE),
        ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",  (0,0), (-1,-1), 12),
        ("BOTTOMPADDING",(0,0),(-1,-1), 12),
        ("LEFTPADDING", (0,0), (-1,-1), 14),
        ("RIGHTPADDING",(0,0), (-1,-1), 14),
        ("LINEBELOW",   (0,0), (-1,-1), 3, RED),
    ]))
    story += [hdr, Spacer(1, 4*mm)]

    # ── Party table ─────────────────────────────────────────────────────────
    plbl = ps("plbl", fontName="Helvetica-Bold", fontSize=7.5, textColor=WHITE)
    pval = ps("pval", fontName="Helvetica",      fontSize=7.5, textColor=BLACK, leading=11)

    def party_cell(d):
        lines = [
            f"<b>Party No.:</b> {d.get('party_no','')}",
            f"<b>Name &amp; Address:</b> {d.get('name','').replace(chr(10),' ')}",
            f"<b>Phone:</b> {d.get('phone','')}   <b>Mobile:</b> {d.get('mobile','')}",
            f"<b>State:</b> {d.get('state_code','')} – {d.get('state','')}",
            f"<b>GST:</b> {d.get('gst','')}   <b>PAN:</b> {d.get('pan','')}",
        ]
        return Paragraph("<br/>".join(lines), pval)

    HALF = USABLE_W / 2
    pt = Table([
        [Paragraph("BILL TO PARTY", plbl), Paragraph("SHIP TO PARTY", plbl)],
        [party_cell(bill_to),              party_cell(ship_to)],
    ], colWidths=[HALF, HALF])
    pt.setStyle(TableStyle([
        ("BACKGROUND",   (0,0), (-1,0), DGRAY),
        ("TEXTCOLOR",    (0,0), (-1,0), WHITE),
        ("BACKGROUND",   (0,1), (-1,1), LGRAY),
        ("BOX",          (0,0), (-1,-1), 0.8, MLGRAY),
        ("INNERGRID",    (0,0), (-1,-1), 0.5, MLGRAY),
        ("VALIGN",       (0,0), (-1,-1), "TOP"),
        ("TOPPADDING",   (0,0), (-1,-1), 7),
        ("BOTTOMPADDING",(0,0), (-1,-1), 7),
        ("LEFTPADDING",  (0,0), (-1,-1), 10),
        ("RIGHTPADDING", (0,0), (-1,-1), 10),
    ]))
    story += [pt, Spacer(1, 5*mm)]

    # ── Main items table ─────────────────────────────────────────────────────
    # CHANGE 4: "Sheet" column removed. Columns: S.No, Product, SKU Code, OD, Inch, MRP, Qty, Total MRP, Dist.Landing, Taxable
    sch  = ps("sch",  fontName="Helvetica-Bold", fontSize=7,   textColor=WHITE, alignment=TA_CENTER)
    cell = ps("c",    fontName="Helvetica",      fontSize=7.5, textColor=BLACK)
    celc = ps("cc",   fontName="Helvetica",      fontSize=7.5, textColor=BLACK, alignment=TA_CENTER)
    celr = ps("cr",   fontName="Helvetica",      fontSize=7.5, textColor=BLACK, alignment=TA_RIGHT)
    celb = ps("cb",   fontName="Helvetica-Bold", fontSize=7.5, textColor=BLACK, alignment=TA_RIGHT)
    skuc = ps("sk",   fontName="Courier",        fontSize=6.5, textColor=DRED)
    qtys_ps = ps("qp",fontName="Helvetica-Bold", fontSize=8,   textColor=RED,   alignment=TA_CENTER)

    heads = ["S.No", "Product", "SKU Code", "OD", "Inch", "MRP (₹)", "Qty", "Total MRP (₹)", "Dist. Landing (₹)", "Taxable (₹)"]
    # Column widths — must sum to USABLE_W
    # S.No=8, Product=68, SKU=48, OD=14, Inch=12, MRP=22, Qty=10, TotalMRP=26, Dist=26, Tax=26  (all mm)
    cw_mm = [8, 68, 48, 14, 12, 22, 10, 26, 26, 26]
    cw = [x * mm for x in cw_mm]
    # Scale to exactly USABLE_W
    scale = USABLE_W / sum(cw)
    cw = [x * scale for x in cw]

    trows = [[Paragraph(h, sch) for h in heads]]
    ln = gm = gd = gt = 0

    # CHANGE 3: Use line_items list (individual detections), NOT quantities dict
    if line_items:
        items_iter = line_items
    else:
        items_iter = [{"sku": s, "qty": q} for s, q in quantities.items() if q > 0]

    for item in items_iter:
        sku = item["sku"]; qty = item["qty"]
        if qty <= 0:
            continue
        info  = sku_master.get(sku, {})
        mi    = mrp_data.get(sku, {})
        mrp   = mi.get("mrp", 0.)
        dist  = mi.get("distributor_landing", 0.)
        tot   = round(mrp * qty, 2)
        tax   = round(dist * qty, 2)
        ln += 1; gm += tot; gd += (mrp - dist) * qty; gt += tax
        trows.append([
            Paragraph(str(ln),             celc),
            Paragraph(info.get("product",""), cell),
            Paragraph(sku,                 skuc),
            Paragraph(info.get("od_size",""),   celc),
            Paragraph(info.get("inch_size",""), celc),
            Paragraph(f"{mrp:,.2f}",       celr),
            Paragraph(str(qty),            qtys_ps),
            Paragraph(f"{tot:,.2f}",       celb),
            Paragraph(f"{dist:,.2f}",      celr),
            Paragraph(f"{tax:,.2f}",       celb),
        ])
    if ln == 0:
        trows.append([Paragraph("No items", cell)] + [""] * 9)

    t2 = Table(trows, colWidths=cw, repeatRows=1)
    ts = [
        ("BACKGROUND",    (0,0),  (-1,0),  DRED),
        ("TEXTCOLOR",     (0,0),  (-1,0),  WHITE),
        ("BOX",           (0,0),  (-1,-1), 0.8,  DGRAY),
        ("INNERGRID",     (0,1),  (-1,-1), 0.3,  MLGRAY),
        ("LINEBELOW",     (0,0),  (-1,0),  1.5,  RED),
        ("VALIGN",        (0,0),  (-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0),  (-1,-1), 4),
        ("BOTTOMPADDING", (0,0),  (-1,-1), 4),
        ("LEFTPADDING",   (0,0),  (-1,-1), 4),
        ("RIGHTPADDING",  (0,0),  (-1,-1), 4),
    ]
    for ri in range(1, len(trows)):
        if ri % 2 == 0:
            ts.append(("BACKGROUND", (0, ri), (-1, ri), LGRAY))
    t2.setStyle(TableStyle(ts))
    story += [t2, Spacer(1, 0)]

    # ── Totals table — full width, spanning all columns, right-aligned block ─
    # CHANGE 1: Totals section spans full table width, formal layout
    lbl_style = ps("tlbl", fontName="Helvetica",      fontSize=8,  textColor=DGRAY,  alignment=TA_RIGHT)
    lbl_bold  = ps("tbb",  fontName="Helvetica-Bold", fontSize=8,  textColor=BLACK,  alignment=TA_RIGHT)
    val_style = ps("tval", fontName="Helvetica-Bold", fontSize=8,  textColor=BLACK,  alignment=TA_RIGHT)
    val_red   = ps("tvr",  fontName="Helvetica-Bold", fontSize=9,  textColor=RED,    alignment=TA_RIGHT)
    val_neg   = ps("tvn",  fontName="Helvetica-Bold", fontSize=8,  textColor=DGRAY,  alignment=TA_RIGHT)

    # 3-column totals spanning full usable width:
    # col0 = filler (left), col1 = label, col2 = value
    LABEL_W = 55 * mm
    VALUE_W = 38 * mm
    FILL_W  = USABLE_W - LABEL_W - VALUE_W

    tot_rows = [
        ["", Paragraph("Gross Total (MRP):", lbl_style),
              Paragraph(f"₹  {gm:,.2f}", val_style)],
        ["", Paragraph("Less Distributor Discount:", lbl_style),
              Paragraph(f"– ₹  {gd:,.2f}", val_neg)],
        ["", Paragraph("Net Taxable Value:", lbl_bold),
              Paragraph(f"₹  {gt:,.2f}", val_red)],
    ]
    tot_tbl = Table(tot_rows, colWidths=[FILL_W, LABEL_W, VALUE_W])
    tot_tbl.setStyle(TableStyle([
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING",    (0,0), (-1,-1), 5),
        ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ("LEFTPADDING",   (0,0), (-1,-1), 6),
        ("RIGHTPADDING",  (0,0), (-1,-1), 6),
        # Box only around label+value columns
        ("BOX",           (1,0), (2,-1),  0.8, MGRAY),
        ("INNERGRID",     (1,0), (2,-1),  0.4, MLGRAY),
        # Highlight net taxable row
        ("BACKGROUND",    (1,2), (2,2),   LGRAY),
        ("LINEABOVE",     (1,2), (2,2),   1.5, RED),
        ("LINEBELOW",     (1,2), (2,2),   1.5, RED),
        # Align right for label & value
        ("ALIGN",         (1,0), (2,-1),  "RIGHT"),
    ]))
    story += [tot_tbl, Spacer(1, 5*mm)]

    # ── Footer ──────────────────────────────────────────────────────────────
    ft = Table([[
        Paragraph("<i>Computer-generated quotation. Subject to change without notice.</i>",
                  ps("ft", fontName="Helvetica-Oblique", fontSize=7, textColor=MGRAY)),
        Paragraph("<b>Authorised Signatory</b>",
                  ps("sg", fontName="Helvetica-Bold", fontSize=8, textColor=BLACK,
                     alignment=TA_RIGHT)),
    ]], colWidths=[USABLE_W * 0.65, USABLE_W * 0.35])
    ft.setStyle(TableStyle([
        ("LINEABOVE",     (0,0), (-1,0), 0.8, MLGRAY),
        ("VALIGN",        (0,0), (-1,0), "MIDDLE"),
        ("TOPPADDING",    (0,0), (-1,0), 6),
        ("BOTTOMPADDING", (0,0), (-1,0), 6),
    ]))
    story.append(ft)

    doc.build(story)
    buf.seek(0)
    return buf.read()


# ═══════════════════════════════════════════════════════════════════════════════
# SESSION STATE
# ═══════════════════════════════════════════════════════════════════════════════

def _ss(k, v):
    if k not in st.session_state: st.session_state[k] = v

for k, v in [
    ("step", 1), ("image_bytes", None), ("ocr_done", False),
    ("quantities", {}), ("match_log", []), ("ocr_grid", []),
    ("ocr_meta", {}), ("ocr_extracted", []), ("line_items", []),
    ("bill_to", {}), ("ship_to", {}), ("party_confirmed", False), ("pdf_bytes", None),
    ("azure_endpoint", AZURE_ENDPOINT), ("azure_key", AZURE_KEY),
]:
    _ss(k, v)

sku_sheets   = load_sku_sheets()
mrp_data     = load_mrp_data()
sku_master   = build_sku_master(sku_sheets)
prefix_index = build_prefix_index(sku_master)
customers    = load_customers()
cust_display = ["— Enter manually —"] + [c["display"] for c in customers]


# ═══════════════════════════════════════════════════════════════════════════════
# APP HEADER
# ═══════════════════════════════════════════════════════════════════════════════

st.markdown("""
<div class="app-header">
  <div class="app-header-badge">🔴</div>
  <div class="app-header-text">
    <h1>Sintex BAPL — Quotation Generator</h1>
    <p>CPVC / UPVC Pipes &amp; Fittings · Chhattisgarh · Any Form Layout</p>
  </div>
</div>""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
# CHANGE 2: TOP NAV TAB BAR — single page, steps unlock sequentially
# Steps unlock: Step 1 always open. Step 2 unlocks after image loaded.
# Step 3 unlocks after OCR done. Step 4 unlocks after party confirmed.
# ═══════════════════════════════════════════════════════════════════════════════

step_labels = ["📷 Capture", "⚡ OCR Review", "👤 Parties", "📄 Download"]
step_unlocked = [
    True,                                    # Step 1: always
    st.session_state.image_bytes is not None,  # Step 2: after image
    st.session_state.ocr_done,               # Step 3: after OCR
    st.session_state.party_confirmed,         # Step 4: after parties confirmed
]
cur = st.session_state.step

# Build HTML nav bar (display only — navigation via buttons inside each step)
nav_html = '<div class="step-navbar">'
for i, lbl in enumerate(step_labels, 1):
    is_active = (i == cur)
    is_done   = (i < cur)
    is_locked = not step_unlocked[i - 1]
    cls = "locked" if is_locked else ("active" if is_active else ("done" if is_done else ""))
    dot_txt = "✓" if is_done else str(i)
    nav_html += f"""
    <div class="step-nav-item {cls}">
      <div class="step-nav-dot">{dot_txt}</div>
      <div class="step-nav-label">{lbl}</div>
    </div>"""
nav_html += "</div>"
st.markdown(nav_html, unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 1 — CAPTURE
# ═══════════════════════════════════════════════════════════════════════════════

def render_step1():
    done = st.session_state.image_bytes is not None
    st.markdown(f"""
    <div class="step-card"><div class="step-card-header">
      <div class="step-number {'done' if done else ''}">{'✓' if done else '1'}</div>
      <div><div class="step-title">Step 1 — Capture Order Form</div>
      <div class="step-subtitle">Upload or photograph any Sintex order form — any layout</div></div>
    </div><div class="step-body">""", unsafe_allow_html=True)

    with st.expander("🔧 Azure OCR Settings", expanded=not st.session_state.azure_key):
        st.markdown('<div class="info-box">Endpoint: <code>https://&lt;resource&gt;.cognitiveservices.azure.com</code><br/>'
                    'Auto-tries Document Intelligence v4 → v3 → Form Recognizer v2.1</div>',
                    unsafe_allow_html=True)
        st.session_state.azure_endpoint = st.text_input(
            "Azure Endpoint", value=st.session_state.azure_endpoint,
            placeholder="https://YOUR-RESOURCE.cognitiveservices.azure.com")
        st.session_state.azure_key = st.text_input(
            "Azure API Key", value=st.session_state.azure_key,
            type="password", placeholder="••••••••••••••••••••••••••••••••")

    st.markdown("""<div class="info-box">
    📋 <b>Any form layout supported.</b> The OCR engine extracts bounding-box positions of every
    word, then <b>spatially reconstructs the table</b> — it finds which column has SKU prefixes,
    which have size headers (15MM/20MM/…), and reads quantities from the right cells.
    SKU prefixes are then matched to the MRP master via prefix+size lookup.</div>""",
    unsafe_allow_html=True)

    tab_cam, tab_up = st.tabs(["📷  Camera", "📁  Upload File"])
    with tab_cam:
        cam = st.camera_input("Order form", label_visibility="collapsed")
        if cam:
            for k in ["quantities", "match_log", "ocr_grid", "ocr_meta", "ocr_extracted", "line_items"]:
                st.session_state[k] = [] if isinstance(st.session_state[k], list) else {}
            st.session_state["image_bytes"] = cam.getvalue()
            st.session_state["ocr_done"]    = False
    with tab_up:
        uf = st.file_uploader("Image", type=["jpg","jpeg","png"], label_visibility="collapsed")
        if uf:
            for k in ["quantities", "match_log", "ocr_grid", "ocr_meta", "ocr_extracted", "line_items"]:
                st.session_state[k] = [] if isinstance(st.session_state[k], list) else {}
            st.session_state["image_bytes"] = uf.getvalue()
            st.session_state["ocr_done"]    = False

    if st.session_state.image_bytes:
        b64 = base64.b64encode(st.session_state.image_bytes).decode()
        st.markdown(f"""<div style="margin:12px 0;border-radius:10px;overflow:hidden;border:2px solid var(--mgray);">
          <img src="data:image/jpeg;base64,{b64}"
               style="width:100%;max-height:320px;object-fit:contain;display:block;background:#000;"/>
        </div>""", unsafe_allow_html=True)
        st.markdown('<div class="success-box">✅ Image loaded — ready for OCR</div>', unsafe_allow_html=True)
        if st.button("Proceed to OCR →  Step 2", key="goto2"):
            st.session_state.step = 2; st.rerun()

    st.markdown("</div></div>", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 2 — OCR + REVIEW
# ═══════════════════════════════════════════════════════════════════════════════

def render_step2():
    # Lock guard
    if not step_unlocked[1]:
        st.markdown('<div class="warn-box">🔒 Complete Step 1 first — upload or capture an image.</div>',
                    unsafe_allow_html=True)
        return

    done = st.session_state.ocr_done
    st.markdown(f"""
    <div class="step-card"><div class="step-card-header">
      <div class="step-number {'done' if done else ''}">{'✓' if done else '2'}</div>
      <div><div class="step-title">Step 2 — OCR · Spatial Reconstruction · SKU Matching</div>
      <div class="step-subtitle">Detects table structure from bounding boxes — works with any form layout</div></div>
    </div><div class="step-body">""", unsafe_allow_html=True)

    if not done:
        c1, c2 = st.columns(2)
        with c1:
            run = st.button("⚡  Run OCR & Match", key="run_ocr",
                            disabled=not (st.session_state.azure_key and st.session_state.azure_endpoint))
        with c2:
            skip = st.button("✏️  Enter SKUs Manually", key="skip_ocr")

        if not st.session_state.azure_key:
            st.markdown('<div class="warn-box">⚠️ Azure key not configured — set in Step 1 or use Manual entry.</div>',
                        unsafe_allow_html=True)

        if run:
            ph = st.empty()
            try:
                ph.info("⏳ 1/4 — Sending image to Azure Document Intelligence…")
                words = run_azure_ocr(st.session_state.image_bytes,
                                      st.session_state.azure_endpoint,
                                      st.session_state.azure_key)

                ph.info(f"🔲 2/4 — Reconstructing table from {len(words)} OCR words…")
                grid = reconstruct_table(words)

                ph.info(f"🔍 3/4 — Analysing table structure ({len(grid)} rows)…")
                res  = analyze_table(grid)

                ph.info("🔗 4/4 — Matching SKU prefixes + sizes to MRP master…")
                qtys, log, line_items = build_quantities(res["rows"], prefix_index, sku_master)

                st.session_state.ocr_grid      = grid
                st.session_state.ocr_meta      = res["meta"]
                st.session_state.ocr_extracted = res["rows"]
                st.session_state.quantities    = qtys
                st.session_state.match_log     = log
                st.session_state.line_items    = line_items
                st.session_state.ocr_done      = True

                nm = sum(1 for m in log if m["status"] == "matched")
                nu = sum(1 for m in log if m["status"] == "unmatched")
                if nm == 0:
                    ph.warning(f"⚠️ Table reconstructed ({len(grid)} rows, {len(res['rows'])} data rows) "
                               f"but 0 SKUs matched. Check Raw OCR Table & Match Log tabs below.")
                else:
                    ph.success(f"✅ Done — {nm} detections matched, {nu} unmatched. Review below.")
                time.sleep(0.8); st.rerun()

            except requests.exceptions.HTTPError as e:
                ph.empty()
                status = e.response.status_code if e.response else "?"
                try:    body = e.response.json().get("error", {}).get("message", "")
                except: body = (e.response.text or "")[:300]
                st.error(f"❌ Azure HTTP {status}: {body or str(e)}")
            except requests.exceptions.ConnectionError:
                ph.empty(); st.error("❌ Cannot connect to Azure — check endpoint URL.")
            except TimeoutError as e:
                ph.empty(); st.error(f"❌ Timeout: {e}")
            except Exception as e:
                ph.empty(); st.error(f"❌ {type(e).__name__}: {e}")

        if skip:
            st.session_state.quantities = {}
            st.session_state.match_log  = []
            st.session_state.ocr_done   = True
            st.rerun()

    # ── Results ──────────────────────────────────────────────────────────
    if done:
        grid = st.session_state.ocr_grid
        meta = st.session_state.ocr_meta
        log  = st.session_state.match_log
        qtys = st.session_state.quantities
        nm = sum(1 for m in log if m["status"] == "matched")
        nu = sum(1 for m in log if m["status"] == "unmatched")

        if nm: st.markdown(f'<div class="success-box">✅ {nm} SKU-size detections matched → {nm} line items in document</div>',
                           unsafe_allow_html=True)
        if nu: st.markdown(f'<div class="warn-box">⚠️ {nu} pair(s) unmatched — check Match Log tab</div>',
                           unsafe_allow_html=True)
        if not nm and not nu:
            st.markdown('<div class="warn-box">No OCR data — use Edit tab to add SKUs manually.</div>',
                        unsafe_allow_html=True)

        tab1, tab2, tab3, tab4 = st.tabs(["📋 Matched Items", "🔲 Raw OCR Table", "🔍 Match Log", "✏️ Edit / Add"])

        # ── Matched Items (CHANGE 3: show all individual line_items, no grouping) ──
        with tab1:
            line_items = st.session_state.line_items
            if line_items:
                h = """<div class="ocr-wrap"><table class="ocr-tbl">
                <thead><tr><th style="text-align:center;width:40px">#</th>
                <th class="L">Product</th><th class="L">SKU</th>
                <th>OD</th><th>Inch</th><th>Qty</th>
                <th>MRP(₹)</th><th>Dist.Landing(₹)</th></tr></thead><tbody>"""
                for idx, item in enumerate(line_items, 1):
                    sku  = item["sku"]; qty = item["qty"]
                    info = sku_master.get(sku, {}); mi = mrp_data.get(sku, {})
                    h += (f'<tr><td style="color:#999;font-size:10px">{idx}</td>'
                          f'<td class="L">{info.get("product", "?")}</td>'
                          f'<td class="M">{sku}</td><td>{info.get("od_size","")}</td>'
                          f'<td>{info.get("inch_size","")}</td>'
                          f'<td><b style="color:#C0211F">{qty}</b></td>'
                          f'<td>₹{mi.get("mrp",0):,.2f}</td>'
                          f'<td>₹{mi.get("distributor_landing",0):,.2f}</td></tr>')
                h += "</tbody></table></div>"
                st.markdown(h, unsafe_allow_html=True)
            else:
                st.markdown('<div class="warn-box">No matched items. Use Edit tab to add manually.</div>',
                            unsafe_allow_html=True)

        # ── Raw OCR Table ─────────────────────────────────────────────────
        with tab2:
            st.caption("Spatially reconstructed table from OCR bounding boxes. "
                       "🟠 = detected SKU column · 🟢 = quantity cells · "
                       "Bold row = detected header.")
            if grid:
                sc   = meta.get("size_cols", {})
                skuc = meta.get("sku_col")
                hrow = meta.get("header_row", 0)

                h = '<div class="raw-wrap"><table class="raw-tbl"><thead><tr>'
                h += "".join(f"<th>C{i}</th>" for i in range(len(grid[0])))
                h += "</tr></thead><tbody>"
                for ri, row in enumerate(grid):
                    h += '<tr class="hdr">' if ri == hrow else "<tr>"
                    for ci, cell in enumerate(row):
                        css = ""
                        if ci == skuc: css = 'class="sku"'
                        elif ci in sc and ri > hrow and _is_qty(cell): css = 'class="qty"'
                        h += f"<td {css}>{cell or '·'}</td>"
                    h += "</tr>"
                h += "</tbody></table></div>"
                st.markdown(h, unsafe_allow_html=True)
                st.markdown(
                    f'<div class="info-box">📊 {len(grid)} rows · {len(grid[0]) if grid else 0} cols · '
                    f'Header row: {hrow} · SKU col: {skuc} · '
                    f'Size cols: {dict(sorted(sc.items()))}</div>', unsafe_allow_html=True)
            else:
                st.info("Run OCR first.")

        # ── Match Log ─────────────────────────────────────────────────────
        with tab3:
            st.caption("Every SKU prefix + size → full SKU match attempt.")
            if log:
                h = """<div class="ocr-wrap"><table class="ocr-tbl">
                <thead><tr><th class="L">Product</th><th class="L">Prefix (OCR)</th>
                <th>Size</th><th>Qty</th><th class="L">Full SKU</th><th>Status</th>
                </tr></thead><tbody>"""
                for m in log:
                    cls = "ok" if m["status"] == "matched" else "no"
                    ico = "✅" if m["status"] == "matched" else "⚠️"
                    h += (f'<tr class="{cls}"><td class="L">{m["product"]}</td>'
                          f'<td class="M">{m["prefix"]}</td><td>{m["size"]}</td>'
                          f'<td><b>{m["qty"]}</b></td><td class="M">{m["full_sku"]}</td>'
                          f'<td>{ico}</td></tr>')
                h += "</tbody></table></div>"
                st.markdown(h, unsafe_allow_html=True)
            else:
                st.info("Run OCR first.")

        # ── Edit / Add ────────────────────────────────────────────────────
        with tab4:
            st.caption("Add or correct SKUs manually.")
            updated = copy.deepcopy(st.session_state.quantities)
            c1, c2, c3 = st.columns([3, 1, 1])
            with c1: si = st.text_input("SKU Code", key="mski", placeholder="e.g. CPF11BV00000015")
            with c2: qi = st.number_input("Qty", min_value=0, value=1, step=1, key="mqi")
            with c3:
                st.markdown("<br/>", unsafe_allow_html=True)
                if st.button("➕ Add", key="add_sku"):
                    st_clean = si.strip().upper()
                    if st_clean in sku_master:
                        updated[st_clean] = updated.get(st_clean, 0) + qi
                        st.session_state.quantities = updated
                        st.success(f"Added {st_clean} × {qi}"); st.rerun()
                    elif st_clean:
                        st.error(f"'{st_clean}' not in master.")

            items = {s: q for s, q in updated.items() if q > 0}
            if items:
                st.markdown('<div class="fsl">Edit Quantities</div>', unsafe_allow_html=True)
                il = list(items.items())
                for i in range(0, len(il), 3):
                    chunk = il[i:i+3]; rcols = st.columns(len(chunk))
                    for col, (sku, qty) in zip(rcols, chunk):
                        info = sku_master.get(sku, {})
                        lbl  = f"{info.get('product', sku)[:20]}\n{info.get('od_size','')}"
                        with col:
                            nq = st.number_input(lbl, min_value=0, value=int(qty),
                                                 step=1, key=f"eq_{sku}")
                            updated[sku] = nq
            st.session_state.quantities = updated

        # ── Totals ────────────────────────────────────────────────────────
        iord  = {s: q for s, q in st.session_state.quantities.items() if q > 0}
        gmrp  = sum(mrp_data.get(s, {}).get("mrp", 0) * q        for s, q in iord.items())
        gdist = sum(mrp_data.get(s, {}).get("distributor_landing", 0) * q for s, q in iord.items())
        disc  = gmrp - gdist
        st.markdown(f"""
        <div class="totals-box">
          <div class="total-row"><span class="total-lbl">Line items</span><span class="total-val">{len(st.session_state.line_items) if st.session_state.line_items else len(iord)}</span></div>
          <div class="total-row"><span class="total-lbl">Gross MRP</span><span class="total-val">₹ {gmrp:,.2f}</span></div>
          <div class="total-row"><span class="total-lbl">Distributor Discount</span><span class="total-val neg">− ₹ {disc:,.2f}</span></div>
          <div class="total-row grand"><span class="total-lbl">Net Distributor Landing</span><span class="total-val">₹ {gdist:,.2f}</span></div>
        </div>""", unsafe_allow_html=True)

        cb, cn = st.columns(2)
        with cb:
            st.markdown('<div class="btn-secondary">', unsafe_allow_html=True)
            if st.button("← Back", key="bk1"): st.session_state.step = 1; st.rerun()
            st.markdown("</div>", unsafe_allow_html=True)
        with cn:
            if st.button("Proceed to Party Details →  Step 3", key="goto3"):
                st.session_state.step = 3; st.rerun()

    st.markdown("</div></div>", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 3 — PARTY DETAILS
# ═══════════════════════════════════════════════════════════════════════════════

def render_step3():
    # Lock guard
    if not step_unlocked[2]:
        st.markdown('<div class="warn-box">🔒 Complete Step 2 (OCR Review) first.</div>',
                    unsafe_allow_html=True)
        return

    done = st.session_state.party_confirmed
    st.markdown(f"""
    <div class="step-card"><div class="step-card-header">
      <div class="step-number {'done' if done else ''}">{'✓' if done else '3'}</div>
      <div><div class="step-title">Step 3 — Dealer / Distributor Details</div>
      <div class="step-subtitle">Bill To &amp; Ship To party information</div></div>
    </div><div class="step-body">""", unsafe_allow_html=True)

    def party_form(pfx, title, emoji):
        st.markdown(f'<div class="party-section"><div class="party-title">{emoji} {title}</div>',
                    unsafe_allow_html=True)
        sel = st.selectbox(f"Lookup ({title})", cust_display,
                           key=f"{pfx}_sel", label_visibility="collapsed")
        d   = next((c for c in customers if c["display"] == sel), {}) if sel != "— Enter manually —" else {}
        c1, c2 = st.columns(2)
        with c1: pno = st.text_input("Party No.", value=d.get("code",""), key=f"{pfx}_pno")
        with c2: gst = st.text_input("GST No.",   value=d.get("gst",""),  key=f"{pfx}_gst")
        na = st.text_area("Name & Address",
                          value=f"{d.get('name','')}\n{d.get('address','')}".strip(),
                          key=f"{pfx}_na", height=80)
        c3, c4 = st.columns(2)
        with c3: ph = st.text_input("Phone",  value=d.get("phone",""),  key=f"{pfx}_ph")
        with c4: mb = st.text_input("Mobile", value=d.get("mobile",""), key=f"{pfx}_mb")
        c5, c6 = st.columns(2)
        with c5: sc = st.text_input("State Code", value=d.get("state_code",""), key=f"{pfx}_sc")
        with c6: st_ = st.text_input("State",      value=d.get("state",""),      key=f"{pfx}_st")
        pan = st.text_input("PAN No.", value=d.get("pan",""), key=f"{pfx}_pan")
        st.markdown("</div>", unsafe_allow_html=True)
        return {"party_no": pno, "name": na, "phone": ph, "mobile": mb,
                "state_code": sc, "state": st_, "gst": gst, "pan": pan}

    bill_to = party_form("bill", "Bill To Party", "🏢")
    ship_to = party_form("ship", "Ship To Party", "🚚")

    cb, cc = st.columns(2)
    with cb:
        st.markdown('<div class="btn-secondary">', unsafe_allow_html=True)
        if st.button("← Back", key="bk2"): st.session_state.step = 2; st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)
    with cc:
        if st.button("Confirm & Generate PDF →  Step 4", key="confirm"):
            st.session_state.bill_to         = bill_to
            st.session_state.ship_to         = ship_to
            st.session_state.party_confirmed = True
            with st.spinner("Generating PDF…"):
                try:
                    pdf = build_pdf(st.session_state.quantities, mrp_data,
                                    bill_to, ship_to, sku_master,
                                    st.session_state.get("line_items", []))
                    st.session_state.pdf_bytes = pdf
                    st.session_state.step = 4; st.rerun()
                except Exception as e:
                    st.error(f"PDF error: {e}")

    st.markdown("</div></div>", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
# STEP 4 — PREVIEW + DOWNLOAD
# CHANGE 5: PDF preview section added before download button
# ═══════════════════════════════════════════════════════════════════════════════

def render_step4():
    # Lock guard
    if not step_unlocked[3]:
        st.markdown('<div class="warn-box">🔒 Complete Step 3 (Party Details) first.</div>',
                    unsafe_allow_html=True)
        return

    st.markdown("""
    <div class="step-card"><div class="step-card-header">
      <div class="step-number done">✓</div>
      <div><div class="step-title">Step 4 — Preview &amp; Download Quotation</div>
      <div class="step-subtitle">Review your Sales Quotation PDF before downloading</div></div>
    </div><div class="step-body">""", unsafe_allow_html=True)

    # Use line_items for totals to match PDF
    line_items = st.session_state.get("line_items", [])
    if line_items:
        items_for_total = line_items
        gmrp  = sum(mrp_data.get(it["sku"], {}).get("mrp", 0) * it["qty"] for it in items_for_total)
        gdist = sum(mrp_data.get(it["sku"], {}).get("distributor_landing", 0) * it["qty"] for it in items_for_total)
        n_lines = len(items_for_total)
    else:
        iord  = {s: q for s, q in st.session_state.quantities.items() if q > 0}
        gmrp  = sum(mrp_data.get(s, {}).get("mrp", 0) * q        for s, q in iord.items())
        gdist = sum(mrp_data.get(s, {}).get("distributor_landing", 0) * q for s, q in iord.items())
        n_lines = len(iord)
    disc = gmrp - gdist

    st.markdown(f"""
    <div class="success-box">✅ Quotation ready — <b>{n_lines} line items</b></div>
    <div class="totals-box">
      <div class="total-row"><span class="total-lbl">Total Line Items</span><span class="total-val">{n_lines}</span></div>
      <div class="total-row"><span class="total-lbl">Gross MRP</span><span class="total-val">₹ {gmrp:,.2f}</span></div>
      <div class="total-row"><span class="total-lbl">Distributor Discount</span><span class="total-val neg">− ₹ {disc:,.2f}</span></div>
      <div class="total-row grand"><span class="total-lbl">Net Taxable Value</span><span class="total-val">₹ {gdist:,.2f}</span></div>
    </div>""", unsafe_allow_html=True)

    bill = st.session_state.bill_to
    ship = st.session_state.ship_to

    # ── CHANGE 5: PDF Preview section ─────────────────────────────────────
    if st.session_state.pdf_bytes:
        st.markdown("""
        <div style="margin:20px 0 8px;">
          <div style="font-size:13px;font-weight:700;color:#1A1A1A;margin-bottom:4px;">
            📄 Document Preview
          </div>
          <div style="font-size:11.5px;color:#6B6B6B;margin-bottom:10px;">
            Review the quotation below. Use the viewer controls to navigate pages.
          </div>
        </div>""", unsafe_allow_html=True)

        b64 = base64.b64encode(st.session_state.pdf_bytes).decode()
        # Embed PDF using an iframe with the data URI
        pdf_display = f"""
        <div class="pdf-preview-wrap">
          <iframe
            src="data:application/pdf;base64,{b64}#toolbar=1&navpanes=1&scrollbar=1&view=FitH"
            width="100%"
            height="620px"
            style="border:none;display:block;"
            type="application/pdf"
          >
            <p style="color:white;padding:20px;text-align:center;">
              Your browser does not support embedded PDF preview.<br/>
              Please use the download button below.
            </p>
          </iframe>
        </div>"""
        st.markdown(pdf_display, unsafe_allow_html=True)

        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

        # ── Download buttons ───────────────────────────────────────────────
        st.markdown(f"""
        <a class="dl-btn" href="data:application/pdf;base64,{b64}" download="sintex_quotation.pdf">
          📥 &nbsp; Download Sales Quotation PDF
        </a>""", unsafe_allow_html=True)

    cb, cn = st.columns(2)
    with cb:
        st.markdown('<div class="btn-secondary">', unsafe_allow_html=True)
        if st.button("← Edit Parties", key="bk3"):
            st.session_state.step = 3; st.session_state.party_confirmed = False; st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)
    with cn:
        if st.button("🔄  New Quotation", key="newq"):
            for k in ["image_bytes", "pdf_bytes"]: st.session_state[k] = None
            for k in ["quantities", "bill_to", "ship_to", "ocr_meta"]: st.session_state[k] = {}
            for k in ["match_log", "ocr_grid", "ocr_extracted", "line_items"]: st.session_state[k] = []
            st.session_state["ocr_done"] = st.session_state["party_confirmed"] = False
            st.session_state["step"] = 1; st.rerun()

    st.markdown("</div></div>", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
# ROUTER — single page, all steps render below nav bar
# CHANGE 2: Each step renders in-place on same page based on session state
# ═══════════════════════════════════════════════════════════════════════════════

s = st.session_state.step
if s == 1:
    render_step1()
elif s == 2:
    render_step2()
elif s == 3:
    render_step3()
elif s == 4:
    render_step4()

st.markdown("""
<div style="text-align:center;padding:32px 0 8px;font-size:11px;color:#AAA;">
  Sintex BAPL Limited &nbsp;·&nbsp; CPVC / UPVC Quotation System &nbsp;·&nbsp; Chhattisgarh
</div>""", unsafe_allow_html=True)