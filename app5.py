import io
import os
import re
import base64
from datetime import date
from difflib import SequenceMatcher

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

try:
    import requests as _requests
    _HAS_REQUESTS = True
except ImportError:
    _HAS_REQUESTS = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, HRFlowable)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    _HAS_REPORTLAB = True
except ImportError:
    _HAS_REPORTLAB = False

st.set_page_config(
    page_title="Sintex BAPL – Quotation Generator",
    page_icon="🔧",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@300;400;500;600;700&family=IBM+Plex+Mono:wght@400;500&display=swap');
:root {
  --navy:#0A2342; --blue:#1565C0; --sky:#1E88E5; --teal:#00796B;
  --gold:#F9A825; --danger:#C62828; --border:#DEE3EC;
  --surface:#F4F6FA; --text:#1A1F36; --muted:#5A6880; --radius:10px;
}
html,body,[class*="css"]{font-family:'IBM Plex Sans',sans-serif!important;background:#F0F2F8!important;color:var(--text);}
#MainMenu,footer,header{visibility:hidden;}
.block-container{padding:1.2rem 1.5rem 5rem!important;max-width:1300px!important;}

.app-header{background:linear-gradient(135deg,#0A2342,#1565C0);border-radius:12px;
  padding:20px 28px;display:flex;align-items:center;gap:18px;margin-bottom:28px;
  box-shadow:0 4px 20px rgba(10,35,66,.28);}
.app-header h1{font-size:21px;font-weight:700;color:#fff;margin:0;}
.app-header p{font-size:12.5px;color:rgba(255,255,255,.6);margin:4px 0 0;}

.sec-hdr{display:flex;align-items:center;gap:14px;padding:14px 22px;
  border-radius:10px;margin:28px 0 16px;}
.sec-hdr.active{background:linear-gradient(135deg,#0A2342,#1565C0);box-shadow:0 3px 14px rgba(10,35,66,.22);}
.sec-hdr.done{background:linear-gradient(135deg,#00695C,#00897B);}
.sec-hdr.locked{background:#E8ECF4;}
.sec-hdr .num{width:30px;height:30px;border-radius:50%;display:inline-flex;
  align-items:center;justify-content:center;font-size:13px;font-weight:800;flex-shrink:0;}
.sec-hdr.active .num,.sec-hdr.done .num{background:rgba(255,255,255,.2);color:#fff;}
.sec-hdr.locked .num{background:#C8D0E0;color:#7A88A0;}
.sec-hdr h2{font-size:15px;font-weight:700;margin:0;}
.sec-hdr.active h2,.sec-hdr.done h2{color:#fff;}
.sec-hdr.locked h2{color:#9AACBF;}
.locked-msg{color:#9AACBF;font-size:13px;padding:0 4px;margin-top:-8px;margin-bottom:20px;}

.ocr-tbl{width:100%;border-collapse:collapse;font-size:12px;margin-top:8px;}
.ocr-tbl th{background:var(--navy);color:#fff;padding:8px 12px;font-weight:700;
  font-size:11px;text-align:left;border:1px solid rgba(255,255,255,.12);white-space:nowrap;}
.ocr-tbl th.r{text-align:right;}
.ocr-tbl td{padding:7px 12px;border:1px solid var(--border);vertical-align:middle;}
.ocr-tbl td.r{text-align:right;font-family:'IBM Plex Mono',monospace;}
.ocr-tbl tr:nth-child(even) td{background:#F7F9FC;}
.sku-mono{font-family:'IBM Plex Mono',monospace;font-size:10.5px;color:var(--blue);}
.fp{font-weight:700;color:var(--navy);}
.ok{color:var(--teal);font-weight:700;font-size:10px;}
.warn{color:#E65100;font-weight:700;font-size:10px;}
.none{color:var(--danger);font-weight:700;font-size:10px;}
.ocr-tbl .muted td{opacity:.45;}
.ocr-tbl tfoot td{background:#EBF3FF!important;font-weight:700;color:var(--navy);border-top:2px solid var(--navy);}

.grp-lbl{display:inline-block;padding:3px 12px;border-radius:20px;font-size:11px;font-weight:700;margin:10px 0 6px;}
.grp-det{background:#E8F5E9;color:#00695C;}
.grp-miss{background:#FFF3E0;color:#E65100;}
.grp-man{background:#EBF3FF;color:#1565C0;}

.tot-row{display:flex;justify-content:space-between;align-items:center;
  padding:9px 0;border-bottom:1px solid var(--border);font-size:13.5px;}
.tot-row:last-child{border:none;padding-top:12px;}
.tot-row.net{font-weight:700;font-size:16px;color:var(--navy);}
.tot-row .val{font-family:'IBM Plex Mono',monospace;font-weight:600;}
.tot-row .neg{color:var(--danger);}
.card{background:#fff;border:1px solid var(--border);border-radius:var(--radius);
  padding:18px 22px;margin-bottom:16px;box-shadow:0 1px 4px rgba(10,35,66,.06);}
.card-title{font-size:12px;font-weight:700;text-transform:uppercase;letter-spacing:.7px;color:var(--navy);margin-bottom:12px;}

.party-box{border:1px solid var(--border);border-radius:8px;padding:14px 16px;background:var(--surface);}
.party-box h4{margin:0 0 10px;font-size:11.5px;font-weight:700;text-transform:uppercase;letter-spacing:.5px;color:var(--navy);}
.p-row{display:flex;gap:6px;margin-bottom:5px;font-size:12.5px;}
.p-lbl{color:var(--muted);min-width:80px;flex-shrink:0;}
.p-val{font-weight:500;color:var(--text);word-break:break-word;}

.stButton>button{font-family:'IBM Plex Sans',sans-serif!important;font-weight:600!important;
  font-size:13.5px!important;border-radius:8px!important;border:none!important;
  background:var(--navy)!important;color:#fff!important;padding:10px 22px!important;
  box-shadow:0 2px 8px rgba(10,35,66,.18)!important;transition:all .15s!important;}
.stButton>button:hover{background:var(--sky)!important;transform:translateY(-1px)!important;}
label{font-size:12px!important;font-weight:600!important;color:var(--muted)!important;}
.stNumberInput>div>div>input{border-radius:6px!important;border:1.5px solid var(--border)!important;
  font-family:'IBM Plex Mono',monospace!important;font-size:13px!important;
  padding:4px 6px!important;text-align:center!important;}

.debug-box{background:#1A1F36;color:#80CBC4;font-family:'IBM Plex Mono',monospace;
  font-size:11px;padding:12px 16px;border-radius:8px;overflow-x:auto;margin-top:8px;
  border:1px solid #2A3A5A;max-height:300px;overflow-y:auto;}
</style>
""", unsafe_allow_html=True)

# ─── Paths ───────────────────────────────────────────────────────────────────
_HERE        = os.path.dirname(os.path.abspath(__file__))
MRP_CSV_PATH  = os.path.join(_HERE, "MRP_State_chhattisghar.csv")
ZSD_CSV_PATH  = os.path.join(_HERE, "ZSD_CUST.csv")
SKU_XLSX_PATH = os.path.join(_HERE, "Sample form for Product list.xlsx")

# ─── Data loaders ────────────────────────────────────────────────────────────
@st.cache_data
def load_mrp():
    df = pd.read_csv(MRP_CSV_PATH)
    df["MRP_clean"] = (df["MRP(ZPR1-933)"].astype(str)
                       .str.replace(",","",regex=False)
                       .pipe(pd.to_numeric,errors="coerce"))
    df["Distributor Landing"] = pd.to_numeric(
        df["Distributor Landing"].astype(str).str.replace(",","",regex=False),errors="coerce")
    return {str(r["Material Number"]).strip(): r.to_dict() for _,r in df.iterrows()}

@st.cache_data
def load_zsd():
    return pd.read_csv(ZSD_CSV_PATH, encoding="latin1")

@st.cache_data
def load_sku_sheets():
    wb  = load_workbook(SKU_XLSX_PATH, read_only=True)
    out = {}
    for sname in wb.sheetnames:
        ws   = wb[sname]
        data = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]
        if len(data)<3: continue
        col_ids    = [str(c).strip() if c is not None else "" for c in data[1][2:]]
        col_labels = [str(c).strip() if c is not None else "" for c in data[2][2:]]
        while col_ids and col_ids[-1]=="": col_ids.pop(); col_labels.pop()
        rows=[]; section="General"
        for raw in data[3:]:
            if not any(c is not None for c in raw): continue
            label = str(raw[0]).strip() if raw[0] is not None else ""
            if label.upper() in ("FITTINGS","PIPES","FITTING SCH 80") and all(c is None for c in raw[2:]):
                section=label.title(); continue
            if not label: continue
            skus=[str(raw[i+2]).strip() if (i+2)<len(raw) and raw[i+2] is not None else None
                  for i in range(len(col_ids))]
            rows.append({"label":label,"section":section,"skus":skus})
        out[sname]={"label":sname,"col_ids":col_ids,"col_labels":col_labels,"rows":rows}
    return out

@st.cache_data
def build_sku_catalogue():
    sheets=load_sku_sheets(); cat={}
    for sname,sheet in sheets.items():
        for ri,row in enumerate(sheet["rows"]):
            for ci,sku in enumerate(row["skus"]):
                if not sku or sku in ("-","None",""): continue
                cat[sku]={"sheet":sname,"row_label":row["label"],"section":row["section"],
                          "col_id":sheet["col_ids"][ci] if ci<len(sheet["col_ids"]) else "",
                          "col_label":sheet["col_labels"][ci] if ci<len(sheet["col_labels"]) else "",
                          "ri":ri,"ci":ci}
    return cat

# ─── SKU matching ─────────────────────────────────────────────────────────────
_OCR_FIXES=str.maketrans({'O':'0','I':'1','S':'5','B':'8','G':'6','Z':'2'})

def _norm(raw):
    s=re.sub(r"[\s\-\._]","",raw).upper()
    m=re.match(r'^([A-Z]+)(.*)$',s)
    return (m.group(1)+m.group(2).translate(_OCR_FIXES)) if m else s.translate(_OCR_FIXES)

@st.cache_data
def _norm_idx(keys:tuple):
    return {_norm(k):k for k in keys}

def match_sku(raw,cat,ni,thr=0.75):
    if not raw: return None,0.0,"none"
    nr=_norm(raw); rl=len(nr)
    if nr in ni: return ni[nr],1.0,"exact"
    if rl>=8:
        cands=[(o,rl/max(len(nc),1)) for nc,o in ni.items() if nc.startswith(nr)]
        if cands:
            best=max(cands,key=lambda x:x[1])
            if best[1]>=0.55: return best[0],min(0.97,best[1]+0.3),"exact"
        sub=[(o,rl/max(len(nc),1)) for nc,o in ni.items() if nr in nc]
        if sub:
            best=max(sub,key=lambda x:x[1])
            if best[1]>=0.50: return best[0],min(0.90,best[1]+0.2),"fuzzy"
    if rl<7 or not(any(c.isdigit() for c in nr) and any(c.isalpha() for c in nr)):
        return None,0.0,"none"
    bk,bs=None,0.0
    for nc,o in ni.items():
        s=SequenceMatcher(None,nr,nc).ratio()
        if s>bs: bs=s; bk=o
    if bs>=thr: return bk,bs,"fuzzy"
    return None,bs,"none"

# ─── OCR helpers ─────────────────────────────────────────────────────────────
# FIX: Expanded size tokens to cover fractional inch notation used on order forms
_SIZES_MM  = ["15MM","20MM","25MM","32MM","40MM","50MM","65MM","80MM","100MM"]
_SIZES_INCH = ['1/2"','3/4"','1"','1-1/4"','1-1/2"','2"','2-1/2"','3"','4"',
               "1/2IN","3/4IN","1IN","1.25IN","1.5IN","2IN",
               "15","20","25","32","40","50"]
_SIZES = _SIZES_MM + _SIZES_INCH

# Map common inch notations back to MM for consistency
_INCH_TO_MM = {
    '1/2"':"15MM",'3/4"':"20MM",'1"':"25MM",'1-1/4"':"32MM",
    '1-1/2"':"40MM",'2"':"50MM",'2-1/2"':"65MM",'3"':"80MM",'4"':"100MM",
    "1/2IN":"15MM","3/4IN":"20MM","1IN":"25MM","1.25IN":"32MM",
    "1.5IN":"40MM","2IN":"50MM",
    "15":"15MM","20":"20MM","25":"25MM","32":"32MM","40":"40MM","50":"50MM",
}

def _normalise_size(s):
    """Normalise any size token to the MM form."""
    u = s.upper().strip()
    return _INCH_TO_MM.get(u, u)


def _lp(tok):
    """Is this token likely a (partial) SKU code?"""
    t = tok.strip()
    return (
        len(t) >= 5 and                          # FIX: lowered from 6 → catches 5-char prefixes
        any(c.isdigit() for c in t) and
        any(c.isalpha() for c in t) and
        not t.isalpha() and
        t.upper() not in _SIZES and
        not re.match(r'^\d+[\.,]\d+$', t)        # exclude decimal numbers like "140.00"
    )


def _is_qty(tok):
    """
    FIX (CORE): A token is a candidate quantity if it is a plain integer
    OR starts with digits and ends with common suffixes (Nos, Pcs, No).
    Rejects prices (contain '.') and pure alphabetic tokens.
    """
    t = tok.strip()
    if not t:
        return False, 0
    # Plain integer
    if re.match(r'^\d+$', t):
        v = int(t)
        # Reject very large numbers that are likely MRP/price codes
        if v > 9999:
            return False, 0
        return True, v
    # Integer with trailing unit suffix  e.g. "10Nos", "5 Pcs"
    m = re.match(r'^(\d+)\s*(?:nos?|pcs?|no\.?|units?)$', t, re.IGNORECASE)
    if m:
        return True, int(m.group(1))
    return False, 0


def _xc(p):
    if not p: return 0.0
    xs=[v["x"] for v in p] if isinstance(p[0],dict) else [p[i] for i in range(0,len(p),2)]
    return (min(xs)+max(xs))/2

def _yc(p):
    if not p: return 0.0
    ys=[v["y"] for v in p] if isinstance(p[0],dict) else [p[i] for i in range(1,len(p),2)]
    return (min(ys)+max(ys))/2


def _rows(words, gap=0.030):
    """
    FIX: increased default gap from 0.015 → 0.025 so words on the same
    physical line (slight y-jitter from OCR) are grouped together.
    """
    if not words:
        return []
    words = sorted(words, key=lambda w: w["y"])
    rows = []; cur = [words[0]]
    for w in words[1:]:
        if w["y"] - cur[-1]["y"] > gap:
            rows.append(cur); cur = [w]
        else:
            cur.append(w)
    rows.append(cur)
    return rows


def _geo(words):
    """
    Improved geometric parser.

    Strategy:
    1. Try to find a size-header row (≥2 recognised size tokens on one row).
    2. If found → use column-position assignment to map qty cells to sizes.
    3. If not found → fall back to _fb_enhanced() for row-by-row parsing.

    FIX: Also tries fractional-inch and bare-number size tokens, not just MM.
    """
    if not words:
        return []
    rs = _rows(words)

    # ── 1. Locate header row ──────────────────────────────────────────────
    cx = {}; hr = None
    for row in rs:
        texts = [w["text"].upper().strip() for w in row]
        # Check against all known size variants
        size_hits = [w for w in row if w["text"].upper().strip() in _SIZES]
        if len(size_hits) >= 2:
            hr = row
            for w in size_hits:
                norm_sz = _normalise_size(w["text"])
                cx[norm_sz] = w["x"]
            break

    res = []; seen = set()

    if cx:
        # ── 2. Geometric / column-assignment mode ─────────────────────────
        ss = sorted(cx, key=lambda s: cx[s])
        xs = [cx[s] for s in ss]
        cw = (max(xs) - min(xs)) / max(len(xs) - 1, 1) if len(xs) > 1 else 0.05

        def asgn(x):
            bd, bs = float("inf"), None
            for s in ss:
                d = abs(cx[s] - x)
                if d < bd:
                    bd = d; bs = s
            return bs if bd <= max(cw * 0.75, 0.03) else None

        # SKU tokens live to the left of the first size column
        xmax = min(xs) - 0.01

        for row in rs:
            if row is hr:
                continue
            rw = sorted(row, key=lambda w: w["x"])
            raw_line = " ".join(w["text"] for w in rw)

            sku_cands = [w for w in rw if w["x"] <= xmax and _lp(w["text"])]
            if not sku_cands:
                continue

            rsku = sku_cands[0]["text"]

            # FIX: use _is_qty() instead of bare regex
            qt = []
            for w in rw:
                if w["x"] > xmax:
                    ok, v = _is_qty(w["text"])
                    if ok:
                        qt.append((w, v))

            if not qt:
                k = (rsku.upper(), "", raw_line[:20])
                if k not in seen:
                    seen.add(k)
                    res.append({"raw_sku": rsku, "raw_size": "", "qty": 0, "raw_line": raw_line})
                continue

            for q_word, qv in qt:
                sz = asgn(q_word["x"])
                if sz is None:
                    continue
                k = (rsku.upper(), sz)
                if k in seen:
                    continue
                seen.add(k)
                res.append({"raw_sku": rsku, "raw_size": sz, "qty": qv, "raw_line": raw_line})
    else:
        # ── 3. No header — row-by-row fallback ───────────────────────────
        res = _fb_enhanced(words)

    return res


def _fb_enhanced(words):
    """
    FIX (CORE REWRITE): Enhanced fallback parser for forms without a size header.

    For each row:
      • Find the leftmost SKU-like token.
      • Find size tokens on the row.
      • Find ALL qty-like tokens to the RIGHT of the SKU token using _is_qty().
      • Match each qty to its nearest size token (if any).
      • Record one entry per (sku, size/positional-index).

    Key improvements over original:
      - Uses _is_qty() which handles "10Nos", "5Pcs" etc.
      - Does NOT exclude qty tokens that are also to the right of a size token
        — many forms print: SKU ... SIZE ... QTY
      - Records qty even when it appears anywhere to the right of the SKU,
        not just immediately adjacent.
    """
    rs = _rows(words)
    res = []; seen = set()

    for row in rs:
        rw = sorted(row, key=lambda w: w["x"])
        raw_line = " ".join(w["text"] for w in rw)

        # ── Find SKU-like tokens ──────────────────────────────────────────
        sku_toks = [w for w in rw if _lp(w["text"])]
        if not sku_toks:
            continue

        rsku = sku_toks[0]["text"]
        sku_x = sku_toks[0]["x"]

        # ── Find size tokens anywhere on this row ─────────────────────────
        size_toks = [w for w in rw if w["text"].upper().strip() in _SIZES]

        # ── Find qty tokens to the RIGHT of the SKU token ─────────────────
        qty_toks = []
        for w in rw:
            if w["x"] <= sku_x:
                continue
            ok, v = _is_qty(w["text"])
            if ok:
                qty_toks.append((w, v))

        if not qty_toks:
            # Record the SKU with qty=0 (user can edit manually)
            k = rsku.upper()
            if k not in seen:
                seen.add(k)
                sz = _normalise_size(size_toks[0]["text"]) if size_toks else ""
                res.append({"raw_sku": rsku, "raw_size": sz, "qty": 0, "raw_line": raw_line})
            continue

        if size_toks:
            # Match each qty to its nearest size token by x distance
            for q_word, qv in qty_toks:
                nearest_sz = min(size_toks, key=lambda w: abs(w["x"] - q_word["x"]))
                sz = _normalise_size(nearest_sz["text"])
                k = (rsku.upper(), sz)
                if k in seen:
                    continue
                seen.add(k)
                res.append({"raw_sku": rsku, "raw_size": sz, "qty": qv, "raw_line": raw_line})
        else:
            # No size info on this row — record each qty with a positional suffix
            for qi, (q_word, qv) in enumerate(qty_toks):
                k = (rsku.upper(), f"_pos{qi}")
                if k in seen:
                    continue
                seen.add(k)
                res.append({"raw_sku": rsku, "raw_size": "", "qty": qv, "raw_line": raw_line})

    return res


# ─── Azure OCR drivers ────────────────────────────────────────────────────────

def _extract_words_from_di(res):
    """Extract normalised word dicts from Document Intelligence response."""
    words = []
    for pg in res.get("analyzeResult", {}).get("pages", []):
        pw = pg.get("width", 1) or 1
        ph = pg.get("height", 1) or 1
        for w in pg.get("words", []):
            t = w.get("content", "").strip()
            if t:
                words.append({
                    "text": t,
                    "x": _xc(w.get("polygon", [])) / pw,
                    "y": _yc(w.get("polygon", [])) / ph,
                })
    return words


def _extract_words_from_cv(res):
    """Extract normalised word dicts from Computer Vision Read response."""
    words = []
    for pg in res.get("analyzeResult", {}).get("readResults", []):
        pw = pg.get("width", 1) or 1
        ph = pg.get("height", 1) or 1
        for ln in pg.get("lines", []):
            for w in ln.get("words", []):
                t = w.get("text", "").strip()
                if not t:
                    continue
                bb = w.get("boundingBox", [])
                xs = [bb[i] for i in range(0, len(bb), 2)] if bb else [0]
                ys = [bb[i] for i in range(1, len(bb), 2)] if bb else [0]
                words.append({
                    "text": t,
                    "x": ((min(xs) + max(xs)) / 2) / pw,
                    "y": ((min(ys) + max(ys)) / 2) / ph,
                })
    return words


def _poll(op_url, key, retries=25, delay=1.5):
    """Poll an Azure async operation URL until succeeded or timeout."""
    import time
    hdr = {"Ocp-Apim-Subscription-Key": key}
    for _ in range(retries):
        time.sleep(delay)
        r = _requests.get(op_url, headers=hdr, timeout=15)
        res = r.json()
        if res.get("status") == "succeeded":
            return res
    return res  # return last result even if not succeeded


def _di(ib, ep, key):
    """Document Intelligence (Form Recognizer) driver."""
    url = (ep.rstrip("/") +
           "/formrecognizer/documentModels/prebuilt-read:analyze"
           "?api-version=2023-07-31")
    h = {"Ocp-Apim-Subscription-Key": key, "Content-Type": "application/octet-stream"}
    r = _requests.post(url, headers=h, data=ib, timeout=30)
    r.raise_for_status()
    res = _poll(r.headers["Operation-Location"], key)
    return _geo(_extract_words_from_di(res))


def _cv(ib, ep, key):
    """Computer Vision v3.2 Read driver."""
    url = ep.rstrip("/") + "/vision/v3.2/read/analyze"
    h = {"Ocp-Apim-Subscription-Key": key, "Content-Type": "application/octet-stream"}
    r = _requests.post(url, headers=h, data=ib, timeout=30)
    r.raise_for_status()
    res = _poll(r.headers["Operation-Location"], key)
    return _geo(_extract_words_from_cv(res))


def run_ocr(ib, ep, key):
    if not _HAS_REQUESTS:
        return []
    try:
        return _di(ib, ep, key)
    except Exception as e:
        s = getattr(getattr(e, "response", None), "status_code", None)
        if s in (401, 403, 404):
            return _cv(ib, ep, key)
        raise


# ─── Resolve OCR rows → catalogue + MRP ──────────────────────────────────────

def resolve_rows(raw, cat, mrp):
    """
    Map each raw OCR row to a catalogue entry and MRP price.

    FIX: De-duplication now uses (raw_sku, raw_size, qty) so the same SKU
    appearing with different sizes, or with genuine different quantities,
    is kept as separate line items.
    """
    ni = _norm_idx(tuple(cat.keys()))
    res = []
    seen = set()

    for r in raw:
        rs  = r["raw_sku"]
        rz  = r.get("raw_size", "")
        q   = r["qty"]
        raw_line = r.get("raw_line", "")

        dk = (rs.upper(), rz.upper(), q)
        if dk in seen:
            continue
        seen.add(dk)

        ms, cf, mt = match_sku(rs, cat, ni)

        # FIX: If we have a size, try to find the exact size-variant SKU
        if rz and ms:
            sn = re.sub(r"[^0-9]", "", rz)          # e.g. "15MM" → "15"
            bn = _norm(rs)
            hits = [
                (s, e) for s, e in cat.items()
                if _norm(s).startswith(bn) and
                re.sub(r"[^0-9]", "", e.get("col_label", "")) == sn
            ]
            if hits:
                ms = hits[0][0]; mt = "exact"; cf = 1.0

        info = mrp.get(ms, {}) if ms else {}
        mv   = info.get("MRP_clean") or 0.0
        lv   = info.get("Distributor Landing") or mv
        ce   = cat.get(ms, {}) if ms else {}

        res.append({
            "raw_sku":    rs,
            "raw_size":   rz,
            "matched_sku": ms or "—",
            "match_type":  mt,
            "confidence":  round(cf * 100, 1),
            "qty":         q,
            "sheet":       ce.get("sheet", "—"),
            "row_label":   ce.get("row_label", "—"),
            "col_label":   ce.get("col_label", rz),
            "mrp":         mv,
            "landing":     lv,
            "amount":      round(lv * q, 2),
            "ri":          ce.get("ri"),
            "ci":          ce.get("ci"),
        })

    return res


# ─── PDF generation ───────────────────────────────────────────────────────────
def gen_pdf(rows, mrp_lkp, bill, ship, dealer):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=10*mm, rightMargin=10*mm,
                            topMargin=12*mm, bottomMargin=12*mm)
    navy = colors.HexColor("#0A2342")
    ts   = ParagraphStyle
    story = []

    hdr = Table([[
        Paragraph("SINTEX BAPL LIMITED",
                  ts("t", fontName="Helvetica-Bold", fontSize=14,
                     textColor=colors.white, alignment=TA_CENTER)),
        Paragraph("Kutesar Road, Raipur, Chhattisgarh – 492101<br/>GSTIN: 22AADCB1921F1ZE",
                  ts("s", fontName="Helvetica", fontSize=8,
                     textColor=colors.white, alignment=TA_CENTER)),
    ]], colWidths=["40%","60%"])
    hdr.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1),navy),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("TOPPADDING",(0,0),(-1,-1),8),
        ("BOTTOMPADDING",(0,0),(-1,-1),8),
        ("LEFTPADDING",(0,0),(-1,-1),10),
    ]))
    story.append(hdr)
    story.append(Spacer(1, 4*mm))

    qt = Table([[
        Paragraph("<b>QUOTATION</b>",
                  ts("q", fontName="Helvetica-Bold", fontSize=13, textColor=navy)),
        Paragraph(f"<b>Date:</b> {date.today().strftime('%d-%m-%Y')}",
                  ts("d", fontName="Helvetica", fontSize=8, alignment=TA_RIGHT)),
    ]], colWidths=["60%","40%"])
    qt.setStyle(TableStyle([
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("LINEBELOW",(0,0),(-1,-1),0.5,navy),
    ]))
    story.append(qt)
    story.append(Spacer(1, 3*mm))

    def pl(d, title):
        return "<br/>".join(
            [f"<b>{title}</b>"] +
            [f"<b>{k}:</b> {v}" for k, v in d.items() if v]
        )

    if any(dealer.values()):
        dt = Table([[
            Paragraph(pl(dealer, "DEALER / DISTRIBUTOR"),
                      ts("pt", fontName="Helvetica", fontSize=7.5, leading=11))
        ]], colWidths=["100%"])
        dt.setStyle(TableStyle([
            ("BOX",(0,0),(-1,-1),0.5,navy),
            ("VALIGN",(0,0),(-1,-1),"TOP"),
            ("TOPPADDING",(0,0),(-1,-1),5),
            ("BOTTOMPADDING",(0,0),(-1,-1),5),
            ("LEFTPADDING",(0,0),(-1,-1),8),
            ("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#EBF3FF")),
        ]))
        story.append(dt)
        story.append(Spacer(1, 3*mm))

    pt = Table([[
        Paragraph(pl(bill, "BILL TO PARTY"),
                  ts("b", fontName="Helvetica", fontSize=7.5, leading=11)),
        Paragraph(pl(ship, "SHIP TO PARTY"),
                  ts("sh", fontName="Helvetica", fontSize=7.5, leading=11)),
    ]], colWidths=["50%","50%"])
    pt.setStyle(TableStyle([
        ("BOX",(0,0),(-1,-1),0.5,navy),
        ("INNERGRID",(0,0),(-1,-1),0.5,colors.HexColor("#C5D0E0")),
        ("VALIGN",(0,0),(-1,-1),"TOP"),
        ("TOPPADDING",(0,0),(-1,-1),6),
        ("BOTTOMPADDING",(0,0),(-1,-1),6),
        ("LEFTPADDING",(0,0),(-1,-1),8),
        ("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#F4F6FA")),
    ]))
    story.append(pt)
    story.append(Spacer(1, 5*mm))

    trows = [["#","Item / Description","SKU Code","Size","MRP (₹)","Qty","Rate (₹)","Amount (₹)"]]
    gt = 0.0
    for i, r in enumerate([x for x in rows if x["qty"] > 0 and x["matched_sku"] != "—"], 1):
        trows.append([
            str(i),
            Paragraph(r["row_label"], ts("rl", fontName="Helvetica", fontSize=7)),
            Paragraph(f'<font name="Courier" size="6.5">{r["matched_sku"]}</font>',
                      ts("sk", fontName="Helvetica", fontSize=7)),
            r["col_label"],
            f"{r['mrp']:,.2f}",
            str(r["qty"]),
            f"{r['landing']:,.2f}",
            f"{r['amount']:,.2f}",
        ])
        gt += r["amount"]

    n = len(trows) + 1
    trows.append([
        "","","","","","",
        Paragraph("<b>GRAND TOTAL</b>",
                  ts("g", fontName="Helvetica-Bold", fontSize=8)),
        Paragraph(f"<b>₹ {gt:,.2f}</b>",
                  ts("gv", fontName="Helvetica-Bold", fontSize=8, alignment=TA_RIGHT)),
    ])

    lt = Table(trows,
               colWidths=[8*mm,55*mm,38*mm,16*mm,18*mm,12*mm,18*mm,22*mm],
               repeatRows=1)
    lt.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),navy),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,0),7.5),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),
        ("ALIGN",(1,1),(1,-1),"LEFT"),
        ("ALIGN",(2,1),(2,-1),"LEFT"),
        ("FONTSIZE",(0,1),(-1,-1),7),
        ("ROWBACKGROUNDS",(0,1),(-1,-2),[colors.white,colors.HexColor("#F4F8FF")]),
        ("GRID",(0,0),(-1,-1),0.35,colors.HexColor("#C5D0E0")),
        ("TOPPADDING",(0,0),(-1,-1),3),
        ("BOTTOMPADDING",(0,0),(-1,-1),3),
        ("BACKGROUND",(0,n-1),(-1,n-1),colors.HexColor("#EBF3FF")),
        ("LINEABOVE",(0,n-1),(-1,n-1),1,navy),
    ]))
    story.append(lt)
    story.append(Spacer(1, 6*mm))
    story.append(HRFlowable(width="100%", thickness=0.5, color=navy))
    story.append(Spacer(1, 2*mm))
    story.append(Paragraph(
        "Computer-generated quotation. Prices subject to change. Taxes as applicable.",
        ts("ft", fontName="Helvetica", fontSize=7, textColor=colors.grey, alignment=TA_CENTER),
    ))
    doc.build(story)
    return buf.getvalue()


# ─── Session defaults ─────────────────────────────────────────────────────────
for k, v in {
    "image_bytes": None, "ocr_resolved": [], "ocr_done": False,
    "review_confirmed": False, "party_done": False,
    "bill_to": {}, "ship_to": {}, "dealer_info": {},
    "pdf_bytes": None, "azure_endpoint": "", "azure_key": "",
    "_cam_recv_val": "", "manual_rows": [],
    "show_debug": False, "raw_ocr_debug": [],
}.items():
    if k not in st.session_state:
        st.session_state[k] = v

mrp_lookup = load_mrp()
zsd_df     = load_zsd()
sku_cat    = build_sku_catalogue()

# ─── UI helpers ───────────────────────────────────────────────────────────────
def sec_hdr(num, icon, label, state):
    lock = ('<span style="margin-left:auto;font-size:14px;">🔒</span>'
            if state == "locked" else "")
    st.markdown(
        f'<div class="sec-hdr {state}"><span class="num">{num}</span>'
        f'<h2>{icon} {label}</h2>{lock}</div>',
        unsafe_allow_html=True,
    )

def party_html(d, title):
    rows = "".join(
        f'<div class="p-row"><span class="p-lbl">{k}</span>'
        f'<span class="p-val">{v or "—"}</span></div>'
        for k, v in d.items()
    )
    return f'<div class="party-box"><h4>{title}</h4>{rows}</div>'

def tot_card(mrp, land, n):
    disc = mrp - land
    st.markdown(f"""<div class="card" style="max-width:440px;">
      <div class="card-title">📊 Summary</div>
      <div class="tot-row"><span>Active Line Items</span><span class="val">{n}</span></div>
      <div class="tot-row"><span>Gross MRP Value</span><span class="val">₹ {mrp:,.2f}</span></div>
      <div class="tot-row"><span>Distributor Discount</span><span class="val neg">− ₹ {disc:,.2f}</span></div>
      <div class="tot-row net"><span>Net Payable (Landing)</span><span class="val">₹ {land:,.2f}</span></div>
    </div>""", unsafe_allow_html=True)


# ─── App header ───────────────────────────────────────────────────────────────
st.markdown("""<div class="app-header"><div>
  <h1>🔧 Sintex BAPL – Quotation Generator</h1>
  <p>CPVC / UPVC Pipes &amp; Fittings · Chhattisgarh Price List</p>
</div></div>""", unsafe_allow_html=True)

ocr_done   = st.session_state.ocr_done
rev_done   = st.session_state.review_confirmed
party_done = st.session_state.party_done


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — Capture & OCR
# ══════════════════════════════════════════════════════════════════════════════
sec_hdr(1, "📸", "Capture & Run OCR", "done" if ocr_done else "active")

with st.expander("🔑 Azure OCR Credentials", expanded=not ocr_done):
    ep   = st.text_input("Azure Endpoint", value=st.session_state.azure_endpoint,
                         placeholder="https://YOUR_RESOURCE.cognitiveservices.azure.com",
                         key="az_ep")
    akey = st.text_input("Azure Key", value=st.session_state.azure_key,
                         type="password", placeholder="32-character key", key="az_key")
    st.session_state.azure_endpoint = ep
    st.session_state.azure_key = akey

img_mode = st.radio(
    "Source", "📷  Camera,📁  Upload".split(","),
    horizontal=True, label_visibility="collapsed", key="img_mode",
)

def _on_upload():
    uf = st.session_state.get("file_up")
    if uf:
        st.session_state.image_bytes = uf.getvalue()
        st.session_state.ocr_done = False
        st.session_state.ocr_resolved = []
        st.session_state.review_confirmed = False
        st.session_state.party_done = False
        st.session_state.pdf_bytes = None
        st.session_state.raw_ocr_debug = []

if img_mode == "📷  Camera":
    import streamlit.components.v1 as components
    CAM = """
<style>
*{box-sizing:border-box;margin:0;padding:0;}body{background:transparent;font-family:sans-serif;}
#cw{width:100%;background:linear-gradient(160deg,#071829,#0d2d56);border-radius:14px;
    overflow:hidden;display:flex;flex-direction:column;align-items:center;}
#video{width:100%;max-height:62vh;object-fit:cover;display:block;}
#canvas{display:none;}#preview{width:100%;display:none;border-bottom:3px solid #1E88E5;}
.tb{display:flex;gap:12px;padding:16px 20px 18px;width:100%;background:rgba(0,0,0,.3);justify-content:center;flex-wrap:wrap;}
.cb{display:inline-flex;align-items:center;justify-content:center;gap:8px;padding:12px 32px;
    border:none;border-radius:10px;font-size:14px;font-weight:700;cursor:pointer;min-width:160px;}
#bc{background:linear-gradient(135deg,#1565C0,#1E88E5);color:#fff;}
#br{background:rgba(255,255,255,.1);color:rgba(255,255,255,.9);border:1.5px solid rgba(255,255,255,.25);display:none;}
.sb{display:flex;align-items:center;gap:10px;background:rgba(0,0,0,.25);border-top:1px solid rgba(255,255,255,.07);
    padding:10px 20px;width:100%;font-size:12.5px;font-weight:500;color:rgba(255,255,255,.7);min-height:42px;}
.sb.ok{background:rgba(0,121,107,.3);color:#80CBC4;}.sb.snd{background:rgba(249,168,37,.12);color:#FFD54F;}
.pd{width:8px;height:8px;border-radius:50%;background:currentColor;flex-shrink:0;animation:p 1.5s infinite;}
@keyframes p{0%,100%{opacity:1;transform:scale(1)}50%{opacity:.3;transform:scale(.65)}}
.sp{width:16px;height:16px;flex-shrink:0;border:2.5px solid currentColor;border-top-color:transparent;
    border-radius:50%;animation:spin .65s linear infinite;}
@keyframes spin{to{transform:rotate(360deg)}}
</style>
<div id="cw">
  <video id="video" autoplay playsinline muted></video>
  <canvas id="canvas"></canvas><img id="preview" alt=""/>
  <div class="tb">
    <button class="cb" id="bc">📷&nbsp;&nbsp;Capture Photo</button>
    <button class="cb" id="br">🔄&nbsp;&nbsp;Retake</button>
  </div>
  <div class="sb" id="sb"><span class="pd" id="si"></span><span id="st">Starting camera…</span></div>
</div>
<script>
(function(){
  const v=document.getElementById('video'),cv=document.getElementById('canvas'),
        pv=document.getElementById('preview'),bc=document.getElementById('bc'),
        br=document.getElementById('br'),sb=document.getElementById('sb'),
        si=document.getElementById('si'),st=document.getElementById('st');
  function ss(m,mode){st.textContent=m;sb.className='sb '+(mode||'');
    si.className='';si.textContent='';
    if(mode==='snd')si.className='sp';else if(mode==='ok')si.textContent='✓';else si.className='pd';}
  async function startCam(){
    try{const s=await navigator.mediaDevices.getUserMedia(
      {video:{facingMode:{ideal:'environment'},width:{ideal:4096},height:{ideal:3072}},audio:false});
      v.srcObject=s;await v.play();ss('Camera ready — frame the order form','');}
    catch(e){ss('Camera error: '+e.message,'');}
  }
  function conv3(data,w,h,k){
    const out=new Uint8ClampedArray(data.length);
    for(let y=1;y<h-1;y++)for(let x=1;x<w-1;x++){
      for(let c=0;c<3;c++){let sum=0;
        for(let ky=-1;ky<=1;ky++)for(let kx=-1;kx<=1;kx++)
          sum+=data[((y+ky)*w+(x+kx))*4+c]*k[(ky+1)*3+(kx+1)];
        out[(y*w+x)*4+c]=Math.max(0,Math.min(255,sum));}out[(y*w+x)*4+3]=255;}return out;}
  function enhance(ctx,w,h){
    const id=ctx.getImageData(0,0,w,h),d=id.data;
    for(let i=0;i<d.length;i+=4)for(let c=0;c<3;c++){
      let v2=Math.round(((d[i+c]/255-.5)*1.25+.5)*255);d[i+c]=Math.max(0,Math.min(255,v2));}
    const k=[0,-1,0,-1,5,-1,0,-1,0],out=ctx.createImageData(w,h);
    out.data.set(conv3(d,w,h,k));ctx.putImageData(out,0,0);}
  bc.addEventListener('click',()=>{
    const w=v.videoWidth||1280,h=v.videoHeight||720;cv.width=w;cv.height=h;
    const ctx=cv.getContext('2d');ctx.drawImage(v,0,0,w,h);enhance(ctx,w,h);
    const b64=cv.toDataURL('image/jpeg',0.97);
    pv.src=b64;v.style.display='none';pv.style.display='block';
    bc.style.display='none';br.style.display='block';
    ss('Sending photo…','snd');
    window.parent.sessionStorage.setItem('sintex_cam_b64',b64);
    window.parent.postMessage({type:'SINTEX_CAM_CAPTURE',data:b64},'*');
    setTimeout(()=>ss('Photo ready — click Run OCR below','ok'),700);});
  br.addEventListener('click',()=>{
    pv.style.display='none';v.style.display='block';bc.style.display='block';br.style.display='none';
    window.parent.sessionStorage.removeItem('sintex_cam_b64');
    window.parent.postMessage({type:'SINTEX_CAM_RETAKE'},'*');ss('Camera ready','');});
  startCam();})();
</script>"""
    components.html(CAM, height=550, scrolling=False)
    components.html("""<script>
(function(){var att=0;function ti(){att++;if(att>120)return;
  var b64=window.parent.sessionStorage.getItem('sintex_cam_b64');
  if(!b64){setTimeout(ti,500);return;}
  var inputs=window.parent.document.querySelectorAll('input[type="text"]'),target=null;
  for(var i=inputs.length-1;i>=0;i--){var inp=inputs[i],st=window.parent.getComputedStyle(inp);
    if(parseFloat(st.height)<5||parseFloat(st.opacity)<0.1){target=inp;break;}}
  if(!target&&inputs.length)target=inputs[inputs.length-1];
  if(!target){setTimeout(ti,500);return;}
  var setter=Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype,'value').set;
  setter.call(target,b64);target.dispatchEvent(new Event('input',{bubbles:true}));
  target.dispatchEvent(new Event('change',{bubbles:true}));
  window.parent.sessionStorage.removeItem('sintex_cam_b64');}
setTimeout(ti,800);})();
</script>""", height=0)
    cam_v = st.text_input("cam_h", value=st.session_state.get("_cam_recv_val", ""),
                          label_visibility="collapsed", key="sintex_cam_recv")
    recv = st.session_state.get("sintex_cam_recv", "")
    if recv and recv.startswith("data:image"):
        import base64 as _b64
        _, enc = recv.split(",", 1)
        raw = _b64.b64decode(enc)
        if raw != st.session_state.image_bytes:
            st.session_state.image_bytes = raw
            st.session_state.ocr_done = False
            st.session_state.ocr_resolved = []
            st.session_state.review_confirmed = False
            st.session_state.party_done = False
            st.session_state.pdf_bytes = None
            st.session_state.raw_ocr_debug = []
            st.session_state["_cam_recv_val"] = recv
            st.rerun()
    components.html("""<script>
(function(){var inputs=window.parent.document.querySelectorAll('input[type="text"]');
if(inputs.length){var last=inputs[inputs.length-1];
  last.style.cssText+='height:1px!important;opacity:0!important;pointer-events:none!important;position:absolute!important;';
  var wrap=last.closest('[data-testid="stTextInput"]');
  if(wrap)wrap.style.cssText+='height:0!important;overflow:hidden!important;margin:0!important;padding:0!important;'}})();
</script>""", height=0)
else:
    st.file_uploader("Upload image", type=["jpg","jpeg","png"],
                     label_visibility="collapsed", key="file_up", on_change=_on_upload)

if st.session_state.image_bytes:
    with st.expander("📸 View image", expanded=False):
        st.image(st.session_state.image_bytes, use_container_width=True)

    if not ocr_done:
        if st.button("🔍  Run OCR", key="run_ocr"):
            _ep = st.session_state.azure_endpoint.strip()
            _k  = st.session_state.azure_key.strip()
            if not _ep or not _k:
                st.error("Enter Azure Endpoint and Key above.", icon="🔑")
            else:
                with st.spinner("Running OCR… 10–20 seconds"):
                    try:
                        raws = run_ocr(st.session_state.image_bytes, _ep, _k)
                        # Store raw results for debug panel
                        st.session_state.raw_ocr_debug = raws
                        res  = resolve_rows(raws, sku_cat, mrp_lookup)
                        st.session_state.ocr_resolved     = res
                        st.session_state.ocr_done         = True
                        st.session_state.review_confirmed = False
                        st.session_state.party_done       = False
                        st.session_state.pdf_bytes        = None
                        st.rerun()
                    except Exception as exc:
                        st.error(f"OCR failed: {exc}", icon="❌")
    else:
        n_m = sum(1 for r in st.session_state.ocr_resolved if r["match_type"] != "none")
        n_t = len(st.session_state.ocr_resolved)
        n_q = sum(1 for r in st.session_state.ocr_resolved if r["qty"] > 0)
        st.success(f"✅ OCR complete — {n_t} rows detected, {n_m} matched, {n_q} with quantity.", icon="🔍")

        # ── Debug panel ──────────────────────────────────────────────────
        with st.expander("🐛 Debug: Raw OCR output", expanded=False):
            st.caption("Shows exactly what the OCR geometry parser returned before SKU matching.")
            raw_dbg = st.session_state.get("raw_ocr_debug", [])
            if raw_dbg:
                dbg_lines = "\n".join(
                    f"row {i+1:>2}: raw_sku={r['raw_sku']!r:25s}  "
                    f"raw_size={r.get('raw_size',''):8s}  qty={r['qty']:>4}  "
                    f"line={r.get('raw_line','')[:60]!r}"
                    for i, r in enumerate(raw_dbg)
                )
                st.markdown(f'<div class="debug-box"><pre>{dbg_lines}</pre></div>',
                            unsafe_allow_html=True)
            else:
                st.info("No debug data — re-run OCR to populate.")

        if st.button("🔄 Re-run OCR", key="rerun_ocr"):
            st.session_state.ocr_done         = False
            st.session_state.ocr_resolved     = []
            st.session_state.review_confirmed = False
            st.session_state.party_done       = False
            st.session_state.raw_ocr_debug    = []
            st.rerun()
else:
    st.markdown("""<div style='background:#F4F6FA;border:2px dashed #DEE3EC;border-radius:10px;
      padding:36px;text-align:center;margin:14px 0;color:#5A6880;'>
      <div style='font-size:36px;margin-bottom:8px;'>🖼️</div>
      <b style='font-size:14px;'>No image yet</b>
      <div style='font-size:12px;margin-top:4px;'>Capture or upload an image above to begin</div>
    </div>""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — Review & Edit
# ══════════════════════════════════════════════════════════════════════════════
if not ocr_done:
    sec_hdr(2, "📋", "Review & Edit OCR Results", "locked")
    st.markdown('<p class="locked-msg">Complete Step 1 — run OCR — to unlock this section.</p>',
                unsafe_allow_html=True)
else:
    sec_hdr(2, "📋", "Review & Edit OCR Results", "done" if rev_done else "active")
    resolved = st.session_state.ocr_resolved
    nt = len(resolved)
    nm = sum(1 for r in resolved if r["match_type"] != "none")
    nz = sum(1 for r in resolved if r["qty"] > 0)

    st.markdown("#### 📊 Detected Items — Summary")
    st.caption("Read-only. Edit quantities or SKUs in the section below.")

    if nt == 0:
        st.warning("No recognisable SKU rows found. Add items manually below.", icon="⚠️")
    else:
        st.markdown(f"""<div style="display:flex;gap:12px;margin-bottom:10px;flex-wrap:wrap;font-size:12px;">
          <span style="background:#E8F5E9;color:#00695C;padding:3px 10px;border-radius:20px;font-weight:700;">✓ {nm} matched</span>
          <span style="background:#EBF3FF;color:#1565C0;padding:3px 10px;border-radius:20px;font-weight:700;">📦 {nz} with qty</span>
          <span style="background:#F4F6FA;color:#5A6880;padding:3px 10px;border-radius:20px;font-weight:700;">🔢 {nt} total</span>
        </div>""", unsafe_allow_html=True)

        gmrp = gland = 0.0; tbody = ""
        for i, r in enumerate(resolved):
            mt = r["match_type"]
            bc = "ok" if mt == "exact" else ("warn" if mt == "fuzzy" else "none")
            bi = "✓ Exact" if mt == "exact" else ("~ Fuzzy" if mt == "fuzzy" else "✗ None")
            sku = r["matched_sku"] if r["matched_sku"] != "—" else "—"
            zrow = ' class="muted"' if r["qty"] == 0 else ""
            gmrp  += r["mrp"] * r["qty"]
            gland += r["amount"]
            tbody += f"""<tr{zrow}><td>{i+1}</td>
              <td style="font-weight:500;">{r['row_label']}</td>
              <td><span class="sku-mono">{sku}</span></td>
              <td>{r.get('col_label') or r.get('raw_size','—')}</td>
              <td><span class="{bc}">{bi}</span></td>
              <td class="r">₹ {r['mrp']:,.2f}</td>
              <td class="r">{r['qty']}</td>
              <td class="r">₹ {r['landing']:,.2f}</td>
              <td class="r fp">₹ {r['amount']:,.2f}</td></tr>"""

        st.markdown(f"""<div style="overflow-x:auto;"><table class="ocr-tbl">
          <thead><tr><th>#</th><th>Product Name</th><th>SKU Code</th><th>Size</th>
          <th>Match</th><th class="r">MRP (₹)</th><th class="r">Qty</th>
          <th class="r">Rate (₹)</th><th class="r">Final Price (₹)</th></tr></thead>
          <tbody>{tbody}</tbody>
          <tfoot><tr><td colspan="5" style="text-align:right;font-weight:700;">TOTALS</td>
          <td class="r">₹ {gmrp:,.2f}</td>
          <td class="r">{sum(r['qty'] for r in resolved)}</td>
          <td class="r">—</td>
          <td class="r fp">₹ {gland:,.2f}</td></tr></tfoot>
        </table></div>""", unsafe_allow_html=True)

    st.markdown("#### ✏️ Edit Items")
    st.caption("Items with detected qty appear first, then zero/unmatched. Edit SKU, size, or qty as needed.")

    _ni  = _norm_idx(tuple(sku_cat.keys()))
    det  = [r for r in resolved if r["qty"] > 0 and r["matched_sku"] != "—"]
    und  = [r for r in resolved if not (r["qty"] > 0 and r["matched_sku"] != "—")]

    def col_hdrs():
        hc = st.columns([0.25,1.0,1.6,0.9,0.85,1.0,1.0,0.85])
        for c, l in zip(hc, ["#","Raw OCR","SKU Code ✏","Size ✏","MRP","Rate","Qty ✏","Amount"]):
            c.markdown(
                f"<div style='font-size:10.5px;font-weight:700;color:#0A2342;"
                f"background:#EBF3FF;border-radius:4px;padding:3px 8px;'>{l}</div>",
                unsafe_allow_html=True,
            )

    def erow(r, idx, pfx):
        cols = st.columns([0.25,1.0,1.6,0.9,0.85,1.0,1.0,0.85])
        cols[0].markdown(
            f"<div style='font-size:11px;color:#9BAECC;padding:8px 2px;text-align:center;'>{idx+1}</div>",
            unsafe_allow_html=True,
        )
        cols[1].markdown(
            f"<div style='font-size:10px;color:#5A6880;padding:4px 2px;word-break:break-all;'>"
            f"<code>{r['raw_sku']}</code></div>",
            unsafe_allow_html=True,
        )
        ns = cols[2].text_input("SKU",
                                value=r["matched_sku"] if r["matched_sku"] != "—" else "",
                                label_visibility="visible", key=f"{pfx}_s_{idx}")
        nz = cols[3].text_input("Size", value=r.get("raw_size", ""),
                                label_visibility="visible", key=f"{pfx}_z_{idx}",
                                placeholder="15MM")
        chg = (ns and ns != r.get("matched_sku","—")) or nz != r.get("raw_size","")
        if chg:
            pr = ns if ns else r["raw_sku"]
            ms2, cf2, mt2 = match_sku(pr, sku_cat, _ni)
            if ms2:
                sn   = re.sub(r"[^0-9]","", nz or r.get("raw_size",""))
                bn   = _norm(pr)
                hits = [(s, e) for s, e in sku_cat.items()
                        if _norm(s).startswith(bn) and
                        re.sub(r"[^0-9]","", e.get("col_label","")) == sn]
                if hits:
                    ms2, ce2 = hits[0]; mt2 = "exact"; cf2 = 1.0
                else:
                    ce2 = sku_cat[ms2]
                info2 = mrp_lookup.get(ms2, {})
                r = {**r,
                     "matched_sku": ms2,
                     "raw_size":    nz or r.get("raw_size",""),
                     "match_type":  mt2,
                     "confidence":  round(cf2*100,1),
                     "sheet":       ce2.get("sheet","—"),
                     "row_label":   ce2.get("row_label","—"),
                     "col_label":   ce2.get("col_label","—"),
                     "mrp":         info2.get("MRP_clean") or 0.0,
                     "landing":     info2.get("Distributor Landing") or 0.0,
                     "ri":          ce2.get("ri"), "ci": ce2.get("ci")}
            else:
                r = {**r, "matched_sku": ns or "—", "raw_size": nz,
                     "match_type": "none", "confidence": 0.0,
                     "mrp": 0.0, "landing": 0.0, "ri": None, "ci": None}

        cols[4].markdown(
            f"<div style='font-size:11px;font-family:monospace;padding:8px 2px;text-align:right;'>"
            f"₹ {r['mrp']:,.2f}</div>",
            unsafe_allow_html=True,
        )
        cols[5].markdown(
            f"<div style='font-size:11px;font-family:monospace;padding:8px 2px;text-align:right;'>"
            f"₹ {r['landing']:,.2f}</div>",
            unsafe_allow_html=True,
        )
        nq  = cols[6].number_input("Qty", value=int(r["qty"]), min_value=0, step=1,
                                   label_visibility="visible", key=f"{pfx}_q_{idx}")
        amt = round(r["landing"] * nq, 2)
        cols[7].markdown(
            f"<div style='font-size:12px;font-family:monospace;font-weight:700;color:#0A2342;"
            f"padding:8px 2px;text-align:right;'>₹ {amt:,.2f}</div>",
            unsafe_allow_html=True,
        )
        return {**r, "qty": nq, "amount": amt}

    col_hdrs()
    updated = []

    if det:
        st.markdown('<span class="grp-lbl grp-det">✅ Detected with Quantity</span>',
                    unsafe_allow_html=True)
        for i, r in enumerate(det):
            updated.append(erow(r, i, "dt"))

    if und:
        st.markdown('<span class="grp-lbl grp-miss">⚠️ Zero Qty / Unmatched — Review These</span>',
                    unsafe_allow_html=True)
        for i, r in enumerate(und):
            updated.append(erow(r, len(det)+i, "ud"))

    st.markdown('<span class="grp-lbl grp-man">➕ Add Items Manually</span>',
                unsafe_allow_html=True)
    nm_rows = []
    for mi, mr in enumerate(st.session_state.manual_rows):
        mc1, mc2, mc3 = st.columns([2,1,1])
        sv  = mc1.text_input("SKU Code", value=mr.get("sku",""), key=f"ms_{mi}",
                              label_visibility="visible")
        szv = mc2.text_input("Size", value=mr.get("size",""), key=f"mz_{mi}",
                              label_visibility="visible", placeholder="15MM")
        qv  = mc3.number_input("Qty", value=int(mr.get("qty",0)), min_value=0, step=1,
                                key=f"mq_{mi}", label_visibility="visible")
        nm_rows.append({"sku": sv, "size": szv, "qty": qv})
    st.session_state.manual_rows = nm_rows

    ac, apc, _ = st.columns([1,1,4])
    with ac:
        if st.button("➕ Add Row", key="add_row"):
            st.session_state.manual_rows = nm_rows + [{"sku":"","size":"","qty":0}]
            st.rerun()
    with apc:
        if st.button("✅ Apply Manual", key="apm"):
            for mr in st.session_state.manual_rows:
                s = mr["sku"].strip()
                if not s:
                    continue
                ms2, cf2, mt2 = match_sku(s, sku_cat, _ni)
                ce2   = sku_cat.get(ms2, {}) if ms2 else {}
                info2 = mrp_lookup.get(ms2, {}) if ms2 else {}
                mv2   = info2.get("MRP_clean") or 0.0
                lv2   = info2.get("Distributor Landing") or mv2
                updated.append({
                    "raw_sku":    s,
                    "raw_size":   mr.get("size",""),
                    "matched_sku": ms2 or "—",
                    "match_type":  mt2,
                    "confidence":  round(cf2*100,1),
                    "qty":         mr["qty"],
                    "sheet":       ce2.get("sheet","—"),
                    "row_label":   ce2.get("row_label","—"),
                    "col_label":   ce2.get("col_label","—"),
                    "mrp":         mv2,
                    "landing":     lv2,
                    "amount":      round(lv2 * mr["qty"], 2),
                    "ri":          ce2.get("ri"),
                    "ci":          ce2.get("ci"),
                })
            st.session_state.ocr_resolved = updated
            st.session_state.manual_rows  = []
            st.success("Applied.", icon="✅")
            st.rerun()

    if updated:
        st.session_state.ocr_resolved = updated

    cur  = st.session_state.ocr_resolved
    cm   = sum(r["mrp"]     * r["qty"] for r in cur if r["matched_sku"] != "—")
    cl   = sum(r["landing"] * r["qty"] for r in cur if r["matched_sku"] != "—")
    cn   = sum(1 for r in cur if r["qty"] > 0 and r["matched_sku"] != "—")
    st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)
    tot_card(cm, cl, cn)

    st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)
    if not rev_done:
        if st.button("✔️  Confirm Review & Proceed to Party Details", key="confirm_rev"):
            st.session_state.review_confirmed = True
            st.rerun()
    else:
        st.success("✅ Review confirmed.", icon="✔️")
        if st.button("↩ Edit Again", key="edit_again"):
            st.session_state.review_confirmed = False
            st.session_state.party_done       = False
            st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# STEP 3 — Party Details
# ══════════════════════════════════════════════════════════════════════════════
if not rev_done:
    sec_hdr(3, "🏪", "Party & Customer Details", "locked")
    st.markdown('<p class="locked-msg">Confirm the review above to unlock this section.</p>',
                unsafe_allow_html=True)
else:
    sec_hdr(3, "🏪", "Party & Customer Details", "done" if party_done else "active")

    st.markdown("#### 🏪 Dealer / Distributor")
    da1, da2 = st.columns(2)
    dn = da1.text_input("Dealer / Distributor Name", key="dealer_name")
    dc = da2.text_input("Dealer Code", key="dealer_code")
    da = st.text_input("Address", key="dealer_address")
    da3, da4 = st.columns(2)
    dp  = da3.text_input("Phone", key="dealer_phone")
    dg  = da4.text_input("GST No.", key="dealer_gst")
    da5, da6 = st.columns(2)
    ds  = da5.text_input("State", key="dealer_state")
    dpn = da6.text_input("PAN No.", key="dealer_pan")
    st.session_state.dealer_info = {
        "Name": dn, "Code": dc, "Address": da,
        "Phone": dp, "GST No.": dg, "State": ds, "PAN No.": dpn,
    }

    st.markdown("---")
    st.markdown("#### 👤 Customer Lookup")
    copts = ["— Select —"] + [
        f'{r["Customer Code"]} | {r["Customer Name"]}'
        for _, r in zsd_df.iterrows()
    ]
    sel = st.selectbox("🔍 Search Customer", copts, key="zsd_sel")

    def zfill(pfx, row):
        addr = " ".join(filter(None, [
            str(row.get(k,"") or "")
            for k in ["Address 1","Address 2","Address 3","City"]
        ]))
        return {
            f"{pfx}_party_no":   str(row.get("Customer Code","") or ""),
            f"{pfx}_party_name": str(row.get("Customer Name","")  or ""),
            f"{pfx}_address":    addr.strip(),
            f"{pfx}_phone":      str(row.get("Telephone","")      or ""),
            f"{pfx}_mobile":     str(row.get("Mobile No.","")     or ""),
            f"{pfx}_state_code": str(row.get("State Code","")     or ""),
            f"{pfx}_state":      str(row.get("State Code Desc.","") or ""),
            f"{pfx}_gst":        str(row.get("GST Number","")     or ""),
            f"{pfx}_pan":        str(row.get("PAN No.","")        or ""),
        }

    if sel != copts[0]:
        code = sel.split("|")[0].strip()
        mrow = zsd_df[zsd_df["Customer Code"].astype(str) == code]
        if not mrow.empty:
            for k, v in zfill("bill", mrow.iloc[0]).items():
                if k not in st.session_state or not st.session_state[k]:
                    st.session_state[k] = v
            st.toast("Bill-to filled from master.", icon="✅")

    st.markdown("#### 📋 Bill To Party")
    b1, b2 = st.columns(2)
    bpno = b1.text_input("Party No.",   key="bill_party_no")
    bpnm = b2.text_input("Party Name",  key="bill_party_name")
    badr = st.text_input("Address",     key="bill_address")
    b3, b4 = st.columns(2)
    bph = b3.text_input("Phone",  key="bill_phone")
    bmo = b4.text_input("Mobile", key="bill_mobile")
    b5, b6 = st.columns(2)
    bsc = b5.text_input("State Code", key="bill_state_code")
    bst = b6.text_input("State",      key="bill_state")
    b7, b8 = st.columns(2)
    bgs = b7.text_input("GST No.",  key="bill_gst")
    bpa = b8.text_input("PAN No.",  key="bill_pan")
    st.session_state.bill_to = {
        "Party No.": bpno, "Name": bpnm, "Address": badr,
        "Phone": bph, "Mobile": bmo,
        "State Code": bsc, "State": bst,
        "GST No.": bgs, "PAN No.": bpa,
    }

    same = st.checkbox("Ship-to same as Bill-to", key="same_ship")
    if same:
        for k, v in [
            ("ship_party_no", bpno), ("ship_party_name", bpnm),
            ("ship_address",  badr), ("ship_phone",  bph),
            ("ship_mobile",   bmo),  ("ship_state_code", bsc),
            ("ship_state",    bst),  ("ship_gst",  bgs),
            ("ship_pan",      bpa),
        ]:
            st.session_state[k] = v

    st.markdown("#### 🚚 Ship To Party")
    s1, s2 = st.columns(2)
    spno = s1.text_input("Party No. ",  key="ship_party_no")
    spnm = s2.text_input("Party Name ", key="ship_party_name")
    sadr = st.text_input("Address ",    key="ship_address")
    s3, s4 = st.columns(2)
    sph = s3.text_input("Phone  ",  key="ship_phone")
    smo = s4.text_input("Mobile  ", key="ship_mobile")
    s5, s6 = st.columns(2)
    ssc = s5.text_input("State Code  ", key="ship_state_code")
    sst = s6.text_input("State  ",      key="ship_state")
    s7, s8 = st.columns(2)
    sgs = s7.text_input("GST No.  ", key="ship_gst")
    spa = s8.text_input("PAN No.  ", key="ship_pan")
    st.session_state.ship_to = {
        "Party No.": spno, "Name": spnm, "Address": sadr,
        "Phone": sph, "Mobile": smo,
        "State Code": ssc, "State": sst,
        "GST No.": sgs, "PAN No.": spa,
    }

    st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)
    if not party_done:
        if st.button("✔️  Save Party Details & Proceed to Download", key="save_party"):
            st.session_state.party_done = True
            st.rerun()
    else:
        st.success("✅ Party details saved.", icon="✔️")
        if st.button("↩ Edit Party Details", key="edit_party"):
            st.session_state.party_done = False
            st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# STEP 4 — Download
# ══════════════════════════════════════════════════════════════════════════════
if not party_done:
    sec_hdr(4, "⬇️", "Download Quotation", "locked")
    st.markdown('<p class="locked-msg">Save party details above to unlock downloads.</p>',
                unsafe_allow_html=True)
else:
    sec_hdr(4, "⬇️", "Download Quotation", "active")
    rows_dl = [r for r in st.session_state.ocr_resolved
               if r["qty"] > 0 and r["matched_sku"] != "—"]

    if not rows_dl:
        st.warning("No valid line items. Go back to the review step to add items.", icon="⚠️")
    else:
        pc1, pc2 = st.columns(2)
        with pc1:
            st.markdown(party_html(st.session_state.bill_to, "BILL TO PARTY"),
                        unsafe_allow_html=True)
        with pc2:
            st.markdown(party_html(st.session_state.ship_to, "SHIP TO PARTY"),
                        unsafe_allow_html=True)
        if any(st.session_state.dealer_info.values()):
            st.markdown(party_html(st.session_state.dealer_info, "DEALER / DISTRIBUTOR"),
                        unsafe_allow_html=True)

        st.markdown("#### 📦 Final Order Lines")
        ld = []; gm = gl = 0.0
        for r in rows_dl:
            gm += r["mrp"] * r["qty"]
            gl += r["landing"] * r["qty"]
            ld.append({
                "Item":       r["row_label"],
                "SKU Code":   r["matched_sku"],
                "Size":       r["col_label"],
                "MRP (₹)":   round(r["mrp"], 2),
                "Qty":        r["qty"],
                "Rate (₹)":  round(r["landing"], 2),
                "Amount (₹)": round(r["amount"], 2),
            })
        dfl = pd.DataFrame(ld)
        st.dataframe(dfl, use_container_width=True, hide_index=True,
                     column_config={
                         "MRP (₹)":   st.column_config.NumberColumn(format="₹ %.2f"),
                         "Rate (₹)":  st.column_config.NumberColumn(format="₹ %.2f"),
                         "Amount (₹)": st.column_config.NumberColumn(format="₹ %.2f"),
                     })
        tot_card(gm, gl, len(rows_dl))

        st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)
        d1, d2, d3, _ = st.columns([1,1,1,2])

        cb = io.StringIO()
        dfl.to_csv(cb, index=False)
        d1.download_button("⬇ CSV", data=cb.getvalue().encode(),
                           file_name=f"sintex_{date.today()}.csv",
                           mime="text/csv", key="dl_csv")

        xb = io.BytesIO()
        with pd.ExcelWriter(xb, engine="openpyxl") as w:
            pd.DataFrame(
                [{"Field": f"Dealer - {k}", "Value": v}
                 for k, v in st.session_state.dealer_info.items()] +
                [{"Field": f"Bill To - {k}", "Value": v}
                 for k, v in st.session_state.bill_to.items()] +
                [{"Field": f"Ship To - {k}", "Value": v}
                 for k, v in st.session_state.ship_to.items()]
            ).to_excel(w, sheet_name="Party Details", index=False)
            dfl.to_excel(w, sheet_name="Quotation Lines", index=False)
            pd.DataFrame([
                {"Description": "Gross MRP",    "Amount (₹)": round(gm, 2)},
                {"Description": "Discount",     "Amount (₹)": round(gm-gl, 2)},
                {"Description": "Net Payable",  "Amount (₹)": round(gl, 2)},
            ]).to_excel(w, sheet_name="Summary", index=False)
        d2.download_button("⬇ Excel", data=xb.getvalue(),
                           file_name=f"sintex_{date.today()}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dl_excel")

        if _HAS_REPORTLAB:
            if st.button("🖨️ Generate PDF", key="gen_pdf"):
                with st.spinner("Generating PDF…"):
                    try:
                        st.session_state.pdf_bytes = gen_pdf(
                            rows_dl, mrp_lookup,
                            st.session_state.bill_to,
                            st.session_state.ship_to,
                            st.session_state.dealer_info,
                        )
                    except Exception as e:
                        st.error(f"PDF failed: {e}")
            if st.session_state.pdf_bytes:
                d3.download_button("⬇ PDF", data=st.session_state.pdf_bytes,
                                   file_name=f"sintex_{date.today()}.pdf",
                                   mime="application/pdf", key="dl_pdf")
        else:
            st.caption("PDF requires `pip install reportlab`")

        st.markdown("<div style='height:18px'></div>", unsafe_allow_html=True)
        if st.button("🔄  Start New Quotation", key="new_quot"):
            keep = {"azure_endpoint", "azure_key", "_cam_recv_val"}
            for k in [k for k in list(st.session_state.keys()) if k not in keep]:
                del st.session_state[k]
            st.rerun()